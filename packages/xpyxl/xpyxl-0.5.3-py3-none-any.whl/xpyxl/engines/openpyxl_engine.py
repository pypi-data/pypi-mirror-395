"""OpenPyXL rendering engine implementation."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import TYPE_CHECKING, Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..styles import to_argb
from .base import EffectiveStyle, Engine, SaveTarget

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

__all__ = ["OpenpyxlEngine"]


class OpenpyxlEngine(Engine):
    """Rendering engine using openpyxl."""

    def __init__(self) -> None:
        super().__init__()
        self._workbook = Workbook()
        # Remove default sheet created by openpyxl
        default_sheet = self._workbook.active
        if default_sheet is not None:
            self._workbook.remove(default_sheet)
        self._current_sheet: Worksheet | None = None
        # Cache style objects to avoid duplicates
        self._style_cache: dict[tuple[Any, ...], dict[str, Any]] = {}
        # Cache color conversions
        self._color_cache: dict[str, str] = {}
        # Cache column letters
        self._column_letter_cache: dict[int, str] = {}

    def create_sheet(self, name: str) -> None:
        self._current_sheet = self._workbook.create_sheet(title=name)

    def write_cell(
        self,
        row: int,
        col: int,
        value: object,
        style: EffectiveStyle,
        border_fallback_color: str,
    ) -> None:
        if self._current_sheet is None:
            raise RuntimeError("No sheet created. Call create_sheet() first.")

        cell = self._current_sheet.cell(row=row, column=col)
        cell.value = value  # type: ignore[assignment]
        self._apply_style(cell, style, border_fallback_color)

    def _get_cached_color(self, color: str) -> str:
        """Get cached ARGB color or convert and cache it."""
        if color not in self._color_cache:
            self._color_cache[color] = to_argb(color)
        return self._color_cache[color]

    def _get_cached_styles(
        self, effective: EffectiveStyle, border_fallback_color: str
    ) -> dict[str, Any]:
        """Get or create cached style objects for the given style."""
        # Create a hashable key from style properties
        cache_key = (
            effective.font_name,
            effective.font_size,
            effective.bold,
            effective.italic,
            effective.text_color,
            effective.fill_color,
            effective.horizontal_align,
            effective.vertical_align,
            effective.indent,
            effective.wrap_text,
            effective.shrink_to_fit,
            effective.number_format,
            effective.border,
            effective.border_color or border_fallback_color
            if effective.border
            else None,
            effective.border_top,
            effective.border_bottom,
            effective.border_left,
            effective.border_right,
        )

        if cache_key in self._style_cache:
            return self._style_cache[cache_key]

        # Create style objects
        text_color_argb = self._get_cached_color(effective.text_color)
        font = Font(
            name=effective.font_name,
            size=effective.font_size,
            bold=effective.bold,
            italic=effective.italic,
            color=text_color_argb,
        )

        fill: PatternFill | None = None
        if effective.fill_color:
            fill_color_argb = self._get_cached_color(effective.fill_color)
            fill = PatternFill(
                fill_type="solid",
                start_color=fill_color_argb,
                end_color=fill_color_argb,
            )

        alignment: Alignment | None = None
        align_kwargs: dict[str, Any] = {}
        if effective.horizontal_align:
            align_kwargs["horizontal"] = effective.horizontal_align
        if effective.vertical_align:
            align_kwargs["vertical"] = effective.vertical_align
        if effective.indent is not None:
            align_kwargs["indent"] = effective.indent
        if effective.wrap_text:
            align_kwargs["wrap_text"] = True
        if effective.shrink_to_fit:
            align_kwargs["shrink_to_fit"] = True
        if align_kwargs:
            align_kwargs.setdefault("vertical", "bottom")
            alignment = Alignment(**align_kwargs)  # type: ignore[arg-type]
        elif effective.wrap_text or effective.shrink_to_fit:
            alignment = Alignment(
                wrap_text=True if effective.wrap_text else None,
                shrink_to_fit=True if effective.shrink_to_fit else None,
            )

        border: Border | None = None
        if effective.border:
            border_color = effective.border_color or border_fallback_color
            border_color_argb = self._get_cached_color(border_color)

            def build_side(enabled: bool) -> Side | None:
                if not enabled:
                    return None
                return Side(style=effective.border, color=border_color_argb)

            explicit = (
                effective.border_top
                or effective.border_bottom
                or effective.border_left
                or effective.border_right
            )
            if explicit:
                border = Border(
                    left=build_side(effective.border_left),
                    right=build_side(effective.border_right),
                    top=build_side(effective.border_top),
                    bottom=build_side(effective.border_bottom),
                )
            else:
                side = build_side(True)
                border = Border(left=side, right=side, top=side, bottom=side)

        styles = {
            "font": font,
            "fill": fill,
            "alignment": alignment,
            "border": border,
            "number_format": effective.number_format,
        }

        self._style_cache[cache_key] = styles
        return styles

    def _apply_style(
        self, cell: object, effective: EffectiveStyle, border_fallback_color: str
    ) -> None:
        """Apply style to an openpyxl cell."""
        styles = self._get_cached_styles(effective, border_fallback_color)

        cell.font = styles["font"]  # type: ignore[attr-defined]

        if styles["fill"]:
            cell.fill = styles["fill"]  # type: ignore[attr-defined]

        if styles["alignment"]:
            cell.alignment = styles["alignment"]  # type: ignore[attr-defined]
        elif cell.alignment is None and (  # type: ignore[attr-defined]
            effective.wrap_text or effective.shrink_to_fit
        ):
            # Handle edge case where alignment wasn't cached but wrap/shrink is needed
            cell.alignment = Alignment(  # type: ignore[attr-defined]
                wrap_text=True if effective.wrap_text else None,
                shrink_to_fit=True if effective.shrink_to_fit else None,
            )

        if styles["number_format"]:
            cell.number_format = styles["number_format"]  # type: ignore[attr-defined]

        if styles["border"]:
            cell.border = styles["border"]  # type: ignore[attr-defined]

    def set_column_width(self, col: int, width: float) -> None:
        if self._current_sheet is None:
            raise RuntimeError("No sheet created. Call create_sheet() first.")

        if col not in self._column_letter_cache:
            self._column_letter_cache[col] = get_column_letter(col)
        letter = self._column_letter_cache[col]
        self._current_sheet.column_dimensions[letter].width = max(width, 8.0)

    def set_row_height(self, row: int, height: float) -> None:
        if self._current_sheet is None:
            raise RuntimeError("No sheet created. Call create_sheet() first.")

        self._current_sheet.row_dimensions[row].height = height

    def fill_background(
        self,
        color: str,
        max_row: int,
        max_col: int,
    ) -> None:
        if self._current_sheet is None:
            raise RuntimeError("No sheet created. Call create_sheet() first.")

        fill_color = self._get_cached_color(color)
        sheet_fill = PatternFill(
            fill_type="solid", start_color=fill_color, end_color=fill_color
        )
        # Reuse the same PatternFill object for all cells (openpyxl supports this)
        for row in self._current_sheet.iter_rows(
            min_row=1, max_row=max_row, min_col=1, max_col=max_col
        ):
            for cell in row:
                cell.fill = sheet_fill

    def save(self, target: SaveTarget | None = None) -> bytes | None:
        if target is None:
            buffer = BytesIO()
            self._workbook.save(buffer)
            return buffer.getvalue()

        if isinstance(target, (str, Path)):
            self._workbook.save(str(target))
        else:
            self._workbook.save(target)
        return None

from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.utils.exceptions import InvalidFileException

from w.services.abstract_service import AbstractService
from w.services.technical.filesystem_service import FilesystemService
from w.exceptions import ValidationError
from w.services.technical.list_service import ListService
import pandas as pd


class ExcelService(AbstractService):
    @classmethod
    def is_excel(cls, filename):
        """Check if filename is excel (from extension)"""
        return FilesystemService.has_suffix(filename, [".xls", ".xlsx"])

    @classmethod
    def check_is_excel(cls, filename) -> Path:
        """
        Check if filename is excel (from extension) or raise RuntimeError
        """
        if cls.is_excel(filename):
            return filename
        raise RuntimeError(f"{Path(filename).name} is not a valid excel file")

    @classmethod
    def open_workbook(cls, filename, **options):
        """
        Open filename is valid excel (check exists too)

        Args:
            filename(str): excel full path filename
            **options:
             - read_only(bool): default False, optimised for reading, content cannot be
               edited
             - keep_vba(bool): default False, preserve vba content
             - data_only(bool): default False, controls whether cells with formulae have
               either the formula (default) or the value stored the last time Excel read
               the sheet
            - keep_links: default True, whether links to external workbooks should be
              preserved.

        Returns:
            Workbook

        Raises:
            RuntimeError: if filename does not exists or is not valid excel
        """
        FilesystemService.check_file_exists(filename)
        cls.check_is_excel(filename)
        try:
            return openpyxl.load_workbook(filename, **options)
        except InvalidFileException:
            raise RuntimeError(f"{Path(filename).name} is not a valid excel file")

    @classmethod
    def load(cls, excel_filename, mapping_columns, sheet_name=None) -> list:
        """
        Load active or specific sheet of excel file

        Args:
            excel_filename(str|Path): excel file
            mapping_columns(dict): csv header columns mapping to wanted attributes
            sheet_name(str) : sheet name (default load active sheet)

        Returns:
            list: [{"<mapping name>": < row col value>, ...}, ...]

        Raises:
            RuntimeError :
                - filename is not csv
                - filename does not exists
                - incorrect or missing header
        """
        wb = cls.open_workbook(excel_filename, read_only=True, data_only=True)

        if (
            sheet_name is not None and sheet_name not in wb.sheetnames
        ):  # pragma: no cover (todo one day)
            raise ValidationError(f"{sheet_name} missing in sheet names")

        sheet = wb[sheet_name] if sheet_name is not None else wb.active

        rows = sheet.iter_rows()
        headers = [
            c.value.strip() if c.value else f"None_{i}"
            for i, c in enumerate(next(rows))
        ]
        final_headers, final_mapping = cls._get_final_headers_and_mapping(
            headers, mapping_columns
        )
        result = []
        for row in rows:
            row_data = {
                final_mapping[final_headers[i]]: cell.value
                for i, cell in enumerate(row)
                if final_headers[i] in final_mapping
            }
            if any(row_data.values()):
                result.append(row_data)
        return result

    @classmethod
    def dump(cls, stream, headers: dict, rows: list[dict]):
        wb = Workbook()
        ws = wb.active
        ws.append(list(headers.values()))

        excel_rows = []
        for row in rows:
            excel_rows.append([row[key] for key in headers.keys()])
        df = pd.DataFrame(excel_rows, columns=list(headers.values()))
        return df.to_excel(stream, index=False, engine="openpyxl")

    @classmethod
    def _get_final_headers_and_mapping(
        cls, headers: list, mapping_columns: dict
    ) -> tuple:
        """Compare excel headers to mapping columns
        and return final headers and mapping"""
        if len(headers) > len(mapping_columns):
            mapping_columns = cls._fill_mapping_columns_with_headers(
                headers, mapping_columns
            )
        else:
            for i, required_header in enumerate(mapping_columns.keys()):
                if required_header != headers[i]:
                    headers = ListService.convert_elements_to_string(headers)
                    raise ValidationError(
                        f"incorrect or missing header, "
                        f"expected '{';'.join(list(mapping_columns.keys()))}' "
                        f"got '{';'.join(headers)}'"
                    )
        return headers, mapping_columns

    @staticmethod
    def _fill_mapping_columns_with_headers(
        headers: list, mapping_columns: dict
    ) -> dict:
        """Treat the case where some headers are empty"""
        new_mapping_column = {}
        for i, header_name in enumerate(headers):
            if header_name in mapping_columns.keys():
                new_mapping_column[header_name] = mapping_columns[header_name]
            else:
                new_mapping_column[header_name] = f"None_{i}"

        return new_mapping_column

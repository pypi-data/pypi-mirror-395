from __future__ import annotations
from abc import abstractmethod
from collections.abc import Callable
from io import BytesIO as BytesIO_1
import openpyxl as openpyxl_1
from typing import (Any, Protocol)
from ..fable_library.reflection import (TypeInfo, union_type, class_type)
from ..fable_library.types import (Array, Union)
from ..fable_library.util import (IEnumerable_1, IEnumerable)

def Helper_writeBytes(bytes: bytearray, path: str) -> None:
    
    # Write the bytes data to the output file path using shutil
    with open(path, 'wb') as output_file:
        output_file.write(bytes)
  


def _expr38() -> TypeInfo:
    return union_type("Fable.Openpyxl.CellType", [], CellType, lambda: [[], [], [], [], [], []])


class CellType(Union):
    def __init__(self, tag: int, *fields: Any) -> None:
        super().__init__()
        self.tag: int = tag or 0
        self.fields: Array[Any] = list(fields)

    @staticmethod
    def cases() -> list[str]:
        return ["Float", "Integer", "String", "Boolean", "DateTime", "Empty"]


CellType_reflection = _expr38

def CellType_fromCellType_Z721C83C5(cell_type: str) -> CellType:
    if cell_type == "int":
        return CellType(1)

    elif cell_type == "float":
        return CellType(0)

    elif cell_type == "str":
        return CellType(2)

    elif cell_type == "bool":
        return CellType(3)

    elif cell_type == "datetime":
        return CellType(4)

    elif cell_type == "NoneType":
        return CellType(5)

    else: 
        raise Exception(("Unknown cell type of type: \'" + cell_type) + "\'")



class Cell(Protocol):
    @property
    @abstractmethod
    def cell_type(self) -> str:
        ...

    @property
    @abstractmethod
    def value(self) -> Any:
        ...

    @value.setter
    @abstractmethod
    def value(self, __arg0: Any) -> None:
        ...


class Table(Protocol):
    @property
    @abstractmethod
    def display_name(self) -> str:
        ...

    @display_name.setter
    @abstractmethod
    def display_name(self, __arg0: str) -> None:
        ...

    @property
    @abstractmethod
    def header_row_count(self) -> int:
        ...

    @header_row_count.setter
    @abstractmethod
    def header_row_count(self, __arg0: int) -> None:
        ...

    @property
    @abstractmethod
    def id(self) -> int:
        ...

    @id.setter
    @abstractmethod
    def id(self, __arg0: int) -> None:
        ...

    @property
    @abstractmethod
    def name(self) -> str:
        ...

    @name.setter
    @abstractmethod
    def name(self, __arg0: str) -> None:
        ...

    @property
    @abstractmethod
    def ref(self) -> str:
        ...

    @ref.setter
    @abstractmethod
    def ref(self, __arg0: str) -> None:
        ...


class TableMap(Protocol):
    @abstractmethod
    def Item(self, __arg0: str) -> Table:
        ...

    @abstractmethod
    def delete(self, displayName: str) -> None:
        ...

    @abstractmethod
    def items(self) -> Array[tuple[str, str]]:
        ...

    @abstractmethod
    def values(self) -> Array[Table]:
        ...


class Worksheet(Protocol):
    @abstractmethod
    def add_table(self, __arg0: Table) -> None:
        ...

    @abstractmethod
    def append(self, __arg0: Array[Any]) -> None:
        ...

    @abstractmethod
    def delete_cols(self, start_index: int, count: int) -> None:
        ...

    @abstractmethod
    def delete_rows(self, start_index: int, count: int) -> None:
        ...

    @abstractmethod
    def delete_table(self, displayName: str) -> None:
        ...

    @property
    @abstractmethod
    def columns(self) -> Array[Array[Cell]]:
        ...

    @property
    @abstractmethod
    def rows(self) -> Array[Array[Cell]]:
        ...

    @property
    @abstractmethod
    def table_count(self) -> int:
        ...

    @property
    @abstractmethod
    def tables(self) -> TableMap:
        ...

    @property
    @abstractmethod
    def title(self) -> str:
        ...

    @title.setter
    @abstractmethod
    def title(self, __arg0: str) -> None:
        ...

    @property
    @abstractmethod
    def values(self) -> Array[Array[Any]]:
        ...

    @abstractmethod
    def insert_cols(self, __arg0: int) -> None:
        ...

    @abstractmethod
    def insert_rows(self, __arg0: int) -> None:
        ...

    @abstractmethod
    def iter_cols(self, min_row: int, max_col: int, max_row: int, action: Callable[[Array[Cell]], None]) -> None:
        ...

    @abstractmethod
    def iter_rows(self, min_row: int, max_col: int, max_row: int, action: Callable[[Array[Cell]], None]) -> None:
        ...


class Workbook(IEnumerable_1, IEnumerable[Any]):
    @abstractmethod
    def Item(self, __arg0: str) -> Worksheet:
        ...

    @abstractmethod
    def copy_worksheet(self, __arg0: Worksheet) -> Worksheet:
        ...

    @abstractmethod
    def create_sheet(self, __arg0: str, position: int | None) -> Worksheet:
        ...

    @property
    @abstractmethod
    def active(self) -> Worksheet:
        ...

    @property
    @abstractmethod
    def iso_dates(self) -> bool:
        ...

    @iso_dates.setter
    @abstractmethod
    def iso_dates(self, __arg0: bool) -> None:
        ...

    @property
    @abstractmethod
    def sheetnames(self) -> Array[str]:
        ...

    @property
    @abstractmethod
    def template(self) -> bool:
        ...

    @template.setter
    @abstractmethod
    def template(self, __arg0: bool) -> None:
        ...

    @property
    @abstractmethod
    def worksheets(self) -> Array[Worksheet]:
        ...

    @worksheets.setter
    @abstractmethod
    def worksheets(self, __arg0: Array[Worksheet]) -> None:
        ...

    @abstractmethod
    def index(self, __arg0: Worksheet) -> int:
        ...

    @abstractmethod
    def remove(self, __arg0: Worksheet) -> None:
        ...


class BytesIO(Protocol):
    @abstractmethod
    def ToFile(self, path: str) -> None:
        ...

    @property
    @abstractmethod
    def x(self) -> str:
        ...

    @abstractmethod
    def getbuffer(self) -> Any:
        ...

    @abstractmethod
    def getvalue(self) -> bytearray:
        ...


class OpenPyXL(Protocol):
    @abstractmethod
    def Workbook(self) -> Workbook:
        ...


class WorkbookStatic(Protocol):
    @abstractmethod
    def create(self) -> Workbook:
        ...


class TableStatic(Protocol):
    @abstractmethod
    def create(self, displayName: str, ref: str) -> Table:
        ...


class BytesIOStatic(Protocol):
    pass

openpyxl: OpenPyXL = openpyxl_1

def _expr39() -> TypeInfo:
    return class_type("Fable.Openpyxl.Xlsx", None, Xlsx)


class Xlsx:
    ...

Xlsx_reflection = _expr39

def Xlsx_readFile_Z721C83C5(path: str) -> Workbook:
    return openpyxl.load_workbook(path)


def Xlsx_read_Z3F6BC7B1(bytes: bytearray) -> Workbook:
    return openpyxl.load_workbook(bytes)


def Xlsx_load_4E60E31B(buffer: Any=None) -> Workbook:
    return openpyxl.load_workbook(buffer)


def Xlsx_writeFile_A9E7E13(wb: Workbook, path: str) -> None:
    wb.save(path)


def Xlsx_write_Z22AEA2D8(wb: Workbook) -> bytearray:
    output: BytesIO = BytesIO_1(None)
    wb.save(output)
    return output.getvalue()


def Xlsx_writeBuffer_Z22AEA2D8(wb: Workbook) -> Any:
    output: BytesIO = BytesIO_1(None)
    wb.save(output)
    return output.getbuffer()


__all__ = ["Helper_writeBytes", "CellType_reflection", "CellType_fromCellType_Z721C83C5", "openpyxl", "Xlsx_reflection", "Xlsx_readFile_Z721C83C5", "Xlsx_read_Z3F6BC7B1", "Xlsx_load_4E60E31B", "Xlsx_writeFile_A9E7E13", "Xlsx_write_Z22AEA2D8", "Xlsx_writeBuffer_Z22AEA2D8"]


import os, sys, csv, shutil, networkx, warnings, pyodbc, urllib
import numpy as np
import pandas as pd
import polars as pl
from time import time
from tqdm import tqdm
from io import StringIO
import multiprocessing as mp
from openpyxl import workbook
from . import Ex_Zeus_Logs as ED
from . import Ex_Zeus_Defs as ECA
from types import CoroutineType
from datetime import timedelta, datetime
from sqlalchemy import create_engine, event
from networkx.algorithms.components.connected import connected_components 
from networkx.algorithms.centrality.current_flow_betweenness import edge_current_flow_betweenness_centrality

def help():
    print(f"""Hello {os.getlogin().title()}!\U0001F60A,

Thank you for choosing the Notification_List package. We sincerely appreciate your support.

Should you require any assistance or have any questions, please do not hesitate to reach out to -
Ranjeet Aloriya at +91 940.660.6239 or ranjeet.aloriya@gmail.com.
We are here to help!

Cheers!
Ranjeet Aloriya""")

def csvtsv_to_excel(folder_path):
    files = os.listdir(folder_path)
    i = 0
    for file in files:
        filename = os.path.join(folder_path, file)
        if os.path.isfile(filename):
            if file.endswith('.csv'):
                df = pd.read_csv(filename, dtype=str, encoding='latin')
            elif file.endswith('.tsv'):
                df = pd.read_csv(filename, dtype=str, delimiter='\t', encoding='latin')
            else:
                continue  # skip non-csv/tsv files

            i += 1
            output_file = os.path.splitext(filename)[0] + ".xlsx"
            df.to_excel(output_file, index=False)
            sys.stdout.write(f"\rFile No. {i} - {file} Processing")
            sys.stdout.flush()

    print(f"\nConversion completed. {i} files processed.")

def copy_files(file):
    df = pl.read_csv(file)
    i = 0
    for row in df.iter_rows():
        source_folder = row[1]
        file_name = row[0]
        destination_folder = row[2]
        source_path = os.path.join(source_folder, file_name)
        destination_path = os.path.join(destination_folder, file_name)
        os.makedirs(destination_folder, exist_ok=True)
        try:
            shutil.copy2(source_path, destination_path)
        except:
           pass
        i +=1
        sys.stdout.write(f"\rFiles Copied - {i}/{df.height}         ")
        sys.stdout.flush()

def move_files(file):
    df = pl.read_csv(file)
    i = 0
    for row in df.iter_rows():
        source_folder = row[1]
        file_name = row[0]
        destination_folder = row[2]
        source_path = os.path.join(source_folder, file_name)
        destination_path = os.path.join(destination_folder, file_name)
        os.makedirs(destination_folder, exist_ok=True)
        try:
            shutil.move(source_path, destination_path)
        except:
           pass
        i +=1
        sys.stdout.write(f"\rFiles Moved - {i}/{df.height}         ")
        sys.stdout.flush()

def copy_files_without_ext(csv_file):
    df = pl.read_csv(csv_file)
    not_found = []
    total = len(df)
    copied_count = 0
 
    for i, row in enumerate(df.iter_rows(), start=1):
        file_name, source_folder, destination_folder = map(str, row)
        found = False
 
        for root, dirs, files in os.walk(source_folder):
            for f in files:
                name, ext = os.path.splitext(f)
                if name.lower() == file_name.lower():
                    os.makedirs(destination_folder, exist_ok=True)
                    shutil.copy2(os.path.join(root, f), os.path.join(destination_folder, f))
                    copied_count += 1
                    found = True
                    break
            if found:
                break
 
        if not found:
            not_found.append([file_name])
        sys.stdout.write(f"\rProgress: {i}/{total} processed, {copied_count} copied")
        sys.stdout.flush()
 
    if not_found:
        ts = datetime.now().strftime("%m%d%y%H%M%S")
        nf_file = f"Not_Found_Copying_{ts}.csv"
        pl.DataFrame(not_found, schema=["FileName"]).write_csv(nf_file)
        print(f"\nSummary: {copied_count}/{total} copied, {len(not_found)} not found (saved in {nf_file})")
    else:
        print(f"\nSummary: All {total} files copied successfully ✅")

def df_info(file):
    df = pl.read_csv(file)
    data = []
    for col in df.columns:
        dtype = df.schema[col]
        non_null_count = len(df[col].drop_nulls())
        unique_count = df[col].n_unique()
        data.append({
            "Column Name": col,
            "Data Type": str(dtype),
            "Non-Null Count": non_null_count,
            "Unique Count": unique_count
        })
    df = pl.DataFrame(data)

def get_segment(f, sep = '~'):
    with open(f, 'r', encoding='ascii', errors='ignore') as file:
        data = file.read()
        data = data.replace("\n\n", "")
        data = data.replace("\n", "")
        segments = data.split(sep)
        df = pl.DataFrame({'Segment': segments})
        df = df.filter(pl.col("Segment").str.contains("*", literal=True))
    return df

def arrange_segment(df, column_name = "Segment", sep = "*"):
    rows = []
    current = {}
    for value in df[column_name]:
        prefix, data = value.split(sep, 1)
        if prefix in current:
            rows.append(current)
            current = {}
        current[prefix] = data
    if current:
        rows.append(current)
    df = pl.DataFrame(rows).fill_null("")
    return df

def split_columns(df, sep='\\*'):
    df = df.to_pandas()
    for column in df.columns:
        df[column] = df[column].fillna('')
        max_splits = df[column].str.count(sep).max() + 1
        max_splits = int(max_splits)
        new_columns = df[column].str.split(sep, expand=True)
        new_column_names = [f"{column}_{i+1}" for i in range(max_splits)]
        new_columns.columns = new_column_names
        df = df.drop(column, axis=1).join(new_columns)
    df = pl.from_pandas(df)
    return df

def split_full_name(df, full_name, suffixes):
    suffixes = suffixes
    def clean_and_split(text):
        return text.replace(",", "").split()
    def extract_suffix(words):
        for i, word in enumerate(words):
            if word.upper() in suffixes:
                return word.upper(), words[:i] + words[i+1:]
        return "", words
    def parse_name(part1, part2):
        first = middle = last = suffix = ""
        if part2:
            last_words = clean_and_split(part1)
            suffix, last_words = extract_suffix(last_words)
            last = " ".join(last_words)
            name_words = clean_and_split(part2)
            sfx2, name_words = extract_suffix(name_words)
            suffix = suffix or sfx2
            if name_words:
                first = name_words[0]
                if len(name_words) > 1:
                    middle = " ".join(name_words[1:])
        else:
            words = clean_and_split(part1)
            suffix, words = extract_suffix(words)
            if len(words) == 1:
                first = words[0]
            elif len(words) == 2:
                first, last = words
            elif len(words) > 2:
                first = words[0]
                last = words[-1]
                middle = " ".join(words[1:-1])
        return [first, middle, last, suffix]
    return (
        df
        .with_columns(pl.col(full_name).str.split_exact(",", 1).alias("_split"))
        .with_columns([
            pl.col("_split").struct.field("field_0").str.strip_chars().alias("_part1"),
            pl.col("_split").struct.field("field_1").str.strip_chars().fill_null("").alias("_part2")
        ])
        .with_columns(pl.struct(["_part1", "_part2"]).map_elements(
            lambda row: parse_name(row["_part1"], row["_part2"]),
            return_dtype=pl.List(pl.Utf8)
        ).alias("_parsed"))
        .with_columns([
            pl.col("_parsed").list.get(0).alias("split_first_name"),
            pl.col("_parsed").list.get(1).alias("split_middle_name"),
            pl.col("_parsed").list.get(2).alias("split_last_name"),
            pl.col("_parsed").list.get(3).alias("split_suffix"),
        ])
        .drop(["_split", "_part1", "_part2", "_parsed"])
    )

def ra_replace_chars(df, column, cleaning_dict):
    col_expr = pl.col(column)
    for pattern, replacement in cleaning_dict.items():
        col_expr = col_expr.str.replace_all(pattern, replacement)
    return df.with_columns(col_expr.str.strip_chars().alias(column))

def excel_compile_without_header(path, f):
    my_df = pl.DataFrame()
    file = os.path.join(path, f)
    sheets = pl.read_excel(file, has_header=False, sheet_id=0, raise_if_empty=False, infer_schema_length=0)
    for sheet in sheets.keys():
        df = pl.read_excel(file, has_header=False, sheet_name = sheet, raise_if_empty=False, infer_schema_length=0)
        df = df.with_columns(pl.lit(f).alias('FileName'))
        df = df.with_columns(pl.lit(sheet).alias('SheetName'))
        df = df.select(['FileName', 'SheetName']+[col for col in df.columns if col not in ['FileName', 'SheetName']])
        my_df = pl.concat([my_df, df], how='diagonal')
    return my_df

def csv_compile(path, f):
    my_df = pl.DataFrame()
    file = os.path.join(path, f)
    df = pl.read_csv(file, raise_if_empty=False, infer_schema_length=0)
    df = df.with_columns(pl.lit(f).alias('FileName'))
    my_df = pl.concat([my_df, df], how='diagonal')
    return my_df
    
def parquet_compile(path, f):
    my_df = pl.DataFrame()
    file = os.path.join(path, f)
    df = pl.read_parquet(file)
    df = df.with_columns(pl.lit(f).alias('FileName'))
    my_df = pl.concat([my_df, df], how='diagonal')
    return my_df

def batch_processing(path, processing_function, b):
    files = [f for f in os.listdir(path) if os.path.isfile(os.path.join(path, f))]
    j = 0
    for i in range(0, len(files), b):
        batch = files[i:i+b]
        my_df = pl.DataFrame()
        j += 1
        k = 0
        for f in batch:
            try:
                df = processing_function(path, f)
                my_df = pl.concat([my_df, df], how='diagonal')
                k += 1
                sys.stdout.write(f"\rFile No. {k} - Processed of Batch No. {j}          ")
                sys.stdout.flush()
            except Exception as e:
                sys.stdout.write(f"\r⚠️ Skipping file due to error: {f}{e}          ")
                sys.stdout.flush()
                continue
        sys.stdout.write(f"\rBatch No. {j} - Processed                                                 ")
        sys.stdout.flush()
        batch_number = f"{j:03d}"
        func_name = processing_function.__name__
        output_folder = os.path.join(path, 'output1')
        os.makedirs(output_folder, exist_ok=True)
        output_file_path = os.path.join(output_folder, f'{func_name}_Batch_{batch_number}.parquet')
        my_df.write_parquet(output_file_path)
    sys.stdout.write(f"\rAll Batches are Processed                                                 ")
    sys.stdout.flush()

def table_from_sql(server, database, table):
    my_df = pl.DataFrame()
    connection = pyodbc.connect(
        f'DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};DATABASE={database};Trusted_Connection=yes;'
    )
    
    sys.stdout.write(f"\rConnection successful!      ")
    sys.stdout.flush()
    query = f"select * from {table}"
    for df in pl.read_database(query, connection=connection, iter_batches=True, batch_size=10000, infer_schema_length=0):
        my_df = pl.concat([my_df, df], how = 'diagonal')
        sys.stdout.write(f"\rDownloading Raws: {my_df.shape[0]}      ")
        sys.stdout.flush()
    sys.stdout.write(f"\rSaved as polars DataFrame - Total Raws: {my_df.shape[0]}      ")
    sys.stdout.flush()
    return my_df
    
def table_to_sql(server, database, table, df):
    params = urllib.parse.quote_plus(
        f"DRIVER={{ODBC Driver 17 for SQL Server}};"
        f"SERVER={server};"
        f"DATABASE={database};"
        "Trusted_Connection=yes;"
    )
    
    engine = create_engine(f"mssql+pyodbc:///?odbc_connect={params}")
    df = df.to_pandas()
    df.to_sql(table, con=engine, index=False, if_exists="replace", schema="dbo")
    sys.stdout.write(f"\rSaved {table} in {database} - Total Raws: {len(df)}      ")
    sys.stdout.flush()

tstart_time = None
tstart2_time = None

def analyze_files_func(dataworker):
    Datapath_Out = dataworker["datapathOut"]
    Datapath_Out = dataworker["StatFile"]
    NbProcesses_int = dataworker["ProcessNo"]
    sys.stdout.write(f"\rProcess: " + str(NbProcesses_int) + " | Files: " + str(len(dataworker["FilesIn"])))
    sys.stdout.flush()
    filecnt_int = 0
    fileprintcnt_int = 100
    stat_dict = {}
    with open(Datapath_Out + "Basket_Set_"+ NbProcesses_int +".dat",'w') as file:
        file.write("\xfeGroupNo\xfe\x14\xfeTable_Columns\xfe\n")
    for filename in tqdm(dataworker["FilesIn"]):
        filecnt_int +=1
        processed_file = False
        curfilename = str(filename[filename.rfind('\\')+1:filename.rfind(".",filename.rfind('\\')+1)])
        filename_value = str(curfilename)
        fileext_value = str(filename[filename.rfind('.')+1:])
        try:
            if fileext_value.lower() in ['xls','xlsx','xlsm','xlsb']:
                    temp_stat_dict = ECA.eca_excel_funct(filename,Datapath_Out,NbProcesses_int)
                    processed_file = True
            elif fileext_value.lower() in ['csv','tsv','txt']:
                    temp_stat_dict = ECA.eca_delimited_file_func(filename,Datapath_Out,NbProcesses_int)
                    processed_file = True
            elif only_tables_bool == False:
                    temp_stat_dict = ECA.eca_text_func(filename,Datapath_Out,NbProcesses_int)
                    processed_file = True
        except Exception as e:
            temp_stat_dict = {
            "FileName": filename_value
            ,"Extension": fileext_value
            ,"Sheet Name": np.nan
            ,"First Row Columns": np.nan
            ,"Table Columns": np.nan
            ,"Table Columns Mapping": np.nan
            ,"Key Columns": np.nan
            ,"Table Number of Rows": np.nan
            ,"Total Number of Columns": np.nan
            ,"Total Number of Rows": np.nan
            ,"Processing Notes": str(e)
            ,"Table Group": np.nan
            ,"Name Identifier": np.nan
            ,"Name Count": np.nan
            ,"Address Identifier": np.nan
            ,"Address Count": np.nan
            ,"DOB Identifier": np.nan
            ,"DOB Count": np.nan
            ,"SSN Identifier": np.nan
            ,"SSN Count": np.nan
            ,"TIN Identifier": np.nan
            ,"TIN Count": np.nan
            ,"Phone Number Identifier": np.nan
            ,"Phone Count": np.nan
            ,"Driver's License Identifier": np.nan
            ,"Driver's License Count": np.nan
                                }
            sys.stdout.write(f"\r{filename,str(e)}")
            sys.stdout.flush()


        if processed_file == True:
            if len(stat_dict) == 0:
                    for k,v in temp_stat_dict.items():
                        stat_dict[k] = v
            else:
                    for k,v in temp_stat_dict.items():
                        for n in v:
                            stat_dict[k].append(n)
    return stat_dict


def excel_analysis(Datapath_In, analyze_files_func):
    global tstart_time, tstart2_time
    tstart_time = time()
    tstart2_time = time()
    warnings.filterwarnings('ignore')
    maxInt =  500000
    csv.field_size_limit(2147483647)
    Datapath_Out = os.path.join(Datapath_In, "Output\\")
    NbProcesses_int = 1
    only_tables_bool = False 
    if not os.path.exists(Datapath_In):
        sys.stdout.write(f"\rFile Location Error")
        sys.stdout.flush()
    if not os.path.exists(Datapath_Out):
        os.makedirs(Datapath_Out)
    if not os.path.exists(Datapath_Out):
        os.makedirs(Datapath_Out)

    def printelaps():
        global tstart_time
        global tstart2_time
        sys.stdout.write(f"\rDuration: " + str(timedelta(seconds=(time() - tstart_time))).split('.')[0] + " | Difference: " + str(timedelta(seconds=(time() - tstart2_time))).split('.')[0] ); tstart2_time = time()
        sys.stdout.flush()


    def printelaps2():
        global tstart_time
        global tstart2_time
        sys.stdout.write(f"\rDuration: " + str(timedelta(seconds=(time() - tstart_time))).split('.')[0] + " | Difference: " + str(timedelta(seconds=(time() - tstart2_time))).split('.')[0] ); tstart2_time = time()
        sys.stdout.flush()


    def to_graph(l):
        G = networkx.Graph()
        for part in l:
        
            G.add_nodes_from(part)
        
            G.add_edges_from(to_edges(part))
        return G



    def to_edges(l):
        
        
        it = iter(l)
        last = next(it)

        for current in it:
            yield last, current
            last = current  



    if __name__ == "__main__":
        printelaps()
        starttime_str = str(datetime.now())
        pool = mp.Pool(NbProcesses_int)

        runn_all_files_bool = True
        files_to_do_list = []
        if runn_all_files_bool == False:
            with open(r'L:\Intrasite\iDAT - Team\Akshat\Excel_Extracted_Sets\ex_zeus\Cache_ECA_POC.txt\\','r') as file:
                for line in file:
                        files_to_do_list.append(line.strip('\n').lower())

        if not os.path.exists(Datapath_In):
            sys.stdout.write(f"\rFile Location Error")
            sys.stdout.flush()
            exit()
        if not os.path.exists(Datapath_Out):
            os.makedirs(Datapath_Out)
        if not os.path.exists(Datapath_Out):
            os.makedirs(Datapath_Out)
        filelist_dict = {} 
        
        for i in range(NbProcesses_int):
            filelist_dict[i] = []
        i = 0

        if NbProcesses_int > 1:
            sys.stdout.write(f"\rProcessing ...")
            sys.stdout.flush()
            total_bytes_int = 0
            for root, subdirs, files in os.walk(Datapath_In):
                for filename in files:
                        if (filename.lower() in files_to_do_list) or (runn_all_files_bool == True):
                            total_bytes_int += os.path.getsize(os.path.join(root, filename))
            total_bytes_ratio_int = total_bytes_int / NbProcesses_int
            sys.stdout.write(f"\rTotal bytes: ",total_bytes_int)
            sys.stdout.flush()
            sys.stdout.write(f"\rTotal bytes per process: ",total_bytes_ratio_int)
            sys.stdout.flush()
            current_bytes_int = 0
            for root, subdirs, files in os.walk(Datapath_In):
                for filename in files:
                        if (filename.lower() in files_to_do_list) or (runn_all_files_bool == True):
                            filelist_dict[i].append(os.path.join(root, filename))
                            current_bytes_int += os.path.getsize(os.path.join(root, filename))
                            if current_bytes_int > total_bytes_ratio_int and i < NbProcesses_int:
                                current_bytes_int = 0
                                i += 1
        else:
            for root, subdirs, files in os.walk(Datapath_In):
                for filename in files:
                        if (filename[0:filename.rfind(".")].lower() in files_to_do_list) or (runn_all_files_bool == True):
                            filelist_dict[i].append(os.path.join(root, filename))
                            i +=1
                            if i > NbProcesses_int-1:
                                i = 0
        dataworker = []       
        for i in range(NbProcesses_int):
            dataworker.append( {
                "FilesIn": filelist_dict[i],
                "datapathOut": Datapath_Out,
                "StatFile": Datapath_Out,
                "ProcessNo": str(i+1),
                    })

        sd = pool.map(analyze_files_func,dataworker)

        fullstat_dict = {}
        for s in sd:
            for k,v in s.items():
                if k in fullstat_dict.keys():
                        fullstat_dict[k] += v
                else:
                        fullstat_dict[k] = v

        
        df = pd.DataFrame.from_dict(fullstat_dict,orient='columns')
        groupno_dict = {}
        groupno_matching_dict = {}

        sys.stdout.write(f"\rCreating Baskets ...")
        sys.stdout.flush()
        for i in range(1,NbProcesses_int+1):
            with open(Datapath_Out + "Basket_Set_"+ str(i) +".dat",'r') as file:
                file_opened_bool = False
                while file_opened_bool == False:
                        try:
                            csv.field_size_limit(maxInt)
                            r = csv.reader(file, delimiter='\x14', quotechar='\xfe')
                            file_opened_bool = True
                        except:
                            maxInt = int(maxInt/10)
                lncnt_int = 1
                for line in r:
                        if lncnt_int == 1:lncnt_int +=1;continue
                        lncnt_int +=1
                        if len(groupno_dict) == 0:
                            groupno_dict[line[0]] = line[1]
                            groupno_matching_dict[line[0]] = [line[0]]
                        else:
                            if not line[1] in groupno_dict.values():
                                groupno_dict[line[0]] = line[1]
                                groupno_matching_dict[line[0]] = [line[0]]
                            else:
                                for k,v in groupno_dict.items():
                                    if v == line[1]:
                                            groupno_matching_dict[k] += [line[0]]
                                            break
            os.remove(Datapath_Out + "Basket_Set_"+ str(i) +".dat")
        
        if not os.path.exists(Datapath_Out + "Baskets_Extracted"):
            os.makedirs(Datapath_Out + "Baskets_Extracted")
        started_groups_list = []

        with open(Datapath_Out + "All_Basket_Cols.dat",'w+') as file:
            file.write("\xfeGroupNo\xfe\x14\xfeTable_Columns\xfe\n")
            for key, value in groupno_dict.items():
                file.write("\xfe" + key + "\xfe\x14\xfe" + value + "\xfe\n")
    
        
        sys.stdout.write(f"\rMoving Baskets ...")
        sys.stdout.flush()
    
        for stackedfile in os.listdir(Datapath_Out):
            if stackedfile.find("Basket_Number") > -1:
                if stackedfile.find("Basket_Number") > -1:
                    if NbProcesses_int > 1:
                        sys.stdout.write(f"\rCombining Baskets ...")
                        sys.stdout.flush()
                        curfilegroup = str(stackedfile[stackedfile.rfind('_')+1:stackedfile.rfind(".",stackedfile.rfind('_')+1)])
                        for key, value in groupno_matching_dict.items():
                                if curfilegroup in value:
                                    if key in started_groups_list:
                                        concat_df = pd.read_csv(Datapath_Out + stackedfile,delimiter=',', dtype=str, keep_default_na=False, skipinitialspace=True,encoding='utf-8')
                                        with open(Datapath_Out + "Baskets_Extracted\\" + "Basket_Number" + "_" + str(key) + ".csv", 'a+', newline ='\n',encoding='utf-16') as file:
                                            concat_df.to_csv(file,index=False,header=False,encoding='utf-16',quoting=csv.QUOTE_ALL)
                                        os.remove(Datapath_Out + stackedfile)
                                    else:
                                        concat_df = pd.read_csv(Datapath_Out + stackedfile,delimiter=',', dtype=str, keep_default_na=False, skipinitialspace=True,encoding='utf-8-sig')
                                        concat_df.to_csv(Datapath_Out + "Baskets_Extracted\\" + "Basket_Number" + "_" + str(key) + ".csv",header=True,index=False,quoting=csv.QUOTE_ALL,encoding='utf-16')
                                        os.remove(Datapath_Out + stackedfile)
                                        started_groups_list.append(key)
                                    break
                    else:
        
                        shutil.move(Datapath_Out + stackedfile, Datapath_Out + "Baskets_Extracted\\" + stackedfile)
    
        sys.stdout.write(f"\rProcessing to file")
        sys.stdout.flush()
    
        for index, row in df.iterrows():
            if row['Table Columns Mapping'] in groupno_dict.values():
                for k,v in groupno_dict.items():
                        if v == row['Table Columns Mapping']:
                            df.at[index,'Table Group'] = k
        groupcnt_col_int = df.columns.get_loc("Table Group") + 1
        df.insert(groupcnt_col_int,"Group Count",'')
        df["Group Count"] = df.groupby('Table Group')['Table Group'].transform('count')

        
        tempdf = df[['Table Group','FileName']].copy()
        tempdf['Table Group'] = tempdf['Table Group'].replace('',np.nan)
        tempdf = tempdf.dropna()
        tempdf = tempdf[tempdf['Table Group'] != ""]
        if len(tempdf) > 0:
            tempdf = tempdf.groupby(['Table Group'])['FileName'].agg(';'.join).reset_index()
            tempdf['INTT_index'] = tempdf['FileName'].apply(lambda x: x.split(';'))
            masterlist = tempdf['INTT_index'].tolist()
            tempdf2 = pd.DataFrame(columns = ['INTT_explode','INTT_index'])
            tempdf2['INTT_explode'] = masterlist
            tempdf2['INTT_explode'] = tempdf2['INTT_explode'].apply(lambda x: set(x))
            sets = tempdf2['INTT_explode'].tolist()
            G = to_graph(sets)
            sets = (list(connected_components(G)))

            df.insert(groupcnt_col_int+1,"WorkFlow Groups",'')
            df.insert(groupcnt_col_int+2,"WorkFlow Group Count",'')
            df.insert(groupcnt_col_int+3,"Files Represnted in Workflow Group",'')
            grpno = 1
            grouping_dict = {}
            for s in sets:
                for i in s:
                        if not i in grouping_dict.keys():
                            cntrl_list = df[df['FileName'] == i].index.tolist()
                            for x in cntrl_list:
                                df.at[x,'WorkFlow Groups'] = grpno
                            grouping_dict[i] = grpno
                grpno +=1
            df["WorkFlow Group Count"] = df.groupby('WorkFlow Groups')['WorkFlow Groups'].transform('count')
            filecount_dict = df.groupby('WorkFlow Groups')['FileName'].apply(list).to_dict()
            df["Files Represnted in Workflow Group"] = df["WorkFlow Groups"].apply(lambda x: len(set(filecount_dict[x])))

            del tempdf2
        del tempdf
        
        df.sort_values(['WorkFlow Groups','Table Group','FileName'], ascending=[True,True,True], inplace=True)
        df.to_excel(Datapath_Out+"!Header_Analysis.xlsx",sheet_name="Header_Analysis",header=True,index=False)
        df=pd.read_excel(Datapath_Out+"!Header_Analysis.xlsx", dtype=str)
        df=df.rename(columns={'FileName':'Doc_ID','Extension':'File_Extension','Sheet Name':'Sheet_Name','Table Columns Mapping':'Columns_Combined','Key Columns':'Columns_Identified','Total Number of Columns':'Column_Count','Total Number of Rows':'Row_Count','Processing Notes':'Comments','Table Group':'Basket_Number'})
        df=df.drop(['Group Count', 'WorkFlow Groups', 'WorkFlow Group Count', 'Files Represnted in Workflow Group', 'First Row Columns', 'Table Columns', 'Table Number of Rows', 'Name Identifier', 'Name Count', 'Address Identifier', 'Address Count', 'DOB Identifier', 'DOB Count', 'SSN Identifier', 'SSN Count', 'TIN Identifier', 'TIN Count', 'Phone Number Identifier', 'Phone Count', "Driver's License Identifier", "Driver's License Count"], axis=1)
        df.to_excel(Datapath_Out+"!Header_Analysis.xlsx",sheet_name="Header_Analysis",header=True,index=False)\
        

        os.remove(Datapath_Out+"ConCat_Columns.csv")
        os.remove(Datapath_Out+"All_Basket_Cols.dat")
    
        sys.stdout.write(f"\rStart Time:  " + starttime_str + "   |  End Time:  " + str(datetime.now()))
        sys.stdout.flush()
        printelaps2()
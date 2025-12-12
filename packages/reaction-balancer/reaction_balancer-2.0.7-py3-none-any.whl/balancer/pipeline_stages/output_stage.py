from __future__ import annotations
from threading import Thread
from queue import Queue
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Iterator
from pathlib import Path
from datetime import datetime
import errno

from ..pipeline_contracts import Context, Bundle, Style, Cell, Block
from ..errors import InputError, OutputError, InitHint


OUTPUT_DIR = "./outputs/"
MIN_COL_WIDTH = 4

class _Sentinel:
    pass

STOP = _Sentinel()
ERROR = _Sentinel()

def block_producer(q:Queue, blocks: Iterator[Block]):
    try:
        for block in blocks:
            q.put(block)
    except InputError as e:
        print(e)
        q.put(ERROR)
    except InitHint as e:
        print(e)
        q.put(ERROR)
    q.put(STOP)



def block_consumer(q:Queue, context:Context):
    wb = Workbook()
    ws = wb.active
    ws.title = "Recipe"
    col_width = {}
    while True:
        block = q.get()
        if block is STOP:
            q.task_done()
            set_col_width(ws,col_width)
            output_basename = context.title
            save(wb, output_basename)
            break
        elif block is ERROR:
            q.task_done()
            break
        
        write(block, ws, col_width)
        q.task_done()



def write(block:Block,ws:Worksheet, col_width:dict[int,float]):
    col = block.anchor_col
    if block.width:
        col_width[col] = max(col_width.get(col, MIN_COL_WIDTH) , block.width)
    for offset, cell_desc in enumerate(block.column):
        row = block.anchor_row + offset
        cell = ws.cell(row=row,column=col,value=cell_desc.value)

        cell.number_format = cell_desc.number_format

        if cell_desc.style:
            cell.font = cell_desc.style.font
            cell.alignment = cell_desc.style.alignment
            cell.border = cell_desc.style.border

def set_col_width(ws:Worksheet, col_width:dict[int,float]):
    for col,width in col_width.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def save(wb:Workbook, basename: str):
    if basename is None:
        filename = f"{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    else:
        filename = f"{basename}.xlsx"
    output_path = Path(OUTPUT_DIR) / filename
    try:
        try:
            wb.save(output_path)
            return
        except FileNotFoundError:
            try:
                output_path.parent.mkdir(parents=True, exist_ok=True)
            except Exception:
                raise OutputError(
                    output_path,
                    explanation=f"Failed to create directory '{output_path.parent}'.",
                    hint="Check if you have permissions.",
                ) 
            wb.save(output_path)
    except OSError as e:
        # no enough permission
        if e.errno in {errno.EACCES, errno.EPERM}: 
            raise OutputError( output_path, 
                                explanation="Permission denied. The file may be open in Excel or you lack write permission.", 
                                hint="Close Excel/WPS or check your permissions.", ) 
        # illegal path
        if e.errno in {errno.EINVAL, errno.ENAMETOOLONG, errno.ENOENT}:
            raise OutputError( output_path,
                                hint="This path is invalid on this operating system.",)
        # other error
        raise OutputError(
            output_path,
            explanation=f"OS error ({e.errno}): {e.strerror}",
            hint="Check disk space, filesystem type, or path validity.",
)

class OutputStage:

    def process(self, bundle:Bundle)->Bundle:
        blocks = bundle.stream
        context = bundle.context

        q = Queue(maxsize=128)
        t1 = Thread(target=block_producer, args=(q,blocks), daemon=True)
        t2 = Thread(target=block_consumer, args=(q,context), daemon=True)
        t1.start()
        t2.start()

        t2.join()


        return Bundle(bundle.context, None)
        

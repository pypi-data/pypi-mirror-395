"""
Sheet-level extraction utilities for Template Sense.

This module provides utilities that extract raw, structured sheet data from the
ExcelWorkbook abstraction. These utilities provide row/column access in normalized
Python structures (lists of lists, dictionaries, etc.) for the extraction pipeline
and AI model to later analyze.

This module focuses ONLY on reading sheet data in a predictable format:
- No AI integration
- No heuristics
- No header detection

All utilities work with the ExcelWorkbook adapter (not openpyxl directly) and return
clean, normalized data with no openpyxl objects.

Functions:
    extract_raw_grid: Extract complete 2D grid of cell values from a sheet
    extract_non_empty_rows: Filter out completely empty rows from a grid
    extract_non_empty_columns: Filter out completely empty columns from a grid
    get_used_range: Get the actual used range excluding empty borders

Usage Example:
    from template_sense.file_loader import load_excel_file
    from template_sense.adapters.excel_adapter import ExcelWorkbook
    from template_sense.extraction.sheet_extractor import extract_raw_grid, get_used_range

    raw_workbook = load_excel_file(Path("template.xlsx"))
    workbook = ExcelWorkbook(raw_workbook)

    # Extract complete grid
    grid = extract_raw_grid(workbook, "Sheet1")

    # Get used range
    min_row, max_row, min_col, max_col = get_used_range(workbook, "Sheet1")
"""

import logging
from typing import Any

from template_sense.adapters.excel_adapter import ExcelWorkbook
from template_sense.errors import ExtractionError

# Set up module logger
logger = logging.getLogger(__name__)


def extract_raw_grid(workbook: ExcelWorkbook, sheet_name: str) -> list[list[Any]]:
    """
    Extract complete 2D grid of cell values from a sheet.

    Returns a list of lists where each inner list represents a row of cell values.
    All cell values are Python primitives (str, int, float, None, datetime, etc.)
    with no openpyxl objects.

    IMPORTANT: Hidden rows and columns are automatically excluded from the grid.

    Args:
        workbook: ExcelWorkbook instance to extract from
        sheet_name: Name of the sheet to extract

    Returns:
        2D list of cell values (list of rows). Empty list if sheet is empty.
        Hidden rows are excluded. Hidden columns are excluded from each row.

    Raises:
        ExtractionError: If sheet_name does not exist or extraction fails

    Example:
        >>> grid = extract_raw_grid(workbook, "Sheet1")
        >>> print(grid)
        [['Header1', 'Header2'], ['Value1', 'Value2'], ['Value3', 'Value4']]
    """
    logger.debug("Extracting raw grid from sheet '%s'", sheet_name)

    try:
        # Get hidden rows and columns
        hidden_rows = workbook.get_hidden_rows(sheet_name)
        hidden_cols = workbook.get_hidden_columns(sheet_name)

        # Use workbook.iter_rows to get all rows
        # This already returns primitive values, not Cell objects
        all_rows = list(workbook.iter_rows(sheet_name))

        # Filter out hidden rows and hidden columns
        grid = []
        for row_idx, row in enumerate(all_rows, start=1):
            # Skip hidden rows
            if row_idx in hidden_rows:
                logger.debug("Skipping hidden row %d", row_idx)
                continue

            # Filter out hidden columns from this row
            if hidden_cols:
                # Convert column indices to letters to check if hidden
                from openpyxl.utils import get_column_letter

                filtered_row = []
                for col_idx, cell_value in enumerate(row, start=1):
                    col_letter = get_column_letter(col_idx)
                    if col_letter not in hidden_cols:
                        filtered_row.append(cell_value)
                    else:
                        logger.debug("Skipping hidden column %s in row %d", col_letter, row_idx)
                grid.append(filtered_row)
            else:
                # No hidden columns, use row as-is
                grid.append(row)

        logger.info(
            "Extracted raw grid from sheet '%s': %d rows (excluded %d hidden rows, %d hidden columns)",
            sheet_name,
            len(grid),
            len(hidden_rows),
            len(hidden_cols),
        )
        return grid

    except ExtractionError:
        # Re-raise ExtractionError from workbook (sheet doesn't exist)
        raise

    except Exception as e:
        logger.error(
            "Unexpected error extracting grid from sheet '%s': %s",
            sheet_name,
            str(e),
        )
        raise ExtractionError(
            extraction_type="grid",
            reason=f"Failed to extract grid from sheet '{sheet_name}'",
        ) from e


def extract_non_empty_rows(grid: list[list[Any]]) -> list[list[Any]]:
    """
    Filter out completely empty rows from a grid.

    A row is considered empty if all its cells are None or empty strings.

    Args:
        grid: 2D list of cell values (list of rows)

    Returns:
        New 2D list with only non-empty rows. Empty list if all rows are empty.

    Example:
        >>> grid = [['A', 'B'], [None, None], ['C', 'D'], ['', '']]
        >>> extract_non_empty_rows(grid)
        [['A', 'B'], ['C', 'D']]
    """
    logger.debug("Filtering non-empty rows from grid with %d rows", len(grid))

    non_empty_rows = []

    for row in grid:
        # Check if row has any non-empty values
        if any(cell not in (None, "") for cell in row):
            non_empty_rows.append(row)

    logger.info(
        "Filtered grid: %d non-empty rows out of %d total rows",
        len(non_empty_rows),
        len(grid),
    )
    return non_empty_rows


def extract_non_empty_columns(grid: list[list[Any]]) -> list[list[Any]]:
    """
    Filter out completely empty columns from a grid.

    A column is considered empty if all its cells are None or empty strings.

    Args:
        grid: 2D list of cell values (list of rows)

    Returns:
        New 2D list with only non-empty columns. Empty list if all columns are empty
        or if input grid is empty.

    Example:
        >>> grid = [['A', None, 'B'], ['C', None, 'D']]
        >>> extract_non_empty_columns(grid)
        [['A', 'B'], ['C', 'D']]
    """
    logger.debug("Filtering non-empty columns from grid with %d rows", len(grid))

    if not grid:
        logger.info("Grid is empty, returning empty list")
        return []

    # Handle case where all rows are empty lists
    if all(len(row) == 0 for row in grid):
        logger.info("All rows are empty, returning empty list")
        return []

    # Get maximum column count to handle inconsistent row lengths
    max_cols = max(len(row) for row in grid)

    # Identify non-empty columns
    non_empty_col_indices = []

    for col_idx in range(max_cols):
        # Check if this column has any non-empty values
        has_non_empty = False

        for row in grid:
            # Handle rows that are shorter than max_cols
            if col_idx < len(row):
                cell_value = row[col_idx]
                if cell_value not in (None, ""):
                    has_non_empty = True
                    break

        if has_non_empty:
            non_empty_col_indices.append(col_idx)

    # Build new grid with only non-empty columns
    result = []
    for row in grid:
        new_row = []
        for col_idx in non_empty_col_indices:
            # Handle rows that are shorter than max_cols
            if col_idx < len(row):
                new_row.append(row[col_idx])
            else:
                new_row.append(None)
        result.append(new_row)

    logger.info(
        "Filtered grid: %d non-empty columns out of %d total columns",
        len(non_empty_col_indices),
        max_cols,
    )
    return result


def get_used_range(workbook: ExcelWorkbook, sheet_name: str) -> tuple[int, int, int, int]:
    """
    Get the actual used range (min_row, max_row, min_col, max_col) excluding empty borders.

    Returns 1-based indices (Excel convention) for the first and last non-empty rows
    and columns. For empty sheets, returns (1, 1, 1, 1).

    Args:
        workbook: ExcelWorkbook instance to extract from
        sheet_name: Name of the sheet to analyze

    Returns:
        Tuple of (min_row, max_row, min_col, max_col) in 1-based indices.
        Returns (1, 1, 1, 1) for empty sheets.

    Raises:
        ExtractionError: If sheet_name does not exist or extraction fails

    Example:
        >>> min_row, max_row, min_col, max_col = get_used_range(workbook, "Sheet1")
        >>> print(f"Used range: R{min_row}:R{max_row}, C{min_col}:C{max_col}")
        Used range: R1:R10, C1:C5
    """
    logger.debug("Calculating used range for sheet '%s'", sheet_name)

    try:
        # Extract raw grid
        grid = extract_raw_grid(workbook, sheet_name)

        # Handle empty sheet
        if not grid or all(len(row) == 0 for row in grid):
            logger.info("Sheet '%s' is empty, returning (1, 1, 1, 1)", sheet_name)
            return (1, 1, 1, 1)

        # Find first non-empty row (1-based)
        min_row = None
        for row_idx, row in enumerate(grid, start=1):
            if any(cell not in (None, "") for cell in row):
                min_row = row_idx
                break

        # Find last non-empty row (1-based)
        max_row = None
        for row_idx in range(len(grid), 0, -1):
            row = grid[row_idx - 1]  # Convert to 0-based index
            if any(cell not in (None, "") for cell in row):
                max_row = row_idx
                break

        # If no non-empty rows found
        if min_row is None or max_row is None:
            logger.info(
                "Sheet '%s' has no non-empty cells, returning (1, 1, 1, 1)",
                sheet_name,
            )
            return (1, 1, 1, 1)

        # Get maximum column count
        max_cols = max(len(row) for row in grid)

        # Find first non-empty column (1-based)
        min_col = None
        for col_idx in range(max_cols):
            for row in grid:
                if col_idx < len(row) and row[col_idx] not in (None, ""):
                    min_col = col_idx + 1  # Convert to 1-based
                    break
            if min_col is not None:
                break

        # Find last non-empty column (1-based)
        max_col = None
        for col_idx in range(max_cols - 1, -1, -1):
            for row in grid:
                if col_idx < len(row) and row[col_idx] not in (None, ""):
                    max_col = col_idx + 1  # Convert to 1-based
                    break
            if max_col is not None:
                break

        # If no non-empty columns found (shouldn't happen if we found non-empty rows)
        if min_col is None or max_col is None:
            logger.info(
                "Sheet '%s' has no non-empty cells, returning (1, 1, 1, 1)",
                sheet_name,
            )
            return (1, 1, 1, 1)

        logger.info(
            "Used range for sheet '%s': R%d:R%d, C%d:C%d",
            sheet_name,
            min_row,
            max_row,
            min_col,
            max_col,
        )
        return (min_row, max_row, min_col, max_col)

    except ExtractionError:
        # Re-raise ExtractionError from extract_raw_grid
        raise

    except Exception as e:
        logger.error(
            "Unexpected error calculating used range for sheet '%s': %s",
            sheet_name,
            str(e),
        )
        raise ExtractionError(
            extraction_type="used_range",
            reason=f"Failed to calculate used range for sheet '{sheet_name}'",
        ) from e


__all__ = [
    "extract_raw_grid",
    "extract_non_empty_rows",
    "extract_non_empty_columns",
    "get_used_range",
]

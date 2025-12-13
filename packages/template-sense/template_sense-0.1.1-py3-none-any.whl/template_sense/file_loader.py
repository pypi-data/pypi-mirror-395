"""
File loader utility for Template Sense.

This module provides safe, consistent Excel file loading functionality.
It abstracts away raw file loading so that downstream extraction, AI analysis,
and mapping never have to deal directly with filesystem concerns.

Functions:
    load_excel_file: Load an Excel file into an openpyxl Workbook object
    load_excel_sheet_names: Get list of sheet names from an Excel file

Usage Example:
    from pathlib import Path
    from template_sense.file_loader import load_excel_file, load_excel_sheet_names

    file_path = Path("template.xlsx")
    workbook = load_excel_file(file_path)
    sheet_names = load_excel_sheet_names(file_path)
"""

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from template_sense.errors import FileValidationError
from template_sense.file_validation import validate_supported_file


def load_excel_file(path: Path) -> Workbook:
    """
    Load an Excel file into an openpyxl Workbook object.

    This function validates the file and loads it into memory using openpyxl.
    It uses data_only=True to ensure formulas are evaluated and only their
    computed values are loaded.

    Args:
        path: Path to the Excel file (.xlsx or .xls)

    Returns:
        An openpyxl Workbook object containing the loaded Excel data

    Raises:
        FileValidationError: If file validation fails or file cannot be opened
        UnsupportedFileTypeError: If file is not a supported Excel format

    Example:
        >>> from pathlib import Path
        >>> workbook = load_excel_file(Path("template.xlsx"))
        >>> print(workbook.sheetnames)
        ['Sheet1', 'Sheet2']
    """
    # Ensure path is a Path object
    if not isinstance(path, Path):
        path = Path(path)

    # Validate file before attempting to load
    validate_supported_file(path)

    # Load the Excel file
    try:
        # data_only=True: Load computed values instead of formulas
        # This ensures consistent behavior across different environments
        return load_workbook(filename=path, data_only=True)
    except FileNotFoundError as e:
        # This should not happen since validate_supported_file checks existence
        raise FileValidationError(
            reason="File not found during loading",
            file_path=str(path),
        ) from e
    except PermissionError as e:
        raise FileValidationError(
            reason="Permission denied - cannot read file",
            file_path=str(path),
        ) from e
    except Exception as e:
        # Catch any other openpyxl errors (corrupt file, unexpected format, etc.)
        raise FileValidationError(
            reason=f"Failed to load Excel file: {str(e)}",
            file_path=str(path),
        ) from e


def load_excel_sheet_names(path: Path) -> list[str]:
    """
    Get list of sheet names from an Excel file.

    This is a convenience function that loads the workbook and returns
    the list of sheet names. Useful for quick inspection without processing
    the full workbook data.

    Args:
        path: Path to the Excel file (.xlsx or .xls)

    Returns:
        List of sheet names in the order they appear in the workbook

    Raises:
        FileValidationError: If file validation fails or file cannot be opened
        UnsupportedFileTypeError: If file is not a supported Excel format

    Example:
        >>> from pathlib import Path
        >>> sheet_names = load_excel_sheet_names(Path("template.xlsx"))
        >>> print(sheet_names)
        ['Sheet1', 'Invoice', 'Summary']
    """
    # Load the workbook
    workbook = load_excel_file(path)

    # Extract sheet names
    sheet_names = workbook.sheetnames

    # Close the workbook to free resources
    workbook.close()

    return sheet_names


__all__ = [
    "load_excel_file",
    "load_excel_sheet_names",
]

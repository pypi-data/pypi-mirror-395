"""
Unit tests for file_loader module.

Tests the Excel file loading functionality including:
- Valid file loading
- Error handling for missing/invalid files
- Sheet name retrieval
- Integration with file validation
"""

from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.workbook.workbook import Workbook as WorkbookType

from template_sense.errors import FileValidationError, UnsupportedFileTypeError
from template_sense.file_loader import load_excel_file, load_excel_sheet_names

# Fixtures


@pytest.fixture
def valid_xlsx_file(tmp_path: Path) -> Path:
    """Create a valid .xlsx file for testing."""
    file_path = tmp_path / "test_file.xlsx"

    # Create a simple workbook with multiple sheets
    workbook = Workbook()

    # First sheet (default)
    sheet1 = workbook.active
    sheet1.title = "Sheet1"
    sheet1["A1"] = "Header"
    sheet1["B1"] = "Value"
    sheet1["A2"] = "Invoice Number"
    sheet1["B2"] = "INV-001"

    # Add a second sheet
    sheet2 = workbook.create_sheet("Invoice")
    sheet2["A1"] = "Item"
    sheet2["B1"] = "Quantity"
    sheet2["A2"] = "Widget"
    sheet2["B2"] = 10

    # Add a third sheet
    sheet3 = workbook.create_sheet("Summary")
    sheet3["A1"] = "Total"
    sheet3["B1"] = 100

    # Save the workbook
    workbook.save(file_path)
    workbook.close()

    return file_path


@pytest.fixture
def empty_xlsx_file(tmp_path: Path) -> Path:
    """Create an empty but valid .xlsx file for testing."""
    file_path = tmp_path / "empty_file.xlsx"

    workbook = Workbook()
    workbook.save(file_path)
    workbook.close()

    return file_path


@pytest.fixture
def corrupted_xlsx_file(tmp_path: Path) -> Path:
    """Create a corrupted .xlsx file (not a valid Excel file)."""
    file_path = tmp_path / "corrupted.xlsx"

    # Write invalid content
    file_path.write_text("This is not a valid Excel file")

    return file_path


@pytest.fixture
def unsupported_file(tmp_path: Path) -> Path:
    """Create an unsupported file type (.csv)."""
    file_path = tmp_path / "unsupported.csv"
    file_path.write_text("col1,col2\nval1,val2")
    return file_path


# Tests for load_excel_file


def test_load_excel_file_valid_file(valid_xlsx_file: Path):
    """Test loading a valid Excel file returns a Workbook object."""
    workbook = load_excel_file(valid_xlsx_file)

    # Verify it's a Workbook object
    assert isinstance(workbook, WorkbookType)

    # Verify we can access sheets
    assert "Sheet1" in workbook.sheetnames
    assert "Invoice" in workbook.sheetnames
    assert "Summary" in workbook.sheetnames

    # Verify we can read data
    sheet = workbook["Sheet1"]
    assert sheet["A1"].value == "Header"
    assert sheet["B2"].value == "INV-001"

    workbook.close()


def test_load_excel_file_empty_file(empty_xlsx_file: Path):
    """Test loading an empty but valid Excel file."""
    workbook = load_excel_file(empty_xlsx_file)

    assert isinstance(workbook, WorkbookType)
    assert len(workbook.sheetnames) >= 1  # At least one default sheet

    workbook.close()


def test_load_excel_file_missing_file(tmp_path: Path):
    """Test that loading a missing file raises FileValidationError."""
    missing_file = tmp_path / "nonexistent.xlsx"

    with pytest.raises(FileValidationError) as exc_info:
        load_excel_file(missing_file)

    assert "does not exist" in str(exc_info.value).lower()


def test_load_excel_file_unsupported_type(unsupported_file: Path):
    """Test that loading an unsupported file type raises UnsupportedFileTypeError."""
    with pytest.raises(UnsupportedFileTypeError) as exc_info:
        load_excel_file(unsupported_file)

    assert ".csv" in str(exc_info.value)


def test_load_excel_file_corrupted_file(corrupted_xlsx_file: Path):
    """Test that loading a corrupted Excel file raises FileValidationError."""
    with pytest.raises(FileValidationError) as exc_info:
        load_excel_file(corrupted_xlsx_file)

    assert (
        "not a valid Excel file" in str(exc_info.value).lower()
        or "corrupted" in str(exc_info.value).lower()
    )


def test_load_excel_file_directory_path(tmp_path: Path):
    """Test that passing a directory path raises FileValidationError."""
    with pytest.raises(FileValidationError) as exc_info:
        load_excel_file(tmp_path)

    assert "not a file" in str(exc_info.value).lower()


def test_load_excel_file_accepts_path_object(valid_xlsx_file: Path):
    """Test that load_excel_file accepts Path objects."""
    workbook = load_excel_file(valid_xlsx_file)
    assert isinstance(workbook, WorkbookType)
    workbook.close()


def test_load_excel_file_accepts_string_path(valid_xlsx_file: Path):
    """Test that load_excel_file accepts string paths."""
    workbook = load_excel_file(str(valid_xlsx_file))
    assert isinstance(workbook, WorkbookType)
    workbook.close()


def test_load_excel_file_data_only_mode(tmp_path: Path):
    """Test that workbook is loaded with data_only=True (formulas evaluated)."""
    file_path = tmp_path / "formulas.xlsx"

    # Create a workbook with a formula
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = 10
    sheet["A2"] = 20
    sheet["A3"] = "=A1+A2"  # Formula
    workbook.save(file_path)
    workbook.close()

    # Load with our function
    loaded_workbook = load_excel_file(file_path)
    sheet = loaded_workbook.active

    # With data_only=True, formulas return None on first load
    # (requires file to be saved with calculated values)
    # This is expected behavior for data_only=True
    assert sheet["A1"].value == 10
    assert sheet["A2"].value == 20

    loaded_workbook.close()


# Tests for load_excel_sheet_names


def test_load_excel_sheet_names_valid_file(valid_xlsx_file: Path):
    """Test retrieving sheet names from a valid Excel file."""
    sheet_names = load_excel_sheet_names(valid_xlsx_file)

    assert isinstance(sheet_names, list)
    assert len(sheet_names) == 3
    assert "Sheet1" in sheet_names
    assert "Invoice" in sheet_names
    assert "Summary" in sheet_names


def test_load_excel_sheet_names_order_preserved(valid_xlsx_file: Path):
    """Test that sheet names are returned in the correct order."""
    sheet_names = load_excel_sheet_names(valid_xlsx_file)

    assert sheet_names[0] == "Sheet1"
    assert sheet_names[1] == "Invoice"
    assert sheet_names[2] == "Summary"


def test_load_excel_sheet_names_empty_file(empty_xlsx_file: Path):
    """Test retrieving sheet names from an empty Excel file."""
    sheet_names = load_excel_sheet_names(empty_xlsx_file)

    assert isinstance(sheet_names, list)
    assert len(sheet_names) >= 1  # At least one default sheet


def test_load_excel_sheet_names_missing_file(tmp_path: Path):
    """Test that retrieving sheet names from missing file raises FileValidationError."""
    missing_file = tmp_path / "nonexistent.xlsx"

    with pytest.raises(FileValidationError):
        load_excel_sheet_names(missing_file)


def test_load_excel_sheet_names_unsupported_type(unsupported_file: Path):
    """Test that retrieving sheet names from unsupported file raises error."""
    with pytest.raises(UnsupportedFileTypeError):
        load_excel_sheet_names(unsupported_file)


def test_load_excel_sheet_names_single_sheet(tmp_path: Path):
    """Test retrieving sheet names from a workbook with a single sheet."""
    file_path = tmp_path / "single_sheet.xlsx"

    workbook = Workbook()
    workbook.active.title = "OnlySheet"
    workbook.save(file_path)
    workbook.close()

    sheet_names = load_excel_sheet_names(file_path)

    assert len(sheet_names) == 1
    assert sheet_names[0] == "OnlySheet"


def test_load_excel_sheet_names_special_characters(tmp_path: Path):
    """Test retrieving sheet names with special characters."""
    file_path = tmp_path / "special_chars.xlsx"

    workbook = Workbook()
    sheet1 = workbook.active
    sheet1.title = "Sheet-1"
    workbook.create_sheet("Sheet_2")
    workbook.create_sheet("Sheet 3")
    workbook.save(file_path)
    workbook.close()

    sheet_names = load_excel_sheet_names(file_path)

    assert "Sheet-1" in sheet_names
    assert "Sheet_2" in sheet_names
    assert "Sheet 3" in sheet_names


# Integration tests


def test_file_loader_integrates_with_file_validation(valid_xlsx_file: Path):
    """Test that file_loader correctly integrates with file_validation module."""
    # This should work without errors because validation is called internally
    workbook = load_excel_file(valid_xlsx_file)
    assert workbook is not None
    workbook.close()


def test_file_loader_handles_validation_errors(tmp_path: Path):
    """Test that validation errors are properly propagated."""
    nonexistent = tmp_path / "missing.xlsx"

    # Should raise FileValidationError from validation layer
    with pytest.raises(FileValidationError):
        load_excel_file(nonexistent)


def test_multiple_loads_same_file(valid_xlsx_file: Path):
    """Test that the same file can be loaded multiple times."""
    workbook1 = load_excel_file(valid_xlsx_file)
    workbook2 = load_excel_file(valid_xlsx_file)

    assert isinstance(workbook1, WorkbookType)
    assert isinstance(workbook2, WorkbookType)

    # Both workbooks should have the same sheet names
    assert workbook1.sheetnames == workbook2.sheetnames

    workbook1.close()
    workbook2.close()

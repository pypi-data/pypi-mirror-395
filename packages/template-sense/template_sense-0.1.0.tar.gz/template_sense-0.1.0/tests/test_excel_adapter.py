"""
Unit tests for excel_adapter module.

Tests the ExcelWorkbook abstraction layer including:
- Sheet name retrieval
- Cell value access
- Row iteration
- Row and column counts
- Error handling for invalid sheets
- Ensuring openpyxl objects are not exposed
"""

from pathlib import Path

import pytest
from openpyxl import Workbook

from template_sense.adapters.excel_adapter import ExcelWorkbook
from template_sense.errors import ExtractionError, FileValidationError
from template_sense.file_loader import load_excel_file

# Fixtures


@pytest.fixture
def multi_sheet_workbook(tmp_path: Path) -> tuple[Path, Workbook]:
    """Create a multi-sheet workbook for testing."""
    file_path = tmp_path / "multi_sheet.xlsx"

    workbook = Workbook()

    # First sheet
    sheet1 = workbook.active
    sheet1.title = "Sheet1"
    sheet1["A1"] = "Header1"
    sheet1["B1"] = "Header2"
    sheet1["A2"] = "Value1"
    sheet1["B2"] = "Value2"
    sheet1["A3"] = "Value3"
    sheet1["B3"] = "Value4"

    # Second sheet
    sheet2 = workbook.create_sheet("Invoice")
    sheet2["A1"] = "Invoice Number"
    sheet2["B1"] = "Date"
    sheet2["A2"] = "INV-001"
    sheet2["B2"] = "2025-01-15"

    # Third sheet
    sheet3 = workbook.create_sheet("Summary")
    sheet3["A1"] = "Total"
    sheet3["B1"] = 100

    workbook.save(file_path)
    raw_workbook = load_excel_file(file_path)

    yield file_path, raw_workbook

    raw_workbook.close()


@pytest.fixture
def empty_sheet_workbook(tmp_path: Path) -> tuple[Path, Workbook]:
    """Create a workbook with an empty sheet."""
    file_path = tmp_path / "empty_sheet.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "EmptySheet"
    # Don't add any data

    workbook.save(file_path)
    raw_workbook = load_excel_file(file_path)

    yield file_path, raw_workbook

    raw_workbook.close()


@pytest.fixture
def valid_template_workbook() -> Workbook:
    """Load the existing valid_template.xlsx fixture."""
    file_path = Path("tests/fixtures/valid_template.xlsx")
    raw_workbook = load_excel_file(file_path)

    yield raw_workbook

    raw_workbook.close()


# Tests for __init__


def test_excel_workbook_init_valid(multi_sheet_workbook):
    """Test initializing ExcelWorkbook with a valid workbook."""
    _, raw_workbook = multi_sheet_workbook
    wb = ExcelWorkbook(raw_workbook)

    assert wb is not None
    assert hasattr(wb, "_workbook")


def test_excel_workbook_init_none_raises_error():
    """Test that initializing with None raises FileValidationError."""
    with pytest.raises(FileValidationError) as exc_info:
        ExcelWorkbook(None)

    assert "cannot be none" in str(exc_info.value).lower()


# Tests for get_sheet_names


def test_get_sheet_names_multi_sheet(multi_sheet_workbook):
    """Test retrieving sheet names from a multi-sheet workbook."""
    _, raw_workbook = multi_sheet_workbook
    wb = ExcelWorkbook(raw_workbook)

    sheet_names = wb.get_sheet_names()

    assert isinstance(sheet_names, list)
    assert len(sheet_names) == 3
    assert "Sheet1" in sheet_names
    assert "Invoice" in sheet_names
    assert "Summary" in sheet_names


def test_get_sheet_names_order_preserved(multi_sheet_workbook):
    """Test that sheet names are returned in correct order."""
    _, raw_workbook = multi_sheet_workbook
    wb = ExcelWorkbook(raw_workbook)

    sheet_names = wb.get_sheet_names()

    assert sheet_names[0] == "Sheet1"
    assert sheet_names[1] == "Invoice"
    assert sheet_names[2] == "Summary"


def test_get_sheet_names_single_sheet(valid_template_workbook):
    """Test retrieving sheet names from a single-sheet workbook."""
    wb = ExcelWorkbook(valid_template_workbook)

    sheet_names = wb.get_sheet_names()

    assert isinstance(sheet_names, list)
    assert len(sheet_names) == 1
    assert "Test Sheet" in sheet_names


# Tests for iter_rows


def test_iter_rows_basic(multi_sheet_workbook):
    """Test iterating over rows in a sheet."""
    _, raw_workbook = multi_sheet_workbook
    wb = ExcelWorkbook(raw_workbook)

    rows = list(wb.iter_rows("Sheet1"))

    assert len(rows) == 3
    assert rows[0] == ["Header1", "Header2"]
    assert rows[1] == ["Value1", "Value2"]
    assert rows[2] == ["Value3", "Value4"]


def test_iter_rows_with_min_row(multi_sheet_workbook):
    """Test iterating rows starting from a specific row."""
    _, raw_workbook = multi_sheet_workbook
    wb = ExcelWorkbook(raw_workbook)

    rows = list(wb.iter_rows("Sheet1", min_row=2))

    assert len(rows) == 2
    assert rows[0] == ["Value1", "Value2"]
    assert rows[1] == ["Value3", "Value4"]


def test_iter_rows_with_max_row(multi_sheet_workbook):
    """Test iterating rows up to a specific row."""
    _, raw_workbook = multi_sheet_workbook
    wb = ExcelWorkbook(raw_workbook)

    rows = list(wb.iter_rows("Sheet1", max_row=2))

    assert len(rows) == 2
    assert rows[0] == ["Header1", "Header2"]
    assert rows[1] == ["Value1", "Value2"]


def test_iter_rows_returns_values_not_cells(multi_sheet_workbook):
    """Test that iter_rows returns primitive values, not openpyxl Cell objects."""
    _, raw_workbook = multi_sheet_workbook
    wb = ExcelWorkbook(raw_workbook)

    rows = list(wb.iter_rows("Sheet1"))

    # Check first row
    assert isinstance(rows[0], list)
    assert isinstance(rows[0][0], str)  # Primitive type, not Cell
    assert isinstance(rows[0][1], str)


def test_iter_rows_invalid_sheet_raises_error(multi_sheet_workbook):
    """Test that iterating over non-existent sheet raises ExtractionError."""
    _, raw_workbook = multi_sheet_workbook
    wb = ExcelWorkbook(raw_workbook)

    with pytest.raises(ExtractionError) as exc_info:
        list(wb.iter_rows("NonExistentSheet"))

    assert "does not exist" in str(exc_info.value).lower()
    assert "NonExistentSheet" in str(exc_info.value)


def test_iter_rows_empty_sheet(empty_sheet_workbook):
    """Test iterating over an empty sheet."""
    _, raw_workbook = empty_sheet_workbook
    wb = ExcelWorkbook(raw_workbook)

    rows = list(wb.iter_rows("EmptySheet"))

    # Empty sheets may still have one empty row
    assert isinstance(rows, list)


# Tests for get_cell_value


def test_get_cell_value_basic(multi_sheet_workbook):
    """Test retrieving cell values using 1-based indexing."""
    _, raw_workbook = multi_sheet_workbook
    wb = ExcelWorkbook(raw_workbook)

    # Test various cells (1-based indexing)
    assert wb.get_cell_value("Sheet1", 1, 1) == "Header1"
    assert wb.get_cell_value("Sheet1", 1, 2) == "Header2"
    assert wb.get_cell_value("Sheet1", 2, 1) == "Value1"
    assert wb.get_cell_value("Sheet1", 2, 2) == "Value2"


def test_get_cell_value_different_sheet(multi_sheet_workbook):
    """Test retrieving cell values from different sheets."""
    _, raw_workbook = multi_sheet_workbook
    wb = ExcelWorkbook(raw_workbook)

    assert wb.get_cell_value("Invoice", 1, 1) == "Invoice Number"
    assert wb.get_cell_value("Invoice", 2, 1) == "INV-001"
    assert wb.get_cell_value("Summary", 1, 2) == 100


def test_get_cell_value_returns_primitive(multi_sheet_workbook):
    """Test that get_cell_value returns primitive values, not Cell objects."""
    _, raw_workbook = multi_sheet_workbook
    wb = ExcelWorkbook(raw_workbook)

    value = wb.get_cell_value("Sheet1", 1, 1)

    assert isinstance(value, str)  # Primitive type
    assert value == "Header1"


def test_get_cell_value_none_for_empty_cell(multi_sheet_workbook):
    """Test that empty cells return None."""
    _, raw_workbook = multi_sheet_workbook
    wb = ExcelWorkbook(raw_workbook)

    # Access a cell beyond the data
    value = wb.get_cell_value("Sheet1", 10, 10)

    assert value is None


def test_get_cell_value_invalid_sheet_raises_error(multi_sheet_workbook):
    """Test that accessing cell in non-existent sheet raises ExtractionError."""
    _, raw_workbook = multi_sheet_workbook
    wb = ExcelWorkbook(raw_workbook)

    with pytest.raises(ExtractionError) as exc_info:
        wb.get_cell_value("NonExistentSheet", 1, 1)

    assert "does not exist" in str(exc_info.value).lower()


# Tests for get_row_count


def test_get_row_count_basic(multi_sheet_workbook):
    """Test getting row count for a sheet with data."""
    _, raw_workbook = multi_sheet_workbook
    wb = ExcelWorkbook(raw_workbook)

    row_count = wb.get_row_count("Sheet1")

    assert row_count == 3


def test_get_row_count_different_sheets(multi_sheet_workbook):
    """Test getting row count for different sheets."""
    _, raw_workbook = multi_sheet_workbook
    wb = ExcelWorkbook(raw_workbook)

    assert wb.get_row_count("Sheet1") == 3
    assert wb.get_row_count("Invoice") == 2
    assert wb.get_row_count("Summary") == 1


def test_get_row_count_empty_sheet(empty_sheet_workbook):
    """Test getting row count for an empty sheet."""
    _, raw_workbook = empty_sheet_workbook
    wb = ExcelWorkbook(raw_workbook)

    row_count = wb.get_row_count("EmptySheet")

    # Empty sheets in openpyxl have max_row=1 even with no data
    assert row_count >= 0


def test_get_row_count_invalid_sheet_raises_error(multi_sheet_workbook):
    """Test that getting row count for non-existent sheet raises ExtractionError."""
    _, raw_workbook = multi_sheet_workbook
    wb = ExcelWorkbook(raw_workbook)

    with pytest.raises(ExtractionError) as exc_info:
        wb.get_row_count("NonExistentSheet")

    assert "does not exist" in str(exc_info.value).lower()


# Tests for get_column_count


def test_get_column_count_basic(multi_sheet_workbook):
    """Test getting column count for a sheet with data."""
    _, raw_workbook = multi_sheet_workbook
    wb = ExcelWorkbook(raw_workbook)

    col_count = wb.get_column_count("Sheet1")

    assert col_count == 2


def test_get_column_count_different_sheets(multi_sheet_workbook):
    """Test getting column count for different sheets."""
    _, raw_workbook = multi_sheet_workbook
    wb = ExcelWorkbook(raw_workbook)

    assert wb.get_column_count("Sheet1") == 2
    assert wb.get_column_count("Invoice") == 2
    assert wb.get_column_count("Summary") == 2


def test_get_column_count_empty_sheet(empty_sheet_workbook):
    """Test getting column count for an empty sheet."""
    _, raw_workbook = empty_sheet_workbook
    wb = ExcelWorkbook(raw_workbook)

    col_count = wb.get_column_count("EmptySheet")

    # Empty sheets in openpyxl have max_column=1 even with no data
    assert col_count >= 0


def test_get_column_count_invalid_sheet_raises_error(multi_sheet_workbook):
    """Test that getting column count for non-existent sheet raises ExtractionError."""
    _, raw_workbook = multi_sheet_workbook
    wb = ExcelWorkbook(raw_workbook)

    with pytest.raises(ExtractionError) as exc_info:
        wb.get_column_count("NonExistentSheet")

    assert "does not exist" in str(exc_info.value).lower()


# Tests for close


def test_close_workbook(multi_sheet_workbook):
    """Test closing the workbook."""
    _, raw_workbook = multi_sheet_workbook
    wb = ExcelWorkbook(raw_workbook)

    # Should not raise any errors
    wb.close()


# Integration tests


def test_excel_workbook_with_valid_template_fixture(valid_template_workbook):
    """Test ExcelWorkbook with the existing valid_template.xlsx fixture."""
    wb = ExcelWorkbook(valid_template_workbook)

    # Test sheet names
    sheet_names = wb.get_sheet_names()
    assert "Test Sheet" in sheet_names

    # Test row iteration
    rows = list(wb.iter_rows("Test Sheet"))
    assert len(rows) >= 2
    assert rows[0][0] == "Invoice Number"
    assert rows[1][0] == "INV-001"

    # Test cell access
    assert wb.get_cell_value("Test Sheet", 1, 1) == "Invoice Number"
    assert wb.get_cell_value("Test Sheet", 2, 1) == "INV-001"

    # Test row/column counts
    assert wb.get_row_count("Test Sheet") >= 2
    assert wb.get_column_count("Test Sheet") >= 2


def test_excel_workbook_does_not_expose_openpyxl_objects(multi_sheet_workbook):
    """Test that ExcelWorkbook never exposes openpyxl Cell or Worksheet objects."""
    _, raw_workbook = multi_sheet_workbook
    wb = ExcelWorkbook(raw_workbook)

    # Check iter_rows returns primitives
    for row in wb.iter_rows("Sheet1"):
        assert isinstance(row, list)
        for cell_value in row:
            # Should be primitive types (str, int, float, None, etc.), not Cell
            assert not hasattr(cell_value, "value")  # Cell objects have .value attribute

    # Check get_cell_value returns primitives
    cell_value = wb.get_cell_value("Sheet1", 1, 1)
    assert not hasattr(cell_value, "value")

    # Check get_sheet_names returns list of strings
    sheet_names = wb.get_sheet_names()
    for name in sheet_names:
        assert isinstance(name, str)


def test_multiple_operations_on_same_workbook(multi_sheet_workbook):
    """Test that multiple operations can be performed on the same ExcelWorkbook instance."""
    _, raw_workbook = multi_sheet_workbook
    wb = ExcelWorkbook(raw_workbook)

    # Multiple operations
    sheet_names = wb.get_sheet_names()
    rows = list(wb.iter_rows("Sheet1"))
    cell_value = wb.get_cell_value("Invoice", 1, 1)
    row_count = wb.get_row_count("Summary")
    col_count = wb.get_column_count("Sheet1")

    # All should succeed
    assert len(sheet_names) == 3
    assert len(rows) == 3
    assert cell_value == "Invoice Number"
    assert row_count == 1
    assert col_count == 2

    wb.close()


# Tests for hidden content filtering


@pytest.fixture
def hidden_content_workbook(tmp_path: Path) -> tuple[Path, Workbook]:
    """Create a workbook with hidden sheets, rows, and columns."""
    file_path = tmp_path / "hidden_content.xlsx"

    workbook = Workbook()

    # First sheet (visible) with hidden row and column
    sheet1 = workbook.active
    sheet1.title = "VisibleSheet"
    sheet1["A1"] = "H1"
    sheet1["B1"] = "H2"
    sheet1["C1"] = "H3_Hidden"
    sheet1["D1"] = "H4"
    sheet1["A2"] = "V1"
    sheet1["B2"] = "V2"
    sheet1["C2"] = "V3_Hidden"
    sheet1["D2"] = "V4"
    sheet1["A3"] = "V5_HiddenRow"
    sheet1["B3"] = "V6_HiddenRow"
    sheet1["C3"] = "V7_HiddenRow"
    sheet1["D3"] = "V8_HiddenRow"
    sheet1["A4"] = "V9"
    sheet1["B4"] = "V10"
    sheet1["C4"] = "V11_Hidden"
    sheet1["D4"] = "V12"

    # Hide row 3 (1-based)
    sheet1.row_dimensions[3].hidden = True

    # Hide column C
    sheet1.column_dimensions["C"].hidden = True

    # Second sheet (hidden)
    sheet2 = workbook.create_sheet("HiddenSheet")
    sheet2["A1"] = "Should not be visible"
    sheet2.sheet_state = "hidden"

    workbook.save(file_path)
    raw_workbook = load_excel_file(file_path)

    yield file_path, raw_workbook

    raw_workbook.close()


class TestHiddenContentFiltering:
    """Test that hidden sheets, rows, and columns are properly filtered."""

    def test_hidden_sheets_excluded(self, hidden_content_workbook):
        """Hidden sheets should not appear in get_sheet_names()."""
        _, raw_workbook = hidden_content_workbook
        wb = ExcelWorkbook(raw_workbook)

        sheet_names = wb.get_sheet_names()

        # Should only include visible sheets
        assert len(sheet_names) == 1
        assert "VisibleSheet" in sheet_names
        assert "HiddenSheet" not in sheet_names

    def test_get_hidden_rows_returns_correct_set(self, hidden_content_workbook):
        """get_hidden_rows() should return set of hidden row numbers."""
        _, raw_workbook = hidden_content_workbook
        wb = ExcelWorkbook(raw_workbook)

        hidden_rows = wb.get_hidden_rows("VisibleSheet")

        assert isinstance(hidden_rows, set)
        assert 3 in hidden_rows
        assert 1 not in hidden_rows
        assert 2 not in hidden_rows
        assert 4 not in hidden_rows

    def test_get_hidden_columns_returns_correct_set(self, hidden_content_workbook):
        """get_hidden_columns() should return set of hidden column letters."""
        _, raw_workbook = hidden_content_workbook
        wb = ExcelWorkbook(raw_workbook)

        hidden_cols = wb.get_hidden_columns("VisibleSheet")

        assert isinstance(hidden_cols, set)
        assert "C" in hidden_cols
        assert "A" not in hidden_cols
        assert "B" not in hidden_cols
        assert "D" not in hidden_cols

    def test_is_row_hidden_detects_correctly(self, hidden_content_workbook):
        """is_row_hidden() should correctly identify hidden rows."""
        _, raw_workbook = hidden_content_workbook
        wb = ExcelWorkbook(raw_workbook)

        # Row 3 is hidden
        assert wb.is_row_hidden("VisibleSheet", 3) is True

        # Other rows are visible
        assert wb.is_row_hidden("VisibleSheet", 1) is False
        assert wb.is_row_hidden("VisibleSheet", 2) is False
        assert wb.is_row_hidden("VisibleSheet", 4) is False

    def test_is_column_hidden_detects_correctly(self, hidden_content_workbook):
        """is_column_hidden() should correctly identify hidden columns."""
        _, raw_workbook = hidden_content_workbook
        wb = ExcelWorkbook(raw_workbook)

        # Column C is hidden
        assert wb.is_column_hidden("VisibleSheet", "C") is True

        # Other columns are visible
        assert wb.is_column_hidden("VisibleSheet", "A") is False
        assert wb.is_column_hidden("VisibleSheet", "B") is False
        assert wb.is_column_hidden("VisibleSheet", "D") is False

    def test_hidden_rows_utility_with_no_hidden_rows(self, multi_sheet_workbook):
        """get_hidden_rows() should return empty set when no rows are hidden."""
        _, raw_workbook = multi_sheet_workbook
        wb = ExcelWorkbook(raw_workbook)

        hidden_rows = wb.get_hidden_rows("Sheet1")

        assert isinstance(hidden_rows, set)
        assert len(hidden_rows) == 0

    def test_hidden_columns_utility_with_no_hidden_columns(self, multi_sheet_workbook):
        """get_hidden_columns() should return empty set when no columns are hidden."""
        _, raw_workbook = multi_sheet_workbook
        wb = ExcelWorkbook(raw_workbook)

        hidden_cols = wb.get_hidden_columns("Sheet1")

        assert isinstance(hidden_cols, set)
        assert len(hidden_cols) == 0

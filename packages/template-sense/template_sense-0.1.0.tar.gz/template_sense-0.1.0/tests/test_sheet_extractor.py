"""
Unit tests for sheet_extractor module.

Tests the sheet-level extraction utilities including:
- Extracting raw grids from sheets
- Filtering non-empty rows
- Filtering non-empty columns
- Calculating used ranges
- Error handling for invalid sheets
- Ensuring no openpyxl objects are exposed
"""

from pathlib import Path

import pytest
from openpyxl import Workbook

from template_sense.adapters.excel_adapter import ExcelWorkbook
from template_sense.errors import ExtractionError
from template_sense.extraction.sheet_extractor import (
    extract_non_empty_columns,
    extract_non_empty_rows,
    extract_raw_grid,
    get_used_range,
)
from template_sense.file_loader import load_excel_file

# Fixtures


@pytest.fixture
def simple_grid_workbook(tmp_path: Path) -> tuple[Path, Workbook]:
    """Create a workbook with a simple 3x3 grid."""
    file_path = tmp_path / "simple_grid.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SimpleGrid"

    # Create 3x3 grid
    sheet["A1"] = "Header1"
    sheet["B1"] = "Header2"
    sheet["C1"] = "Header3"
    sheet["A2"] = "Value1"
    sheet["B2"] = "Value2"
    sheet["C2"] = "Value3"
    sheet["A3"] = "Value4"
    sheet["B3"] = "Value5"
    sheet["C3"] = "Value6"

    workbook.save(file_path)
    raw_workbook = load_excel_file(file_path)

    yield file_path, raw_workbook

    raw_workbook.close()


@pytest.fixture
def empty_sheet_workbook(tmp_path: Path) -> tuple[Path, Workbook]:
    """Create a workbook with an empty sheet."""
    file_path = tmp_path / "empty_sheet.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "EmptySheet"
    # Don't add any data

    workbook.save(file_path)
    raw_workbook = load_excel_file(file_path)

    yield file_path, raw_workbook

    raw_workbook.close()


@pytest.fixture
def trailing_empty_workbook(tmp_path: Path) -> tuple[Path, Workbook]:
    """Create a workbook with data surrounded by empty rows/columns."""
    file_path = tmp_path / "trailing_empty.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TrailingEmpty"

    # Add data in middle with empty borders
    sheet["B2"] = "A"
    sheet["C2"] = "B"
    sheet["B3"] = "C"
    sheet["C3"] = "D"

    # Leave row 1, column A, and rows/columns after C3 empty

    workbook.save(file_path)
    raw_workbook = load_excel_file(file_path)

    yield file_path, raw_workbook

    raw_workbook.close()


@pytest.fixture
def mixed_empty_workbook(tmp_path: Path) -> tuple[Path, Workbook]:
    """Create a workbook with mixed empty and non-empty rows/columns."""
    file_path = tmp_path / "mixed_empty.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "MixedEmpty"

    # Row 1: has data
    sheet["A1"] = "Header1"
    sheet["B1"] = "Header2"

    # Row 2: empty

    # Row 3: has data
    sheet["A3"] = "Value1"
    sheet["B3"] = "Value2"

    # Column C: empty in all rows

    # Column D: has data
    sheet["D1"] = "Col4"
    sheet["D3"] = "Val4"

    workbook.save(file_path)
    raw_workbook = load_excel_file(file_path)

    yield file_path, raw_workbook

    raw_workbook.close()


@pytest.fixture
def merged_cells_workbook(tmp_path: Path) -> tuple[Path, Workbook]:
    """Create a workbook with merged cells."""
    file_path = tmp_path / "merged_cells.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "MergedCells"

    # Add merged cells
    sheet.merge_cells("A1:B1")
    sheet["A1"] = "Merged Header"

    sheet["A2"] = "Data1"
    sheet["B2"] = "Data2"

    workbook.save(file_path)
    raw_workbook = load_excel_file(file_path)

    yield file_path, raw_workbook

    raw_workbook.close()


# Tests for extract_raw_grid


def test_extract_raw_grid_simple(simple_grid_workbook):
    """Test extracting a simple 3x3 grid."""
    _, raw_workbook = simple_grid_workbook
    wb = ExcelWorkbook(raw_workbook)

    grid = extract_raw_grid(wb, "SimpleGrid")

    assert isinstance(grid, list)
    assert len(grid) == 3
    assert grid[0] == ["Header1", "Header2", "Header3"]
    assert grid[1] == ["Value1", "Value2", "Value3"]
    assert grid[2] == ["Value4", "Value5", "Value6"]


def test_extract_raw_grid_empty_sheet(empty_sheet_workbook):
    """Test extracting from an empty sheet."""
    _, raw_workbook = empty_sheet_workbook
    wb = ExcelWorkbook(raw_workbook)

    grid = extract_raw_grid(wb, "EmptySheet")

    assert isinstance(grid, list)
    # Empty sheets may have one empty row from openpyxl
    # We just verify it's a list and doesn't crash


def test_extract_raw_grid_returns_primitives(simple_grid_workbook):
    """Test that extract_raw_grid returns primitive values, not Cell objects."""
    _, raw_workbook = simple_grid_workbook
    wb = ExcelWorkbook(raw_workbook)

    grid = extract_raw_grid(wb, "SimpleGrid")

    for row in grid:
        assert isinstance(row, list)
        for cell in row:
            # Should be primitive types (str, int, float, None, etc.)
            # Not openpyxl Cell objects (which have .value attribute)
            assert not hasattr(cell, "value")


def test_extract_raw_grid_invalid_sheet_raises_error(simple_grid_workbook):
    """Test that extracting from non-existent sheet raises ExtractionError."""
    _, raw_workbook = simple_grid_workbook
    wb = ExcelWorkbook(raw_workbook)

    with pytest.raises(ExtractionError) as exc_info:
        extract_raw_grid(wb, "NonExistentSheet")

    assert "does not exist" in str(exc_info.value).lower()


def test_extract_raw_grid_with_trailing_empty(trailing_empty_workbook):
    """Test extracting grid with trailing empty rows/columns."""
    _, raw_workbook = trailing_empty_workbook
    wb = ExcelWorkbook(raw_workbook)

    grid = extract_raw_grid(wb, "TrailingEmpty")

    # Grid includes empty rows/columns as openpyxl sees them
    assert isinstance(grid, list)
    assert len(grid) >= 3  # Should have at least 3 rows


def test_extract_raw_grid_with_merged_cells(merged_cells_workbook):
    """Test extracting grid with merged cells."""
    _, raw_workbook = merged_cells_workbook
    wb = ExcelWorkbook(raw_workbook)

    grid = extract_raw_grid(wb, "MergedCells")

    assert isinstance(grid, list)
    assert len(grid) >= 2
    # Merged cells: only top-left cell has value, others are None
    assert grid[0][0] == "Merged Header"


# Tests for extract_non_empty_rows


def test_extract_non_empty_rows_basic():
    """Test filtering non-empty rows from a grid."""
    grid = [
        ["Header1", "Header2"],
        ["Value1", "Value2"],
        [None, None],
        ["Value3", "Value4"],
        ["", ""],
    ]

    result = extract_non_empty_rows(grid)

    assert len(result) == 3
    assert result[0] == ["Header1", "Header2"]
    assert result[1] == ["Value1", "Value2"]
    assert result[2] == ["Value3", "Value4"]


def test_extract_non_empty_rows_all_empty():
    """Test filtering when all rows are empty."""
    grid = [[None, None], ["", ""], [None, ""]]

    result = extract_non_empty_rows(grid)

    assert result == []


def test_extract_non_empty_rows_all_non_empty():
    """Test filtering when all rows have data."""
    grid = [["A", "B"], ["C", "D"], ["E", "F"]]

    result = extract_non_empty_rows(grid)

    assert len(result) == 3
    assert result == grid


def test_extract_non_empty_rows_partial_empty_cells():
    """Test that rows with at least one non-empty cell are kept."""
    grid = [
        ["A", None],
        [None, "B"],
        [None, None],
        ["", "C"],
    ]

    result = extract_non_empty_rows(grid)

    assert len(result) == 3
    assert result[0] == ["A", None]
    assert result[1] == [None, "B"]
    assert result[2] == ["", "C"]


def test_extract_non_empty_rows_empty_grid():
    """Test filtering an empty grid."""
    grid = []

    result = extract_non_empty_rows(grid)

    assert result == []


# Tests for extract_non_empty_columns


def test_extract_non_empty_columns_basic():
    """Test filtering non-empty columns from a grid."""
    grid = [
        ["A", None, "B", ""],
        ["C", None, "D", ""],
    ]

    result = extract_non_empty_columns(grid)

    assert len(result) == 2
    assert result[0] == ["A", "B"]
    assert result[1] == ["C", "D"]


def test_extract_non_empty_columns_all_empty():
    """Test filtering when all columns are empty."""
    grid = [
        [None, "", None],
        [None, "", None],
    ]

    result = extract_non_empty_columns(grid)

    assert result == [[], []]


def test_extract_non_empty_columns_all_non_empty():
    """Test filtering when all columns have data."""
    grid = [["A", "B", "C"], ["D", "E", "F"]]

    result = extract_non_empty_columns(grid)

    assert len(result) == 2
    assert result[0] == ["A", "B", "C"]
    assert result[1] == ["D", "E", "F"]


def test_extract_non_empty_columns_partial_empty_cells():
    """Test that columns with at least one non-empty cell are kept."""
    grid = [
        ["A", None, "", "D"],
        [None, None, "C", "E"],
    ]

    result = extract_non_empty_columns(grid)

    # Column 0: ["A", None] - has A, keep
    # Column 1: [None, None] - all empty, remove
    # Column 2: ["", "C"] - has C, keep
    # Column 3: ["D", "E"] - has both, keep
    # Result should have 3 columns: 0, 2, 3
    assert len(result) == 2  # 2 rows
    assert len(result[0]) == 3  # 3 columns
    assert result[0] == ["A", "", "D"]
    assert result[1] == [None, "C", "E"]


def test_extract_non_empty_columns_empty_grid():
    """Test filtering an empty grid."""
    grid = []

    result = extract_non_empty_columns(grid)

    assert result == []


def test_extract_non_empty_columns_inconsistent_row_lengths():
    """Test filtering with inconsistent row lengths."""
    grid = [
        ["A", "B", "C"],
        ["D"],
        ["E", "F"],
    ]

    result = extract_non_empty_columns(grid)

    # All columns have at least one non-empty value
    assert len(result) == 3
    assert result[0] == ["A", "B", "C"]
    assert result[1] == ["D", None, None]  # Padded with None
    assert result[2] == ["E", "F", None]  # Padded with None


def test_extract_non_empty_columns_empty_rows():
    """Test filtering when grid has empty row lists."""
    grid = [[], [], []]

    result = extract_non_empty_columns(grid)

    assert result == []


# Tests for get_used_range


def test_get_used_range_simple(simple_grid_workbook):
    """Test getting used range for a simple 3x3 grid."""
    _, raw_workbook = simple_grid_workbook
    wb = ExcelWorkbook(raw_workbook)

    min_row, max_row, min_col, max_col = get_used_range(wb, "SimpleGrid")

    assert min_row == 1
    assert max_row == 3
    assert min_col == 1
    assert max_col == 3


def test_get_used_range_empty_sheet(empty_sheet_workbook):
    """Test getting used range for an empty sheet."""
    _, raw_workbook = empty_sheet_workbook
    wb = ExcelWorkbook(raw_workbook)

    min_row, max_row, min_col, max_col = get_used_range(wb, "EmptySheet")

    # Empty sheets return (1, 1, 1, 1)
    assert min_row == 1
    assert max_row == 1
    assert min_col == 1
    assert max_col == 1


def test_get_used_range_with_trailing_empty(trailing_empty_workbook):
    """Test getting used range excluding empty borders."""
    _, raw_workbook = trailing_empty_workbook
    wb = ExcelWorkbook(raw_workbook)

    min_row, max_row, min_col, max_col = get_used_range(wb, "TrailingEmpty")

    # Data is in B2:C3, so:
    # min_row=2, max_row=3, min_col=2, max_col=3
    assert min_row == 2
    assert max_row == 3
    assert min_col == 2
    assert max_col == 3


def test_get_used_range_with_mixed_empty(mixed_empty_workbook):
    """Test getting used range with mixed empty rows/columns."""
    _, raw_workbook = mixed_empty_workbook
    wb = ExcelWorkbook(raw_workbook)

    min_row, max_row, min_col, max_col = get_used_range(wb, "MixedEmpty")

    # Data spans rows 1-3 and columns 1,2,4 (A,B,D)
    # min_row=1, max_row=3, min_col=1, max_col=4
    assert min_row == 1
    assert max_row == 3
    assert min_col == 1
    assert max_col == 4


def test_get_used_range_returns_1_based_indices(simple_grid_workbook):
    """Test that get_used_range returns 1-based indices (Excel convention)."""
    _, raw_workbook = simple_grid_workbook
    wb = ExcelWorkbook(raw_workbook)

    min_row, max_row, min_col, max_col = get_used_range(wb, "SimpleGrid")

    # All indices should be >= 1 (1-based)
    assert min_row >= 1
    assert max_row >= 1
    assert min_col >= 1
    assert max_col >= 1


def test_get_used_range_invalid_sheet_raises_error(simple_grid_workbook):
    """Test that getting used range for non-existent sheet raises ExtractionError."""
    _, raw_workbook = simple_grid_workbook
    wb = ExcelWorkbook(raw_workbook)

    with pytest.raises(ExtractionError) as exc_info:
        get_used_range(wb, "NonExistentSheet")

    assert "does not exist" in str(exc_info.value).lower()


def test_get_used_range_single_cell(tmp_path: Path):
    """Test getting used range for a sheet with single cell."""
    file_path = tmp_path / "single_cell.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SingleCell"
    sheet["C3"] = "SingleValue"

    workbook.save(file_path)
    raw_workbook = load_excel_file(file_path)
    wb = ExcelWorkbook(raw_workbook)

    min_row, max_row, min_col, max_col = get_used_range(wb, "SingleCell")

    # Single cell at C3 (row=3, col=3)
    assert min_row == 3
    assert max_row == 3
    assert min_col == 3
    assert max_col == 3

    raw_workbook.close()


# Integration tests


def test_full_extraction_workflow(mixed_empty_workbook):
    """Test a complete extraction workflow with all utilities."""
    _, raw_workbook = mixed_empty_workbook
    wb = ExcelWorkbook(raw_workbook)

    # 1. Extract raw grid
    grid = extract_raw_grid(wb, "MixedEmpty")
    assert len(grid) >= 3

    # 2. Filter non-empty rows
    non_empty_rows = extract_non_empty_rows(grid)
    assert len(non_empty_rows) == 2  # Rows 1 and 3 have data

    # 3. Filter non-empty columns
    non_empty_cols = extract_non_empty_columns(grid)
    # Columns A, B, D have data (column C is empty)
    assert all(len(row) == 3 for row in non_empty_cols)

    # 4. Get used range
    min_row, max_row, min_col, max_col = get_used_range(wb, "MixedEmpty")
    assert min_row == 1
    assert max_row == 3
    assert min_col == 1
    assert max_col == 4


def test_no_openpyxl_objects_in_results(simple_grid_workbook):
    """Test that no openpyxl objects leak into any extraction results."""
    _, raw_workbook = simple_grid_workbook
    wb = ExcelWorkbook(raw_workbook)

    # Extract raw grid
    grid = extract_raw_grid(wb, "SimpleGrid")
    for row in grid:
        for cell in row:
            assert not hasattr(cell, "value")  # No Cell objects

    # Filter non-empty rows
    non_empty_rows = extract_non_empty_rows(grid)
    for row in non_empty_rows:
        for cell in row:
            assert not hasattr(cell, "value")

    # Filter non-empty columns
    non_empty_cols = extract_non_empty_columns(grid)
    for row in non_empty_cols:
        for cell in row:
            assert not hasattr(cell, "value") if cell is not None else True

    # Get used range returns tuple of ints
    result = get_used_range(wb, "SimpleGrid")
    assert isinstance(result, tuple)
    assert all(isinstance(val, int) for val in result)


def test_extraction_with_valid_template_fixture():
    """Test extraction utilities with the existing valid_template.xlsx fixture."""
    file_path = Path("tests/fixtures/valid_template.xlsx")
    raw_workbook = load_excel_file(file_path)
    wb = ExcelWorkbook(raw_workbook)

    # Extract raw grid
    grid = extract_raw_grid(wb, "Test Sheet")
    assert len(grid) >= 2
    assert grid[0][0] == "Invoice Number"

    # Filter non-empty rows
    non_empty_rows = extract_non_empty_rows(grid)
    assert len(non_empty_rows) >= 2

    # Get used range
    min_row, max_row, min_col, max_col = get_used_range(wb, "Test Sheet")
    assert min_row >= 1
    assert max_row >= 2
    assert min_col >= 1
    assert max_col >= 2

    raw_workbook.close()


# Tests for hidden content filtering integration


@pytest.fixture
def hidden_content_workbook_for_extraction(tmp_path: Path) -> tuple[Path, Workbook]:
    """Create a workbook with hidden rows and columns for extraction testing."""
    file_path = tmp_path / "hidden_extraction.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TestSheet"

    # Create 4x4 grid
    sheet["A1"] = "H1"
    sheet["B1"] = "H2"
    sheet["C1"] = "H3_Hidden"
    sheet["D1"] = "H4"
    sheet["A2"] = "V1"
    sheet["B2"] = "V2"
    sheet["C2"] = "V3_Hidden"
    sheet["D2"] = "V4"
    sheet["A3"] = "V5_HiddenRow"
    sheet["B3"] = "V6_HiddenRow"
    sheet["C3"] = "V7_HiddenRow"
    sheet["D3"] = "V8_HiddenRow"
    sheet["A4"] = "V9"
    sheet["B4"] = "V10"
    sheet["C4"] = "V11_Hidden"
    sheet["D4"] = "V12"

    # Hide row 3
    sheet.row_dimensions[3].hidden = True

    # Hide column C
    sheet.column_dimensions["C"].hidden = True

    workbook.save(file_path)
    raw_workbook = load_excel_file(file_path)

    yield file_path, raw_workbook

    raw_workbook.close()


def test_extract_raw_grid_excludes_hidden_rows(hidden_content_workbook_for_extraction):
    """extract_raw_grid() should exclude hidden rows from the grid."""
    _, raw_workbook = hidden_content_workbook_for_extraction
    wb = ExcelWorkbook(raw_workbook)

    grid = extract_raw_grid(wb, "TestSheet")

    # Should have 3 rows (rows 1, 2, 4), not 4 (row 3 is hidden)
    assert len(grid) == 3
    assert grid[0][0] == "H1"  # Row 1
    assert grid[1][0] == "V1"  # Row 2
    assert grid[2][0] == "V9"  # Row 4 (row 3 was hidden)


def test_extract_raw_grid_excludes_hidden_columns(hidden_content_workbook_for_extraction):
    """extract_raw_grid() should exclude hidden columns from each row."""
    _, raw_workbook = hidden_content_workbook_for_extraction
    wb = ExcelWorkbook(raw_workbook)

    grid = extract_raw_grid(wb, "TestSheet")

    # Each row should have 3 values (columns A, B, D), not 4 (column C is hidden)
    assert all(len(row) == 3 for row in grid)
    assert grid[0] == ["H1", "H2", "H4"]  # Row 1, column C excluded
    assert grid[1] == ["V1", "V2", "V4"]  # Row 2, column C excluded
    assert grid[2] == ["V9", "V10", "V12"]  # Row 4, column C excluded


def test_extract_raw_grid_with_both_hidden_rows_and_columns(
    hidden_content_workbook_for_extraction,
):
    """extract_raw_grid() should handle both hidden rows and columns simultaneously."""
    _, raw_workbook = hidden_content_workbook_for_extraction
    wb = ExcelWorkbook(raw_workbook)

    grid = extract_raw_grid(wb, "TestSheet")

    # Should have 3 rows, each with 3 columns
    assert len(grid) == 3
    for row in grid:
        assert len(row) == 3

    # Verify row 3 data is not in grid
    for row in grid:
        assert "V5_HiddenRow" not in row
        assert "V6_HiddenRow" not in row
        assert "V8_HiddenRow" not in row

    # Verify column C data is not in grid
    for row in grid:
        assert "H3_Hidden" not in row
        assert "V3_Hidden" not in row
        assert "V11_Hidden" not in row


def test_extract_raw_grid_with_no_hidden_content(simple_grid_workbook):
    """extract_raw_grid() should work normally when no content is hidden."""
    _, raw_workbook = simple_grid_workbook
    wb = ExcelWorkbook(raw_workbook)

    grid = extract_raw_grid(wb, "SimpleGrid")

    # Should have all 3 rows and 3 columns
    assert len(grid) == 3
    assert all(len(row) == 3 for row in grid)


def test_get_used_range_with_hidden_content(hidden_content_workbook_for_extraction):
    """get_used_range() should calculate range based on visible content only."""
    _, raw_workbook = hidden_content_workbook_for_extraction
    wb = ExcelWorkbook(raw_workbook)

    min_row, max_row, min_col, max_col = get_used_range(wb, "TestSheet")

    # Hidden content is excluded by extract_raw_grid, so used range reflects visible content
    # Visible rows: 1, 2, 4 (but indices are renumbered in grid: 1, 2, 3)
    # Visible columns: A, B, D (but indices are renumbered in grid: 1, 2, 3)
    assert min_row == 1
    assert max_row == 3  # 3 visible rows
    assert min_col == 1
    assert max_col == 3  # 3 visible columns


def test_extract_raw_grid_with_complex_hidden_content():
    """Test with complex fixture having multiple hidden sheets, rows, and columns (BAT-49).

    This test uses hidden_content_test.xlsx which replicates the structure of real
    invoice templates but with sanitized dummy data. It has:
    - 3 sheets total: 1 visible ("CO"), 2 hidden ("Invoice (CO)", "Price List")
    - Hidden row 39 in visible sheet
    - Hidden columns L, M in hidden sheet
    """
    file_path = Path("tests/fixtures/hidden_content_test.xlsx")
    if not file_path.exists():
        pytest.skip("hidden_content_test.xlsx fixture not found")

    raw_workbook = load_excel_file(file_path)
    wb = ExcelWorkbook(raw_workbook)

    # Test 1: Verify only visible sheets are returned
    sheet_names = wb.get_sheet_names()
    assert "CO" in sheet_names
    assert "Invoice (CO)" not in sheet_names  # Hidden sheet
    assert "Price List" not in sheet_names  # Hidden sheet

    # Test 2: Verify hidden row 39 is detected
    hidden_rows = wb.get_hidden_rows("CO")
    assert 39 in hidden_rows

    # Test 3: Extract grid and verify hidden row is excluded
    grid = extract_raw_grid(wb, "CO")

    # Original sheet has 45 rows, with row 39 hidden
    # After filtering, should have 44 rows (45 - 1 header excluded from some counts, but we include it)
    # Grid should have 44 visible rows (1-38, 40-45)
    assert isinstance(grid, list)
    assert len(grid) == 44  # 45 total rows - 1 hidden row (39)

    # Row 39 had specific data pattern - verify it's not in extracted grid
    # We look for the row that would be row 39 (with value "INV-1039")
    row_39_data = [cell for row in grid for cell in row if cell == "INV-1039"]
    assert len(row_39_data) == 0, "Hidden row 39 data should not appear in extracted grid"

    # Test 4: Verify hidden columns in hidden sheet
    hidden_cols = wb.get_hidden_columns("Invoice (CO)")
    assert "L" in hidden_cols
    assert "M" in hidden_cols

    raw_workbook.close()

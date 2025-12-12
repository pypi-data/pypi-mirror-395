# -*- coding: utf-8 -*-

from os.path import basename
import json, pymysql, re, os
from typing import Union, List, Dict, Any
from loguru import logger
from sshtunnel import SSHTunnelForwarder
from pymysql import Error
from .翻译 import translate
from .其他 import 英文表头对应中文
import sys

# from random import randint
# from hashlib import md5
# from requests import get

# def translate(text):
#     """外语翻译成汉语
#     :param text:要翻译成汉语的外语
#     :return: 返回汉语
#     """
#     try:
#         appid = '20220712001270949'
#         q = str(text)
#         secret_key = 'ODwtPobgXFes3sBML_NM'
#         salt = str(randint(1000000000, 9999999999))
#         m = md5()
#         m.update((appid + q + salt + secret_key).encode("utf8"))
#         s = f'http://api.fanyi.baidu.com/api/trans/vip/translate?q={q}&from=en&to=zh&appid={appid}&salt={salt}&sign={m.hexdigest()}'
#         result = get(s).json()["trans_result"][0]['dst']
#         return result
#     except:
#         return text

def prvadepl(num):
    """保留有效的小数位
    :param num:
    :return: 删除无效的小数位的数字
    """
    return eval(str(num).rstrip('0').rstrip('.')) if type(num) == float else num


# 删除字符串首尾看不见的字符，将中文括号改为英文括号，以及其他符号的替换
def optstr(values: Union[str, list, dict]):
    def format_str(s):
        s = s.strip()
        return s.replace('（', '(').replace('）', ')').replace('\t', '').replace('\n', '').replace(r'\n', '').replace(
            '\r', '').replace(' ', '').replace('\xa0', ' ').replace('★', ' ').replace('☆', ' ').replace('▲',
                                                                                                        ' ').replace(
            '\u3000', '').replace('／', '/').replace('［','[').replace('］',']').replace('，',',').replace('．','.')

    if not values:
        return values
    elif type(values) == str:
        return format_str(values)
    elif type(values) == list:
        for i, value in enumerate(values):
            value = prvadepl(value)
            if type(value) == str:
                values[i] = format_str(value)
    elif type(values) == dict:
        for i, value in enumerate(values.values()):
            value = prvadepl(value)
            if type(value) == str:
                values.values()[i] = format_str(value)
    return values


def read_excel(file_path):
    """读取.xlsx、.xls以及htm格式的文件数据，默认读取第一个sheet
    :param file_path:文件路径
    :return:表格数据
    """
    datas = []
    if file_path.endswith('.xlsx'):
        from openpyxl import load_workbook
        import zipfile
        try:
            workbook = load_workbook(filename=file_path, read_only=True)
            sheet_names = workbook.sheetnames
            sheet = workbook[sheet_names[0]]
            for row in sheet.iter_rows():
                datas.append([prvadepl(cell.value) if isinstance(cell.value, float) else cell.value for cell in row])
        except zipfile.BadZipFile:
            from txdpy import webptablesl
            with open(file_path, 'r', encoding='utf-8') as f:
                datas = webptablesl(f.read(), '//table')
    elif file_path.endswith('.xls'):
        import xlrd
        workbook = xlrd.open_workbook(file_path)
        sheet = workbook.sheet_by_index(0)
        for row_idx in range(sheet.nrows):
            datas.append([prvadepl(cell) if isinstance(cell, float) else cell for cell in sheet.row_values(row_idx)])
    if datas:
        i = [i for i, l in enumerate(datas[0] + [None]) if not l and datas[0][i - 1]]

        data = []
        if i and len(datas[0]) > 100:
            i = i[-1]
            for l in datas:
                if any(l): data.append(l[:i])
        else:
            for l in datas:
                if any(l): data.append(l)
        return data
    return []


def gen_excel(data: list, save_path, header=None, is_optstr=True):
    """生成.xlsx数据文件
    :param data:二维列表数据
    :param save_path:保存为.xlsx文件路径
    :param header:表头以列表形式传入，非必要参数
    :param is_optstr:文本格式整理默认为 True
    """
    import xlsxwriter
    from tqdm import tqdm
    if header:
        data = list(data)
        data.insert(0, header)
    workbook = xlsxwriter.Workbook(save_path if save_path.endswith('.xlsx') else f'{save_path}.xlsx')
    sheet = workbook.add_worksheet()
    for row_num, row_data in enumerate(tqdm(data, desc=f'{basename(save_path)}')):
        for col_num, col_value in enumerate(row_data):
            if is_optstr and type(col_value) is str:
                col_value = col_value.strip().replace('\n', '')
            elif type(col_value) in [dict, list]:
                try:
                    col_value = json.dumps(col_value, ensure_ascii=False)
                except Exception as e:
                    print(e)
                    col_value = str(col_value)
            elif type(col_value) in [set, tuple]:
                col_value = str(col_value)
            sheet.write(row_num, col_num, col_value)
    workbook.close()


class ReadData:
    """
    读取mysql和.xlsx文件数据，提供一些方便的方法
    """

    def __init__(self, xlsx_mysql_sql: Union[str, List[List]], needful_field: List[str] = None, select_sql=None,
                 replace_th=True, mysql_type=1):
        """
        :param name:二维列表数据或者.xlsx文件路径或者mysql数据库数据表名称信息关键字或者mysql查询语句  比如：[['姓名','年齡'],['张三',26]]/内蒙古计划.xlsx/陕西计划、院校数据/select * from bk_school
        :param needful_field:需要保留的数据字段,只支持中文,表头字段名称会被(强制)自动转为中文表头名称,sql语句查询有筛选字段时,不支持保留字段  比如：['院校名称','专业名称']
        :param select_sql:查询数据库数据表名时筛选部分的sql语句  比如：year=2022 and school_name like '%北%大学'
        :param replace_th:是否替换表头字段为中文，默认为True
        :param mysql_type:1是正式库，2是测试库
        """
        self.data = []  # 所有数据以二维列表存放，不包括表头
        self.replace_th = replace_th
        self.th_dict = None
        with open('c:/mysql_config.json', 'r', encoding='utf-8') as f:
            self.mysql_config = json.load(f)
        self.connect_mysql()  # 连接mysql数据库

        data_info = xlsx_mysql_sql
        self.xlsx_mysql_sql = xlsx_mysql_sql
        self.mysql_type = mysql_type
        if not data_info:
            raise ValueError(f'传入参数为空值，xlsx_mysql_sql：{xlsx_mysql_sql}')
        if type(data_info) in (list, tuple):
            if all(type(sub_lst) in (list, tuple) for sub_lst in data_info):
                self.data = data_info
                self.data[0] = [str(f) for f in self.data[0]]  # 数据第一行必须作为列索引且为字段名称类型字符串
            else:
                raise ValueError('数据必须是二维列表/元组')
        elif data_info.endswith('.xlsx') or data_info.endswith('.xls'):  # 判断是否为.xlsx文件读取数据
            self.read_excel(data_info)

        else:
            self.read_mysql(data_info, select_sql)  # 当做查询数据库数据

        if not self.data:
            raise ValueError('表格数据为空')

        self.columns = self.replace_th_name(self.data[0])  # 替换表头字段名称为中文
        self.data = self.data[1:]  # 所有数据以二维列表存放，不包括表头

        self.columns_index = {field: i for i, field in enumerate(self.columns)}  # 创建列索引

        if needful_field:  # 判断是否有需要保留的字段，按保留的字段保留数据
            self.reserve_needful_field(needful_field)

        for k, v in self.columns_index.items():
            try:
                exec(f'self.{k}={v}')
            except:
                pass

        self.len_row = len(self.data)  # 行数
        self.len_col = len(self.columns)  # 列数

    def connect_mysql(self):
        self.db = pymysql.connect(host=self.mysql_config["1"]['host'], port=3306, user='root',
                                  password=self.mysql_config["1"]['password'],
                                  database=self.mysql_config["1"]['database'])
        self.cursor = self.db.cursor()
        self.cursor.execute("""select * from 公司表名称对应关系""")
        self.table_dict = {k: v for k, v in (v[0].split(':') for v in self.cursor.fetchall())}

    def select_mysql_data(self, sql, table_name=None):
        if not table_name: table_name = self.extract_table_name(sql)
        if table_name in self.table_dict.values():
            database = 'bk179saas'
            if any([x in self.xlsx_mysql_sql for x in
                    ['计划', '智能库', '专业分']]) and '单招' not in self.xlsx_mysql_sql:
                database = 'bk179saas_test'
            if self.mysql_type == 2:
                database = 'bk179saas_test'
            if 'information_schema.COLUMNS' in sql:
                sql = sql.replace('需要替换的数据库名称', database)
            db = pymysql.connect(host=self.mysql_config["2"]['host'], port=3306, user=self.mysql_config["2"]['user'],
                                 password=self.mysql_config["2"]['passwd'], database=database)
            cursor = db.cursor()
            cursor.execute(sql)
            data = cursor.fetchall()
            db.close()
            cursor.close()
            return data
        if 'information_schema.COLUMNS' in sql:
            sql = sql.replace('需要替换的数据库名称', self.mysql_config["1"]['database'])
        self.cursor.execute(sql)
        return self.cursor.fetchall()

    def get_th(self, table_name):  # 获取表头字段
        table_name = self.table_dict.get(table_name, table_name)
        table_header = self.select_mysql_data(f"""SELECT COLUMN_NAME   
                                FROM information_schema.COLUMNS
                                WHERE TABLE_SCHEMA = '需要替换的数据库名称' AND TABLE_NAME = '{table_name}'   
                                ORDER BY ORDINAL_POSITION ASC;""", table_name)
        return [k[0] for k in table_header]

    def replace_th_name(self, lit):  # 替换表头名称
        if not self.replace_th: return lit
        # if not self.th_dict:
            # self.th_dict = {k: v for k, v in (v[0].split(':') for v in self.select_mysql_data("""select * from 表头字段对应名称"""))}
            # self.th_dict = 英文表头对应中文字典
        # new_name = []
        # names = []
        # for name in lit:
        #     if name in names:
        #         name += str(names.count(name))
        #     new_name.append(self.th_dict.get(name, name))
        #     names.append(name)
        return 英文表头对应中文(lit)

    # 提取数据表名称
    def extract_table_name(self, sql):
        return re.search(r'from\s([^\s]+)', sql, re.IGNORECASE).group(1).split('.')[-1].strip('`')

    def read_mysql(self, name, select_sql):
        if name[:6].lower() == 'select':
            table_name = self.extract_table_name(name)
            self.table_name = table_name
            table_name = self.table_dict.get(table_name, table_name)
            sql = re.sub(r'(?i)(FROM\s+)[^\s]+', r'\1' + f'`{table_name}`', name, 1)
            logger.info(sql)
            table_th = re.search(r'select\s(.+?)\sfrom', name, re.IGNORECASE).group(1)
            if table_th == '*':
                self.data.append(self.replace_th_name(self.get_th(table_name)))
            else:
                self.data.append(
                    self.replace_th_name(table_th.replace('`', '').replace('"', '').replace("'", '').split(',')))
            self.data += [list(v) for v in self.select_mysql_data(sql)]
        else:
            self.table_name = name
            table_name = self.table_dict.get(name, name)
            self.data.append(self.replace_th_name(self.get_th(table_name)))
            sql = f"""select * from `{table_name}`{' where ' + select_sql if select_sql else ''}"""
            # logger.info(sql)
            self.data += [list(v) for v in self.select_mysql_data(sql)]

    # 读取execl数据
    def read_excel(self, file_path):
        """
        :param file_path:文件路径
        return:表格数据
        """
        self.data = read_excel(file_path)
        self.data[0] = [str(f) for f in self.data[0]]  # 数据第一行必须作为列索引且为字段名称类型字符串

    def reserve_needful_field(self, needful_field):
        """
        保留数据所需字段
        """
        self.columns = [self.columns[self.columns_index[field]] for field in needful_field]

        data = []
        for row in self.data:
            col = []
            for field in needful_field:
                col.append(row[self.columns_index[field]])
            data.append(col)
        self.data = data
        self.columns_index = {field: i for i, field in enumerate(self.columns)}  # 创建列索引

    def sel_cl(self, row_mark: str = None, column_mark: str = None, da_ts=1):
        """
        筛选行和列
        :param row_mark:行索引只支持自然数 比如'1,10-20'
        :param column_mark:列索引支持自然数和字段名 比如'1,10-20','1,院校名称,专业代码-最低分,14'
        :param da_ts:默认为1横向显示数据，2为纵向显示数据
        """
        is_num = lambda s: type(s) == int or re.search('^([0-9]+)$', str(s))

        def diiini(i):
            if i == '':
                i = self.len_col
            elif not is_num(i):
                i = self.columns_index.get(i)
                if not i:
                    raise ValueError(f'传入索引错误{i}')
            return int(i)

        def index_split(mark, typ):
            rcis = []
            for i1 in mark.split(','):
                i2 = i1.split('-')
                if len(i2) == 1:
                    rcis.append(diiini(i2[0]) if typ == 'col' else int(i2[0]))
                else:
                    [rcis.append(i) for i in range((diiini(i2[0]) if typ == 'col' else int(i2[0] or 0)), (
                        diiini(i2[1]) if typ == 'col' else int(i2[1] or self.len_row)) + 1 or 0)]
            return sorted(list(set(rcis)))

        ris = index_split(row_mark, 'row') if row_mark else range(self.len_row + 1)
        cis = index_split(column_mark, 'col') if column_mark else range(self.len_col + 1)

        if da_ts == 1:
            for ri in ris:
                if ri < self.len_row - 1:
                    row_data = []
                    for ci in cis:
                        if ci < self.len_col:
                            row_data.append(self.data[ri + 1][ci])
                    yield row_data
        else:
            for ci in cis:
                if ci < self.len_col:
                    row_data = []
                    for ri in ris:
                        if ri < self.len_row - 1:
                            row_data.append(self.data[ri + 1][ci])
                    yield row_data

    def group(self, *args):  # 分组键值应该使用元祖类型
        data = {}
        for row in self.data:
            key = tuple([row[self.columns_index[arg]] for arg in args])
            if len(key) == 1:
                key = key[0]
            if key in data:
                data[key].append(row)
            else:
                data[key] = [row]
        return data

    def group1(self, *args):
        data = {}
        for row in self.data:
            key = ','.join([str(row[self.columns_index[arg]]) for arg in args])
            if key in data:
                data[key].append(row)
            else:
                data[key] = [row]
        return data

    def save(self, fp=None):
        if not fp:
            fp = self.table_name
        gen_excel(self.data, fp, self.columns)

    def __del__(self):
        if self.db:
            try:
                self.db.close()
                self.cursor.close()
            except:
                pass


class MysqlConn:
    """
    连接Mysql数据库
    """

    def __init__(self, database, config: list = []):
        """连接Mysql数据库
        :param database 数据库名
        :param config 数据库配置[host, port, user, password]
        """
        self.config = config
        self.database = database
        self.table_tf = False
        self.cursor = None
        self.db = None
        self.conn_mysql()

    def __enter__(self):
        return self

    def conn_mysql(self):
        if self.config:
            config = {
                'host': self.config[0],
                'port': self.config[1],
                'user': self.config[2],
                'password': self.config[3],
                'db': self.database,
            }
            self.db = pymysql.connect(**config)  # 连接数据库

            self.cursor = self.db.cursor()  # 创建游标，用于执行SQL语句
            logger.info("连接数据库成功。")

        else:
            with open('c:/mysql_config.json', 'r', encoding='utf-8') as f:
                self.mysql_config = json.load(f)
            try:
                self.db = pymysql.connect(host=self.mysql_config["2"]['host'], port=3306,
                                          user=self.mysql_config["2"]['user'],
                                          password=self.mysql_config["2"]['passwd'], database=self.database)
                self.cursor = self.db.cursor()
                logger.success("成功连接到MySQL数据库")
            except Error as e:
                logger.error(f"连接失败: {translate(e)}")
                sys.exit()

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()

    def close(self):
        if self.cursor:
            self.cursor.close()
        if self.db:
            self.db.close()
        logger.info(f'已断开与MySQL数据库的连接')

    # 获取所有字段
    def get_table_columns(self, table_name):
        """  mysql获取表字段，并按照列的位置排序
        :param table_name 表名
        :return: 列表格式字段名称，按列位置排序，列表格式
        """
        sql = """
            SELECT COLUMN_NAME 
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s
            ORDER BY ORDINAL_POSITION
            """
        self.cursor.execute(sql, (self.database, table_name))
        columns = [column[0] for column in self.cursor.fetchall()]

        return columns

    def create_table(self, table_name, columns, data_pass):
        """ mysql创建数据表
        :param table_name 表名，注意：若数据库中有同名数据表，会先删除数据库中同名数据表
        :param columns 字段名称, 当传入列表时，默认所有字段都为varchar(225);当传入字典时，可自定义字段类型，例如：{"name":"varchar(225)", "age":"int"}
        """
        items = {"院校名称": "varchar(225)", "专业名称": "text", "年份": "varchar(30)", "院校代码": "varchar(30)",
                 "专业代码": "varchar(30)"
            , "批次": "varchar(30)", "科类": "varchar(30)", "最低分": "varchar(30)", "平均分": "varchar(30)",
                 "最低位次": "varchar(30)"
            , "计划数": "varchar(30)", "录取数": "varchar(30)", "学费": "varchar(30)", "学制": "varchar(30)",
                 "专业组": "varchar(30)"
            , "专业备注": "text", "院校计划数": "varchar(30)", "专业组计划数": "varchar(30)", "再选": "varchar(30)",
                 "选科": "varchar(30)"}
        if isinstance(data_pass, dict):
            set_clause = ','.join([row + " " + items.get(row, "varchar(500)") for row in data_pass])

        elif isinstance(data_pass[0], dict):
            set_clause = ','.join([row + " " + items.get(row, "varchar(500)") for row in data_pass[0]])

        else:
            if not isinstance(columns, list):
                raise ValueError("columns未设置字段名称")
            set_clause = ','.join([row + " " + items.get(row, "varchar(500)") for row in columns])

        sql = f"CREATE TABLE {table_name} ({set_clause})"  # ID INT AUTO_INCREMENT PRIMARY KEY,
        logger.info(f"创建数据表sql如下: {sql}")

        try:
            self.cursor.execute(sql)
            logger.success(f"{table_name}创建成功！！！")
        except Exception as e:
            logger.error(f"数据表{table_name}创建失败！！！{e}")
            sys.exit()

    def execute_sql(self, sql, params):
        if isinstance(params[0], tuple):
            if self.cursor.executemany(sql, params):
                self.db.commit()
        else:
            if self.cursor.execute(sql, params):
                self.db.commit()

    def insert(self, table_name, data_pass: Union[dict, List[dict], list, Dict[str, List[Any]]] = None,
               replace: Union[dict] = None, list_increment_id: Union[int] = None, list_insert_field: Union[list] = None,
               columns=None):
        """  mysql插入数据/替换数据
        :param table_name 表名
        :param data_pass 接收插入数据，支持以下格式之一： - 单个字典 - 单个列表 - 列表中包含多个列表 - 列表中包含多个字典 - 字典中的值为列表，是所有需要插入的数据 注意：data_pass参数与replace参数互斥，若同时提供，将以replace` 数据为准进行操作。
        :param replace 插入并更新数据，支持以下格式之一：- 单个字典 - 列表中包含多个列表 - 列表中包含多个字典 - 单个列表
        :param list_increment_id 以列表格式插入所有字段数据时，如果数据中没有自增的值，list_increment_id = 数据库中字段的位置(int)
        :param list_insert_field 指定列表插入时的字段名称，例如：["字段1", "字段2"]， 不与list_increment_id参数同时使用
        """

        if not self.table_tf:
            self.cursor.execute(f"SHOW TABLES LIKE '{table_name}'")
            if not self.cursor.fetchone():
                self.create_table(table_name, columns, data_pass)
            self.table_tf = True

        if (not replace and not data_pass) or not self.cursor:
            logger.error(f'数据列表为空或连接未建立')
            return

        def detailed_data(keys):
            item = {}
            item["columns"] = ', '.join(keys)
            item["placeholders"] = ', '.join(['%s'] * len(keys))
            if replace:
                statement = ""
                for key in keys:
                    statement += f"{key} = VALUES({key}), "
                item["statement"] = statement[:-2] + ";"
            return item

        try:
            # - 插入并更新数据
            if replace:
                # -单个字典
                if isinstance(replace, dict):
                    params = tuple(replace.values())
                    item = detailed_data(replace.keys())

                    sql = f"INSERT INTO user_pass ({item['columns']})	VALUES({item['placeholders']}) ON DUPLICATE KEY UPDATE {item['statement']}"
                    self.execute_sql(sql, params)
                    logger.success(f"数据插入成功")

                if isinstance(replace, list):
                    # - 多个字典
                    if isinstance(replace[0], dict):
                        keys = set().union(*replace)
                        item = detailed_data(keys)
                        params_list = []
                        for row in replace:
                            params = tuple(row.get(key, None) for key in keys)  # get方法可以处理键不存在的情况
                            params_list.append(params)
                        params = tuple(params_list)

                        sql = f"INSERT INTO user_pass ({item['columns']})	VALUES({item['placeholders']}) ON DUPLICATE KEY UPDATE {item['statement']}"

                        self.execute_sql(sql, params)
                        logger.success(f"数据插入成功")

                    # - 多个列表
                    elif isinstance(replace[0], list):
                        keys = self.get_table_columns(table_name)
                        item = detailed_data(keys)
                        sql = f"INSERT INTO user_pass ({item['columns']})	VALUES({item['placeholders']}) ON DUPLICATE KEY UPDATE {item['statement']}"
                        params = tuple(tuple(row) for row in replace)
                        self.execute_sql(sql, params)
                        logger.success(f"数据插入成功")

                    # - 单个列表
                    else:
                        keys = self.get_table_columns(table_name)
                        item = detailed_data(keys)
                        sql = f"INSERT INTO user_pass ({item['columns']})	VALUES({item['placeholders']}) ON DUPLICATE KEY UPDATE {item['statement']}"
                        params = tuple(replace)
                        self.execute_sql(sql, params)
                        logger.success(f"数据插入成功")

            # - 插入数据
            else:
                if isinstance(data_pass, list):

                    # - 多个字典
                    if isinstance(data_pass[0], dict):
                        keys = set().union(*data_pass)
                        item = detailed_data(keys)
                        sql = f"INSERT INTO {table_name} ({item['columns']}) VALUES ({item['placeholders']})"
                        params = tuple(tuple(item.get(key) for key in keys) for item in data_pass)
                        self.execute_sql(sql, params)
                        logger.success(f"{self.cursor.rowcount} 条数据插入成功")

                    # - 多条列表
                    elif isinstance(data_pass[0], list):  # data_pass 为多条的 list
                        if list_increment_id:
                            for item in data_pass:
                                item.insert(list_increment_id - 1, 0)

                        placeholders = ', '.join(['%s'] * len(data_pass[0]))
                        sql = f"INSERT INTO {table_name} VALUES ({placeholders})"

                        if list_insert_field:
                            field = ','.join(list_insert_field)
                            sql = f"INSERT INTO {table_name}({field}) VALUES ({placeholders})"

                        params = tuple(tuple(data) for data in data_pass)

                        self.execute_sql(sql, params)
                        logger.success(f'{self.cursor.rowcount} 条数据插入成功')

                    # - 一个列表
                    else:
                        if list_increment_id:
                            data_pass.insert(list_increment_id - 1, 0)

                        placeholders = ', '.join(['%s'] * len(data_pass))
                        sql = f"INSERT INTO {table_name} VALUES ({placeholders})"
                        if list_insert_field:
                            field = ','.join(list_insert_field)
                            sql = f"INSERT INTO {table_name}({field}) VALUES ({placeholders})"

                        self.execute_sql(sql, tuple(data_pass))
                        logger.success(f'{self.cursor.rowcount} 条数据插入成功')

                # - 单个字典
                elif isinstance(data_pass, dict):
                    keys = ','.join(data_pass.keys())
                    values = ','.join(['%s'] * len(data_pass))
                    sql = f'INSERT INTO {table_name}({keys}) VALUES({values})'
                    if self.cursor.execute(sql, tuple(data_pass.values())):
                        self.db.commit()

                    logger.success(f"{self.cursor.rowcount} 条数据插入成功")

        except Error as e:
            e = str(e)
            if 'Duplicate entry' in e:
                e = e.replace('Duplicate entry', '重复条目')
                logger.error(f"插入数据时发生错误: {e}")
            else:
                e = translate(e)
                logger.error(f"插入数据时发生错误: {e}")
                sys.exit()

    def update(self, table_name, match_field, data_pass: Union[dict, List[dict]]):
        """  mysql修改数据
        :param table_name 表名
        :param match_field 进行匹配的字段
        :param data_pass 参数接收，支持以下格式之一：- 字典 - 列表中嵌套字典 格式
        \nupdate('bk_major','major_ljdm',{'major_ljdm':row[rd.major_ljdm],'evaluation':row[rd.expert_comment]})
        \nupdate('bk_major',['major_ljdm'],{'major_ljdm':row[rd.major_ljdm],'evaluation':row[rd.expert_comment]})
        """
        if not data_pass or not self.cursor:
            logger.error(f'数据列表为空或连接未建立')
            sys.exit()

        try:
            # - 多条字典
            if isinstance(data_pass, list):
                for item in data_pass:
                    if match_field not in item:
                        raise ValueError(f"数据中必须包含替换字段 '{match_field}'")

                all_params_list = []
                for data_item in data_pass:
                    update_fields = [k for k in data_item.keys() if k != match_field]
                    params_for_item = list(data_item[field] for field in update_fields) + [data_item[match_field]]
                    all_params_list.append(params_for_item)

                set_clause = ', '.join([f"{field} = %s" for field in update_fields])
                sql = f"UPDATE {table_name} SET {set_clause} WHERE {match_field} = %s"
                try:
                    self.execute_sql(sql, tuple(all_params_list))
                    logger.success(f"{self.cursor.rowcount} 条数据更新成功")
                except Exception as e:
                    self.db.rollback()
                    logger.error(f"数据更新失败: {translate(e)}")
                    sys.exit()

            # - 单个字典
            elif isinstance(data_pass, dict):
                update_fields = [k for k in data_pass.keys() if k != match_field]
                set_clause = ', '.join([f"{field} = %s" for field in update_fields])

                sql = f"UPDATE {table_name} SET {set_clause} WHERE {match_field} = %s"
                params = tuple([data_pass[field] for field in update_fields] + [data_pass[match_field]])

                self.execute_sql(sql, params)
                logger.success(f"{self.cursor.rowcount} 条数据更新成功")

        except Error as e:
            logger.error(f"插入数据时发生错误: {translate(e)}")
            sys.exit()

    def select(self, table_name, fields=None, select_sql=None, output_format=None):
        """  mysql查询数据
        :param table_name 表名
        :param fields 默认查询所有字段, 自定义字段列表传入, fields=['COUNT(*)']查询数据数量
        :param select_sql 自定义筛选内容，例如 select_sql= "where name='内容1'"
        :param output_format 输出格式，默认输出为list，output_format='dict'输出格式为字典
        """

        def execute_select(sql):
            self.cursor.execute(sql)
            results = self.cursor.fetchall()
            return results

        select_clause = '*' if fields is None else ', '.join(fields)
        sql = f"SELECT {select_clause} FROM {table_name}"

        # - 数据计数
        if fields is not None and len(fields) == 1 and fields[0] == 'COUNT(*)':
            base_query = f"SELECT COUNT(*) FROM {table_name}"
            count_result = execute_select(base_query)[0][0]
            logger.info(f"已查询出所有数据，有{count_result}条")
            return count_result

        # - 查询指定数据
        elif fields is not None and isinstance(fields, list):
            # 指定字段
            if select_sql:
                sql = sql + " " + select_sql

            results = execute_select(sql)
            mysql_data = [fields] + [list(row) for row in results]

            if output_format == "dict":
                mysql_data = [{column: value for column, value in zip(mysql_data[0], row)} for row in mysql_data[1:]]
            return mysql_data

        # - 查询全部数据
        else:
            if select_sql:
                sql = sql + " " + select_sql
            print(sql)
            columns = self.get_table_columns(table_name)
            results = execute_select(sql)
            logger.info(f"已查询出所有数据，有{len(results)}条")
            mysql_data = [columns] + [list(row) for row in results]

            if output_format == "dict":
                mysql_data = [{column: value for column, value in zip(mysql_data[0], row)} for row in mysql_data[1:]]
            return mysql_data

    def 执行自定义sql语句(self, sql语句):
        self.cursor.execute(sql语句)
        if not sql语句.lower().startswith("select"):
            self.db.commit()
        logger.success(f'{sql语句} 执行完成！')

    def save(self, table_name, path=""):
        """  mysql存储数据为excel
        :param table_name 表名
        :param path 存储路径，默认存储在当前文件同级目录，以表名命名
        """
        # 执行SQL查询
        sql = f"SELECT * FROM {table_name}"
        self.cursor.execute(sql)

        # 获取所有行数据
        rows = self.cursor.fetchall()
        # 获取所有列名
        columns = self.get_table_columns(table_name)

        basename = os.path.basename(path)
        base, ext = os.path.splitext(basename)
        if not path:
            current_file_path = os.path.abspath(__file__)
            path = os.path.dirname(current_file_path)

        if ext != "":
            gen_excel(rows, path, columns)
        else:
            path = path + f"/{table_name}.xlsx"
            gen_excel(rows, path, columns)

        logger.success(f"数据存储成功！--{path}")
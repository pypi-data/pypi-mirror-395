import io
from typing import List, Union

from openpyxl import Workbook, utils
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet


class PlanilhaHelper:

    def __init__(self):
        """ Inicializa a classe Planilha criando um novo Workbook. 
            A aba padrão que é criada automaticamente é removida para começar com uma planilha vazia.
        """
        self.__wb: Workbook = Workbook()
        self.__wb.remove(self.__wb.active) # remove a aba ativa por default.

    def __obter_aba(self, nome: str) -> Worksheet:
        """ Obtém a aba com o nome especificado. Se a aba não existir, cria uma nova.

            :param nome: O nome da aba.
            :return: A aba obtida ou criada.
        """
        if nome in self.__wb.sheetnames:
            return self.__wb[nome]
        else:
            return self.__wb.create_sheet(title=nome)
    
    def criar_aba(self, nome_aba: str):
        """ Cria uma nova aba com o nome especificado.

            :param nome_aba: O nome da nova aba.
            :return: A instância atual da classe PlanilhaHelper.
        """
        self.__obter_aba(nome_aba)
        return self

    def mesclar_colunas(self, nome_aba: str, start_row: int, start_column: int, end_row: int, end_column: int):
        """ Mesclar colunas

            :param nome_aba: O nome da aba onde a linha será criada.
            :param start_row: número do começo da linha 
            :param start_column: número do começo da coluna
            :param end_row: número do fim da linha
            :param end_column: número do fim da coluna
            :return: A instância atual da classe PlanilhaHelper.
        """
        aba = self.__obter_aba(nome_aba)
        aba.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)
        return self

    def criar_linha(self, nome_aba: str, num_linha: int, valores: Union[List[str], List[int], List[float]], negrito: bool = False):
        """ Cria uma nova linha na aba especificada.

            :param nome_aba: O nome da aba onde a linha será criada.
            :param num_linha: número da linha onde será inseridos os valores
            :param valores: A lista de valores para a linha.
            :param negrito: Se True, a linha será em negrito.
            :return: A instância atual da classe PlanilhaHelper.
        """
        aba = self.__obter_aba(nome_aba)
        for coluna in range(1, len(valores) + 1):
            celula = aba.cell(row=num_linha, column=coluna, value=valores[coluna - 1])
            if negrito:
                celula.font = Font(bold=True)

        return self

    def criar_coluna(self, nome_aba: str, num_coluna: int, valores: Union[List[str], List[int], List[float]], negrito: bool = False):
        """ Cria uma nova coluna na aba especificada.

            :param nome_aba: O nome da aba onde a coluna será criada.
            :param num_coluna: número da coluna onde será inseridos os valores
            :param valores: A lista de valores para a coluna.
            :param negrito: Se True, a coluna será em negrito.
            :return: A instância atual da classe PlanilhaHelper.
        """
        aba = self.__obter_aba(nome_aba)
        for linha in range(1, len(valores) + 1):
            celula = aba.cell(row=linha, column=num_coluna, value=valores[linha - 1])
            if negrito:
                celula.font = Font(bold=True)

        return self

    def criar_celula(self, nome_aba: str, linha: int, coluna: int, valor: Union[str, int, float], largura: int = None, negrito: bool = False, centralizar: bool = False, hyperlink: str = None):
        """ Cria uma nova célula na aba especificada.

            :param nome_aba: O nome da aba onde a célula será criada.
            :param linha: A linha onde a célula será criada.
            :param coluna: A coluna onde a célula será criada.
            :param valor: O valor para a célula.
            :param largura: valor da largura da célula
            :param negrito: Se True, a célula será em negrito.
            :param centralizar: Se True, o conteúdo da célula é centralizado.
            :param hyperlink: Se fornecido, a célula terá um hyperlink para a aba especificada.
            :return: A instância atual da classe PlanilhaHelper.
        """
        aba = self.__obter_aba(nome_aba)
        celula = aba.cell(row=linha, column=coluna, value=valor)

        alignment = Alignment()
        font = Font()

        if len(str(valor)) > 50:
            alignment.wrap_text = True
        if largura:
            aba.column_dimensions[celula.column_letter].width = largura
            alignment.vertical = 'top'
        if centralizar:
            alignment.horizontal = 'center'
        if negrito:
            font.bold = True
        if hyperlink is not None:
            celula.hyperlink = f"#{hyperlink}!A1"  # Cria um hyperlink para a célula A1 na aba especificada

        celula.font = font
        celula.alignment = alignment
        return self

    def redimensionar_coluna(self, nome_aba: str, num_coluna: int, tamanho: int):
        """ Redimensiona a coluna especificada na aba especificada.

            :param nome_aba: O nome da aba onde a coluna será redimensionada.
            :param num_coluna: O número da coluna a ser redimensionada.
            :param tamanho: O novo tamanho da coluna.
            :return: A instância atual da classe PlanilhaHelper.
        """
        aba = self.__obter_aba(nome_aba)
        coluna = utils.get_column_letter(num_coluna)
        aba.column_dimensions[coluna].width = tamanho
        return self

    def redimensionar_linha(self, nome_aba: str, num_linha: int, altura: float):
        """ Redimensiona a linha especificada na aba especificada.

            :param nome_aba: O nome da aba onde a linha será redimensionada.
            :param num_linha: O número da linha a ser redimensionada.
            :param altura: A nova altura da linha.
            :return: A instância atual da classe PlanilhaHelper.
        """
        aba = self.__obter_aba(nome_aba)
        aba.row_dimensions[num_linha].height = altura
        return self

    def criar_hyperlink(self, nome_aba_origem: str, nome_celula: str, nome_aba_referenciada: str):
        """ Cria um hyperlink para uma outra aba

            :param nome_aba_origem: O nome da aba de origem.
            :param nome_celula: nome da célula, ex.: 'A1', 'D7', 'C9'.
            :param nome_aba_referenciada: O nome da aba referenciada
            :return: A instância atual da classe PlanilhaHelper.
        """
        aba: Worksheet = self.__obter_aba(nome_aba_origem)
        aba[nome_celula].hyperlink = f"#{nome_aba_referenciada}!A1"
        return self

    def obter_buffer(self) -> io.BytesIO:
        """ Retorna o buffer da planilha como um objeto io.BytesIO.

            :return: O buffer da planilha.
        """
        buffer = io.BytesIO()
        self.__wb.save(buffer)  # Salva a planilha no buffer
        buffer.seek(0)  # Retorna o ponteiro do buffer para o início
        return buffer

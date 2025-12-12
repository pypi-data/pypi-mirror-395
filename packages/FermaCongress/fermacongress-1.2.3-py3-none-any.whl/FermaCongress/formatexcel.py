import os
import pandas as pd
from tqdm import tqdm
from typing import Optional
from dateutil import parser
from datetime import datetime, date, time
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment



def format(input_path: Optional[str] = None, dataframe: Optional[pd.DataFrame] = None, output_path: Optional[str] = None, input_sheet: Optional[str] = None, output_sheet: Optional[str] = None) -> None:
    """
    Format Excel/CSV files or DataFrames with consistent styling and data cleaning.
    
    Args:
        input_path: Path to input .xlsx or .csv file [mutually exclusive with 'dataframe' parameter]
        dataframe: Pandas DataFrame to format [mutually exclusive with 'input_path' parameter]
        output_path: Path for output .xlsx file [required]
        input_sheet: Specific sheet name to read from Excel input [optional]
        output_sheet: Name for output sheet [default: "Sheet"]
    """

    # ========================================
    # HELPER FUNCTIONS
    # ========================================
    
    def clean_illegal_characters(value):
        """Remove illegal XML characters that would corrupt Excel files."""
        if isinstance(value, str):
            return ILLEGAL_CHARACTERS_RE.sub("", value)
        return value

    def is_date(value):
        """Check if value is a date/time object."""
        return isinstance(value, (datetime, date, time))

    def is_numeric(value):
        """Check if value is numeric (int or float)."""
        return isinstance(value, (int, float))


    # ========================================
    # INPUT VALIDATION
    # ========================================
    
    # Validate mutual exclusivity of input sources
    if input_path and dataframe is not None:
        raise ValueError("Provide either 'input_path' or 'dataframe', not both.")

    # Validate at least one input is provided
    if input_path is None and dataframe is None:
        raise ValueError("Either 'input_path' or 'dataframe' must be provided.")

    # Validate output path is specified
    if output_path is None:
        raise ValueError("The 'output_path' must be specified.")

    # Validate output path has .xlsx extension
    if not str(output_path).lower().endswith('.xlsx'):
        raise ValueError("The 'output_path' must be an .xlsx file.")
    
    # Validate input file exists (if file path provided)
    if input_path is not None and not os.path.exists(input_path):
        raise FileNotFoundError(f"The Input File: '{input_path}' does not exist.")

    # ========================================
    # FILE LOADING & DATA PREPARATION
    # ========================================

    if input_path:
        file_extension = os.path.splitext(input_path)[1].lower()

        # -------------------- EXCEL FILE --------------------
        if file_extension == '.xlsx':
            wb = load_workbook(input_path)

            # Select specific sheet or use active sheet
            if input_sheet:
                if input_sheet not in wb.sheetnames:
                    raise ValueError(
                        f"The sheet '{input_sheet}' does not exist in the input file. "
                        f"Available sheets: {', '.join(wb.sheetnames)}"
                    )
                ws = wb[input_sheet]
            else:
                ws = wb.active
            
            # Remove all other sheets (keep only target sheet)
            for sheet_name in list(wb.sheetnames):
                if sheet_name != ws.title:
                    del wb[sheet_name]
            
            # Apply illegal character cleaning and formula protection to Excel files
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        # Clean illegal characters
                        cleaned_value = clean_illegal_characters(cell.value)
                        cell.value = cleaned_value
                        
                        # Force string data type for values starting with formula characters
                        if cleaned_value and cleaned_value[0] in ('=', '+', '-', '@'):
                            cell.data_type = 's'  # Force as string, not formula

        
        # -------------------- CSV FILE --------------------
        elif file_extension == '.csv':
            df = pd.read_csv(input_path)
            df = df.apply(lambda col: col.map(clean_illegal_characters))

            # Create new workbook and select active sheet
            wb = Workbook()
            ws = wb.active
            
            # Write cells explicitly to prevent formula interpretation
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.value = value
                    
                    # Force string data type for values starting with formula characters
                    # This prevents Excel from interpreting them as formulas
                    if isinstance(value, str) and value and value[0] in ('=', '+', '-', '@'):
                        cell.data_type = 's'  # 's' = string type
                        
        else:
            raise ValueError(
                f"Unsupported file extension '{file_extension}'. "
                "The input file must be either .xlsx or .csv."
            )

    # -------------------- DATAFRAME INPUT -------------------- 
    elif dataframe is not None:

        # Validate DataFrame is not empty
        if dataframe.empty:
            raise ValueError("The provided DataFrame is empty. Cannot format an empty DataFrame.")
        
        # Apply illegal character cleaning
        dataframe = dataframe.apply(lambda col: col.map(clean_illegal_characters))
        
        # Create new workbook and select active sheet
        wb = Workbook()
        ws = wb.active
        
        # Write cells explicitly to prevent formula interpretation
        for r_idx, row in enumerate(dataframe_to_rows(dataframe, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value
                
                # Force string data type for values starting with formula characters
                if isinstance(value, str) and value and value[0] in ('=', '+', '-', '@'):
                    cell.data_type = 's'
    

    # ========================================
    # SHEET NAMING
    # ========================================

    ws.title = output_sheet if output_sheet else "Sheet"
    

    # ========================================
    # HEADER PROCESSING
    # ========================================

    # Validate worksheet is not empty
    if ws.max_row < 1:
        raise ValueError("The worksheet is empty. Cannot format an empty worksheet.")
    
    # Retrieve and normalize header row with robust None/empty handling
    header_row = []
    for idx, cell in enumerate(ws[1], start=1):
        if isinstance(cell.value, str) and cell.value.strip():
            # Normal string header - lowercase and strip whitespace
            header_row.append(cell.value.lower().strip())
        elif cell.value is not None:
            # Non-string, non-None value (e.g., number) - convert to string
            header_row.append(str(cell.value).lower().strip())
        else:
            # Empty or None header - use default name with index
            header_row.append(f"Unnamed_column_{idx}")
    

    # ========================================
    # CONFIGURATION
    # ========================================

    # Total steps for progress bar
    total_steps = 9

    # Track columns with unmatched date formats
    unmatched_columns = set()

    # Column Configuration - All lowercase to match normalized headers
    # Using sets for O(1) lookup performance

    date_columns = {'start_date', 'start_time', 'end_date', 'end_time', 'date of posting', 'date', 'start date', 'start time', 'end date', 'end time'}
    id_columns = {'id', 'int id', 'internal_id', 'internal id', 'session_id', 'session id', 'display_id', 'display id', 'abstract_id', 'abstract id'}
    
    title_columns = {'abstract', 'title', 'tweet text', 'abstract title', 'session title', 'abstract_title', 'session_title', 'old abstract',
                     'old title', 'full abstract', 'full_abstract_text', 'full abstract text', 'old session title', 'old abstract title'}
    url_columns = {'abstract link', 'url', 'link', 'abstract_link', 'ferma link', 'tweet url'}

    # Initialize the progress bar
    with tqdm(total=total_steps, desc="Formatting Progress") as pbar:
        trim_count = 0
        
        # -------------------------------------------------------------------------------------------------
        # Step 1: Trimming Leading and Trailing Whitespaces
        # -------------------------------------------------------------------------------------------------
        try:
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if isinstance(cell.value, str):
                        trimmed_value = cell.value.strip()
                        if trimmed_value != cell.value:
                            cell.value = trimmed_value
                            trim_count += 1
            pbar.update(1)
        except Exception as e:
            print(f"• Error in Trimming Whitespace: {e}")


        # -------------------------------------------------------------------------------------------------
        # Step 2: Sort by Priority and Apply Conditional Formatting
        """
        Sorting Order: Very High → High → Internal → Medium → Low → Not Relevant

        Colors:
        - Very High: Light Blue (A5B3F7)
        - High: Light Green (39C7A5)
        - Medium: Yellow (FFCA42)
        - Low: Pink (EA4970)
        - Not Relevant: Light Grey (A2A2A7)
        - Internal: No Background Color
        """
        # -------------------------------------------------------------------------------------------------
        
        try:
            if 'priority' in header_row and ws.max_row >= 2:
                priority_col_index = header_row.index('priority') + 1
                priority_col_letter = get_column_letter(priority_col_index)
                
                # Define priority order and colors
                priority_order = ['Very High', 'High', 'Internal', 'Medium', 'Low', 'Not Relevant']
                priority_sort_map = {priority: index for index, priority in enumerate(priority_order)}
                
                colors = {
                    'Very High': 'A5B3F7',
                    'High': '39C7A5',
                    'Medium': 'FFCA42',
                    'Low': 'EA4970',
                    'Not Relevant': 'A2A2A7',
                    'Internal': None
                }
                
                # Sort rows by priority
                rows = []
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                    rows.append(list(row))
                
                rows.sort(key=lambda row: priority_sort_map.get(row[priority_col_index - 1], float('inf')))
                
                # Write sorted data back
                for row_idx, row_data in enumerate(rows, start=2):
                    for col_idx, value in enumerate(row_data, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Apply direct cell formatting based on priority value
                for row_idx in range(2, ws.max_row + 1):
                    priority_cell = ws.cell(row=row_idx, column=priority_col_index)
                    priority_value = priority_cell.value
                    
                    if priority_value in colors and colors[priority_value]:
                        fill = PatternFill(start_color=colors[priority_value], 
                                        end_color=colors[priority_value], 
                                        fill_type="solid")
                        priority_cell.fill = fill
            
            pbar.update(1)
            
        except Exception as e:
            print(f"• Error in sorting and conditional formatting: {e}")


        # -------------------------------------------------------------------------------------------------
        # Step 3: Applying Formatting for the Data Rows
        """
        - Font Name: Manrope
        - Font Size: 10
        - Wrap Text - True
        - Row Height: 48.75
        
        """
        # -------------------------------------------------------------------------------------------------
        try:
            if ws.max_row < 2:
                pbar.update(1)
            
            else:
                data_font = Font(name="Manrope", size=10)

                # Set row heights
                for row_idx in range(2, ws.max_row + 1):
                    ws.row_dimensions[row_idx].height = 48.75

                # Apply cell formatting
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.font = data_font
                
                pbar.update(1)
        except Exception as e:
            print(f"• Error in formatting data rows: {e}")
            
        
        
        # -------------------------------------------------------------------------------------------------
        # Step 4: Applying Column Alignment Based on Column
        """
        - Center Alignment (Center, Center): Numeric Columns, 'priority'
        - Left Alignment (Left, Top): Text Columns
        - Right Alignment (Right, Bottom): Date Columns
        """
        # -------------------------------------------------------------------------------------------------

        try:
            if ws.max_row < 2:
                pbar.update(1)
            
            else:
                center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                date_alignment = Alignment(horizontal="right", vertical="bottom", wrap_text=True)

                
                for col_index, col in enumerate(ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=1):
                    col_name = header_row[col_index - 1] if col_index - 1 < len(header_row) else ""

                    # Collect non-empty values for content checking
                    non_none_values = [cell.value for cell in col if cell.value is not None]
                    
                    # Determine alignment: column name first, then content type
                    if col_name in id_columns or col_name == 'priority':
                        alignment = center_alignment
                    elif col_name in date_columns:
                        alignment = date_alignment
                    elif non_none_values and all(is_numeric(val) for val in non_none_values):
                        alignment = center_alignment
                    elif non_none_values and all(is_date(val) for val in non_none_values):
                        alignment = date_alignment
                    else:
                        alignment = left_alignment

                    # Apply alignment to all cells in column
                    for cell in col:
                        cell.alignment = alignment

                pbar.update(1)
        except Exception as e:
            print(f"• Error in column alignment: {e}")
            
            
            
        # -------------------------------------------------------------------------------------------------
        # Step 5: Applying Column Widths
        """
        Default Width: 21
        Custom Widths:
        - Title Columns: 60
        - ID Columns: 12
        - Date Columns: 22
        """
        # -------------------------------------------------------------------------------------------------

        try:
            default_width = 21

            column_width_mapping = {**dict.fromkeys(title_columns, 60),
                                    **dict.fromkeys(id_columns, 12),
                                    **dict.fromkeys(date_columns, 22)}
            
            # Apply column widths
            for col_index in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_index)
                col_name = header_row[col_index - 1] if col_index - 1 < len(header_row) else ""
                
                # Use custom width if column name matches, otherwise use default
                width = column_width_mapping.get(col_name, default_width)
                ws.column_dimensions[col_letter].width = width
            pbar.update(1)
            
        except Exception as e:
            print(f"• Error in applying column widths: {e}")

        
        # -------------------------------------------------------------------------------------------------
        # Step 6: Applying Formatting in Hyperlinks
        """
        - Font Name: Manrope
        - Font Size: 10
        - Font Color: Blue (0000FF)
        - Underline: Single
        """
        # -------------------------------------------------------------------------------------------------

        try:
            if ws.max_row < 2:
                pbar.update(1)
            else:
                hyperlink_font = Font(name="Manrope", size=10, color="0000FF", underline="single")

                # Find hyperlink columns using set intersection for efficiency
                hyperlink_columns = []
                for col_name in url_columns:
                    if col_name in header_row:
                        hyperlink_columns.append(header_row.index(col_name) + 1)

                # Apply hyperlink formatting
                for col_index in hyperlink_columns:
                    for cell in ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=col_index, max_col=col_index):
                        for c in cell:
                            if c.value and isinstance(c.value, str):
                                url_value = c.value.strip()
                                if url_value.startswith(('http://', 'https://')):
                                    try:
                                        c.hyperlink = url_value
                                        c.font = hyperlink_font
                                    except Exception:
                                        # Skip if hyperlink fails
                                        pass

                pbar.update(1)

        except Exception as e:
            print(f"• Error in hyperlink formatting: {e}")
            
        
        # -------------------------------------------------------------------------------------------------
        # Step 7: Applying Date and Time Formatting
        # -------------------------------------------------------------------------------------------------

        try:
            if ws.max_row < 2:
                pbar.update(1)
            else:
                columns_formats = {
                    'date': {'initial': "%Y-%m-%d %H:%M:%S", 'desired': 'dddd, dd mmmm yyyy'},
                    'start_time': {'initial': "%Y-%m-%d %H:%M:%S", 'desired': 'hh:mm am/pm'},
                    'end_time': {'initial': "%Y-%m-%d %H:%M:%S", 'desired': 'hh:mm am/pm'},
                    'start time': {'initial': "%Y-%m-%d %H:%M:%S", 'desired': 'hh:mm am/pm'},
                    'end time': {'initial': "%Y-%m-%d %H:%M:%S", 'desired': 'hh:mm am/pm'},
                    'start_date': {'initial': '%Y-%m-%d %H:%M:%S', 'desired': 'yyyy-mm-dd hh:mm:ss'},
                    'end_date': {'initial': '%Y-%m-%d %H:%M:%S', 'desired': 'yyyy-mm-dd hh:mm:ss'},
                    'date of posting': {'initial': "%Y-%m-%d %H:%M:%S", 'desired': 'dd-mm-yyyy'},
                    'created_at': {'initial': "%Y-%m-%d %H:%M:%S", 'desired': 'dd-mm-yyyy'},
                }

                for column_name, format_info in columns_formats.items():
                    if column_name not in header_row:
                        continue
                    column_index = header_row.index(column_name) + 1

                    initial_format = format_info.get('initial')
                    desired_format = format_info.get('desired')

                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=column_index, max_col=column_index):
                        cell = row[0]
                        if cell.value:
                            try:
                                if not isinstance(cell.value, datetime):
                                    try:
                                        dt = datetime.strptime(str(cell.value), initial_format) if initial_format else parser.parse(str(cell.value))
                                    except ValueError:
                                        unmatched_columns.add(column_name)
                                        continue
                                else:
                                    dt = cell.value
                                cell.value = dt
                                cell.number_format = desired_format

                            except Exception:
                                unmatched_columns.add(column_name)
                
                pbar.update(1)

        except Exception as e:
            print(f"• Error in date formatting: {e}")
            
            
        # -------------------------------------------------------------------------------------------------
        # Step 8: Applying All Borders
        # -------------------------------------------------------------------------------------------------

        try:
            border = Border(left=Side(style="thin"), right=Side(style="thin"),
                            top=Side(style="thin"), bottom=Side(style="thin"))
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = border
            
            pbar.update(1)
            
        except Exception as e:
            print(f"• Error in applying borders: {e}")

        
        # -------------------------------------------------------------------------------------------------        
        # Step 9: Applying Formatting for the Header Row
        """
        - Font Name: Playfair Display Black
        - Font Size: 11
        - Font Color: White (FFFFFF)
        - Bold: True
        - Background Color: Blue (1E41EB)
        - Alignment: Horizontal - Center, Vertical - Center
        - Wrap Text - True
        - Row Height: 38
        """
        # -------------------------------------------------------------------------------------------------

        try:
            if ws.max_row >= 1:
                header_font = Font(name="Playfair Display Black", size=11, bold=True, color="FFFFFF")
                header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                header_fill = PatternFill(start_color="1E41EB", end_color="1E41EB", fill_type="solid")

                for cell in ws[1]:
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.fill = header_fill

                ws.row_dimensions[1].height = 38

            pbar.update(1)
            
        except Exception as e:
            print(f"• Error in header row formatting: {e}")

        # -------------------------------------------------------------------------------------------------

    wb.save(output_path)
    if trim_count > 0:
        print(f"\n• Trimmed {trim_count} cells")
    if unmatched_columns:
        print(f"• Unmatched Formats Detected: {list(unmatched_columns)}")
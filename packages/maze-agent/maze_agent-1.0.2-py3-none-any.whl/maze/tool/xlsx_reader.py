from maze import task
import openpyxl


@task(
    inputs=["file_path"],
    outputs=["result"],
)
def xlsx_reader(params):
    file_path = params.get("file_path")
    
    if not file_path:
        return {"result": None, "error": "Missing required parameter: file_path"}
    
    try:
        workbook = openpyxl.load_workbook(file_path)
        
        sheets_data = {}
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_data = []
            for row in sheet.iter_rows(values_only=True):
                sheet_data.append(list(row))
            sheets_data[sheet_name] = sheet_data
            
        return {"result": sheets_data}
    except Exception as e:
        return {"result": None, "error": str(e)}
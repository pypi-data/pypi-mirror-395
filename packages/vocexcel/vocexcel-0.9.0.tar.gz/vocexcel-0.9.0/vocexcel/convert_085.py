from pathlib import Path
from typing import Literal as TypeLiteral
from typing import Optional

from dateutil.parser import parse as date_parser
from openpyxl.styles import Alignment, Font
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pyshacl import validate as shacl_validate
from rdflib import DCTERMS, BNode, Graph, Literal, Namespace, URIRef
from rdflib.namespace import OWL, PROV, RDF, RDFS, SDO, SKOS, XSD

from vocexcel.convert_070 import (
    extract_additions_concept_properties as extract_additions_concept_properties_070,
)
from vocexcel.convert_070 import extract_collections as extract_collections_070
from vocexcel.utils import (
    RDF_FILE_ENDINGS,
    STATUSES,
    VOCDERMODS,
    ConversionError,
    ShaclValidationError,
    add_top_concepts,
    bind_namespaces,
    fill_cell_with_list_of_curies,
    load_workbook,
    make_agent,
    make_iri,
    return_error,
    split_and_tidy_to_iris,
    split_and_tidy_to_strings,
    xl_hyperlink,
)

DATAROLES = Namespace("https://linked.data.gov.au/def/data-roles/")


def extract_prefixes(
    sheet: Worksheet,
    error_format: TypeLiteral["python", "cmd", "json"] = "python",
) -> dict[str, Namespace] | str | None:
    prefixes = {}
    i = 3
    while True:
        ns = sheet[f"B{i}"].value
        if ns is None:
            break
        else:
            if not ns.startswith("http"):
                error = ConversionError(
                    f"Your namespace value on sheet Prefixes, cell C{i} is invalid. It must start with 'http'"
                )
                return return_error(error, error_format)

            pre = sheet[f"A{i}"].value
            proper_pre = str(pre).strip(":") + ":" if pre is not None else ":"
            prefixes[proper_pre] = ns

        i += 1

    return prefixes


def extract_concept_scheme(
    sheet: Worksheet,
    prefixes,
    template_version="0.8.5",
    error_format: TypeLiteral["python", "cmd", "json"] = "python",
) -> tuple[Graph, str]:
    iri_s = sheet["B3"].value
    title = sheet["B4"].value
    description = sheet["B5"].value
    created = sheet["B6"].value
    modified = sheet["B7"].value
    creator = sheet["B8"].value
    publisher = sheet["B9"].value
    custodian = sheet["B10"].value
    version = str(sheet["B11"].value).strip("'")
    history_note = sheet["B12"].value
    citation = sheet["B13"].value
    derived_from = sheet["B14"].value
    voc_der_mod = sheet["B15"].value
    themes = split_and_tidy_to_strings(sheet["B16"].value)
    status = sheet["B17"].value
    if template_version in ["0.8.5.GA", "0.9.0.GA"]:
        catalogue_pid = sheet["B18"].value

    if iri_s is None:
        error = ConversionError(
            "Your vocabulary has no IRI. Please add it to the Concept Scheme sheet"
        )
        return return_error(error, error_format)
    else:
        iri = make_iri(iri_s, prefixes)

    if title is None:
        error = ConversionError(
            "Your vocabulary has no title. Please add it to the Concept Scheme sheet"
        )
        return return_error(error, error_format)

    if description is None:
        error = ConversionError(
            "Your vocabulary has no description. Please add it to the Concept Scheme sheet"
        )
        return return_error(error, error_format)

    if created is None:
        error = ConversionError(
            "Your vocabulary has no created date. Please add it to the Concept Scheme sheet"
        )
        return return_error(error, error_format)

    if modified is None:
        error = ConversionError(
            "Your vocabulary has no modified date. Please add it to the Concept Scheme sheet"
        )
        return return_error(error, error_format)

    if creator is None:
        error = ConversionError(
            "Your vocabulary has no creator. Please add it to the Concept Scheme sheet"
        )
        return return_error(error, error_format)

    if publisher is None:
        error = ConversionError(
            "Your vocabulary has no publisher. Please add it to the Concept Scheme sheet"
        )
        return return_error(error, error_format)

    if history_note is None:
        error = ConversionError(
            "Your vocabulary has no History Note statement. Please add it to the Concept Scheme sheet"
        )
        return return_error(error, error_format)

    # citation

    if derived_from is not None:
        if voc_der_mod is None:
            error = ConversionError(
                "If you supply a 'Derived From' value - IRI of another vocab - "
                "you must also supply a 'Derivation Mode' value"
            )
            return return_error(error, error_format)

        if voc_der_mod not in VOCDERMODS:
            error = ConversionError(
                f"You have supplied a vocab derivation mode for your vocab of {voc_der_mod} but it is not recognised. "
                f"If supplied, it must be one of {', '.join(VOCDERMODS.keys())}"
            )
            return return_error(error, error_format)

        derived_from = make_iri(derived_from, prefixes)

    # keywords

    if status is not None and status not in STATUSES:
        error = ConversionError(
            f"You have supplied a status for your vocab of {status} but it is not recognised. "
            f"If supplied, it must be one of {', '.join(STATUSES.keys())}"
        )
        return return_error(error, error_format)

    if template_version in ["0.8.5.GA", "0.9.0.GA"]:
        if catalogue_pid is None or not str(catalogue_pid).startswith(
            "https://pid.geoscience.gov.au/"
        ):
            error = ConversionError(
                "All GA vocabularies must have an eCat ID starting https://pid.geoscience.gov.au/dataset/..., assigned in the Concept Scheme metadata"
            )
            return return_error(error, error_format)

    g = Graph(bind_namespaces="rdflib")
    g.add((iri, RDF.type, SKOS.ConceptScheme))
    g.add((iri, SKOS.prefLabel, Literal(title, lang="en")))
    g.add((iri, SKOS.definition, Literal(description, lang="en")))
    g.add((iri, SDO.dateCreated, Literal(created.date(), datatype=XSD.date)))
    g.add((iri, SDO.dateModified, Literal(modified.date(), datatype=XSD.date)))
    g += make_agent(creator, SDO.creator, prefixes, iri)
    g += make_agent(publisher, SDO.publisher, prefixes, iri)

    if custodian is not None:
        for _custodian in split_and_tidy_to_strings(custodian):
            g += make_agent(_custodian, DATAROLES.custodian, prefixes, iri)
            g.bind("DATAROLES", DATAROLES)

    if version is not None:
        g.add((iri, SDO.version, Literal(str(version))))
        g.add((iri, OWL.versionIRI, URIRef(iri + "/" + str(version))))

    g.add((iri, SKOS.historyNote, Literal(history_note)))

    if citation is not None:
        if str(citation).startswith("http"):
            val = Literal(citation, datatype=XSD.anyURI)
        else:
            val = Literal(citation)

        g.add((iri, SDO.citation, val))

    if derived_from is not None:
        qd = BNode()
        g.add((iri, PROV.qualifiedDerivation, qd))
        g.add((qd, PROV.entity, URIRef(derived_from)))
        g.add((qd, PROV.hadRole, URIRef(VOCDERMODS[voc_der_mod])))

    if themes is not None:
        for theme in themes:
            try:
                theme = make_iri(theme, prefixes)
            except ConversionError:
                theme = Literal(theme)
            g.add((iri, SDO.keywords, theme))

    if status is not None:
        g.add((iri, SDO.status, URIRef(STATUSES[status])))

    if template_version in ["0.8.5.GA", "0.9.0.GA"]:
        g.add((iri, SDO.identifier, Literal(catalogue_pid, datatype=XSD.anyURI)))

    bind_namespaces(g, prefixes)
    g.bind("", Namespace(str(iri) + "/"))
    return g, iri


def extract_concepts(
    sheet: Worksheet,
    prefixes,
    cs_iri,
    error_format: TypeLiteral["python", "cmd", "json"] = "python",
):
    g = Graph(bind_namespaces="rdflib")
    i = 4
    while True:
        # get values
        iri_s = sheet[f"A{i}"].value
        pref_label = sheet[f"B{i}"].value
        definition = sheet[f"C{i}"].value
        alt_labels = sheet[f"D{i}"].value
        narrower = sheet[f"E{i}"].value
        history_note = sheet[f"F{i}"].value
        citation = sheet[f"G{i}"].value
        is_defined_by = sheet[f"H{i}"].value
        status = sheet[f"I{i}"].value
        example = sheet[f"J{i}"].value
        image_url = sheet[f"K{i}"].value
        image_embedded = sheet[f"L{i}"].value

        # check values
        if iri_s is None:
            break

        iri = make_iri(iri_s, prefixes)

        if pref_label is None:
            error = ConversionError(
                f"You must provide a Preferred Label for Concept {iri_s}"
            )
            return return_error(error, error_format)

        if definition is None:
            error = ConversionError(
                f"You must provide a Definition for Concept {iri_s}"
            )
            return return_error(error, error_format)

        if status is not None and status not in STATUSES:
            error = ConversionError(
                f"You have supplied a status for your Concept of {status} but it is not recognised. "
                f"If supplied, it must be one of {', '.join(STATUSES.keys())}"
            )
            return return_error(error, error_format)

        if image_url is not None:
            if not image_url.startswith("http"):
                error = ConversionError(
                    "If supplied, an Image URL must start with 'http'"
                )
                return return_error(error, error_format)

        # TODO: test embedded mage is not too large
        # image_embedded
        if image_embedded not in [None, "#VALUE!"]:
            error = ConversionError(
                "The Image Embedded colum, you must only insert images or leave it blank. "
                f"You have an unexpected value in Cell L{i}"
            )
            return return_error(error, error_format)

        # ignore example Concepts
        if iri_s in [
            "http://example.com/earth-science",
            "http://example.com/atmospheric-science",
            "http://example.com/geology",
        ]:
            continue

        # create Graph
        g.add((iri, RDF.type, SKOS.Concept))
        g.add((iri, SKOS.inScheme, cs_iri))

        if "@" in pref_label:
            val, lang = pref_label.strip().split("@")
            g.add((iri, SKOS.prefLabel, Literal(val, lang=lang)))
        else:
            g.add((iri, SKOS.prefLabel, Literal(pref_label.strip(), lang="en")))
        g.add((iri, SKOS.definition, Literal(definition.strip(), lang="en")))

        if alt_labels is not None:
            for al in split_and_tidy_to_strings(alt_labels):
                g.add((iri, SKOS.altLabel, Literal(al, lang="en")))

        if narrower is not None:
            for n in split_and_tidy_to_iris(narrower, prefixes):
                g.add((iri, SKOS.narrower, n))
                g.add((n, SKOS.broader, iri))

        if history_note is not None:
            g.add((iri, SKOS.historyNote, Literal(history_note.strip())))

        if citation is not None:
            for _citation in split_and_tidy_to_strings(citation):
                g.add(
                    (iri, SDO.citation, Literal(_citation.strip(), datatype=XSD.anyURI))
                )

        if is_defined_by is not None:
            g.add((iri, RDFS.isDefinedBy, URIRef(is_defined_by.strip())))
        else:
            g.add((iri, RDFS.isDefinedBy, cs_iri))

        if status is not None:
            g.add((iri, SDO.status, URIRef(STATUSES[status])))

        if example is not None:
            g.add((iri, SKOS.example, Literal(str(example).strip())))

        if image_url is not None:
            g.add(
                (iri, SDO.image, Literal(str(image_url).strip(), datatype=XSD.anyURI))
            )

        if image_embedded == "#VALUE!":
            g.add((iri, SDO.image, Literal(f"Image at L{i}")))

        i += 1

    bind_namespaces(g, prefixes)
    return g


def extract_collections(
    sheet: Worksheet,
    prefixes,
    cs_iri,
    error_format: TypeLiteral["python", "cmd", "json"] = "python",
):
    return extract_collections_070(sheet, prefixes, cs_iri)


def extract_additions_concept_properties(
    sheet: Worksheet,
    prefixes,
    error_format: TypeLiteral["python", "cmd", "json"] = "python",
):
    return extract_additions_concept_properties_070(sheet, prefixes)


def excel_to_rdf(
    wb: Workbook,
    output_file: Optional[Path] = None,
    template_version: str = "0.8.5",
    output_format: TypeLiteral[
        "graph",
        "rdf",
    ] = "rdf",
    error_format: TypeLiteral["python", "cmd", "json"] = "python",
):
    """Converts an Excel workbook file to RDF

    Parameters:
        excel_file: The Excel workbook file to convert
        output_file: Optional. If set, the RDF output will be written to this file in the turtle format.
        output_format: Optional, default rdf. rdf: return serialized RDF in the turtle format, graph: return an RDFLib Graph object
        error_format: Optional, default python. the kind of errors to return: python is Python, cmd is command line-formatted string, json is stringified JSON

    Returns:
        output_format or an error in one of the error_formats
    """

    if not isinstance(wb, Workbook):
        error = ValueError(
            "Files for conversion to Excel must end with one of the RDF file formats: '{}'".format(
                "', '".join(RDF_FILE_ENDINGS.keys())
            )
        )
        return return_error(error, error_format)

    if output_file is not None:
        if not Path(output_file).suffix in RDF_FILE_ENDINGS:
            error = ValueError(
                f"If specifying output_file, it must have a file suffix in {', '.join(RDF_FILE_ENDINGS)}, not {output_file}."
            )
            return return_error(error, error_format)

    if template_version not in ["0.8.5", "0.8.5.GA"]:
        error = ValueError(
            f"This converter can only handle templates with versions 0.8.x or 0.8.x.GA, not {template_version}"
        )
        return return_error(error, error_format)

    allowed_output_formats = ["graph", "rdf"]
    if output_format not in allowed_output_formats:
        error = ValueError(
            f"If specifying output_format, it be in {', '.join(allowed_output_formats)}, not {output_format}."
        )
        return return_error(error, error_format)

    allowed_error_formats = ["python", "cmd", "json"]
    if error_format not in allowed_error_formats:
        error = ValueError(
            f"error_format must be on of {', '.join(allowed_error_formats)}, not {error_format}"
        )
        return return_error(error, error_format)

    prefixes = extract_prefixes(wb["Prefixes"], error_format)
    if not isinstance(prefixes, dict):
        return prefixes

    x = extract_concept_scheme(
        wb["Concept Scheme"], prefixes, template_version, error_format
    )
    if isinstance(x, tuple):
        cs, cs_iri = x
    else:
        return x

    cons = extract_concepts(wb["Concepts"], prefixes, cs_iri)
    if not isinstance(cons, Graph):
        return cons

    cols = extract_collections(wb["Collections"], prefixes, cs_iri)
    if not isinstance(cols, Graph):
        return cols

    extra = extract_additions_concept_properties(
        wb["Additional Concept Properties"], prefixes
    )
    if not isinstance(extra, Graph):
        return extra

    g = cs + cons + cols + extra
    g = add_top_concepts(g)
    g.bind("cs", cs_iri)

    # validate the RDF file
    shacl_graph = Graph().parse(Path(__file__).parent / "vocpub-5.1.ttl")
    v = shacl_validate(g, shacl_graph=shacl_graph, allow_warnings=True)
    if not v[0]:
        return return_error(ShaclValidationError(v[2], v[1]), error_format)

    if output_file is not None:
        g.serialize(destination=str(output_file), format="longturtle")
    else:  # print to std out
        if output_format == "graph":
            return g
        else:
            return g.serialize(format="longturtle")


def rdf_to_excel(
    rdf_file: Path,
    output_file: Optional[Path] = None,
    template_version="0.8.5",
    output_format: TypeLiteral["blob", "file"] = "file",
    error_format: TypeLiteral["python", "cmd", "json"] = "python",
):
    """Converts RDF files to Excel workbooks.

    Parameters:
        rdf_file: Required. An RDF file in one of the common formats understood by RDFLib
        output_file: Optional, default none. A name for an Excel file to output. Must end in .xlsx. Not used if output_format set to blob
        template_version: Optional, default 0.8.5. Currently only 0.8.5 and 0.8.5.GA are supported
        output_format: Optional, default file. Whether to return a binary blob (openpyxl Workbook instance) or write results to file.
        error_format: Optional, default python. the kind of errors to return: python is Python, cmd is command line-formatted string, json is stringified JSON

    Returns:
        output_format or an error in one of the error_formats
    """

    # value checkers
    if not rdf_file.name.endswith(tuple(RDF_FILE_ENDINGS.keys())):
        error = ValueError(
            "Files for conversion to Excel must end with one of the RDF file formats: '{}'".format(
                "', '".join(RDF_FILE_ENDINGS.keys())
            )
        )
        return return_error(error, error_format)

    if output_file is not None:
        if not Path(output_file).suffix == ".xlsx":
            error = ValueError("If specifying an output_file, it must end with .xlsx")
            return return_error(error, error_format)

    if template_version not in ["0.8.5", "0.8.5.GA"]:
        error = ValueError(
            f"This converter can only handle templates with versions 0.8.5 or 0.8.5.GA, not {template_version}"
        )
        return return_error(error, error_format)

    allowed_output_formats = ["blob", "file"]
    if output_format not in allowed_output_formats:
        error = ValueError(
            f"output_format must be on of {', '.join(allowed_output_formats)}, not {output_format}"
        )
        return return_error(error, error_format)

    allowed_error_formats = ["python", "cmd", "json"]
    if error_format not in allowed_error_formats:
        error = ValueError(
            f"error_format must be on of {', '.join(allowed_error_formats)}, not {error_format}"
        )
        return return_error(error, error_format)

    # load the RDF file
    try:
        g = Graph().parse(str(rdf_file), format=RDF_FILE_ENDINGS[rdf_file.suffix])
    except Exception as e:
        return return_error(e, error_format)

    g.bind("ex", "http://example.com/")
    ns = g.namespace_manager

    # validate the RDF file
    shacl_graph = Graph().parse(Path(__file__).parent / "vocpub-5.1.ttl")
    v = shacl_validate(g, shacl_graph=shacl_graph, allow_warnings=True)
    if not v[0]:
        return return_error(ShaclValidationError(v[2], v[1]), error_format)

    # load the template
    fn = (
        "VocExcel-template-085-GA.xlsx"
        if template_version == "0.8.5.GA"
        else "VocExcel-template-085.xlsx"
    )
    wb = load_workbook(Path(__file__).parent / "templates" / fn)

    # Concept Scheme
    ws = wb["Concept Scheme"]
    cs_iri = g.value(predicate=RDF.type, object=SKOS.ConceptScheme)
    ws["B3"] = cs_iri
    ws["B4"] = g.value(subject=cs_iri, predicate=SKOS.prefLabel)
    ws["B5"] = g.value(subject=cs_iri, predicate=SKOS.definition)
    ws["B5"].alignment = Alignment(wrap_text=True)
    ws["B6"] = date_parser(
        str(
            g.value(subject=cs_iri, predicate=SDO.dateCreated)
            or g.value(subject=cs_iri, predicate=DCTERMS.created)
        )
    )
    ws["B7"] = date_parser(
        str(
            g.value(subject=cs_iri, predicate=SDO.dateModified)
            or g.value(subject=cs_iri, predicate=DCTERMS.modified)
        )
    )
    xl_hyperlink(
        ws["B8"],
        g.value(subject=cs_iri, predicate=SDO.creator)
        or g.value(subject=cs_iri, predicate=DCTERMS.creator),
    )
    ws["B9"] = g.value(subject=cs_iri, predicate=SDO.publisher) or g.value(
        subject=cs_iri, predicate=DCTERMS.publisher
    )
    # custodian
    for o in g.objects(subject=cs_iri, predicate=PROV.qualifiedAttribution):
        for p, o2 in g.predicate_objects(subject=o):
            if p == DATAROLES.custodian:
                ws["B10"] = str(o2)
    ws["B11"] = g.value(subject=cs_iri, predicate=SDO.version) or g.value(
        subject=cs_iri, predicate=OWL.versionInfo
    )
    ws["B12"] = g.value(subject=cs_iri, predicate=SKOS.historyNote)
    ws["B13"] = g.value(subject=cs_iri, predicate=SDO.citation)
    for o in g.objects(subject=cs_iri, predicate=PROV.qualifiedDerivation):
        for p, o2 in g.predicate_objects(subject=o):
            if p == PROV.entity:
                ws["B14"] = str(o2)
            if p == PROV.hadRole:
                for k, v in VOCDERMODS.items():
                    if v == str(o2):
                        ws["B15"] = k
    ws["B16"] = ", ".join(
        [str(x) for x in g.objects(subject=cs_iri, predicate=SDO.keywords)]
    )
    ws["B17"] = str(g.value(subject=cs_iri, predicate=SDO.status)).split("/")[-1]
    if template_version == "0.8.5.GA":
        ws["B18"] = str(g.value(subject=cs_iri, predicate=SDO.identifier))

    # Concepts
    ws = wb["Concepts"]
    r = 4
    cs = sorted(list(g.subjects(predicate=RDF.type, object=SKOS.Concept)))
    for c in cs:
        xl_hyperlink(ws[f"A{r}"], ns.curie(c))
        ws[f"B{r}"] = g.value(subject=c, predicate=SKOS.prefLabel)
        ws[f"B{r}"].font = Font(size=14)
        ws[f"C{r}"] = g.value(subject=c, predicate=SKOS.definition)
        ws[f"C{r}"].alignment = Alignment(wrap_text=True)
        ws[f"C{r}"].font = Font(size=14)

        alt_labels = []
        for alt_label in g.objects(subject=c, predicate=SKOS.altLabel):
            alt_labels.append(alt_label)
        ws[f"D{r}"] = ",\n".join(alt_labels)
        ws[f"D{r}"].font = Font(size=14)

        for s, o in g.subject_objects(SKOS.broader):
            g.add((o, SKOS.narrower, s))
        narrowers = []
        for narrower in g.objects(subject=c, predicate=SKOS.narrower):
            narrowers.append(narrower)
        ws[f"E{r}"] = ",\n".join([ns.curie(x) for x in narrowers])
        ws[f"E{r}"].font = Font(size=14)

        hn = g.value(subject=c, predicate=SKOS.historyNote)
        if hn is not None:
            ws[f"F{r}"] = hn
            ws[f"F{r}"].font = Font(size=14)

        cit = g.value(subject=c, predicate=SDO.citation)
        if cit is not None:
            ws[f"G{r}"] = hn
            ws[f"G{r}"].font = Font(size=14)

        is_defined_by = g.value(subject=c, predicate=RDFS.isDefinedBy)
        if is_defined_by != "" and is_defined_by != cs_iri:
            xl_hyperlink(ws[f"H{r}"], is_defined_by)

        status = g.value(subject=c, predicate=SDO.status)
        if status is not None:
            for k, v in STATUSES.items():
                if v == str(status):
                    ws[f"I{r}"] = k
                    ws[f"I{r}"].font = Font(size=14)

        eg = g.value(subject=c, predicate=SKOS.example)
        if eg is not None:
            ws[f"J{r}"] = eg
            ws[f"J{r}"].font = Font(size=14)

        img = g.value(subject=c, predicate=SDO.image)
        if img is not None:
            xl_hyperlink(ws[f"K{r}"], img)

        r += 1

    r = 4
    cs = sorted(list(g.subjects(predicate=RDF.type, object=SKOS.Concept)))
    for c in cs:
        # Concepts - Additional Properties
        if (
            c,
            SKOS.relatedMatch
            | SKOS.closeMatch
            | SKOS.exactMatch
            | SKOS.narrowMatch
            | SKOS.broadMatch
            | SKOS.notation,
            None,
        ) in g:
            ws = wb["Additional Concept Properties"]
            xl_hyperlink(ws[f"A{r}"], ns.curie(c))

            fill_cell_with_list_of_curies(f"B{r}", ws, g, c, SKOS.relatedMatch)
            fill_cell_with_list_of_curies(f"C{r}", ws, g, c, SKOS.closeMatch)
            fill_cell_with_list_of_curies(f"D{r}", ws, g, c, SKOS.exactMatch)
            fill_cell_with_list_of_curies(f"E{r}", ws, g, c, SKOS.narrowMatch)
            fill_cell_with_list_of_curies(f"F{r}", ws, g, c, SKOS.broadMatch)

            notations = []
            datatypes = []
            for i in g.objects(subject=c, predicate=SKOS.notation):
                notations.append(str(i))
                datatypes.append(i.datatype)
            ws[f"G{r}"] = ",\n".join(notations)
            ws[f"G{r}"].font = Font(size=14)
            ws[f"H{r}"] = ",\n".join(datatypes)
            ws[f"H{r}"].font = Font(size=14)

            r += 1

    # Collections
    ws = wb["Collections"]
    r = 4
    cols = sorted(list(g.subjects(predicate=RDF.type, object=SKOS.Collection)))
    for col in cols:
        xl_hyperlink(ws[f"A{r}"], ns.curie(col))
        ws[f"B{r}"] = g.value(subject=col, predicate=SKOS.prefLabel)
        ws[f"B{r}"].font = Font(size=14)
        ws[f"C{r}"] = g.value(subject=col, predicate=SKOS.definition)

        fill_cell_with_list_of_curies(f"D{r}", ws, g, col, SKOS.member)

        hn = g.value(subject=col, predicate=SKOS.historyNote)
        if hn is not None:
            ws[f"F{r}"] = hn
            ws[f"F{r}"].font = Font(size=14)

        r += 1

    # Namespaces
    ws = wb["Prefixes"]
    common_prefixes = [
        "ex",
        "brick",
        "csvw",
        "dc",
        "dcat",
        "dcmitype",
        "dcterms",
        "dcam",
        "doap",
        "foaf",
        "geo",
        "odrl",
        "org",
        "prof",
        "prov",
        "qb",
        "schema",
        "sh",
        "skos",
        "sosa",
        "ssn",
        "time",
        "vann",
        "void",
        "wgs",
        "owl",
        "rdf",
        "rdfs",
        "xsd",
        "xml",
    ]
    r = 4
    for pre, ns in g.namespaces():
        if pre not in common_prefixes:
            ws[f"A{r}"] = pre
            ws[f"A{r}"].font = Font(size=14)
            xl_hyperlink(ws[f"B{r}"], ns)
            r += 1

    if output_format == "blob":
        return wb
    else:
        # save the output
        if output_file is None:
            wb.save(str(rdf_file.with_suffix(".xlsx")))
        else:
            wb.save(str(output_file))

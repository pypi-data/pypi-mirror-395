"""
Comparison functions for Excel spreadsheets
"""
import numpy as np
import shutil as sh
import openpyxl as xl 
import matplotlib as mpl

from zdev.colors import cRed, cBlue, cCitrus, cGreen

hx = lambda x: mpl.colors.rgb2hex(x).strip('#')  
# Note: Modified conversion since 'xl.styles' uses hex codes w/o initial '#' (e.g. 'ff0000')!


# INTERNAL PARAMETERS & DEFAULTS

_MAX_ROW_SCAN = int(1e4)

_CS_CHANGES = {
    'font':  xl.styles.Font(italic=True, color=hx(cRed)),  
    'fill':  xl.styles.PatternFill(fill_type='solid', bgColor='000000', fgColor=hx(cCitrus)),
    }
    
_CS_INSERTS = {
    'font':  xl.styles.Font(italic=True, color=hx(cBlue)), 
    'fill':  xl.styles.PatternFill(fill_type='solid', bgColor='000000', fgColor='F0FBFF'),
    }

_CS_MOVES = {
    'font':  xl.styles.Font(italic=True, color=hx(cGreen)), 
    'fill':  xl.styles.PatternFill(fill_type='solid', bgColor='000000', fgColor='F0FFF0'),
    }


def xl_cmp_workbooks(file1, file2, tab=None, cols=None, mode='first', max_rows=None, 
                     allow_inserts=True, allow_moves=False, save_to_file=True, verbose=True):
    """     
    Compares two Excel workbooks for differences.    
    
    Note: Both workbooks are expected to have the same column headers in the first line of the 
    respective 'tab'. Moreover, it is advised to keep the 'allow' flags active, since 
    deactivating these may render unexpected results if the 2nd workbook nevertheless exhibits
    such modifications.    
    
    Args:
        file 1 / 2 : Filenames of Excel workbooks (*.xlsx) to compare.        
        tab: Tab to compare, if desired. Defaults to 'None' (i.e. use first sheet).
        cols (list of int|str, optional): List of columns that should be used for the row-wise 
            comparison. The elements can be given as either 'int' (1-based) or as 'str' in 
            which case the 1st row in the spreadsheet is expected to contain column headers. 
            Defaults to 'None' (i.e. all columns will be checked).
        mode (str, optional): Mode of comparison for changes w/ options 'first'|'all'. Defaults 
            to 'first' to enable faster processing.
        max_rows 
        allow_inserts (bool, optional): Switch for taking inserted rows into account. Defaults 
            to 'True'. 
        allow_moves (bool, optional): Switch for taking moved rows into account. Defaults to 
            'True'. 
        save_to_file (bool, optional): Swicth to store results to a comparison files w/ all 
            modifications being highlighted appropriately ("_cmp.xlsx").
        verbose (bool, optional): Switch to enable debugging messages. Defaults to 'False'.                   
        
    Returns:
        stat (dict): Statistics on modifications found in the sheet of 'file2' acc. to the 
            categories 'ok' / 'changed' / 'inserted' / 'moved'.
        track_changes (list of 2-tuples): All changes that found in the sheet of 'file2'. The
            format of each item is given by a 2-tuple (see also 'xl_cmp_rows()'):
            (row, row2), [(col1, text, text_changed) [, ..., (colN, text, text_changed)]]
        track_inserts (list): Indices of all rows that have been newly inserted in the sheet 
            of 'file2'.
        track_moves ()
            
    """
    
    # open workbooks & select spreadsheets
    WB1 = xl.load_workbook(file1, data_only=True)
    WB2 = xl.load_workbook(file2, data_only=True)
    if (tab is None):
        sheet = WB1.sheetnames[0]
    else:
        sheet = tab
    sh1 = WB1[sheet]
    sh2 = WB2[sheet]
    
    # init
    if (cols is None):
        cols_to_be_checked = list(np.arange(1, 1+sh1.max_column, dtype=int))
    elif (type(cols[0]) is int):
        cols_to_be_checked = cols
    elif (type(cols[0]) is str):
        headers = []
        for c in range(1, 1+sh1.max_column):
            headers.append(sh1.cell(1,c).value)
        cols_to_be_checked = []
        for item in cols:
            cols_to_be_checked.append(1+headers.index(item)) 
    if (max_rows is None):
        max_rows = sh1.max_row
    else:
        max_rows = min(max_rows, sh1.max_row)        
    
    # init
    num_ok = 0
    num_changed = 0
    num_moved = 0
    num_inserted = 0
    track_changes = []  # i.e. actual changes in row items from 'sh1' to 'sh2' (if not mapped to shifts) 
    track_moves = []    # i.e. rows moved in position from 'sh1' to 'sh2'
    track_inserts = []  # i.e. additions to 'sh2' only!
         
    # cycle through all rows (of 'sh1')
    for row in range(1+1, 1+max_rows):  
        row2 = row + num_inserted # account for general shift in rows on 'sh2'
        
        # compare corresponding rows in 'sh1' and 'sh2'
        if (verbose):
            print(f"#{row:d} (vs. #{row2:d})")            
        matching, row_changes = xl_cmp_rows(sh1, sh2, [row, row2], cols_to_be_checked, mode)        
        
        if (matching):
            num_ok += 1
            # if (verbose):
            #     print("  -> ok!")        
        
        else:            
            found_inserts = False
            found_move = False
            
            # (i) check for insertions or moves
            if (allow_inserts):
                max_rows = min(row2+_MAX_ROW_SCAN, sh2.max_row)
                
                # consider scan range (on 'sh2')
                for row2_scan in range(row2+1, 1+max_rows):
                    matching_scan, _ = xl_cmp_rows(sh1, sh2, [row, row2_scan], cols_to_be_checked)
                                        
                    if (matching_scan): # new position of row found, but now...                       
                        
                        if (allow_moves):                            
                            new_inserts = 0 
                            ####### found_inserts = False
                            
                            # ...check on all rows in-between (if they are "known" or inserted?)...
                            for row2_chk in range(row2, row2_scan):
                                for row_orig in range(1, 1+sh1.max_row): # try to find rows ANYWHERE in 'sh1'!
                                    matching_orig, _ = xl_cmp_rows(sh1, sh2, [row_orig, row2_chk])
                                    if (matching_orig):                                                                              
                                        found_move = True # Note: Define as move if at least one known row is found in between!
                                        if (verbose):
                                            print(f"    - found origin of #{row2_chk}")
                                        break
                                if (not matching_orig):                                    
                                    found_inserts = True # Note: Define as insert if none of the rows in-between have been known before!
                                    new_inserts += 1
                                    print(f"    - treat #{row2_chk} as insert!")
                                    
                            if (new_inserts):
                                num_inserted += new_inserts
                            else:
                                num_inserted -= 1 
                            # Note: If no inserts are found, this even has to be reduce by one, 
                            # since then it's a mere move and this is required to enable the 
                            # next proper comparison of 'row' (in 'sh1') to 'row2' ('sh2').
                        
                        # ... or assume rows in-between ARE ALL INSERTED!
                        else:
                            found_inserts = True
                            new_inserts = row2_scan - row2
                            num_inserted += new_inserts                        
                        
                        break # exit 'row2_scan' loop anyway (moved position has been found)   
                
                # result of scan
                if (matching_scan):                    
                    if (found_move):
                        if (verbose):
                            print(f"  -> moved to #{row2_scan}")
                        num_moved += 1 
                        track_moves.append([row, row2_scan])
                    elif (found_inserts):
                        if (verbose):
                            print(f"  -> inserted {new_inserts} new rows (from #{row2} to #{row2_scan-1}) => total: +{num_inserted}")
                        for r in range(row2, row2_scan):
                            track_inserts.append(r) 
                    else:
                        print("THIS SHOULD NOT HAPPEN????")
                # else:
                #     if (verbose):
                #         print(f"  -> no moves/inserts found in scan range (treat as change)")
                    
            # # TODO / FIXME: what if only moves but no inserts are allowd????
            # elif (allow_moves):
            #     pass
            
            # (ii) mark as change
            if ((not allow_inserts) or (not matching_scan)):
                num_changed += 1
                track_changes.append(((row, row2), row_changes))                
                if (verbose):
                    print("  -> changes")
                    for item in row_changes:
                        print(f"     - col {item[0]}:  '{item[1]}'  <-!->  '{item[2]}'  ")
                
        
        
    # cycle through additional rows of 'sh2' 
    # Note: These can only be additions *unless* the san range has been set too small!
    row2 = sh1.max_row+num_inserted
    for r in range(row2+1, 1+sh2.max_row):
        track_inserts.append(r)
    new_inserts = sh2.max_row - (sh1.max_row+num_inserted)
    num_inserted += new_inserts
    if (new_inserts and verbose):
        print(f"Additional {new_inserts} rows at end of 'sh2' (from #{1+row2} to #{sh2.max_row})")
    
    # store results to a new workbook file w/ highlighted modifications
    if (save_to_file):
        
        # make a copy of WB2 & open for operation
        new_fname = file2[:-5]+'_cmp.xlsx'
        sh.copy(file2, new_fname)
        WB_cmp = xl.load_workbook(new_fname, data_only=True)
        sh_cmp = WB_cmp[sheet]
        
        # mark cells as changed
        for item in track_changes:
            row = item[0][1]
            for changes in item[1]:
                col = changes[0]
                sh_cmp.cell(row,col).fill = _CS_CHANGES['fill']
                sh_cmp.cell(row,col).font = _CS_CHANGES['font']
                
        # mark rows as inserted
        for row in track_inserts:
            for col in range(1, 1+sh_cmp.max_column):
                sh_cmp.cell(row,col).fill = _CS_INSERTS['fill']
                sh_cmp.cell(row,col).font = _CS_INSERTS['font']
                
        # mark rows as moved
        for item in track_moves:
             row = item[1]
             for col in range(1, 1+sh_cmp.max_column):
                 sh_cmp.cell(row,col).fill = _CS_MOVES['fill']
                 sh_cmp.cell(row,col).font = _CS_MOVES['font']
            
        # save
        WB_cmp.save(new_fname)
        WB_cmp.close()
        
    # create modification statistics
    stat = {}
    stat['ok'] = num_ok
    stat['changed'] = num_changed
    stat['moved'] = num_moved
    stat['inserted'] = num_inserted 
    
    return stat, track_changes, track_moves, track_inserts




# def xl_cmp_workbooks_WORKING(file1, file2, tab=None, cols=None, mode='first', max_rows=None, 
#                      allow_inserts=True, allow_moves=False, save_to_file=True, verbose=True):
#     """ Retrieves information on available cases (= stationary conditions). 
    
#     Args:
#         --
#         cols (list of int|str, optional): List of columns that should be used for the row-wise 
#             comparison. The elements can be given as either 'int' (1-based) or as 'str' in 
#             which case the 1st row in the spreadsheet is expected to contain column headers. 
#             Defaults to 'None' (i.e. all columns will be checked).
            
#          --- max_cols   
        
#     Returns:
#         cases (dict): All available cases, w/ 'Cn' as keys and descriptive titles (str) as
#             values.
#     """
    
#     # open workbooks & select spreadsheets
#     WB1 = xl.load_workbook(file1, data_only=True)
#     WB2 = xl.load_workbook(file2, data_only=True)
#     if (tab is None):
#         sheet = WB1.sheetnames[0]
#     else:
#         sheet = tab
#     sh1 = WB1[sheet]
#     sh2 = WB2[sheet]
    
#     # init
#     if (cols is None):
#         cols_to_be_checked = list(np.arange(1, 1+sh1.max_column, dtype=int))
#     elif (type(cols[0]) is int):
#         cols_to_be_checked = cols
#     elif (type(cols[0]) is str):
#         headers = []
#         for c in range(1, 1+sh1.max_column):
#             headers.append(sh1.cell(1,c).value)
#         cols_to_be_checked = []
#         for item in cols:
#             cols_to_be_checked.append(1+headers.index(item)) 
#     if (max_rows is None):
#         max_rows = sh1.max_row
#     else:
#         max_rows = min(max_rows, sh1.max_row)        
    
#     # init
#     num_ok = 0
#     num_changed = 0
#     num_shifted = 0
#     num_inserted = 0
#     track_changes = []  # i.e. actual changes in row items from 'sh1' to 'sh2' (if not mapped to shifts) 
#     track_shifts = []   # i.e. rows moved in position from 'sh1' to 'sh2'
#     track_inserts = []  # i.e. additions to 'sh2' only!
         
#     # cycle through all rows (of 'sh1')   
#     for row in range(1+1, 1+max_rows):
        
#         # compare corresponding rows in 'sh1' and 'sh2'
#         row2 = row + num_inserted
#         matching, row_changes = xl_cmp_rows(sh1, sh2, [row, row2], cols_to_be_checked, mode)
#         if (verbose):
#             print(f"#{row:d} (cmp vs. #{row2:d})")        
        
#         # no direct match...
#         if (not matching):
#             found_inserts = False
#             found_shifts = False
            
#             if (allow_inserts):
                
#                 # scan for shifted row (within scan range)...
#                 max_rows = min(row2+_MAX_ROW_SCAN, sh2.max_row)
#                 for row2_scan in range(row2+1, 1+max_rows):                 
#                     matching_scan, _ = xl_cmp_rows(sh1, sh2, [row, row2_scan], cols_to_be_checked, 'first')
                    
#                     if (matching_scan):
#                         if (not allow_moves): # Note: ALL rows in between -> must be inserted!                            
#                             found_inserts = True
#                             new_inserts = row2_scan - row2
#                             num_inserted += new_inserts
#                         else:
#                             old_origins = 0
#                             new_inserts = 0 
#                             found_inserts = False
#                             for row2_chk in range(row2, row2_scan): # look for all rows inbetween... 
#                                 print(f"  -> IN BETWEEN: what is with row {row2_chk}?")
#                                 # ...try to find in rest of table                          
#                                 for lulu in range(1, 1+sh1.max_row): #### SEARCH IN FULL sh1 LIST!!!!!!                                    
#                                     matching_move, _ = xl_cmp_rows(sh1, sh2, [lulu, row2_chk], None, 'first')
#                                     if (matching_move):
#                                         print(f"    -> found origin of #{row2_chk}")
#                                         old_origins += 1
#                                         found_shifts = True
#                                         break
#                                 if (not matching_move):
#                                     new_inserts += 1
#                                     found_inserts = True
#                                     print(f"    -> count #{row2_chk} as insert!")
                                    
#                             print(f"olds: {old_origins} | news: {new_inserts}")                            
#                             if (new_inserts):
#                                 num_inserted += new_inserts
#                             else:
#                                 num_inserted -= 1 
#                                 # Note: Even REDUCE by one, since a mere move (but no inserts, only old has been detected)
#                                 # Note: This is required to enable the next step of 'row' (sh1) vs. 'row2' (sh2) comparison to match!
                        
#                         # num_shifted += 1 # row that WAS really moved
#                         # track_shifts.append([row, row2_scan])
#                         break # exit 'row2_scan' loop (moved position has been found either way)   
                
#                 # result of scan
#                 if (matching_scan):
#                     print(f"  -> shifted to #{row2_scan}")
#                     num_shifted += 1 # row that WAS really moved
#                     track_shifts.append([row, row2_scan])
#                     if (found_inserts):
#                         print(f"  -> {new_inserts} rows inserted (#{row2} - #{row2_scan-1}), total => +{num_inserted}")
#                         for r in range(row2, row2_scan):
#                             track_inserts.append(r)                   
#                 else:
#                     print(f"  -> no shift found in range (-> change!)")
            
#             # ...otherwise mark as change
#             if ((not matching_scan) and (not found_inserts) and (not found_shifts)):
#                 num_changed += 1
#                 track_changes.append(((row, row2), row_changes))                
#                 print("  -> changes")
#                 for item in row_changes:
#                     print(f"       col {item[0]}:  '{item[1]}'  <-!->  '{item[2]}'  ")
                
#         else:
#             num_ok += 1
#             # if (verbose):
#             #     print("  -> ok!")
        
#     # cycle through additional rows of 'sh2' 
#     # (Note: These can only be additions *unless* the _MAX_ROW_SCAN range has been set too small!) 
#     row2 = sh1.max_row+num_inserted
#     for r in range(row2+1, 1+sh2.max_row):
#         track_inserts.append(r)
#     new_inserts = sh2.max_row - (sh1.max_row+num_inserted)
#     num_inserted += new_inserts
#     if (new_inserts and verbose):
#         print(f"Additional {new_inserts} rows at end of 'sh2' (from #{1+row2} to #{sh2.max_row})")
    
#     # store results to a new workbook file = highlighting the changes!
#     if (save_to_file):
        
#         # make a copy of WB2 & open for operation
#         new_fname = file2[:-5]+'_cmp.xlsx'
#         sh.copy(file2, new_fname)
#         WB_cmp = xl.load_workbook(new_fname)#, data_only=True)
#         sh_cmp = WB_cmp[sheet]        
        
#         # mark rows as inserted
#         for row in track_inserts:
#             for col in range(1, 1+sh_cmp.max_column):
#                 # ftsize = sh_cmp.cell(row,col).font['size']
#                 # print(ftsize)
#                 sh_cmp.cell(row,col).fill = _CS_INSERTS['fill']
#                 sh_cmp.cell(row,col).font = _CS_INSERTS['font']
        
#         # mark row as changed
#         for item in track_changes:
#             row = item[0][1]
#             for changes in item[1]:
#                 col = changes[0]
#                 # ftsize = sh_cmp.cell(row,col).font['size']
#                 # print(ftsize)
#                 sh_cmp.cell(row,col).fill = _CS_CHANGES['fill']
#                 sh_cmp.cell(row,col).font = _CS_CHANGES['font']
                
#         # mark shifts
#         for item in track_shifts:
#              row = item[1]
#              print(row)
#              for col in range(1, 1+sh_cmp.max_column):
#                  sh_cmp.cell(row,col).fill = _CS_MOVES['fill']
#                  sh_cmp.cell(row,col).font = _CS_MOVES['font']
            
#         # save   
#         WB_cmp.save(new_fname)
#         WB_cmp.close()
    
    
#     # Note: Actually 'num_shifted' should be added to 'num_ok' since the shifts have been 
#     #       accounted for and insertions are tracked separately!
    
#     stat = {}
#     stat['ok'] = num_ok
#     stat['changed'] = num_changed
#     stat['shifted'] = num_shifted
#     stat['inserted'] = num_inserted    
    
#     return stat, track_changes, track_shifts, track_inserts



                       
                        
        
# TODO: keep 'row' fixed for 'sh1', but compare with next rows in 'sh2' (up to _MAX_ROW_SCAN)
#       iFF successful, store effective shift for any further progress (in 'sh1')
#                       and mark lines of 'sh2' in between as additions/insertions
#
# TODO: how to do this backwards? (i.e. for deletions of 'sh1' contents in 'sh2')?
#       --> not likely for lists with "growing nature"


def xl_cmp_rows(sh1, sh2, rows, cols=None, mode='first'):
    """ Compares row(s) in spreadsheets 'sh1' and 'sh2' for differences (in all defined cols). 
    
    Args:
        sh1 (:obj:): 1st spreadsheet ~ "original" (class 'xl.worksheet.worksheet.Worksheet').
        sh2 (:obj:): 2nd spreadsheet ~ "modified" (class 'xl.worksheet.worksheet.Worksheet').
        rows (int or 2-tuple): Row index/indices of both sheets. If only one (integer) is used,
            it is applied for both 'sh1' and 'sh2'.
        cols (list of int): List of column indices (1-based integer) for which cell-wise 
            comparison between the two sheets is to be performed. Defaults to 'None' (i.e. the
            whole row will be compared). 
        mode (str, optional): Mode of comparison w/ options 'first'|'all'. 'first' will stop
            cell-checking after first mismatch is found, whereas 'all' will continue the check
            for the remaining columns (to include). 
    
    Returns:
        matching (bool): Flag if both rows match completely (for all 'cols').
        changes (list of 3-tuples): List of the first or all mismatches (dep. on 'mode').
        
        [ (row1, row2), (col, value1, value2) [, ... ] ]
    """   
    
    # init & check consistency
    matching = True
    changes = []
    if (type(rows) is int):
        row1 = rows
        row2 = rows
    else:
        [row1, row2] = rows
    if (cols is None):
        cols_to_be_checked = list(np.arange(1, 1+sh1.max_column, dtype=int))
    else:
        cols_to_be_checked = cols
        
    # compare rows of both sheets (cell-wise) & track changes
    for n, col in enumerate(cols_to_be_checked):
        if (sh1.cell(row1,col).value != sh2.cell(row2,col).value):
            matching = False
            changes.append( (col, sh1.cell(row1,col).value, sh2.cell(row2,col).value) )
            if (mode == 'first'):
                pass
            elif (mode == 'all'):
                for nn in range(n+1, len(cols_to_be_checked)):
                    col_next = cols_to_be_checked[nn]
                    if (sh1.cell(row1,col_next).value != sh2.cell(row2,col_next).value):
                        changes.append( (col_next, sh1.cell(row1,col_next).value, sh2.cell(row2,col_next).value) )
            break
        else:
            continue # w/ next row

    return matching, changes

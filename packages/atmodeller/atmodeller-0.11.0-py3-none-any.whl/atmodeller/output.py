#
# Copyright 2024 Dan J. Bower
#
# This file is part of Atmodeller.
#
# Atmodeller is free software: you can redistribute it and/or modify it under the terms of the GNU
# General Public License as published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# Atmodeller is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY; without
# even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU
# General Public License for more details.
#
# You should have received a copy of the GNU General Public License along with Atmodeller. If not,
# see <https://www.gnu.org/licenses/>.
#
"""Output"""

import logging
import pickle
from pathlib import Path
from typing import Optional

import jax.numpy as jnp
import numpy as np
import pandas as pd
from jaxmod.constants import GAS_CONSTANT
from jaxmod.solvers import MultiAttemptSolution
from jaxtyping import Array, Float
from openpyxl.styles import PatternFill

from atmodeller import override
from atmodeller.containers import Parameters
from atmodeller.output_core import Output, broadcast_arrays_in_dict
from atmodeller.type_aliases import NpArray, NpBool, NpFloat, NpInt

logger: logging.Logger = logging.getLogger(__name__)


class OutputDisequilibrium(Output):
    """Output disequilibrium calculations

    Args:
        parameters: Parameters
        solution: Solution
    """

    @override
    def asdict(self) -> dict[str, dict[str, NpArray]]:
        """All outputs in a dictionary, with caching.

        Additionally includes the disequilibrium group, compared to the base class.

        Returns:
            Dictionary of all output
        """
        out: dict[str, dict[str, NpArray]] = super().asdict()

        out["disequilibrium"] = self.disequilibrium_asdict()

        self._cached_dict = out  # Re-cache result for faster re-accessing

        return out

    def disequilibrium_asdict(self) -> dict[str, NpArray]:
        """Gets the reaction disequilibrium as a dictionary.

        Returns:
            Reaction disequilibrium as a dictionary
        """
        reaction_mask: NpBool = self.reaction_mask()
        residual: NpFloat = np.asarray(self.vmapf.objective_function(jnp.asarray(self.solution)))

        # Number of True entries per row (must be same for all rows)
        n_cols: NpInt = reaction_mask.sum(axis=1)[0]
        # logger.debug("n_cols = %s", n_cols)
        # Convert boolean mask to sorted column indices for each row
        col_indices: NpInt = np.argsort(~reaction_mask, axis=1)[:, :n_cols]
        # logger.debug("col_indices = %s", col_indices)
        # Gather the True entries in order
        compressed: NpFloat = np.take_along_axis(residual, col_indices, axis=1)
        # logger.debug("compressed = %s", compressed)

        # To compute the limiting reactant/product in each reaction we need to know the
        # availability of each species. We will ignore condensates later because their stability
        # criteria prevents a simple calculation of what is limiting the reaction.
        number_fraction: NpFloat = self.number_moles / np.sum(
            self.number_moles, axis=1, keepdims=True
        )
        # logger.debug("number_fraction = %s", number_fraction)
        reaction_matrix: NpFloat = self.parameters.species_network.reaction_matrix
        # logger.debug("reaction_matrix = %s", reaction_matrix)

        out: dict[str, NpArray] = {}

        for jj in range(n_cols):
            # logger.debug("Working on reaction %d", jj)
            per_mole_of_reaction: NpFloat = compressed[:, jj] * GAS_CONSTANT * self.temperature
            stoich: NpFloat = reaction_matrix[jj]
            # logger.debug("stoich = %s", stoich)

            # Normalised ratios for limiting species (ignore divide-by-zero warnings)
            with np.errstate(divide="ignore"):
                ratios: NpFloat = np.where(stoich != 0, number_fraction / stoich, np.nan)
            # logger.debug("ratios = %s", ratios)
            limiting: NpFloat = np.full_like(per_mole_of_reaction, np.nan)
            # logger.debug("limiting (full_like) = %s", limiting)

            # Initialise with None placeholders for every row
            limiting_species_names: list[Optional[str]] = [None] * residual.shape[0]
            limiting_species_type: list[Optional[str]] = [None] * residual.shape[0]

            # Backward-favoured: products limit
            mask_back: NpBool = per_mole_of_reaction > 0
            # logger.debug("mask_back = %s", mask_back)
            if np.any(mask_back):
                # Subarray of only product species for backward-favoured reactions
                sub_ratios: NpFloat = ratios[mask_back][:, stoich > 0]
                # Column indices of product species in the full array
                sub_cols: NpInt = np.where(stoich > 0)[0]
                # Value of limiting species
                limiting[mask_back] = np.min(sub_ratios, axis=1)
                # logger.debug("limiting[mask_back] = %s", limiting[mask_back])
                # Column index (within subarray) of limiting species
                min_idx_within: NpInt = np.argmin(sub_ratios, axis=1)
                # Map back to global indices in ratios / species_names
                min_idx_global: NpInt = sub_cols[min_idx_within]
                # Get the actual species names
                for row_idx, species_idx in zip(np.where(mask_back)[0], min_idx_global):
                    limiting_species_names[row_idx] = self.species.species_names[species_idx]
                    limiting_species_type[row_idx] = "Product"
                # logger.debug("limiting_species_names (back) = %s", limiting_species_names)

            # Forward-favoured: reactants limit
            mask_fwd: NpBool = ~mask_back
            # logger.debug("mask_fwd = %s", mask_fwd)
            if np.any(mask_fwd):
                sub_ratios: NpFloat = ratios[mask_fwd][:, stoich < 0]
                sub_cols: NpInt = np.where(stoich < 0)[0]
                # Limiting species is the largest negative ratio among reactants (closest to zero)
                limiting[mask_fwd] = np.max(sub_ratios, axis=1)
                # logger.debug("limiting[mask_fwd] = %s", limiting[mask_fwd])
                max_idx_within: NpInt = np.argmax(sub_ratios, axis=1)
                max_idx_global: NpInt = sub_cols[max_idx_within]
                # Get the actual species names
                for row_idx, species_idx in zip(np.where(mask_fwd)[0], max_idx_global):
                    limiting_species_names[row_idx] = self.species.species_names[species_idx]
                    limiting_species_type[row_idx] = "Reactant"
                # logger.debug("limiting_species_names (fwd) = %s", limiting_species_names)

            # Compute the energy per mole of atmosphere
            energy_per_mol_atmosphere: NpFloat = per_mole_of_reaction * limiting
            logger.debug("energy_per_mol_atmosphere = %s", energy_per_mol_atmosphere)

            out[f"Reaction_{jj}"] = per_mole_of_reaction

            if self.species.gas_only:
                out[f"Reaction_{jj}_per_atmosphere"] = energy_per_mol_atmosphere
                out[f"Reaction_{jj}_limiting_species"] = np.array(limiting_species_names)
                out[f"Reaction_{jj}_limiting_species_role"] = np.array(limiting_species_type)

        return out


class OutputSolution(Output):
    """Output equilibrium solution(s)

    Args:
        parameters: Parameters
        solution: Solution
        multi_attempt_solution: :class:`~jaxmod.solvers.MultiAttemptSolution` object
    """

    def __init__(
        self,
        parameters: Parameters,
        solution: Float[Array, "batch solution"],
        multi_attempt_solution: MultiAttemptSolution,
    ):
        super().__init__(parameters, solution)
        self.multi_attempt_solution: MultiAttemptSolution = multi_attempt_solution

    @override
    def asdict(self) -> dict[str, dict[str, NpArray]]:
        """All outputs in a dictionary, with caching.

        Returns:
            Dictionary of all output
        """
        out: dict[str, dict[str, NpArray]] = super().asdict()

        # Temperature and pressure have already been expanded to the number of solutions
        temperature: NpFloat = out["state"]["temperature"]
        pressure: NpFloat = out["state"]["pressure"]

        out["constraints"] = {}
        out["constraints"] |= broadcast_arrays_in_dict(
            self.parameters.mass_constraints.asdict(), self.number_solutions
        )
        out["constraints"] |= broadcast_arrays_in_dict(
            self.parameters.fugacity_constraints.asdict(temperature, pressure),
            self.number_solutions,
        )
        out["residual"] = self.residual_asdict()  # type: ignore since keys are int

        out["solver"] = {
            "status": np.asarray(self.multi_attempt_solution.solver_success),
            "steps": np.asarray(self.multi_attempt_solution.num_steps),
            "attempts": np.asarray(self.multi_attempt_solution.attempts),
            "converged": np.asarray(self.multi_attempt_solution.converged),
        }

        self._cached_dict = out  # Re-cache result for faster re-accessing

        return out

    @override
    def to_dataframes(self, drop_unsuccessful: bool = False) -> dict[str, pd.DataFrame]:
        """Gets the output in a dictionary of dataframes.

        Args:
            drop_unsuccessful: Drop models that did not solve. Defaults to ``False``.

        Returns:
            Output in a dictionary of dataframes
        """
        dataframes: dict[str, pd.DataFrame] = super().to_dataframes()

        if drop_unsuccessful:
            logger.info("Dropping models that did not solve")
            dataframes: dict[str, pd.DataFrame] = self._drop_unsuccessful_solves(dataframes)

        return dataframes

    @override
    def to_excel(
        self, file_prefix: Path | str = "atmodeller_out", drop_unsuccessful: bool = False
    ) -> None:
        """Writes the output to an Excel file.

        Compared to the base class, this highlights rows where the solver failed to find a
        a solution if ``drop_successful = False``.

        Args:
            file_prefix: Prefix of the output file. Defaults to ``atmodeller_out``.
            drop_unsuccessful: Drop models that did not solve. Defaults to ``False``.
        """
        logger.info("Writing output to excel")
        out: dict[str, pd.DataFrame] = self.to_dataframes(drop_unsuccessful)
        output_file: Path = Path(f"{file_prefix}.xlsx")

        # Convenient to highlight rows where the solver failed to find a solution for follow-up
        # analysis. Define a fill colour for highlighting rows (e.g., yellow)
        highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Get the indices where the successful_solves mask is False
        unsuccessful_indices: NpArray = np.where(
            np.asarray(self.multi_attempt_solution.solver_success) == False  # noqa: E712
        )[0]

        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            for df_name, df in out.items():
                df.to_excel(writer, sheet_name=df_name, index=True)
                sheet = writer.sheets[df_name]

                # Apply highlighting to the rows where the solver failed to find a solution
                for idx in unsuccessful_indices:
                    # Highlight the entire row (starting from index 2 to skip header row)
                    for col in range(1, len(df.columns) + 2):
                        # row=idx+2 because Excel is 1-indexed and row 1 is the header
                        cell = sheet.cell(row=idx + 2, column=col)
                        cell.fill = highlight_fill

        logger.info("Output written to %s", output_file)

    @override
    def to_pickle(
        self, file_prefix: Path | str = "atmodeller_out", drop_unsuccessful: bool = False
    ) -> None:
        """Writes the output to a pickle file.

        Args:
            file_prefix: Prefix of the output file. Defaults to ``atmodeller_out``.
            drop_unsuccessful: Drop models that did not solve. Defaults to ``False``.
        """
        logger.info("Writing output to pickle")
        out: dict[str, pd.DataFrame] = self.to_dataframes(drop_unsuccessful)
        output_file: Path = Path(f"{file_prefix}.pkl")

        with open(output_file, "wb") as handle:
            pickle.dump(out, handle, protocol=pickle.HIGHEST_PROTOCOL)

        logger.info("Output written to %s", output_file)

    def _drop_unsuccessful_solves(
        self, dataframes: dict[str, pd.DataFrame]
    ) -> dict[str, pd.DataFrame]:
        """Drops unsuccessful solves.

        Args:
            dataframes: Dataframes from which to drop unsuccessful models

        Returns:
            Dictionary of dataframes without unsuccessful models
        """
        return {
            key: df.loc[np.asarray(self.multi_attempt_solution.solver_success)]
            for key, df in dataframes.items()
        }

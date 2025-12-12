import unittest
import openpyxl.styles
import openpyxl.utils

from hu2.excel2 import ExcelTable, ExcelStyle, COLOR_GREEN, COLOR_DARKGREEN, COLOR_DARKBLUE

headFont = openpyxl.styles.Font(color=COLOR_GREEN,bold=True,size=16)
headAlignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

headFill2 = openpyxl.styles.GradientFill(stop=('FFFFFF','99ccff','000000'))  #渐变颜色
headFill  = openpyxl.styles.PatternFill('solid',fgColor=COLOR_DARKBLUE,bgColor=COLOR_DARKGREEN) # solid时只有前景色有效
#                                                          前景色                  背景色

class TestExcel(unittest.TestCase):
    def test_excel(self):
        et = ExcelTable('excel.xlsx')
        et.set_column_width({'A': 10, 'B': 20, 'C': 30})
        et.set_head(['Aa', 'Bb', 'Cc'])
        et.set_body([['a', 'b', 'c'], (1, 2, 3)])
        et.set_head_style(ExcelStyle(font=headFont, alignment=headAlignment,fill=headFill,border=True))
        et.set_body_style(ExcelStyle(alignment=headAlignment,border=True))
        et.sheet.row_dimensions[3].height = 40 # 3-第三行，首行为第1行
        et.sheet.title='工作表11' # 设置工作表的名称，默认为 Sheet
        et.save()

    def test_utils(self):
        print(' index=1 -> ',openpyxl.utils.get_column_letter(1))
        print('letter=B -> ',openpyxl.utils.column_index_from_string('B'))

    def test_load_excel(self):
        wb = openpyxl.load_workbook("excel.xlsx")
        sheet = wb.active
        for col in sheet.iter_cols(): # _cells_by_col
            print(type(col),col,type(col[0]), col[0].value)
        for col in sheet.columns: # iter_cols
            print(type(col), col, type(col[0]), col[0].value)

        for row in sheet.iter_rows(): # _cells_by_row
            print(type(row), row, type(row[0]), row[0].value)
        for row1 in sheet.rows: # iter_rows
            print(type(row1), row1, type(row1[0]), row1[0])

        for row in sheet.values: # iter_rows(values_only=True) 返回不是 cell对象 而是 cell.value
            print(type(row), row)

        for t in sheet.tables:  # 插入表格后，可以查看到 不常用
            print(type(t),t)

if __name__ == '__main__':
    unittest.main()
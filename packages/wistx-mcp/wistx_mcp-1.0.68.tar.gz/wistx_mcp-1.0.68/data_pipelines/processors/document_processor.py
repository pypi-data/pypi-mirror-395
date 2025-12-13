"""Document processor using Docling for PDF and web content extraction.

Supports PDF, DOCX, Markdown, TXT, XML, Excel, and CSV formats.
"""

import csv
import tempfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any

import httpx

from ..utils.logger import setup_logger

logger = setup_logger(__name__)

HAS_OPENPYXL = False
HAS_XLRD = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    pass

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    pass

HAS_DOCLING = False
DocumentConverter = None
InputFormat = None
PdfFormatOption = None
PdfParsingOptions = None


class DocumentProcessor:
    """Process PDF and web documents using Docling.

    Extracts structured content from PDFs and web pages for compliance processing.
    """

    def __init__(self):
        """Initialize document processor."""
        self.converter = None
        self._check_docling_availability()

    def _check_docling_availability(self) -> bool:
        """Check if Docling is available at runtime and initialize converter.
        
        Returns:
            True if Docling is available and initialized, False otherwise
        """
        try:
            from docling.document_converter import DocumentConverter as DoclingConverter
            from docling.datamodel.base_models import InputFormat as DoclingInputFormat
            
            self.converter = DoclingConverter(allowed_formats=[DoclingInputFormat.PDF])
            logger.info("Docling converter initialized successfully for PDF and web content processing")
            return True
        except ImportError as e:
            logger.warning("Docling not installed. Install docling: pip install docling")
            logger.warning("PDF processing will be unavailable until Docling is installed.")
            logger.debug("ImportError details: %s", e)
            return False
        except (ValueError, TypeError, KeyError, IOError, AttributeError, RuntimeError) as e:
            logger.error("Failed to initialize Docling converter: %s", e, exc_info=True)
            logger.error("Docling may be installed but converter initialization failed.")
            logger.error("Try: pip install --upgrade docling")
            return False

    def process_pdf(self, pdf_path: Path | str) -> dict[str, Any]:
        """Process a PDF file and extract structured content.

        Args:
            pdf_path: Path to PDF file

        Returns:
            Dictionary with extracted content:
            {
                "text": str,
                "markdown": str,
                "tables": list[dict],
                "metadata": dict
            }
        """
        if self.converter is None:
            logger.error(
                "Docling converter not available. Install docling: pip install docling\n"
                "PDF processing requires Docling for proper extraction."
            )
            return {"text": "", "markdown": "", "tables": [], "metadata": {}}

        pdf_path = Path(pdf_path)

        if not pdf_path.exists():
            logger.error("PDF file not found: %s", pdf_path)
            return {"text": "", "markdown": "", "tables": [], "metadata": {}}

        try:
            logger.info("Processing PDF with Docling: %s", pdf_path)

            result = self.converter.convert(pdf_path)

            text_content = result.document.export_to_text()
            markdown_content = result.document.export_to_markdown()

            tables = []
            for table in result.document.tables:
                tables.append({
                    "content": table.export_to_markdown(),
                    "rows": len(table.rows) if hasattr(table, "rows") else 0,
                })

            metadata = {
                "page_count": len(result.document.pages) if hasattr(result.document, "pages") else 0,
                "file_name": pdf_path.name,
                "file_size": pdf_path.stat().st_size,
            }

            logger.info(
                "Docling extracted %d characters, %d tables from PDF: %s",
                len(text_content),
                len(tables),
                pdf_path.name,
            )

            return {
                "text": text_content,
                "markdown": markdown_content,
                "tables": tables,
                "metadata": metadata,
            }

        except (ValueError, TypeError, KeyError, IOError) as e:
            logger.error("Error processing PDF with Docling %s: %s", pdf_path, e)
            return {"text": "", "markdown": "", "tables": [], "metadata": {}, "error": str(e)}

    def process_web_content(self, html_content: str, url: str) -> dict[str, Any]:
        """Process web content (HTML) using Docling.

        Args:
            html_content: HTML content string
            url: Source URL

        Returns:
            Dictionary with extracted content:
            {
                "text": str,
                "markdown": str,
                "tables": list[dict],
                "metadata": dict
            }
        """
        if self.converter is None:
            logger.debug("Docling not available, falling back to basic HTML parsing")
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(html_content, "html.parser")
            return {
                "text": soup.get_text(separator="\n", strip=True),
                "markdown": "",
                "tables": [],
                "metadata": {"source_url": url, "processed_with": "beautifulsoup"},
            }

        try:
            logger.debug("Processing web content with Docling: %s", url)

            from io import StringIO
            result = self.converter.convert(StringIO(html_content))

            text_content = result.document.export_to_text()
            markdown_content = result.document.export_to_markdown()

            tables = []
            for table in result.document.tables:
                tables.append({
                    "content": table.export_to_markdown(),
                    "rows": len(table.rows) if hasattr(table, "rows") else 0,
                })

            metadata = {
                "source_url": url,
                "processed_with": "docling",
            }

            logger.debug(
                "Docling extracted %d characters, %d tables from web content: %s",
                len(text_content),
                len(tables),
                url,
            )

            return {
                "text": text_content,
                "markdown": markdown_content,
                "tables": tables,
                "metadata": metadata,
            }

        except (ValueError, TypeError, KeyError, IOError) as e:
            logger.warning("Error processing web content with Docling %s: %s", url, e)
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(html_content, "html.parser")
            return {
                "text": soup.get_text(separator="\n", strip=True),
                "markdown": "",
                "tables": [],
                "metadata": {"source_url": url, "processed_with": "beautifulsoup_fallback"},
            }

    def process_document(self, content: str | Path, source_url: str, content_type: str = "auto") -> dict[str, Any]:
        """Process any document type.

        Supports PDF, DOCX, Markdown, TXT, XML, Excel, and CSV formats.

        Args:
            content: Document content (file path or HTML string)
            source_url: Source URL or file path
            content_type: Content type ("pdf", "html", "xml", "excel", "csv", "auto")

        Returns:
            Dictionary with extracted content
        """
        file_path = None
        if isinstance(content, Path):
            file_path = content
        elif isinstance(content, str) and Path(content).exists():
            file_path = Path(content)

        if file_path:
            detected_type = self._detect_file_type(file_path, content_type)
            
            if detected_type == "pdf":
                return self.process_pdf(file_path)
            elif detected_type == "xml":
                return self.process_xml(file_path)
            elif detected_type in ("excel", "xlsx", "xls"):
                return self.process_excel(file_path)
            elif detected_type == "csv":
                return self.process_csv(file_path)
            elif detected_type in ("docx", "markdown", "txt"):
                return self._process_text_file(file_path, detected_type)
            else:
                logger.warning("Unknown file type: %s, falling back to text extraction", detected_type)
                return self._process_text_file(file_path, "txt")
        elif content_type == "html" or content_type == "auto":
            return self.process_web_content(str(content), source_url)
        else:
            logger.warning("Unknown content type: %s, falling back to text extraction", content_type)
            return {
                "text": str(content),
                "markdown": "",
                "tables": [],
                "metadata": {"source_url": source_url, "processed_with": "fallback"},
            }

    def _detect_file_type(self, file_path: Path, provided_type: str = "auto") -> str:
        """Detect file type from extension or provided type.

        Args:
            file_path: File path
            provided_type: Provided content type

        Returns:
            Detected file type string
        """
        if provided_type and provided_type != "auto":
            return provided_type.lower()

        extension = file_path.suffix.lower()
        type_map = {
            ".pdf": "pdf",
            ".docx": "docx",
            ".md": "markdown",
            ".markdown": "markdown",
            ".txt": "txt",
            ".xml": "xml",
            ".xlsx": "excel",
            ".xls": "excel",
            ".csv": "csv",
        }
        return type_map.get(extension, "txt")

    def _process_text_file(self, file_path: Path, file_type: str) -> dict[str, Any]:
        """Process plain text files (TXT, Markdown).

        Args:
            file_path: File path
            file_type: File type

        Returns:
            Dictionary with extracted content
        """
        try:
            content = file_path.read_text(encoding="utf-8")
            markdown = content if file_type == "markdown" else ""
            
            return {
                "text": content,
                "markdown": markdown,
                "tables": [],
                "metadata": {
                    "source_url": str(file_path),
                    "processed_with": "text_reader",
                    "file_type": file_type,
                },
            }
        except Exception as e:
            logger.error("Error processing text file %s: %s", file_path, e)
            return {
                "text": "",
                "markdown": "",
                "tables": [],
                "metadata": {
                    "source_url": str(file_path),
                    "processed_with": "text_reader_error",
                    "error": str(e),
                },
            }

    def process_pdf_directory(self, pdf_dir: Path | str) -> dict[str, dict[str, Any]]:
        """Process all PDF files in a directory.

        Args:
            pdf_dir: Directory containing PDF files

        Returns:
            Dictionary mapping PDF filenames to extracted content
        """
        pdf_dir = Path(pdf_dir)

        if not pdf_dir.exists():
            logger.warning("PDF directory not found: %s", pdf_dir)
            return {}

        pdf_files = list(pdf_dir.glob("*.pdf"))

        if not pdf_files:
            logger.info("No PDF files found in %s", pdf_dir)
            return {}

        logger.info("Processing %d PDF files from %s", len(pdf_files), pdf_dir)

        results = {}

        for pdf_file in pdf_files:
            content = self.process_pdf(pdf_file)
            results[pdf_file.name] = content

        return results

    def process_xml(self, xml_path: Path | str) -> dict[str, Any]:
        """Process XML file.

        Extracts text content and structure from XML files.

        Args:
            xml_path: Path to XML file

        Returns:
            Dictionary with extracted content
        """
        xml_path = Path(xml_path)
        
        try:
            tree = ET.parse(xml_path)
            root = tree.getroot()

            text_parts = []
            markdown_parts = []
            structure = []

            def extract_element(elem: ET.Element, depth: int = 0) -> None:
                """Recursively extract element content."""
                indent = "  " * depth
                tag = elem.tag
                text = (elem.text or "").strip()
                tail = (elem.tail or "").strip()

                if text:
                    text_parts.append(text)
                    markdown_parts.append(f"{indent}**{tag}**: {text}")

                structure.append({
                    "tag": tag,
                    "attributes": elem.attrib,
                    "text": text,
                    "depth": depth,
                })

                for child in elem:
                    extract_element(child, depth + 1)

                if tail:
                    text_parts.append(tail)

            extract_element(root)

            text_content = "\n".join(text_parts)
            markdown_content = "\n\n".join(markdown_parts) if markdown_parts else ""

            namespaces = {}
            if root.tag.startswith("{"):
                namespace_uri = root.tag[1:].split("}")[0]
                namespaces["default"] = namespace_uri

            return {
                "text": text_content,
                "markdown": markdown_content,
                "tables": [],
                "structure": structure,
                "metadata": {
                    "source_url": str(xml_path),
                    "processed_with": "xml_parser",
                    "root_tag": root.tag,
                    "namespaces": namespaces,
                    "element_count": len(structure),
                },
            }
        except ET.ParseError as e:
            logger.error("XML parse error for %s: %s", xml_path, e)
            return {
                "text": "",
                "markdown": "",
                "tables": [],
                "metadata": {
                    "source_url": str(xml_path),
                    "processed_with": "xml_parser_error",
                    "error": f"Parse error: {str(e)}",
                },
            }
        except Exception as e:
            logger.error("Error processing XML file %s: %s", xml_path, e)
            return {
                "text": "",
                "markdown": "",
                "tables": [],
                "metadata": {
                    "source_url": str(xml_path),
                    "processed_with": "xml_parser_error",
                    "error": str(e),
                },
            }

    def process_excel(self, excel_path: Path | str) -> dict[str, Any]:
        """Process Excel file (.xlsx or .xls).

        Extracts data from all sheets and converts to markdown tables.

        Args:
            excel_path: Path to Excel file

        Returns:
            Dictionary with extracted content
        """
        excel_path = Path(excel_path)
        extension = excel_path.suffix.lower()

        if extension == ".xlsx":
            return self._process_xlsx(excel_path)
        elif extension == ".xls":
            return self._process_xls(excel_path)
        else:
            logger.warning("Unsupported Excel format: %s", extension)
            return {
                "text": "",
                "markdown": "",
                "tables": [],
                "metadata": {
                    "source_url": str(excel_path),
                    "processed_with": "excel_parser_error",
                    "error": f"Unsupported format: {extension}",
                },
            }

    def _process_xlsx(self, excel_path: Path) -> dict[str, Any]:
        """Process .xlsx file using openpyxl.

        Args:
            excel_path: Path to .xlsx file

        Returns:
            Dictionary with extracted content
        """
        if not HAS_OPENPYXL:
            logger.warning("openpyxl not installed. Install with: pip install openpyxl")
            return {
                "text": "",
                "markdown": "",
                "tables": [],
                "metadata": {
                    "source_url": str(excel_path),
                    "processed_with": "excel_parser_error",
                    "error": "openpyxl not installed",
                },
            }

        try:
            workbook = openpyxl.load_workbook(excel_path, data_only=True)
            
            text_parts = []
            markdown_parts = []
            sheets_data = []
            total_cells = 0

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                rows = []
                cell_values = []

                for row in sheet.iter_rows(values_only=True):
                    row_data = []
                    row_text = []
                    for cell_value in row:
                        if cell_value is not None:
                            cell_str = str(cell_value)
                            row_data.append(cell_str)
                            row_text.append(cell_str)
                    if row_data:
                        rows.append(row_data)
                        cell_values.extend(row_text)

                if rows:
                    markdown_table = self._rows_to_markdown_table(rows)
                    markdown_parts.append(f"## Sheet: {sheet_name}\n\n{markdown_table}")
                    text_parts.extend(cell_values)

                    sheets_data.append({
                        "name": sheet_name,
                        "rows": rows,
                        "row_count": len(rows),
                        "column_count": len(rows[0]) if rows else 0,
                        "markdown_table": markdown_table,
                    })

                    total_cells += sum(len(row) for row in rows)

            text_content = "\n".join(text_parts)
            markdown_content = "\n\n".join(markdown_parts)

            return {
                "text": text_content,
                "markdown": markdown_content,
                "tables": sheets_data,
                "metadata": {
                    "source_url": str(excel_path),
                    "processed_with": "openpyxl",
                    "sheet_count": len(sheets_data),
                    "total_cells": total_cells,
                    "sheets": [s["name"] for s in sheets_data],
                },
            }
        except Exception as e:
            logger.error("Error processing Excel file %s: %s", excel_path, e)
            return {
                "text": "",
                "markdown": "",
                "tables": [],
                "metadata": {
                    "source_url": str(excel_path),
                    "processed_with": "excel_parser_error",
                    "error": str(e),
                },
            }

    def _process_xls(self, excel_path: Path) -> dict[str, Any]:
        """Process .xls file using xlrd.

        Args:
            excel_path: Path to .xls file

        Returns:
            Dictionary with extracted content
        """
        if not HAS_XLRD:
            logger.warning("xlrd not installed. Install with: pip install xlrd")
            return {
                "text": "",
                "markdown": "",
                "tables": [],
                "metadata": {
                    "source_url": str(excel_path),
                    "processed_with": "excel_parser_error",
                    "error": "xlrd not installed",
                },
            }

        try:
            workbook = xlrd.open_workbook(excel_path)
            
            text_parts = []
            markdown_parts = []
            sheets_data = []
            total_cells = 0

            for sheet_name in workbook.sheet_names():
                sheet = workbook.sheet_by_name(sheet_name)
                rows = []
                cell_values = []

                for row_idx in range(sheet.nrows):
                    row_data = []
                    row_text = []
                    for col_idx in range(sheet.ncols):
                        cell_value = sheet.cell_value(row_idx, col_idx)
                        if cell_value:
                            cell_str = str(cell_value)
                            row_data.append(cell_str)
                            row_text.append(cell_str)
                    if row_data:
                        rows.append(row_data)
                        cell_values.extend(row_text)

                if rows:
                    markdown_table = self._rows_to_markdown_table(rows)
                    markdown_parts.append(f"## Sheet: {sheet_name}\n\n{markdown_table}")
                    text_parts.extend(cell_values)

                    sheets_data.append({
                        "name": sheet_name,
                        "rows": rows,
                        "row_count": len(rows),
                        "column_count": len(rows[0]) if rows else 0,
                        "markdown_table": markdown_table,
                    })

                    total_cells += sum(len(row) for row in rows)

            text_content = "\n".join(text_parts)
            markdown_content = "\n\n".join(markdown_parts)

            return {
                "text": text_content,
                "markdown": markdown_content,
                "tables": sheets_data,
                "metadata": {
                    "source_url": str(excel_path),
                    "processed_with": "xlrd",
                    "sheet_count": len(sheets_data),
                    "total_cells": total_cells,
                    "sheets": [s["name"] for s in sheets_data],
                },
            }
        except Exception as e:
            logger.error("Error processing Excel file %s: %s", excel_path, e)
            return {
                "text": "",
                "markdown": "",
                "tables": [],
                "metadata": {
                    "source_url": str(excel_path),
                    "processed_with": "excel_parser_error",
                    "error": str(e),
                },
            }

    def process_csv(self, csv_path: Path | str) -> dict[str, Any]:
        """Process CSV file.

        Extracts data and converts to markdown table.

        Args:
            csv_path: Path to CSV file

        Returns:
            Dictionary with extracted content
        """
        csv_path = Path(csv_path)
        
        try:
            delimiter = self._detect_csv_delimiter(csv_path)
            
            rows = []
            cell_values = []
            
            with open(csv_path, "r", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f, delimiter=delimiter)
                for row in reader:
                    if row:
                        rows.append(row)
                        cell_values.extend(row)

            headers = rows[0] if rows else []
            data_rows = rows[1:] if len(rows) > 1 else []

            markdown_table = self._rows_to_markdown_table(rows)

            text_content = "\n".join(" | ".join(row) for row in rows)
            markdown_content = markdown_table

            return {
                "text": text_content,
                "markdown": markdown_content,
                "tables": [{
                    "headers": headers,
                    "rows": data_rows,
                    "row_count": len(data_rows),
                    "column_count": len(headers),
                    "markdown_table": markdown_table,
                }],
                "metadata": {
                    "source_url": str(csv_path),
                    "processed_with": "csv_parser",
                    "delimiter": delimiter,
                    "row_count": len(data_rows),
                    "column_count": len(headers),
                },
            }
        except Exception as e:
            logger.error("Error processing CSV file %s: %s", csv_path, e)
            return {
                "text": "",
                "markdown": "",
                "tables": [],
                "metadata": {
                    "source_url": str(csv_path),
                    "processed_with": "csv_parser_error",
                    "error": str(e),
                },
            }

    def _detect_csv_delimiter(self, csv_path: Path) -> str:
        """Detect CSV delimiter by sampling first few lines.

        Args:
            csv_path: Path to CSV file

        Returns:
            Detected delimiter character
        """
        delimiters = [",", ";", "\t", "|"]
        
        try:
            with open(csv_path, "r", encoding="utf-8", errors="replace") as f:
                sample = f.read(1024)
            
            delimiter_counts = {delim: sample.count(delim) for delim in delimiters}
            detected = max(delimiter_counts.items(), key=lambda x: x[1])[0]
            
            return detected if delimiter_counts[detected] > 0 else ","
        except Exception:
            return ","

    def _rows_to_markdown_table(self, rows: list[list[str]]) -> str:
        """Convert rows to markdown table format.

        Args:
            rows: List of rows, each row is a list of cell values

        Returns:
            Markdown table string
        """
        if not rows:
            return ""

        max_cols = max(len(row) for row in rows) if rows else 0
        if max_cols == 0:
            return ""

        normalized_rows = []
        for row in rows:
            normalized = row + [""] * (max_cols - len(row))
            normalized_rows.append([str(cell) for cell in normalized])

        header = normalized_rows[0] if normalized_rows else []
        data_rows = normalized_rows[1:] if len(normalized_rows) > 1 else []

        if not header:
            return ""

        header_row = "| " + " | ".join(header) + " |"
        separator = "| " + " | ".join(["---"] * len(header)) + " |"
        
        data_rows_str = "\n".join("| " + " | ".join(row) + " |" for row in data_rows)

        return f"{header_row}\n{separator}\n{data_rows_str}" if data_rows else header_row

    def extract_compliance_controls_from_pdf(
        self, pdf_path: Path | str, _standard: str
    ) -> list[dict[str, Any]]:
        """Extract compliance controls from a PDF file using Docling + LLM (sync stub).

        Use extract_compliance_controls_from_pdf_async for actual extraction.

        Args:
            pdf_path: Path to PDF file
            _standard: Compliance standard name (unused in sync version)

        Returns:
            Empty list (use async version for actual extraction)
        """
        content = self.process_pdf(pdf_path)

        if not content.get("text"):
            logger.warning("No text extracted from PDF: %s", pdf_path)
            return []

        text = content.get("text", "")
        logger.info("PDF processed with Docling, extracted %d characters from: %s", len(text), pdf_path)

        return []

    async def extract_compliance_controls_from_pdf_async(
        self, pdf_path: Path | str, standard: str
    ) -> list[dict[str, Any]]:
        """Extract compliance controls from PDF using Docling + LLM (async).

        Args:
            pdf_path: Path to PDF file
            standard: Compliance standard name

        Returns:
            List of extracted compliance control dictionaries
        """
        from .llm_extractor import LLMControlExtractor

        content = self.process_pdf(pdf_path)

        if not content.get("text"):
            logger.warning("No text extracted from PDF: %s", pdf_path)
            return []

        text = content.get("text", "")
        markdown = content.get("markdown", "")

        extractor = LLMControlExtractor()
        controls = await extractor.extract_controls(
            content=text,
            standard=standard,
            source_url=str(pdf_path),
            prefer_markdown=True,
            markdown_content=markdown if markdown else None,
        )

        for control in controls:
            control["source_url"] = str(pdf_path)

        logger.info("Extracted %d controls from PDF using Docling + LLM: %s", len(controls), pdf_path)

        return controls

    async def extract_compliance_controls_from_pdf_url_async(
        self, pdf_url: str, standard: str
    ) -> list[dict[str, Any]]:
        """Download and extract compliance controls from a PDF URL using Docling + LLM (async).

        Args:
            pdf_url: URL to PDF file
            standard: Compliance standard name

        Returns:
            List of extracted compliance control dictionaries
        """
        try:
            logger.info("Downloading PDF from URL: %s", pdf_url)

            with httpx.Client(timeout=30.0, follow_redirects=True) as client:
                response = client.get(pdf_url)
                response.raise_for_status()

                if not response.headers.get("content-type", "").startswith("application/pdf"):
                    logger.warning("URL does not appear to be a PDF: %s", pdf_url)
                    return []

                with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_file:
                    tmp_path = Path(tmp_file.name)
                    tmp_path.write_bytes(response.content)
                    logger.info("Downloaded PDF (%d bytes) to temporary file: %s", len(response.content), tmp_path)

                try:
                    from .llm_extractor import LLMControlExtractor

                    content = self.process_pdf(tmp_path)
                    if not content.get("text"):
                        logger.warning("No text extracted from PDF URL: %s", pdf_url)
                        return []

                    text = content.get("text", "")
                    markdown = content.get("markdown", "")

                    extractor = LLMControlExtractor()
                    controls = await extractor.extract_controls(
                        content=text,
                        standard=standard,
                        source_url=pdf_url,
                        prefer_markdown=True,
                        markdown_content=markdown if markdown else None,
                    )

                    for control in controls:
                        control["source_url"] = pdf_url

                    logger.info("Extracted %d controls from PDF URL using Docling + LLM: %s", len(controls), pdf_url)
                    return controls
                finally:
                    if tmp_path.exists():
                        tmp_path.unlink()
                        logger.debug("Cleaned up temporary PDF file: %s", tmp_path)

        except httpx.HTTPError as e:
            logger.error("HTTP error downloading PDF from %s: %s", pdf_url, e)
            return []
        except (ValueError, TypeError, KeyError, IOError, OSError) as e:
            logger.error("Error processing PDF from URL %s: %s", pdf_url, e)
            return []

    def _parse_controls_from_text(
        self, text: str, _markdown: str, standard: str, source_url: str
    ) -> list[dict[str, Any]]:
        """Parse compliance controls from extracted text.

        This is a basic parser. For complex PDFs, LLM extraction should be used.

        Args:
            text: Extracted text content
            _markdown: Extracted markdown content (unused, kept for future LLM processing)
            standard: Compliance standard name
            source_url: Source PDF path/URL

        Returns:
            List of control dictionaries
        """
        controls = []

        lines = text.split("\n")

        current_control: dict[str, Any] | None = None

        for line in lines:
            line = line.strip()

            if not line:
                continue

            control_id = self._extract_control_id(line, standard)

            if control_id:
                if current_control:
                    controls.append(current_control)

                current_control = {
                    "control_id": control_id,
                    "standard": standard,
                    "title": line[:200],
                    "description": "",
                    "source_url": source_url,
                    "requirement": "",
                }
            elif current_control is not None:
                if not current_control["description"]:
                    current_control["description"] = line
                else:
                    current_control["description"] += " " + line

        if current_control:
            controls.append(current_control)

        return controls

    def _extract_control_id(self, line: str, standard: str) -> str | None:
        """Extract control ID from a line of text.

        Args:
            line: Text line
            standard: Compliance standard name

        Returns:
            Control ID or None
        """
        line_upper = line.upper()

        if standard.upper() == "PCI-DSS":
            if "REQUIREMENT" in line_upper or "REQ" in line_upper:
                parts = line.split()
                for i, part in enumerate(parts):
                    if part.upper().startswith("REQ") or part.upper().startswith("REQUIREMENT"):
                        if i + 1 < len(parts):
                            return f"PCI-DSS-{parts[i+1]}"
        elif standard.upper() == "CIS":
            if "CIS" in line_upper or "BENCHMARK" in line_upper:
                parts = line.split()
                for part in parts:
                    if part.startswith(("CIS-", "CIS.", "CIS_")):
                        return part.replace(".", "-").replace("_", "-")

        return None


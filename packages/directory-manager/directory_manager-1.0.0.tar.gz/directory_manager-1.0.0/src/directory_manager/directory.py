from __future__ import annotations
import os, re, sys, csv, time, json, functools
from typing import Union, Literal, List, Dict, Iterable, Optional, Tuple, Any, Callable, overload
from pathlib import Path, PurePath
from shutil import rmtree
from abc import ABC, abstractmethod
from weakref import WeakValueDictionary
from enum import Enum
from openpyxl import Workbook, load_workbook
from PIL import Image, UnidentifiedImageError
from urllib.request import urlopen
from requests.exceptions import RequestException
from pandas import DataFrame, ExcelWriter, Index, read_excel
from colorama import Fore, Style, init
import prettytable as pt
try:
    from ..utils import is_valid_url, NumericCondition, DynamicLogger, default_logger, Numeric, NumericTrue, NOT_FOUND, EMPTY
except ImportError:
    from utils import is_valid_url, NumericCondition, DynamicLogger, default_logger, Numeric, NumericTrue, NOT_FOUND, EMPTY


__all__ = ["NullPath", 
           "DirCache", 
           "Directory",
           "FileParser",
           "JsonParser",
           "ExcelParser",
           "TextParser",
           "WorkbookParser",
           "ImageParser",
           "CsvParser",
           "OperatingSystem",
           "PathManager",
           "DynamicLogger",
           "NumericCondition",
           "Numeric",
           "NumericTrue",
           "NOT_FOUND",
           "EMPTY",
           "PathLike",
           "DirFile",
           "KEYS",
           "TOGGLE_LOCKS",
           "TOGGLE_CACHE"]

__author__ = "ismemo2121"
__version__ = "1.0.0"
__license__ = "MIT" 

#########################################
    #INIT LOGS
#########################################
logger = default_logger(enable_file=False, enable_sysout=False, enable_error=False)
init(autoreset=True)
#########################################
    #GLOBALS
#########################################
PathLike = Union[str, Path, PurePath]
KEYS = Literal["name", "path", "absolute", "exists", "type", "count", "totalcount", "hidden", "is_root", "root", "parent", "depth", "ctime", "lmtime", "latime", "size_b", "size_kb", "size_mb", "mode", "inode", "dev_id", "nlinks"]
RESET_COLOR = Style.RESET_ALL
TEXT_SUFFIX = (".txt",)
JSON_SUFFIX = (".json",)
CSV_SUFFIX = (".csv",)
EXCEL_SUFFIX = (".xlsx", ".xlsm", ".xltx", ".xltm")
IMAGE_SUFFIX = (".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp", ".tif")
#########################################
    #Functions
#########################################
def TOGGLE_LOCKS()->None:
    """
    Enable or disable both reference lookups for Directory and FileParser objects.
    """
    Directory._toggle_lock()
    FileParser._toggle_lock()
    return

def TOGGLE_CACHE()->None:
    """
    Enables or disables cache for all Directory objects, recommended use case would be at the beginning of the program.
    Cache uses more memory for quicker lookup on large Directory trees.
    """
    Directory._toggle_cache()

#########################################
    #OBJECTS
#########################################
class NullPath(Path):
    """
    Used to represent a Directory or FileParser object with no path or an untraced path.

    NullPath is assigned to a Directory or FileParser after removing an instance reference from a parent Directory object and can be unassigned after
    adding the instance back to a Directory object.

    An object with a NullPath is a False object and cannot be modified. 
    """
    def __init__(self, *args):
        super().__init__(*args)
        logger.debug("%s path was set to Null", args[0])

    def __eq__(self, value):
        return super().__eq__(value)

    def __hash__(self):
        return super().__hash__()

    def __bool__(self)->bool:
        return False

    def __str__(self)->str:
        return f"NullPath({super().__str__()})"

class DirCache:
    __initiated : bool = False
    top_down : bool = True
    def __new__(cls, *a, **kw):
        cls.__initiated = True
        return super().__new__(cls)
    
    @classmethod
    def is_initiated(cls)->bool:
        """
        Checks if an instance was created beforehand.
        """
        return cls.__initiated

    def __init__(self, slots : int = 100) -> None:
        """
        Stores paths in cache memory for quick access, made for Directory objects to lookup a sub directory rather than recursively searching for them. 

        **Examples:**
        >>> cache = DirCache(slots=200)
        >>> app_external = Directory(path="hi/bye", cache, enable_cache=True) #directory with external cache reference
        >>> app_internal = Directory(path="bye/hi", enable_cache=True) #directory with internal cache reference
        >>> app_nocache = Directory(path="foo") #disabled cache by default but can be activated at any time
        """
        self.__slots : int = slots
        self.__keys : list[Path] = list()
        self.__cache : Dict[Path, Directory] = dict()
        logger.debug("Initiated cache with %d slots", slots)

    def __getitem__(self, key : Path)->Optional[Directory]:
        """
        Using a Path object to access cache, no cache misses occur when passing the correct key, otherwise compares every path in cache with key.
        """
        if key in self.__cache:
            logger.debug("cache returned '%s'", key)
            return self.__cache.get(key)

        if DirCache.top_down:
            for element in self.__keys:
                if PathManager.relative(element, key):
                    logger.debug("cache returned '%s' while searching for '%s'", element, key)
                    return self.__cache[element]
        else:
            for element in reversed(self.__keys):
                if PathManager.relative(element, key):
                    logger.debug("cache returned '%s' while searching for '%s'", element, key)
                    return self.__cache[element]

        return

    def __setitem__(self, key : PathLike, value : Directory):
        """
        Setting prioritizes paths with more parts within for a quicker lookup.
        """
        path = Path(key)
        if path in self.__cache:
            self.__cache[path] = value
            return None

        if len(self) >= self.__slots:
            logger.warning("Cache overflow, cannot receive more data.")
            return None

        if len(self) == 0:
            self.__keys.insert(0, path)
            self.__cache[path] = value
            return None

        key_length : int = len(path.parts)
        left, right = 0, len(self) - 1
        if key_length <= len(self.__keys[right].parts):
            self.__keys.insert(right + 1, path)
            self.__cache[path] = value
            return None
        elif key_length >= len(self.__keys[left].parts):
            self.__keys.insert(left, path)
            self.__cache[path] = value
            return None

        mid : int = 0
        list_range : int = len(self.__keys[left].parts) - len(self.__keys[right].parts)
        jump : int = 1
        while left <= right:
            mid = (left + right) // 2
            mid_size : int = len(self.__keys[mid].parts)
            if mid_size == key_length:
                break

            diff = key_length - mid_size
            if abs(diff) > list_range * 0.1:
                jump = max(1, (right - left) // 4)
            else:
                jump = 1 

            if mid_size < key_length:
                right -= jump #converging to 0 if bigger
            elif mid_size > key_length:
                left += jump #diverging from 0 if smaller
            
        self.__keys.insert(mid, path)
        self.__cache[path] = value
        return

    def remove(self, *keys : Path)->None:
        """
        Removes a single path from cache, specify an identical Path or string object to select it.
        """
        if len(self) == 0:
            return
        for path in keys:
            if path not in self.__cache:
                continue
            self.__keys.remove(path)
            self.__cache.pop(path)
            logger.debug("'%s' path was removed from cache", path)
        return

    def delete_path(self, key : PathLike)->None:
        """
        Use to remove every reference to key path in cache.
        """
        path = Path(key)
        deleted_keys : List[Path] = list()
        for path_key in self.__keys:
            if PathManager.relative(path_key, path):
                deleted_keys.append(path_key)
                self.__cache.pop(path_key)

        for deleted_key in deleted_keys:
            self.__keys.remove(deleted_key)
            logger.debug("'%s' path was removed from cache", deleted_key)
        return

    def forget_dir(self, dir : Directory)->None:
        """
        Drops dir reference and references to its sub directories inside cache.
        """
        if not isinstance(dir, Directory):
            return
        
        if dir.path not in self.__keys:
            return None
        dirs, _ = dir._items()
        for i_dir in dirs:
            self.__cache.pop(i_dir.path, None)
            self.__keys.remove(i_dir.path)
            
        logger.debug("Directory '%s' was forgotten by cache", dir)
        return
    
    def inherit_dir(self, dir : Directory)->None:
        """
        Caches dir and its sub directories to cache.
        """
        if not isinstance(dir, Directory):
            return
        self[dir.path] = dir
        dirs, _ = dir._items()
        for i_dir in dirs:
            self[i_dir.path] = i_dir
        logger.debug("Directory '%s' was inherited in cache", dir)
        

    def replace(self, old_keys : Iterable[Path], objects : Iterable[Directory])->None:
        """
        Swaps old keys as long as they exist inside of cache, with new lookup keys that reference the same object.
        """
        if len(old_keys) != len(objects):
            logger.critical("None identical lengths when updating cache %s != %s",len(old_keys), len(objects))
            raise Exception(f"None identical lengths when updating cache {len(old_keys)} != {len(objects)}")

        for old, obj in zip(old_keys, objects):
            if isinstance(old, Directory) or not isinstance(obj, Directory):
                logger.debug("skipping cache replacement of '%s' with '%s'", old, obj)
                continue      
            #hashmap          
            ref = self.__cache.pop(old, None)
            self.__cache[obj.path] = obj
            #list
            if ref is not None:
                #existed before
                idx = self.__keys.index(old)
                self.__keys.remove(old)
                self.__keys.insert(idx, obj.path)
            else:
                #new item
                self[obj.path] = obj
            logger.debug("'%s' cache path was updated to '%s' path", old, obj)


    def clear(self)->None:
        self.__keys.clear()
        self.__cache.clear()
        logger.debug("Cleared cache")
        

    def __contains__(self, key : PathLike)->bool:
        return Path(key) in self.__cache

    def __len__(self)->int:
        return len(self.__keys)

    def __repr__(self) -> str:
        return f"keys:{[str(key) for key in self.__keys]}"

class Directory:
    """
    A Directory object can be used for prototyping data and quick creation of directories, used with FileParser object to simulate the files system's paths (excluding shortcuts).
    
    All objects of the class use their Path object as a membership key, two Directory objects with the same absolute path cannot co-exist while the membership lock is enabled:
    
    -Membership lock: safety measure used by class to prevent duplication of absolute paths during initiation and gives membership to existing folders, 
                     hidden objects are not referenced by class therefore can be duplicated, however, duplicated objects are automatically renamed by parent class.

    To disable/enable the lock (True by default), run before creating any instances:
    >>> TOGGLE_LOCKS()

    To disable/enable Directory cache lookup (False by default), add:
    >>> TOGGLE_CACHE()

    To toggle cache's searching priority, run:
    >>> DirCache.top_down = False #True by default, better for large directory trees

    Directory asks before re-writing a directory tree when calling populate, to disable that option:
    >>> Directory.ask_before_del = False 
    """
    __GLOBAL_REFERENCE : WeakValueDictionary[Path, Directory] = WeakValueDictionary() #membership map for paths
    __GLOBAL_LOCK : bool = True #looks up membership map everytime a new instance is created if set
    __PERMISSION : bool = False #set when expecting a change in data structure
    __DEFAULT_TABLE : Tuple[str] = ("name", "path", "type") #for .info table
    __RENAME_DUPLICATES : bool = True #for duplicated names, only necessary if references are blocked
    __CACHE_ENABLE : bool = False #cache call flag
    _CACHE_DEFAULT : int = 1000 #default capacity for cache
    _DIR_COLOR : str = Fore.LIGHTBLACK_EX #default color for directory name in Directory visualizing methods
    _PATHLESS_COLOR : str = Fore.LIGHTRED_EX #default null directory name color when using .info method
    ask_before_del : bool = True
    @classmethod
    def __update(cls, key : Path, new_key : Path)->None:
        """
        Internally managed, updates paths for directories.
        """
        if key == new_key: #no changes
            logger.debug("Path '%s' not updated", key)
            return
        if key in cls.__GLOBAL_REFERENCE: #path changed
            ref = cls.__GLOBAL_REFERENCE.pop(key)
            cls.__GLOBAL_REFERENCE[new_key] = ref
            logger.debug("Path '%s' updated to '%s'", key, new_key)

    @classmethod
    def _blocks_duplicates(cls)->bool:
        """
        To check whether duplicates are handled, default is True.
        """
        return cls.__RENAME_DUPLICATES

    @classmethod
    def _toggle_cache(cls)->None:
        if DirCache.is_initiated():
            logger.error("Cannot toggle cache as its already initialized.")
            return None
        cls.__CACHE_ENABLE = not(cls.__CACHE_ENABLE)
        message = "Enabled Directory cache" if cls.__CACHE_ENABLE else "Disabled Directory cache."
        logger.warning(message)
        return

    @classmethod
    def _toggle_lock(cls)->None:
        """
        For referencing previously created objects for paths, disabling the flag allows for duplicates to be created, only allowed to call before creating any
        Directory instances otherwise raises a ConfigError.
        """
        if cls.__GLOBAL_REFERENCE:
            logger.error("Cannot lock/unlock references with created references.")
            return None
        cls.__GLOBAL_LOCK = not(cls.__GLOBAL_LOCK)
        logger.warning("Directory object memberships lock set to %s", cls.__GLOBAL_LOCK)

    def __new__(cls, path : PathLike = '.', *a, **k):
        if not cls.__GLOBAL_LOCK or k.get("hide", False):
            logger.debug("untraced Directory object with path: '%s'", path)
            return super().__new__(cls) #not tracked

        abs_path = Path(path).absolute()
        if abs_path in cls.__GLOBAL_REFERENCE:
            logger.debug("returned traced Directory object with path: '%s'", abs_path)
            return cls.__GLOBAL_REFERENCE[abs_path]
        
        instance = super().__new__(cls)
        logger.debug("created new directory with path: '%s'", abs_path)
        cls.__GLOBAL_REFERENCE[abs_path] = instance
        return instance

    def __init__(self, path : PathLike = '.', *attachements, **kw):
        """
        Instance that inherits the properties of a directory and references to it's sub paths and designed to manage a path within the script, mainly designed to prototype directories.

        Arguments:
            path: string or Path object pointing to a directory (directory does not have to exist in the files system)
            attachements: mostly used by class to inherit references for cache, only use case is manually passing a DirCache object to a new instance
            create: choose to initiate instance to the files system initially (does nothing if path exists), can be initialized later using .populate method
            color: used in visualization methods (Directory.info, PathManager.indent_directory, PathManager.indent_directory_legacy), uses colorama's package constants by default
            hide: instance cannot be referenced by global map if True is passed (cannot reset)
        

        **Some common methods:**

        1)creating a new directory as a prototype or mirroring an existing directory:
        >>> my_directory = Directory(path="path/something", ...) #initialization
        >>> my_directory.image(...) #cloning all elements in directory if it exists, otherwise proceeds to do nothing
        >>> my_directory.populate(...) #creates all files and directories in path to the files system

        2)adding/removing sub directories and files:
        >>> subs = ("dir1", "dir2", "dir3/sub-dir3")
        >>> my_directory.add_directories(*subs) #adds 4 directories in total
        >>> files = ("new.txt", "file.json", "dir4/FILE")
        >>> my_directory.add_files(*files) #adds 3 files and a directory
        >>> my_directory.remove_directory("dir3") #removes the first found instance of dir3 directory's reference alongside any sub element
        >>> my_directory.remove_file("FILE") #removes the first found instance of FILE
        >>> my_directory.clear() #clears all sub elements

        3)getting elements:

        ***Directories***

        >>> dir1 = my_directory.get_directory("dir1") #returns dir1 Directory object referene or None if not found (only one layer deep search)
        >>> dir2 = my_directory["dir2"] #returns reference of Directory object of dir2 or None if not found
        >>> list_of_dir_paths = my_directory.search_dir("dir4") #searches sub directories only
        
        ***Files***
        
        >>> file = my_directory.get_file("file.json") #returns FileParser Object related to file if it exists one layer deep
        >>> sub_file = my_directory.find_file("file.json") #searches for file in all directory tree
        >>> list_of_file_paths = my_directory.search_file("new.txt") #searches files only
        
        ***Both***
        
        >>> list_of_paths = my_directory.search_name("new.txt") #returns a list of Path objects that point to elements with the name "new.txt" (searches both directories and files)
        
        4)other functions
        >>> my_directory.info(table) #returns a table of information about directory
        >>> my_directory.rename("new_name") #changes directory name
        >>> my_directory.count(...) #not implemented yet
        >>> my_directory.find_by(...) #not implemented yet
        >>> "dir2" in my_directory
        >>> len(my_directory) #all elements count in directory
        >>> my_directory == "something" compates directory with other directory, string or Path (checks if they point to the same name element)
        >>> bool(my_directory) #only False if directory belongs to no root, happens when keeping reference to a directory that was removed by its root so it becomes NullPath
        
        5)properties
        >>> my_directory.path #full path
        >>> my_directory.name #name only
        >>> my_directory.exists #checks if it exists in files system
        >>> my_directory.files #list of directory files
        >>> my_directory.directories #list of sub directories references
        >>> my_directory.is_root #checks if directory is root
        >>> my_directory.root #return root directory object
        >>> my_directory.parent #reference to parent Directory object
        >>> my_directory.absolute #absolute path
        >>> my_directory.color #color of directory name when using info method or indent_directory from PathManager
        """
        if hasattr(self, '_init'):
            return

        if not path:
            logger.critical("Invalid directory path '%s'", path)
            raise Exception(f"Invalid directory path '{path}'")

        self._init = True
        self.__path : Path = Path(path)
        #Cache is initialized either way but not used until the global flag is set
        self.__cache : DirCache = None
        if Directory.__CACHE_ENABLE:
            self.__cache = (attachements[0] if (len(attachements) >= 1 and isinstance(attachements[0], DirCache)) else DirCache(Directory._CACHE_DEFAULT))
        self.__parent : Optional[Directory] = None
        self.__root : Optional[Directory] = None
        #might switch to typelists
        self.__files : List[FileParser] = []
        self.__directories : List[Directory] = []
        #default color
        self.__color : str = kw.get("color", Directory._DIR_COLOR)
        #any extra configs
        if str(self.__path) == '.':
            self.__path = Path.cwd()
        elif str(self.__path) == "..":
            self.__path = Path("..").resolve()
        
        if self.exists and not self.__path.is_dir():
            logger.critical("Invalid path for a directory '%s'", self.__path)
            raise Exception(f"Invalid path for a directory '{self.__path}'")

        operating_sys = PathManager.os()
        if operating_sys == OperatingSystem.WINDOWS:
            for prohibited_path in PathManager.PROHIBITED_PATHS_WIN:
                if PathManager.is_subpath(prohibited_path, self.absolute, True):
                    logger.critical("Selected path is blocked: '%s'", self.absolute)
                    raise Exception(f"Selected path is blocked: '{self.absolute}'")

        elif operating_sys == OperatingSystem.LINUX:
            if not PathManager.is_subpath("/home", self.absolute):
                logger.critical("Selected path is blocked: '%s'", self.absolute)
                raise Exception(f"Selected path is blocked: '{self.absolute}'")
        
        elif operating_sys == OperatingSystem.MACOS:
            if not PathManager.is_subpath("/Users", self.absolute):
                logger.critical("Selected path is blocked: '%s'", self.absolute)
                raise Exception(f"Selected path is blocked: '{self.absolute}'")

        self.__hidden : bool = kw.get("hide", False)
        if kw.get("create", False):
            self.populate(rewrite=False)
        
    def populate(self, rewrite : bool = False, write_files : bool = False)->None:
        """
        Creates the current instance tree on the files system

        Arguments:
            rewrite: erases directory tree from the files system before rewriting it (WARNING: do not set rewrite to True on a sensitive directory, including the working directory path)
            write_files: updates the content of the files on top of creating them in the files system if set, otherwise creates the files only
            
        """
        if not self:
            logger.warning("Directory '%s' is null and cannot be created", self)
            return None
        if self.exists and rewrite:
            flag : str = 'y'
            if Directory.ask_before_del:
                flag = str(input(f"Confirm deleting the directory {self.absolute} before rewriting? [Y][N]"))
            if flag.lower() in ("y", "yes"):
                logger.warning("Rewriting directory tree '%s'", self)
                rmtree(str(self.__path), ignore_errors=True)

        self.__path.mkdir(exist_ok=True)

        for file in self.__files:
            file.write(only_create=not(write_files))
        
        for i_directory in self.__directories:
            i_directory.populate(rewrite=rewrite)
        return

    def image(self, rewrite : bool = True, read_files : bool = False, *ignore_abs : Path)->None:
        """
        Copies the instance's path from file explorer

        Arguments:
            rewrite: wipes out all data in object before the imaging process
            read_files: reads the content of each file in the directory tree (Warning: enabling this flag on a directory with too many file would cause a sudden large increase in memory usage, 
                        do not enable it on a 100+ files directory)
            ignore_abs: absolute paths for files/directories to skip
        """
        if not self.exists:
            logger.warning("Directory '%s' does not exist and can't be imaged", self)
            return None

        if rewrite:
            self.clear(deep_clear=True)

        for root, directories, files in self.__path.walk():
            if root.absolute() in ignore_abs:
                logger.info("Ignored root '%s'", root)
                continue
            current_directory = self[root] #returns self on the first iteration
            if current_directory is None:
                continue

            for i_directory in directories:
                path = root / i_directory
                if path.absolute() in ignore_abs or path.absolute().is_symlink():
                    logger.warning("ignored directory '%s'", path)
                    continue
                dir_temp = Directory(path, self.__cache)
                Directory.__PERMISSION = True
                dir_temp._redirect(current_directory)
                logger.info("Added directory '%s'", dir_temp)

            for file in files:
                path = root / file
                if path.absolute() in ignore_abs or path.absolute().is_symlink():
                    logger.warning("ignored file '%s'", path)
                    continue
                file_temp = FileParser(path)
                Directory.__PERMISSION = True
                file_temp._redirect(current_directory)
                if read_files:
                    file_temp.read()
                logger.info("Added file '%s'", file_temp)
        return

    def rename(self, new_name : Optional[str])->None:
        """
        Renames the instance, passing None prompts to check if current name is a duplicate in parent directory and renames it if so.
        """
        if not self:
            logger.warning("Directory '%s' is null and cannot be renamed", self)
            return None

        if isinstance(new_name, str):
            if PathManager.os() == OperatingSystem.WINDOWS and PathManager.contains_illegal_chars_win(new_name):
                logger.error("'%s' contains illegal windows characters, cannot rename '%s' directory", new_name, self)
                return None
            
            elif PathManager.os() in (OperatingSystem.LINUX, OperatingSystem.MACOS) and PathManager.contains_illegal_chars_psx(new_name):
                logger.error("'%s' contains illegal POSIX characters, cannot rename '%s' directory", new_name, self)
                return None

            if len(Path(new_name).parts) != 1:
                logger.error("name must not include other directory parts.")
                return None
            #actual name, regular duplicate check
            if Directory.__RENAME_DUPLICATES and self.__parent:
                new_name = self.__parent._numeric_name(new_name) #checks if its dup + gives it a proper name if so

        #requesting a duplicate rename
        elif new_name is None:
            if not self.__parent:
                logger.debug("cancelled since parent is %s", self.__parent)
                return None
            new_name = self.__parent._numeric_name(self.name) #get correct name
        
        else: #new_name is neither str or None
            logger.error("Expected None or a string object, got %s instead", type(new_name).__name__)
            return None

        if new_name == self.name: #no changes
            return None

        #create a new path (parent/name)
        old_path = self.__path
        path, _ = PathManager.split(self.__path)
        new_path = Path(path) / new_name if path else Path(new_name)

        #check if paths are referenced or dir is hidden
        if Directory.__GLOBAL_LOCK and not self.__hidden:
            if new_path.absolute() in Directory.__GLOBAL_REFERENCE:
                logger.error("Directory '%s' already has '%s', skipping process..", path, new_name)
                return None
            Directory.__update(old_path.absolute(), new_path.absolute())
            if self.exists:
                self.__path.rename(new_path)

        self.__path = new_path
        if self.is_root:
            if Directory.__CACHE_ENABLE:
                #clear it as all paths are outdated, slowly replaced in directories loop
                self.__cache.clear()
        else: #is sub
            if Directory.__CACHE_ENABLE:
                #sub directories updated here, .replace adds object either way
                self.__cache.replace([old_path], [self])

        #redirecting all subs, not that heavy as it only modifies path variable
        for file in self.__files:
            Directory.__PERMISSION = True
            file._redirect(self)

        for i_directory in self.__directories:
            Directory.__PERMISSION = True
            i_directory._redirect(self)
        return
        
    def add_directories(self, *directories : Union[PathLike, Directory])->None:
        """
        Adds a sub directory to the root path, directories can be string paths or Path objects pointing to the added directory or just its name, can also add multiple directories in a path if they
        don't exist.
        
        Inserting a string/Path creates a new instance internally, and passing a Directory object performs a cut+paste process to sub directory.
        """
        if not self: #instance is null
            logger.warning("Directory '%s' is null and cannot be modified", self)
            return None

        #process
        for i_directory in directories:
            if isinstance(i_directory, PathLike):
                if not i_directory:
                    continue   
                path = PathManager.set(self, i_directory, False) #new path

                if len(path.parts) == 1: #added to this directory
                    path = self.__path / path
                    Directory.__PERMISSION = True
                    i_directory = Directory(path, self.__cache) #returns existing instance if dup name
                    i_directory._redirect(self) #won't do anything if i_dir is referenced in instance, otherwise redirects it to parent dir

                else:
                    i_directory = self.__build(path, True) #same as 1 part process but recursive

                #if global ref is disabled, then dups could exist
                if Directory.__RENAME_DUPLICATES:
                        i_directory.rename(None) #passing None checks if name is dup, does nothing if name isn't duplicated

                #add to cache if flag is set
                if Directory.__CACHE_ENABLE:
                    self.__cache.inherit_dir(i_directory)
                
                logger.info("Added directory '%s' to parent '%s'", i_directory, self)
                continue

            elif isinstance(i_directory, Directory):
                #give the directory a path, a home, and a parent.
                Directory.__PERMISSION = True
                i_directory._redirect(self)

                #rename if dup
                if Directory.__RENAME_DUPLICATES:
                        i_directory.rename(None)

                #add to cache if flag is set
                if Directory.__CACHE_ENABLE:
                    self.__cache.inherit_dir(i_directory)
                logger.info("Added directory '%s' to parent '%s'", i_directory, self)

        return

    def add_files(self, *files : Union[PathLike, FileParser])->None:
        """
        Adds files to directory, a path to a file creates the file instance + sub directories if they don't exist (does not add sub directories if a FileParser is passed).

        Passing an object cuts+pastes the file to this directory.
        """
        if not self:
            logger.warning("Directory '%s' is null and cannot be modified", self)
            return None

        for file in files:
            if isinstance(file, PathLike):
                if not file:
                    continue
                relative_path = PathManager.set(self, file, False) #special path
                file_obj = FileParser(self.__path / relative_path) #getting file object based on suffix
                parent = None
                if len(relative_path.parts) == 1: #added to current dir
                    Directory.__PERMISSION = True
                    file_obj._redirect(self)

                else: #added to sub dir
                    path, _ = PathManager.split(relative_path)
                    parent = self.__build(path, True)
                    Directory.__PERMISSION = True
                    file_obj._redirect(parent)
                Directory.__PERMISSION = False
                #dup case, change name to file(n).abc
                if Directory.__RENAME_DUPLICATES:
                    file_obj.rename(None)
                logger.info("Added file '%s' to parent '%s'", file_obj, parent if parent else self)

            elif isinstance(file, FileParser):
                #changing path + parent
                Directory.__PERMISSION = True
                file._redirect(self)
                Directory.__PERMISSION = False #file redirect does not set off directory permission so we set it off manually

                if Directory.__RENAME_DUPLICATES:
                    file.rename(None)
                logger.info("Added file '%s' to parent '%s'", file, self)

        return
                    
    def remove_directory(self, path : PathLike)->None:
        """
        Removes specified directory path from root directory, current instance and files cannot be removed with this method.
        """
        if not self: #rather not debug changes to a null directory
            logger.warning("Directory '%s' is null and cannot be modified", self)
            return None

        if not path or not isinstance(path, PathLike):
            logger.error("Expected Path or string object, got %s instead", type(path).__name__)
            return None

        path = Path(path)
        rd : Optional[Directory] = self[path]
        if rd is None: #directory not found
            logger.info("could not find the path '%s'", path)
            return None

        if rd == self: #directory is self
            logger.error("cannot remove current directory '%s'", self)
            return None

        if Directory.__CACHE_ENABLE:
            self.__cache.forget_dir(rd)

        Directory.__PERMISSION = True
        if not rd.parent._remove_directory(rd):
            logger.error("could not remove the path '%s'", path)
        return

    def remove_file(self, path : PathLike)->None:
        """
        Removes specified file path from root directory.
        """
        if not self: #idea of a null directory is that its cut and waiting to be placed in a new directory so it shouldn't get modified
            logger.warning("Directory '%s' is null and cannot be modified", self)
            return None

        if not path or not isinstance(path, PathLike):
            logger.error("Expected Path or string object, got %s instead", type(path).__name__)
            return None

        path = Path(path)

        if len(path.parts) == 1: #file path is only the name, removing the first file with that name
            file = self.get_first_file(path.name)
            if file is None:
                logger.info("couldn't find the path '%s'", path)
            else:
                Directory.__PERMISSION = True
                if not file.parent._remove_file(file):
                    logger.error("couldn't remove '%s'", path)
            return None
        
        #file has a path, search for parent directory and remove it from there
        path, name = PathManager.split(path)
        parent = self[path]

        if parent is None:
            logger.info("couldn't find the path '%s' in '%s'", path, self)
            return None
        
        file = parent.get_file(name)
        if file is None:
            logger.info("couldn't find the path '%s' in '%s'", path, parent)
            return None

        Directory.__PERMISSION = True
        if not parent._remove_file(file):
            logger.error("cannot remove the file '%s' from '%s'", file, parent)
        return

    def get_directory(self, dir_name : PathLike)->Optional[Directory]:
        """
        Searches for sub directory one layer deep in directory, provide directory name only.
        """
        if not self: #cannot return off null directory
            logger.warning("Directory '%s' is null and cannot return objects", self)
            return None

        if not dir_name:
            logger.debug("no input")
            return None

        if not isinstance(dir_name, PathLike):
            logger.error("Expected a Path or string object, got '%s' instead", type(dir_name).__name__)
            return None

        path = Path(dir_name)
        if len(path.parts) != 1: #only accepts a name
            logger.error("Expected directory name only, not a path")
            return None

        for i_directory in self.__directories: #yeah
            if path.name == i_directory:
                logger.debug("returned '%s'", i_directory)
                return i_directory

        return None

    def get_file(self, file_name : PathLike)->Optional[FileParser]:
        """
        Searches for file one layer deep in directory.
        """
        if not self:
            logger.warning("Directory '%s' is null and cannot return objects", self)
            return None

        if not file_name:
            logger.debug("no input")
            return None

        if not isinstance(file_name, PathLike):
            logger.error("Expected a Path or string object, got '%s' instead", type(file_name).__name__)
            return None

        path = Path(file_name)
        if len(path.parts) != 1:
            logger.error("Expected file name only, not a path")
            return None

        for file in self.__files:
            if path.name == file:
                logger.debug("returned '%s'", file)
                return file

        return None

    def find_file(self, file_path : PathLike)->Optional[FileParser]:
        """
        Searches for file in directory tree, returns the first instance of the file name if no specific path is provided.
        """
        if not self:
            logger.warning("Directory '%s' is null and cannot return objects", self)
            return None

        if not isinstance(file_path, PathLike):
            logger.error("Expected a Path or string object, got '%s' instead", type(file_path).__name__)
            return None

        if not file_path:
            logger.debug("no input")
            return None
        
        path = Path(file_path)
        if len(path.parts) == 1:
            return self.get_first_file(path)
        
        parent_path, name = PathManager.split(path)
        parent_dir = self[parent_path]
        if parent_dir is None:
            return None
        
        return parent_dir.get_file(name)

    @overload
    def find_by(self, is_type : Literal["Directory", "File"], key : Literal["name"], condition : Union[str, Iterable[str]])->List[Directory, FileParser]: ...
    @overload
    def find_by(self, is_type : Literal["Directory", "File"], key : Literal["suffix"], condition : Union[str, Iterable[str]])->List[Directory, FileParser]: ...
    @overload
    def find_by(self, is_type : Literal["Directory", "File"], key : Literal["exists"], condition : Literal[0, 1])->List[Directory, FileParser]: ...
    @overload
    def find_by(self, is_type : Literal["Directory", "File"], key : Literal["size"], condition : Union[NumericCondition, NumericTrue])->List[Directory, FileParser]: ...
    @overload
    def find_by(self, is_type : Literal["Directory", "File"], key : Literal["depth"], condition : Union[NumericCondition, NumericTrue])->List[Directory, FileParser]: ...
    @overload
    def find_by(self, is_type : Literal["Directory", "File"], key : Literal["sub-count"], condition : Union[NumericCondition, NumericTrue])->List[Directory, FileParser]: ...

    def find_by(self, is_type : Literal["Directory", "File"], key : Literal["name", "exists", "size", "depth", "sub-count", "suffix"], condition : Union[NumericCondition, Numeric, Iterable[str]])->List[Directory, FileParser]:
        """
        Special search looks for directories/files that match the provided condition that is related with input key, returns a list of Directory/FileParser objects.
        
        ***Example 1:***
        >>> condition = ".json"
        >>> results = find_by(is_type="File", key="suffix", condition=condition) #returns all json files in directory

        ***Example 2:***
        >>> condition = NumericCondition(0, 10000)
        >>> results = find_by(is_type="File", key="size", condition=condition) #returns the "type" with size in range [0, 10000] bytes

        ***Example 3:***
        >>> condition = 0
        >>> results = find_by(is_type="Directory", key="sub-count", condition=condition) #returns directories with no sub directories or files

        ***Example 4:***
        >>> condition = 5
        >>> results = find_by(is_type="File", key="depth", condition=condition) #returns the "type" that exists 5 layers down of instance directory

        ***Example 5:***
        >>> condition = "directoryname"
        >>> results = find_by(is_type="Directory", key="name", condition=condition) #returns all directories with the same name(s) provided in condition

        ***Example 6:***
        >>> condition = 0
        >>> results = find_by(is_type="File", key="exists", condition=condition) #returns all files that don't exist in files eplorer but exist in instance (1 to check if they exist)
        """
        result = []
        match key:
            case "name":
                if isinstance(condition, str):
                    condition = [condition]
                if not isinstance(condition, Iterable):
                    logger.error("suffix condition must be a string or an iterable of strings, not %s.", type(condition).__name__)
                    return list()
                result.extend(self.__find_name(is_type, condition))
                for i_directory in self.__directories:
                    result.extend(i_directory.find_by(is_type, key, condition))

            case "exists":
                result.extend(self.__find_exists(is_type, condition))
                for i_directory in self.__directories:
                    result.extend(i_directory.find_by(is_type, key, condition))

            case "suffix":
                if isinstance(condition, str):
                    condition = [condition]
                if not isinstance(condition, Iterable):
                    logger.error("suffix condition must be a string or an iterable of strings, not %s.", type(condition).__name__)
                    return list()
                result.extend(self.__find_suffix(condition))
                for i_directory in self.__directories:
                    result.extend(i_directory.find_by("File", key, condition))

            case "size":
                result.extend(self.__find_size(is_type, condition))
                for i_directory in self.__directories:
                    result.extend(i_directory.find_by(is_type, key, condition))
        
            case "sub-count":
                result.extend(self.__find_subs(condition))
                for i_directory in self.__directories:
                    result.extend(i_directory.find_by(is_type, key, condition))

            case "depth":
                result.extend(self.__find_depth(is_type, condition))
                for i_directory in self.__directories:
                    result.extend(i_directory.find_by(is_type, key, condition))
        return result

    def search_name(self, name : PathLike)->List[Path]:
        """
        returns all paths that lead to name (duplicated names, different paths), can be a file or a sub directory.
        """
        if not self:
            logger.warning("Directory '%s' is null and cannot return objects", self)
            return list()

        if not isinstance(name, PathLike):
            logger.error("Expected string or path, got %s instead.", type(name).__name__)
            return list()

        if not name or len(Path(name).parts) != 1:
            logger.error("Expected a name, not an empty object or a path.")
            return list()

        results = []
        for file in self.__files:
            if name == file:
                results.append(file.path)

        for i_directory in self.__directories:
            if not i_directory:
                continue
            if name == i_directory:
                results.append(i_directory.path)

            results.extend(i_directory.search_name(name))
        return results

    def search_file(self, name : PathLike)->List[Path]:
        """
        Returns all paths that lead to files with the same name as the input name.
        """
        if not self:
            logger.warning("Directory '%s' is null and cannot return objects", self)
            return list()

        if not isinstance(name, PathLike):
            logger.error("Expected string or path, got %s instead.", type(name).__name__)
            return list()

        if not name or len(Path(name).parts) != 1:
            logger.error("Expected a name, not an empty object or a path.")
            return list()

        results = []
        for file in self.__files:
            if file == name:
                results.append(file.path)
        for i_directory in self.__directories:
            results.extend(i_directory.search_file(name))
        return results
        
    def search_dir(self, name : PathLike)->List[Path]:
        """
        Returns all directory paths that point to input name.
        """
        if not self:
            logger.warning("Directory '%s' is null and cannot return objects", self)
            return list()

        if not isinstance(name, PathLike):
            logger.error("Expected string or path, got %s instead.", type(name).__name__)
            return list()

        if not name or len(Path(name).parts) != 1:
            logger.error("Expected a name, not an empty object or a path.")
            return list()

        results = []
        if name == self:
            results.append(self.__path)
        for i_directory in self.__directories:
            results.extend(i_directory.search_dir(name))
        return results

    def find_first(self, name : PathLike)->Optional[Directory]:
        """
        Returns the first result that points to directory path inside the root directory.
        """
        if not self:
            logger.warning("Directory '%s' is null and cannot return objects", self)
            return None

        if not isinstance(name, PathLike) or not name:
            logger.error("Expected a Path or string object, got '%s' instead", type(name).__name__)
            return None

        if not isinstance(name, PathLike):
            logger.error("Expected string or path, got %s instead.", type(name).__name__)
            return list()

        if not name or len(Path(name).parts) != 1:
            logger.error("Expected a name, not an empty object or a path.")
            return list()

        result : Optional[Directory] = None
        for i_directory in self.__directories:
            if i_directory == name:
                return i_directory
            
            result = i_directory.find_first(name)
            if result is not None:
                return result
        
        return result

    def get_first_file(self, name : PathLike)->Optional[FileParser]:
        """
        Removes the first file with specified name, includes sub directories as well.
        """
        if not self:
            logger.warning("Directory '%s' is null and cannot return objects", self)
            return False

        if not isinstance(name, PathLike):
            logger.error("Expected string or path, got %s instead.", type(name).__name__)
            return list()

        if not name or len(Path(name).parts) != 1:
            logger.error("Expected a name, not an empty object or a path.")
            return list()

        for file in self.__files:
            if file == name:
                return file

        for i_dir in self.__directories:
            if not i_dir:
                continue
            file = i_dir.get_first_file(name)
            if file is not None:
                return file
        return

    def clear(self, deep_clear : bool = False)->None:
        """
        Removes all objects in instance, including files and sub directories.

        deep_clear recursively unassigns sub-directories with their objects as well if set, otherwise unassigns sub-directories and files one layer deep only.
        """
        if not self:
            logger.warning("Directory '%s' is null and cannot be modified", self)
            return None

        for file in self.files:
            Directory.__PERMISSION = True
            self._remove_file(file)
        
        for i_directory in self.directories:
            if deep_clear:
                i_directory.clear(deep_clear)
            Directory.__PERMISSION = True
            self._remove_directory(i_directory)
        if Directory.__CACHE_ENABLE and self.is_root:
            self.__cache.clear()
        if len(self) != 0: #shouldn't happen but here we are
            logger.warning("%d.........", len(self))
        return

    def info(self, info_table : Optional[pt.PrettyTable])->pt.PrettyTable:
        """
        Returns a table containing the following columns of information about the instance:

        **name, path, type, count, totalcount, exists, is_root, root, parent, depth, ctime, lmtime, latime, size_b, size_kb, size_mb**
        """
        #create a default table if needed
        if not isinstance(info_table, pt.PrettyTable):
            info_table = pt.PrettyTable(Directory.__DEFAULT_TABLE)
        
        columns = info_table.field_names
        #add name as a default column
        if "name" not in columns: 
            columns.insert(0, "name")
            info_table = pt.PrettyTable(field_names=columns)

        if len(columns) == 1: #no input
            logger.error("insufficient table fields")
            return info_table

        row = self._extract_info(True, *columns)
        info_table.add_row(row)
        
        #can't access null directory
        if not self:
            if len(self.__directories) or len(self.__files):
                info_table.add_row([f"{self.color}...{RESET_COLOR}" * len(columns)])
            return info_table
        
        for file in self.__files:
            info_table.add_row(file._extract_info(True, *columns))

        for info_directory in self.__directories:
            info_table = info_directory.info(info_table=info_table)

        return info_table

    def _extract_info(self, with_color : bool = True, *tags : str)->List[Any]:
        """
        Returns a list of properties based on entered tags, possible tags:

        **path, absolute, type, count, totalcount, exists, hidden, is_root, root, parent, depth, ctime, lmtime, latime, size_b, size_kb, size_mb**
        """
        output : List[Any] = []
        if not self:
            return output

        size_bytes : int = 0
        for field in tags:
            match field.lower():
                case "name":
                    output.append(f"{self.color}{self.name}{RESET_COLOR}" if with_color else f"{self.name}")
                case "path":
                    output.append(f"{self.color}{self.path}{RESET_COLOR}" if with_color else f"{self.path}")
                case "absolute":
                    output.append(f"{self.color}{self.absolute}{RESET_COLOR}" if with_color else f"{self.absolute}")
                case "type":
                    output.append(f"{self.color}{self.type}{RESET_COLOR}" if with_color else f"{self.type}")
                case "count":
                    output.append(f"{self.color}Dirs:{len(self.__directories)}|Files:{len(self.__files)}{RESET_COLOR}" if with_color else f"Dirs:{len(self.__directories)}|Files:{len(self.__files)}")
                case "totalcount":
                    output.append(f"{self.color}{len(self)}{RESET_COLOR}" if with_color else f"{len(self)}")
                case "exists":
                    output.append(f"{self.color}{self.exists}{RESET_COLOR}" if with_color else f"{self.exists}")
                case "hidden":
                    output.append(f"{self.color}{self.hidden}{RESET_COLOR}" if with_color else f"{self.hidden}")
                case "is_root":
                    output.append(f"{self.color}{self.is_root}{RESET_COLOR}" if with_color else f"{self.is_root}")
                case "root":
                    output.append(f"{self.color}{self.root}{RESET_COLOR}" if with_color else f"{self.root}")
                case "parent":
                    output.append(f"{self.color}{self.parent}{RESET_COLOR}" if with_color else f"{self.parent}")
                case "depth":
                    output.append(f"{self.color}{self.depth}{RESET_COLOR}" if with_color else f"{self.depth}")
                case "ctime":
                    output.append(f"{self.color}{self.creation_stime}{RESET_COLOR}" if with_color else f"{self.creation_stime}")
                case "lmtime":
                    output.append(f"{self.color}{self.modified_stime}{RESET_COLOR}" if with_color else f"{self.modified_stime}")
                case "latime":
                    output.append(f"{self.color}{self.accessed_stime}{RESET_COLOR}" if with_color else f"{self.accessed_stime}")
                case "size_b":
                    if not size_bytes:
                        size_bytes = self.true_size
                    output.append(f"{self.color}{size_bytes} Bytes{RESET_COLOR}" if with_color else f"{size_bytes} Bytes")
                case "size_kb":
                    if not size_bytes:
                        size_bytes = self.true_size
                    output.append(f"{self.color}{round(size_bytes / 1024, 3)} KB{RESET_COLOR}" if with_color else f"{round(size_bytes / 1024, 3)} KB")
                case "size_mb":
                    if not size_bytes:
                        size_bytes = self.true_size
                    output.append(f"{self.color}{round(size_bytes / 1024 * 1024, 5)} MB{RESET_COLOR}" if with_color else f"{round(size_bytes / 1024 * 1024, 5)} MB")
                case _:
                    output.append(f"{self.color}{NOT_FOUND}{RESET_COLOR}" if with_color else f"{NOT_FOUND}")
        return output

    def _in_dir(self, path : Union[PathLike, Directory])->bool:
        """
        Checks if directory path exists in root directory.
        """
        if not self:
            logger.warning("Directory '%s' is null and cannot return objects", self)
            return False

        for i_directory in self.__directories:
            if i_directory == path:
                return True
            if i_directory._in_dir(path):
                return True
        return False

    def _in_files(self, key : Union[PathLike, FileParser])->bool:
        """
        Checks if file exists in root directory.
        """
        if not self:
            logger.warning("Directory '%s' is null and cannot return objects", self)
            return False

        if key in self.__files:
            return True
        for i_directory in self.__directories:
            if key in i_directory.files:
                return True
            if i_directory._in_files(key):
                return True
        return False
    
    def _redirect(self, new_master : Directory)->None:
        """
        Cuts the instance directory and pastes it to a new parent of type Directory, requires global permission.
        """
        if not isinstance(new_master, Directory):
            logger.error("Expected Directory object, got '%s' instead", type(new_master).__name__)
            return None
        if not new_master:
            logger.error("Directory '%s' is null and cannot return objects", new_master)
            return None
        if not Directory.__PERMISSION:
            logger.error("Missing permission in Directory")
            return None
        Directory.__PERMISSION = False

        old_path = self.absolute
        #new assignment
        if new_master is not self.__parent:
            if self.__parent is not None:
                if not self.__parent: #checks if parent is null, different in this case than None
                    logger.error("Cannot modify null directory %s", self.__parent)
                    return None

                #parent unassignment
                if Directory.__CACHE_ENABLE:
                    self.__cache.forget_dir(self)
                Directory.__PERMISSION = True
                self.__parent._remove_directory(self)

            #new parent assignment
            self.__path = new_master.path / self.name
            Directory.__PERMISSION = True
            new_master._add_dir(self)

            #syncing cache
            if Directory.__CACHE_ENABLE:
                Directory.__PERMISSION = True
                new_master._sync_ch(self)

        else: #same parent, updated path (likely parent rename call)
            self.__path = new_master.path / self.name
            if self.__path and old_path == self.__path:
                #case when path does not change, happens when trying to add a directory that exists in sub dir while global lock is enabled
                return

            #syncing cache
            if Directory.__CACHE_ENABLE:
                self.__cache.replace([old_path], [self.absolute])

        if Directory.__GLOBAL_LOCK and not self.__hidden:
            Directory.__update(old_path, self.absolute)

        #syncing subs
        for i_directory in self.__directories:
            Directory.__PERMISSION = True
            i_directory._redirect(self)
        
        for file in self.__files:
            Directory.__PERMISSION = True
            file._redirect(self)
        return

    def _sync_ch(self, sub_dir : Directory)->None:
        if not Directory.__PERMISSION:
            logger.error("Missing permission in %s", Directory.__name__)
        sub_dir._set_ch(self.__cache)
        if Directory.__CACHE_ENABLE:
            self.__cache[sub_dir.path] = sub_dir
        return
        
    def _get_ch(self)->Optional[DirCache]:
        """
        Managed inside instance.
        """
        if not Directory.__PERMISSION:
            logger.error("Missing permission in %s", Directory.__name__)
            return None
        Directory.__PERMISSION = False
        return self.__cache

    def _set_ch(self, new : DirCache)->None:
        """
        Managed inside instance.
        """
        if not Directory.__PERMISSION:
            logger.error("Missing permission in %s", Directory.__name__)
            return None
        Directory.__PERMISSION = False
        self.__cache = new
        return

    def _add_dir(self, new_directory : Directory)->None:
        """
        A direct way to add a sub directory, requires a pass from root directory.
        """
        if not self:
            logger.warning("Directory '%s' is null and cannot be modified", self)
            return None

        if not Directory.__PERMISSION:
            logger.error("Missing permission in %s", Directory.__name__)
            return

        Directory.__PERMISSION = False
        self.__directories.append(new_directory)
        new_directory.parent = self
        return
    
    def _add_file(self, new_file : FileParser)->None:
        """
        A direct way to add a file to root directory, requires a pass.
        """
        if not self:
            logger.warning("Directory '%s' is null and cannot be modified", self)
            return None

        if not Directory.__PERMISSION and not new_file.permitted:
            logger.error("Missing permission in %s and %s", Directory.__name__, FileParser.__name__)
            return

        if Directory.__PERMISSION:
            Directory.__PERMISSION = False
        self.__files.append(new_file)
        new_file.parent = self
        return

    def _remove_file(self, file : FileParser)->bool:
        """
        Takes the file name and removes the first instance it finds one layer down.
        """
        if not self:
            logger.warning("Directory '%s' is null and cannot be modified", self)
            return False
        if not Directory.__PERMISSION and not file.permitted:
            logger.error("Missing permission in %s", Directory.__name__)
            return False
        if Directory.__PERMISSION:
            Directory.__PERMISSION = False

        self.__files.remove(file)
        file.parent = None    
        return True

    def _remove_directory(self, directory : Directory)->bool:
        """
        Removes the first instance of the path it finds one layer down in instance.
        """
        if not self:
            logger.warning("Directory '%s' is null and cannot be modified", self)
            return False

        if not Directory.__PERMISSION:
            logger.error("Missing permission in %s", Directory.__name__)
            return False

        Directory.__PERMISSION = False
        self.__directories.remove(directory)
        directory.parent = None
        return True

    def _items(self)->Tuple[List[Directory], List[FileParser]]:
        """
        Returns two arrays of sub Directories and Files respectively.
        """
        if not self:
            logger.warning("Directory '%s' is null and cannot return objects", self)
            return list(), list()
        directories : List[Directory] = [self]
        files : List[FileParser] = [file for file in self.__files]
        if not self:
            return list(), list()
        for i_dir in self.__directories:
            i_dirs, i_files = i_dir._items()
            directories.extend(i_dirs)
            files.extend(i_files)
        return directories, files

    def _numeric_name(self, name : str, is_file : bool = False)->str:
        """
        Gives the duplicated name a proper number.

        Returns the same name if it appears once at most, otherwise renames it with a numerical format (n).
        """
        #how many times name appears on instance
        if not self:
            logger.warning("Directory '%s' is null and cannot be modified", self)
            return str()
        count, flag = self.__count_name(name)
        suffix = str()

        if not flag or count <= 1: #not a dup
            return name

        if is_file:
            sep = name.rfind('.')
            name, suffix = (name[:sep], name[sep:]) if sep != -1 else (name, str())

        if count == 2:
            #no margin of error
            name = name + f"({count - 1})" + suffix
            logger.info("Updated to name '%s'", name)
            return name

        elif count > 2:
            #search for the smallest n
            for n in range(1, count+1):
                temp = name + f"({n})" + suffix
                if not temp in self.__directories and temp not in self.__files:
                    #name available
                    logger.info("Changed name '%s' to '%s'", name, temp)
                    return temp
            else: #impossible case
                logger.critical("No numeric name was found for %s", name)
                raise Exception(f"No numeric name was found for '{name}'")

    def __count_name(self, directory_or_file : str | DirFile)->Tuple[int, bool]:
        """
        Checks how many instances have duplicated name (counting renamed duplicates) and returns the count.
        """
        if isinstance(directory_or_file, (Directory, FileParser)):
            directory_or_file = directory_or_file.name

        count_dups = 0
        #first check if the same name appears more than once
        requires_rename = self.__directories.count(directory_or_file) + self.__files.count(directory_or_file) >= 2
        if requires_rename: #if it does then we check how many previous dups exist
            count_dups = len(list(filter(lambda element: self.__comp(directory_or_file, element), self.__directories + self.__files)))
        logger.debug("found %d duplicate names in '%s'", count_dups, self)
        return count_dups, requires_rename

    def __comp(self, name : str, dir : DirFile)->bool:
        """
        Checks if dir is a duplicate of name.
        """
        content : Optional[Tuple[str, int]] = PathManager.strip(dir)
        if content is None:
            return name == dir.name
        return name == content[0]

    def __build(self, path : PathLike, create_permission : bool = True)->Optional[Directory]:
        """
        Builds a sub path starting from instance directory and returns the last part of the path as a Directory object.
        """
        path_parts = path.parts
        current_directory = self

        for i, part in enumerate(path_parts):
            temp_dir = current_directory.get_directory(part)
            if temp_dir is None:

                if not create_permission:
                    return temp_dir
                
                temp_path = Path(*path_parts[:i + 1])
                temp_dir = Directory(current_directory.path / temp_path.name, self.__cache)
                Directory.__PERMISSION = True
                temp_dir._redirect(current_directory)
            
            current_directory = temp_dir
        return current_directory

    def __find_name(self, _t : Literal["Directory", "File"], names : Iterable[str])->Union[List[FileParser], List[Directory]]:
        if _t == "File":
            return list(filter(lambda file: file.name in names, self.__files))
        elif _t == "Directory":
            return list(filter(lambda i_directory: i_directory.name in names, self.__directories))

    def __find_suffix(self, suffixes : Iterable[str])->List[FileParser]:
        return list(filter(lambda file: file.suffix in suffixes, self.__files))

    def __find_exists(self, _t : Literal["Directory", "File"], condition : Literal[0, 1])->Union[List[FileParser], List[Directory]]:
        if _t == "File":
            return list(filter(lambda file: file.exists if condition else not(file.exists), self.__files))

        elif _t == "Directory":
            return list(filter(lambda i_directory: i_directory.exists if condition else not(i_directory.exists), self.__directories))

    def __find_size(self, _t : Literal["Directory", "File"], condition : Union[NumericCondition, NumericTrue])->Union[List[FileParser], List[Directory]]:
        if _t == "File":
            if isinstance(condition, NumericCondition):
                    return list(filter(lambda file: file.size_bytes in condition, self.__files))
            else:
                    return list(filter(lambda file: file.size_bytes == condition, self.__files))

        elif _t == "Directory":
            if isinstance(condition, NumericCondition):
                    return list(filter(lambda i_directory: i_directory.true_size in condition, self.__directories))
            else:
                    return list(filter(lambda i_directory: i_directory.true_size == condition, self.__directories))

    def __find_subs(self, condition : Union[NumericCondition, NumericTrue])->List[Directory]:
        if isinstance(condition, NumericCondition):
            return list(filter(lambda i_directory: len(i_directory.directories) + len(i_directory.files) in condition, self.__directories))
        else:
            return list(filter(lambda i_directory: len(i_directory.directories) + len(i_directory.files) == condition, self.__directories))

    def __find_depth(self, _t : Literal["Directory", "File"], condition : Union[NumericCondition, NumericTrue])->Union[List[FileParser], List[Directory]]:
        if _t == "File":
            if isinstance(condition, NumericCondition):
                    return list(filter(lambda file: file.depth in condition, self.__files))
            else:
                    return list(filter(lambda file: file.depth == condition, self.__files))

        elif _t == "Directory":
            if isinstance(condition, NumericCondition):
                    return list(filter(lambda i_directory: i_directory.depth in condition, self.__directories))
            else:
                    return list(filter(lambda i_directory: i_directory.depth == condition, self.__directories))

    def __getitem__(self, path : PathLike)->Optional[Directory]:
        """
        Only to search for a folder in directory tree, not used for files and returns the first result it finds.
        For better results pass the relative or absolute path of the directory.
        """
        if not isinstance(path, PathLike) or not path:
            return None

        if not self:
            logger.warning("Directory '%s' is null and cannot return objects.", self)
            return None

        logger.debug("'%s' getter called for '%s'", self, path)

        if str(path) == '.':
            return self if Path.cwd() == self.__path else self[Path.cwd()]

        path = Path(path)

        if path.absolute() == self.absolute:
            return self

        if Directory.__CACHE_ENABLE:
            cache_hit = self.__cache[path]
            if cache_hit is not None:
                if len(cache_hit.absolute.parts) > len(self.absolute.parts): 
                    return cache_hit
            return self if path == self else None

        if path == self:
            return self

        for i_directory in self.__directories:
            if i_directory == path:
                return i_directory
            
            result = i_directory[path]
            if result is not None:
                return result
        return

    def __eq__(self, other : Union[PathLike, Directory])->bool: 
        """
        Checks by comparing paths.
        """
        if isinstance(other, (Directory, PathLike)):
            return PathManager.relative(self, other)

        if other is None:
            return bool(self) == False

        logger.critical("Cannot compare between %s and %s", Directory.__name__, type(other).__name__)
        raise Exception(f"Cannot compare between {Directory.__name__} and {type(other).__name__}")

    def __contains__(self, other : Union[PathLike, Directory, FileParser])->bool:
        """
        Searches in current dir and all sub directories for a file or a directory.
        """
        if isinstance(other, Directory):
            return self._in_dir(other)
        elif isinstance(other, FileParser):
            return self._in_files(other)
        elif isinstance(other, PathLike):
            path = Path(other)
            if self._in_dir(other):
                return True
            else:
                if len(path.parts) == 1:
                    return self._in_files(path.name)
                temp_path = PathManager.set(self, Path(*path.parts[:-1]), False)
                temp = self.__build(temp_path, False)
                if temp is not None:
                    return temp._in_files(path.name)
        
        return False

    def __len__(self)->int:
        """
        Amount of objects in directory including both files and sub directories.
        """
        s : int = len(self.__files) + len(self.__directories)
        for i_dir in self.__directories:
            s += len(i_dir)
        return s

    def __str__(self)->str:
        """
        Represented by path.
        """
        return str(self.__path)

    def __repr__(self)->str:
        """
        Object is represented by the name of the directory.
        """
        return self.__path.name

    def __bool__(self)->bool:
        """
        Boolean value of Directory object is determined by whether path is null or not.
        """
        return bool(self.__path)

    def __del__(self)->None:
        try:
            if sys.meta_path is None:
                return
        except ImportError:
            return

        if Directory.__GLOBAL_REFERENCE.get(self.absolute) is self:
            logger.debug("'%s' freed from reference memory", self)
            Directory.__GLOBAL_REFERENCE.pop(self.absolute)

    @property
    def path(self)->Path:
        """
        Original path for instance.
        """
        return self.__path

    @property
    def absolute(self)->Path:
        return self.__path.absolute()
    
    @property
    def name(self)->str:
        """
        The name of the directory.
        """
        return self.__path.name

    @property
    def exists(self)->bool:
        """
        Returns Path.exists method.
        """
        return self.__path.exists() if self else False

    @property
    def hidden(self)->bool:
        return self.__hidden

    @property
    def color(self)->str:
        """
        Color of directory name shown when using .info method.
        """
        return self.__color if self else Directory._PATHLESS_COLOR
    
    @property
    def is_root(self)->bool:
        """
        Returns True if instance has no parent directory.
        """
        return self.__parent is None if self else False

    @property
    def is_null(self)->bool:
        """
        Checks if directory has no path (has no path -> has no parent + is not root).
        """
        return not(bool(self.__path)) 

    @property
    def files(self)->List[FileParser]:
        """
        Files in directory.
        """
        return [file for file in self.__files] if self else list()
    
    @property
    def directories(self)->List[Directory]:
        """
        Sub directories one layer down.
        """
        return [d for d in self.__directories] if self else list()
    
    @property
    def root(self)->Optional[Directory]:
        """
        Returns the root directory for instance.
        """
        return self.__root if not self.is_root else self
        
    @property
    def parent(self)->Optional[Directory]:
        """
        Equivalent of Path.parent but returns the Directory object that holds this instance's reference (root dir is fatherless).
        """
        return self.__parent

    @property
    def depth(self)->int:
        """
        Depth calculated between instance and root object and not actual root in files system (unless...)
        """
        return 0 if self.is_root or not self else len(PathManager.set(self.root, self, False).parts)

    @property
    def type(self)->str:
        return Directory.__name__

    @property
    def size_bytes(self)->Optional[int]:
        """
        Size of directory instance itself in the files system.
        """
        return self.__path.stat().st_size if self.exists else None
    
    @property
    def true_size(self)->Optional[int]:
        """
        Calculated size in directory tree.
        """
        if not self.exists:
            return 0
        size_bytes = 0
        for file in self.__files:
            if file.exists:
                size_bytes += file.size_bytes
        
        for i_dir in self.__directories:
            size_bytes += i_dir.true_size
        
        return size_bytes

    @property
    def creation_time(self)->Optional[int]:
        """
        Creation time in seconds.
        """
        return self.__path.stat().st_ctime if self.exists else None

    @property
    def creation_stime(self)->Optional[str]:
        """
        Creation time in string format.
        """
        return time.ctime(self.creation_time) if self.exists else None

    @property
    def accessed_time(self)->Optional[int]:
        """
        Last accessed time in seconds.
        """
        return self.__path.stat().st_atime if self.exists else None

    @property
    def accessed_stime(self)->Optional[str]:
        """
        Last accessed time in string format.
        """
        return time.ctime(self.accessed_time) if self.exists else None

    @property
    def modified_time(self)->Optional[int]:
        """
        Last modified time in seconds.
        """
        return self.__path.stat().st_mtime if self.exists else None

    @property
    def modified_stime(self)->Optional[str]:
        """
        Last modified time in string format
        """
        return time.ctime(self.modified_time) if self.exists else None

    @parent.setter
    def parent(self, new : Optional[Directory])->None:
        """
        Internally managed.
        """
        if not isinstance(new, Directory) and new is not None: #new is neither an Directory object nor None
            logger.critical("Expected Directory object, got %s instead", type(new).__name__)
            raise Exception(f"Expected Directory object, got {type(new).__name__} instead")
        
        if new is not None and self.__parent is not None: #new is Directory and parent exists
            logger.critical("Cannot assign directory '%s' to parent '%s' as it already has a parent '%s'.", self, new, self.__parent)
            raise Exception(f"Cannot assign directory '{self}' to parent '{new}' as it already has a parent '{self.__parent}'.")

        if new is None and self.__parent is not None: #new is None and parent is Directory, unassignment case
            if self.__parent.get_directory(self.name):
                logger.critical("Cannot unassign directory '%s' to '%s' while still belonging to parent '%s'.", self, new, self.__parent)
                raise Exception(f"Cannot unassign directory '{self}' to '{new}' while still belonging to parent '{self.__parent}'.")

            self.__parent = None
            self.__root = None
            self.__path = NullPath(self.__path)

        elif isinstance(new, Directory) and self.__parent is None: #new is Directory and parent is None, assignment case
            if new.get_directory(self.name) is None:
                logger.critical("Cannot assign '%s' to new parent '%s' as no confirmation from parent is available to add child directory.", self, new)
                raise Exception(f"Cannot assign '{self}' to new parent '{new}' as no confirmation from parent is available to add child directory.")

            logger.debug("directory '%s' assigned to parent '%s'", self, new)
            self.__parent = new
            self.__root = new.root
    
    @color.setter
    def color(self, new : str)->None:
        self.__color = new
    
    @property
    def permitted(self)->bool:
        """
        Instance can make modifications if True.
        """
        return Directory.__PERMISSION

#########################################
    #FILES PARSERS
#########################################
class FileParser(ABC):
    """
    Base class for file parsers, initiation of abstract class returns a child class instance of the optimal parser type.

    File Parsers share references memory similar but separated of Directory objects, can be toggled using FileParser._toggle_lock at the initial parts of the program. 

    Initializing using FileParser class will return the most suitable available type for a file reference:
    >>> file : TextParser = FileParser(path="file.txt")

    Note that each parser has their own methods and properties, even abstracted methods may require separate arguments, abstracted methods are:
    >>> FileParser.read(overwrite : bool) #main method for data parsing, overwrite forces new data in object.
    >>> FileParser.write() #main data writing method, works as native .write and creates the file if it does not exist.
    >>> FileParser.rename(name : Optional[str]) #used to change the name of the file (only file name and not suffix), handles duplicated names if file has a parent Directory object.

    All abstracted class types share the same properties with a few extras based on the class type.
    """
    __REFERENCES_MAP : WeakValueDictionary[Path, FileParser] = WeakValueDictionary()
    __PERMISSION : bool = False #permission to modify any file
    __GLOBAL_LOCK : bool = True
    _supported_suffixes : Tuple[str] = TEXT_SUFFIX + JSON_SUFFIX + CSV_SUFFIX + EXCEL_SUFFIX + IMAGE_SUFFIX
    _PATHLESS_COLOR : str = Fore.LIGHTRED_EX
    
    @classmethod
    def is_locked(cls)->bool:
        """
        Returns True if global lock is enabled.
        """
        return cls.__GLOBAL_LOCK

    @classmethod
    def _toggle_lock(cls)->None:
        """
        Toggles references lock for FileParser objects (True by default), disabling references lock ignores file memberships and allows for duplicated references to the same file (enabled for security reasons).

        Can only be toggled before creating the first instance or when all referenced instances are garbage collected.
        """
        if cls.__REFERENCES_MAP:
            logger.error("Cannot toggle files references map after creating an instance.")
            return None
        cls.__GLOBAL_LOCK = not(cls.__GLOBAL_LOCK)
        logger.debug("Files reference map was set to '%s'", cls.__GLOBAL_LOCK)
        return

    @classmethod
    def _update(cls, key : Path, new_key : Path)->None:
        """
        Internally managed, updates paths for active files.
        """
        if not cls.__PERMISSION:
            logger.error("Missing permission when attempting to update a file path.")
            return None
        cls.__PERMISSION = False
        if key == new_key:
            return None
        if key in cls.__REFERENCES_MAP:
            ref = cls.__REFERENCES_MAP.pop(key)
        cls.__REFERENCES_MAP[new_key] = ref
        logger.debug("reference '%s' updated in FileParser map", new_key)
        return

    @classmethod
    def _del(cls, abs_key : Path, inst : FileParser)->None:
        """
        Called for garbage collection.
        """
        if not cls.__PERMISSION:
            logger.error("Missing permission when attempting to delete a reference to file.")
            return None
        cls.__PERMISSION = False
        if cls.__REFERENCES_MAP.get(abs_key) is inst:
            cls.__REFERENCES_MAP.pop(abs_key, None)
        return

    @classmethod
    def _contains(cls, abs_path : Path)->bool:
        """
        Checks if absolute path is referenced as an object.
        """
        return abs_path in cls.__REFERENCES_MAP

    @classmethod
    def _get(cls)->Optional[WeakValueDictionary[Path, FileParser]]:
        """
        Internally managed, returns references map.
        """
        if not cls.__PERMISSION:
            logger.error("Missing permission when attmepting to look up references for files.")
            return None
        cls.__PERMISSION = False
        return cls.__REFERENCES_MAP

    @classmethod
    def _correct_parser(cls, suffix : str)->str:
        """
        Returns the name of the available parser for suffix.
        """
        suffix = suffix.lower()
        if suffix in TEXT_SUFFIX:
            return TextParser.__name__
        elif suffix in JSON_SUFFIX:
            return JsonParser.__name__
        elif suffix in EXCEL_SUFFIX:
            return ExcelParser.__name__ + ' | ' + WorkbookParser.__name__
        elif suffix in IMAGE_SUFFIX:
            return ImageParser.__name__
        elif suffix in CSV_SUFFIX:
            return CsvParser.__name__
        else:
            return TextParser.__name__

    def __new__(cls, path : PathLike, *a, **kw)->FileParser:
        abs_path = Path(path).absolute()
        if cls.__GLOBAL_LOCK and not kw.get("hide", False): #global locked and instance is not hidden
            #file lookup if locked to references
            if abs_path in cls.__REFERENCES_MAP:
                logger.debug("Returned existing reference to file '%s'", abs_path)
                return cls.__REFERENCES_MAP.get(abs_path)
        
        #file not found, create file and return it
        if cls is not FileParser: #forced type class
            instance = super().__new__(cls)
        else: #used abstract class instead
            suffix = Path(path).suffix.lower()
            if suffix in TEXT_SUFFIX:
                instance = super().__new__(TextParser)
            elif suffix in JSON_SUFFIX:
                instance = super().__new__(JsonParser)
            elif suffix in EXCEL_SUFFIX:
                instance = super().__new__(ExcelParser)
            elif suffix in IMAGE_SUFFIX:
                instance = super().__new__(ImageParser)
            elif suffix in CSV_SUFFIX:
                instance = super().__new__(CsvParser)
            else:
                instance = super().__new__(TextParser)
        
        if cls.__GLOBAL_LOCK and not kw.get("hide", False): #file not hidden -> save reference
            cls.__REFERENCES_MAP[abs_path] = instance
        logger.debug("Returned new instance for file '%s'", abs_path)
        return instance

    def __init__(self, path : PathLike, **kw):
        """
        Base class for file parser objects in dir module, can be initialized to return a fitting object type for path or directly initialize an
        abstract class for the correct file type:

        >>> file = FileParser(path="file.txt") #returns a TextParser object instead
        
        **More optimal way**
        >>> file = TextParser(path="file.txt")

        **Arguments**
            -path: Path or string object absolute/relative path or name pointing to a file (file does not have to exist)
            -create: creates file object initially if it does not exist.
            -color: color of the name used in visualization methods in Directory and PathManager class, uses colorama's package colors by default.
            -hide: instance becomes invisible to references map and can be duplicated, by default, only one object can point to a certain file.
        
        Non supported types are initiated as text parsers, supported types are:
        >>> .txt ==> TextParser
        >>> .json ==> JsonParser
        >>> .xlsx ==> ExcelParser | WorkbookParser
        >>> .csv ==> CsvParser
        >>> .jpg ==> ImageParser
        """

    @abstractmethod
    def read(self, overwrite : bool = True, *a, **kw)->Any:
        """
        Reads file and returns its content, if overwrite is False and data is already existent, returns current data without reading the file.
        """
    @abstractmethod
    def write(self, only_create : bool = False, *a, **kw)->None:
        """
        Write data to file in the files system, setting only_create creates the file without writing any data.
        """
    @abstractmethod
    def rename(self, new_name : Optional[str])->None:
        """
        Renames the instance, passing None calls for a check in parent directory on whether the file instance has a duplicate name or not and renames it if so.
        """
    @abstractmethod
    def _redirect(self, new_master : Directory)->None:
        """
        Assigns a parent directory to file object or redirects it to a new parent.
        """
    @abstractmethod
    def _extract_info(self, with_color : bool = True, *tags : KEYS)->List[Any]:
        """
        Returns info about the based on input tags.
        """
    @property
    def parent(self)->Optional[Directory]:
        """
        Parent Directory object.
        """
    @property
    def root(self)->Optional[Directory]:
        """
        Root Directory object.
        """
    @property
    def path(self)->Path:
        """
        Path for object.
        """
    @property
    def name(self)->str:
        """
        Name of the file.
        """
    @property
    def type(self)->str:
        """
        Parser type.
        """
    @property
    def suffix(self)->Optional[str]:
        """
        File suffix.
        """
    @property
    def permitted(self)->bool:
        """
        Checks if object has permission to make modifications.
        """
    @property
    def depth(self)->int:
        """
        Depth calculated within root Directory object and file.
        """
    @property
    def absolute(self)->Path:
        """
        Absolute path for file.
        """
    @property
    def exists(self)->bool:
        """
        Checks if file exists in file explorer.
        """
    @property
    def hidden(self)->bool:
        """
        Checks whether object is not referenced by global map.
        """
    @property
    def color(self)->str:
        """
        Color of file name when using .info method from Directory class.
        """
    @property
    def size_bytes(self)->Union[int, Literal[-1]]:
        """
        Size of the file in bytes (returns -1 if it does not exist).
        """
    @property
    def size_kbytes(self)->Union[int, Literal[-1]]:
        """
        Size of the file in Kilo bytes (returns -1 if it does not exist).
        """
    @property
    def size_mbytes(self)->Union[int, Literal[-1]]:
        """
        Size of the file in Mega bytes (returns -1 if it does not exist).
        """
    @property
    def mode(self)->Optional[int]:
        """
        Mode property from os.stat
        """   
    @property
    def inode(self)->Optional[int]:
        """
        Inode property from os.stat
        """
    @property
    def dev_id(self)->Optional[int]:
        """
        device id property from os.stat
        """    
    @property
    def nlinks(self)->Optional[int]:
        """
        nlinks property from os.stat
        """
    @property
    def creation_time(self)->Optional[int]:
        """
        File creation time in seconds.
        """
    @property
    def creation_stime(self)->Optional[str]:
        """
        File creation time in string format.
        """
    @property
    def accessed_time(self)->Optional[int]:
        """
        File last accessed time in seconds.
        """
    @property
    def accessed_stime(self)->Optional[str]:
        """
        File last accessed time in string format.
        """
    @property
    def modified_time(self)->Optional[int]:
        """
        File last modifed time in seconds.
        """
    @property
    def modified_stime(self)->Optional[str]:
        """
        File last modifed time in string format.
        """

class TextParser(FileParser):
    separator : str = '\n'
    _default_color = Fore.CYAN
    def __init__(self, path : PathLike, **kw):
        """
        Prototype parser for text files and default parser object for non supported file extensions, used as a data container in Directory instances for text files or other files that are not supported by
        FileParser objects.

        **Arguments**
            -path: Path or string object absolute/relative path or name pointing to a file (file does not have to exist)
            -create: creates file object initially if it does not exist.
            -color: color of the name used in visualization methods in Directory and PathManager class, uses colorama's package colors by default.
            -hide: instance becomes invisible to references map and can be duplicated, by default, only one object can point to a certain file.

        **Example**
        >>> file = TextParser("file.txt") #or FileParser("file.txt")
        >>> data = file.read(overwrite=True, ...) #read and return data from file if it exists
        >>> file.add_line(row="new data", index=None) #appends data to a new line
        >>> file.write(...) #writes to file and creates if it does not exist
        """
        if hasattr(self, "_init"):
            return
        self._init : bool = True
        self.__path : Path = Path(path)
        self.__data : str = str()
        self.__parent : Optional[Directory] = None
        self.__color : str = kw.get("color", TextParser._default_color)

        if self.__path.suffix.lower() not in TEXT_SUFFIX and self.__path.suffix in FileParser._supported_suffixes:
            logger.critical("Expected a .txt file or any file that is not supported, for %s file use %s instead.", self.__path.suffix, FileParser._correct_parser(self.__path.suffix))
            raise Exception(f"Expected a .txt file or any file that is not supported, for {self.__path.suffix} file use {FileParser._correct_parser(self.__path.suffix)} instead.")
        if self.exists and self.__path.is_dir():
            logger.critical("Path must point to a file, not a directory.")
            raise Exception("Path must point to a file, not a directory.")
        
        self.__hidden : bool = kw.get("hide", False)
        if kw.get("create", False):
            self.__path.open("a").close()
#METHODS
    def read(self, overwrite : bool = True, buffering : int = -1, encoding : Optional[str] = None, errors : Optional[str] = None, newline : Optional[str] = None)->str:
        """
        Fetch data from file it if exists, if data exists in object and overwrite is False then returns data without reading from
        file.
        """
        if not self.exists:
            return self.__data
        if not self.__data or (self.__data and overwrite):
            try:
                with self.__path.open('r', buffering=buffering, encoding=encoding, errors=errors, newline=newline) as file:
                    self.__data = file.read()
                logger.info("data for '%s' updated from file", self)
            except UnicodeDecodeError:
                logger.error("Failed to read file '%s'", self.__path)
        return self.__data

    def write(self, only_create : bool = False, mode : str = 'w', buffering : int = -1, encoding : Optional[str] = None, errors : Optional[str] = None, newline : Optional[str] = None)->None:
        """
        Writes data to file and creates the file if it does not exist.
        """
        if not self:
            logger.error("Missing file '%s' to write to", self)
            return None

        if only_create:
            self.__path.open("a").close()
            return None

        with self.__path.open(mode=mode, buffering=buffering, encoding=encoding, errors=errors, newline=newline) as file:
            file.write(self.__data)
        logger.debug("wrote data into file: '%s'", self)
        return

    def rename(self, new_name : Optional[str])->None:
        if not self:
            logger.error("Cannot modify the file '%s'", self)
            return None

        #string handling case       
        if isinstance(new_name, str): 
            if PathManager.os() == OperatingSystem.WINDOWS and PathManager.contains_illegal_chars_win(new_name):
                logger.error("name %s contains illegal characters.", new_name)
                return None            
            elif PathManager.os() in (OperatingSystem.LINUX, OperatingSystem.MACOS) and PathManager.contains_illegal_chars_psx(new_name):
                logger.error("name %s contains illegal characters.", new_name)
                return None            

            if len(Path(new_name).parts) != 1:
                logger.error("name must not include other directory parts.")

            #if file has parent
            if self.__parent and Directory._blocks_duplicates():
                new_name = self.__parent._numeric_name(new_name+self.suffix, True) #get correct name

            #no parent, add suffix to it
            else:
                new_name = new_name + self.suffix

        #requesting a duplicate rename
        elif new_name is None:
            if self.__parent is None:
                return
            new_name = self.__parent._numeric_name(self.name, True) #get correct name

        else: #new name is neither str or None
            logger.error("Expected a string or None type, got %s instead", type(new_name).__name__)
            return None

        #name did not change
        if new_name == self.name:
            logger.info("New name %s matches the old name, skipping the rest of the process", new_name)
            return None

        #new path conclusion
        path, old_name = PathManager.split(self.__path)
        new_path = Path(path) / new_name if path else new_name
        #check if new path is already occupied
        if FileParser.is_locked() and not self.__hidden:
            #check if new path already exists
            if FileParser._contains(new_path.absolute()):
                logger.error("File with the name '%s' already exists in this directory.", new_name)
                return None

        #updating name
            FileParser._FileParser__PERMISSION = True
            FileParser._update(self.absolute, new_path.absolute())
            if self.exists:
                self.__path.rename(new_path)

        self.__path = new_path
        logger.info("File '%s' updated to '%s'", old_name, new_name)
        return

    def overwrite_content(self, content : str)->None:
        """
        Replaces the object's data with content.
        """
        if not isinstance(content, str):
            logger.error("Expected string content, got %s instead", type(content).__name__)
            return None
        self.__data = content
        logger.info("Updated '%s' file content", self)
        return

    def encrypt_data(self, func : Callable, *args, **kw)->None:
        """
        Takes an encryption function with a return type of string, can also be used to decrypt using this method.

        Function must take a string variable as its first argument
        """
        data = func(self.__data, *args, **kw)
        if not isinstance(data, str):
            logger.error("Expected string content, got %s instead", type(data).__name__)
            return None
        self.__data = data
        logger.info("Encryped '%s' file content", self)
        return

    def add_line(self, row : str, index : Optional[int] = None)->None:
        """
        Adds a line of data to the specified index, otherwise it is appended as the last line.
        """
        if not isinstance(row, str):
            logger.error("Expected string content, got %s instead", type(content).__name__)
            return None

        if not self.__data:
            self.__data = row
            return None

        if index is None:
            self.__data += TextParser.separator + row
            return None
        
        content = self.lines
        content.insert(index, row)
        self.__data = TextParser.separator.join(content)
        return

    @overload
    def remove_row(self, row : str)->None:...
    @overload
    def remove_row(self, row : int)->None:...

    def remove_row(self, row : Union[str, int])->None:
        """
        Removes a row of data from file data, row can be the index or the content of the string line.
        """
        try:
            content = self.lines
            if isinstance(row, int):
                content.pop(row)
            elif isinstance(row, str):
                content.remove(row)
            else:
                raise NotImplementedError
        except (IndexError, ValueError):
            return

        self.overwrite_content(TextParser.separator.join(content))
        return

    def replace(self, old_text : str, new_text : str)->None:
        """
        Replaces old text with new text through the whole file data scope.
        """
        if not self.__data:
            return
        self.__data = self.__data.replace(old_text, new_text)

    def filter(self, pattern : str)->None:
        """
        Removes specified pattern from data.
        """
        self.__data = self.__data.replace(pattern, '')

    def clear(self)->None:
        """
        Clears object data, actual file is unchanged until write method is called.
        """
        self.__data = str()
    
    def _redirect(self, new_master : Directory)->None:
        if not isinstance(new_master, Directory):
            logger.error("expected Directory object, got %s instead.", type(new_master).__name__)
            return None
        if not new_master:
            logger.error("Cannot modify the directory: %s.", new_master)
            return None
        if not new_master.permitted:
            logger.error("Missing permission to redirect '%s'", self)
            return None

        old_path = self.absolute
        if new_master is not self.__parent:
            FileParser._FileParser__PERMISSION = True
            if self.__parent is not None:
                self.__parent._remove_file(self)
            self.__path = new_master.path / self.name
            new_master._add_file(self)
            FileParser._FileParser__PERMISSION = False
        else:
            self.__path = new_master.path / self.name
            if self.__path and old_path == self.__path:
                #case when dir already exists in parent + path does not change
                logger.error("File with the name '%s' already exists in directory '%s'.", self.name, new_master)
                return None

        if FileParser.is_locked() and not self.__hidden:
            FileParser._FileParser__PERMISSION = True
            FileParser._update(old_path, self.__path.absolute())
        
        logger.info("Added file '%s' to directory '%s'", self.name, self.__parent)
        return

    def _extract_info(self, with_color : bool = True, *tags : KEYS)->List[Any]:
        """
        Returns an iterable of properties based on input tags, possible tags:

        **name, path, type, exists, hidden, parent, root, absolute, depth, size_b, size_kb, size_mb, ctime, lmtime, latime, mode, inode, dev_id, nlinks**
        """
        output = []
        if not self:
            return output

        for field in tags:
            match field.lower():
                case "name":
                    output.append(f"{self.color}{self.name}{RESET_COLOR}" if with_color else f"{self.name}")
                case "path":
                    output.append(f"{self.color}{self.path}{RESET_COLOR}" if with_color else f"{self.path}")
                case "type":
                    output.append(f"{self.color}{self.type}{RESET_COLOR}" if with_color else f"{self.type}")
                case "exists":
                    output.append(f"{self.color}{self.exists}{RESET_COLOR}" if with_color else f"{self.exists}")
                case "hidden":
                    output.append(f"{self.color}{self.hidden}{RESET_COLOR}" if with_color else f"{self.hidden}")
                case "root":
                    output.append(f"{self.color}{self.root}{RESET_COLOR}" if with_color else f"{self.root}")
                case "parent":
                    output.append(f"{self.color}{self.parent}{RESET_COLOR}" if with_color else f"{self.parent}")
                case "absolute":
                    output.append(f"{self.color}{self.absolute}{RESET_COLOR}" if with_color else f"{self.absolute}")
                case "depth":
                    output.append(f"{self.color}{self.depth}{RESET_COLOR}" if with_color else self.depth)
                case "size_b":
                    output.append(f"{self.color}{self.size_bytes} B{RESET_COLOR}" if with_color else self.size_bytes)
                case "size_kb":
                    output.append(f"{self.color}{self.size_kbytes} KB{RESET_COLOR}" if with_color else self.size_kbytes)
                case "size_mb":
                    output.append(f"{self.color}{self.size_mbytes} MB{RESET_COLOR}" if with_color else self.size_mbytes)
                case "ctime":
                    output.append(f"{self.color}{self.creation_stime}{RESET_COLOR}" if with_color else f"{self.creation_stime}")
                case "lmtime":
                    output.append(f"{self.color}{self.modified_stime}{RESET_COLOR}" if with_color else f"{self.modified_stime}")
                case "latime":
                    output.append(f"{self.color}{self.accessed_stime}{RESET_COLOR}" if with_color else f"{self.accessed_stime}")
                case "mode":
                    output.append(f"{self.color}{self.mode}{RESET_COLOR}" if with_color else self.mode)
                case "inode":
                    output.append(f"{self.color}{self.inode}{RESET_COLOR}" if with_color else self.inode)
                case "dev_id":
                    output.append(f"{self.color}{self.dev_id}{RESET_COLOR}" if with_color else self.dev_id)
                case "nlinks":
                    output.append(f"{self.color}{self.nlinks}{RESET_COLOR}" if with_color else self.nlinks)
                case _:
                    output.append(f"{self.color}{NOT_FOUND}{RESET_COLOR}" if with_color else f"{NOT_FOUND}")

        return output

#MAGIC
    def __repr__(self)->str:
        return self.__path.name

    def __str__(self)->str:
        return str(self.__path)
    
    def __len__(self)->int:
        return len(self.lines)
    
    def __bool__(self)->bool:
        return bool(self.__path)

    def __eq__(self, other : Union[FileParser, PathLike])->bool:
        if isinstance(other, (FileParser, PathLike)):
            return PathManager.relative(self, other)
        elif other is None:
            return bool(self) == False
        logger.critical("Cannot compare %s type with %s type.", FileParser.__name__, type(other).__name__)
        raise Exception(f"Cannot compare {FileParser.__name__} type with {type(other).__name__} type.")

    def __del__(self)->None:
        try:
            if sys.meta_path is None:
                return
        except ImportError:
            return
        FileParser._FileParser__PERMISSION = True
        FileParser._del(self.__path.absolute(), self)

#SPECIAL
    @property
    def lines(self)->List[str]:
        """
        Returns all lines in data as a list.
        """
        return self.__data.split('\n') if self.__data else list()

#PROPERTIES
    @property
    def path(self)->Path:
        return self.__path
    @property
    def name(self)->str:
        return self.__path.name
    @property
    def type(self)->str:
        return f"{self.suffix} File"
    @property
    def suffix(self)->str:
        return self.__path.suffix
    @property
    def root(self)->Optional[Directory]:
        return self.__parent.root if self.__parent else None
    @property
    def depth(self)->int:
        return 0 if not self else len(PathManager.set(self.parent.root, self.parent, False).parts)
    @property
    def absolute(self)->Path:
        return self.__path.absolute()
    @property
    def exists(self)->bool:
        return self.__path.exists() if self else False
    @property
    def hidden(self)->bool:
        return self.__hidden
    @property
    def color(self)->str:
        return self.__color if self else FileParser._PATHLESS_COLOR
    @property
    def permitted(self)->bool:
        return FileParser._FileParser__PERMISSION
    @property
    def parent(self)->Optional[Directory]:
        """
        Equivalent of Path.parent but returns the Directory object that holds this instance's reference (root dir is fatherless).
        """
        return self.__parent
    @parent.setter
    def parent(self, new : Optional[Directory])->None:
        """
        Internally managed.
        """
        if not isinstance(new, Directory) and new is not None:
            logger.critical("Expected Directory object, got %s instead.", type(new).__name__)
            raise Exception(f"Expected Directory object, got {type(new).__name__} instead.")

        if new is not None and self.__parent is not None:
            logger.critical("Cannot assign parent to '%s' as it already has a parent '%s'.", self.name, self.__parent)
            raise Exception(f"Cannot assign parent to '{self.name}' as it already has a parent '{self.__parent}'.")

        if new is None and self.__parent is not None:
            if self.__parent.get_file(self.name):
                logger.critical("Failed to remove file '%s' from parent '%s'.", self.name, self.__parent)
                raise Exception(f"Failed to remove file '{self.name}' from parent '{self.__parent}'.")
            self.__parent = None
            self.__path = NullPath(self.absolute)

        elif isinstance(new, Directory) and self.__parent is None:
            if not new.get_file(self.name):
                logger.critical("Failed to assign '%s' to parent '%s'.", self.name, new)
                raise Exception(f"Failed to assign '{self.name}' to parent '{new}'.")

            self.__parent = new
    @property
    def size_bytes(self)->int|Literal[-1]:
        return self.__path.stat().st_size if self.exists else -1
    @property
    def size_kbytes(self)->int|Literal[-1]:
        return round(self.size_bytes /1024, 3) if self.exists else -1
    @property
    def size_mbytes(self)->int|Literal[-1]:
        return round(self.size_bytes /(1024 * 1024), 3) if self.exists else -1
    @property
    def mode(self)->Optional[int]:
        return self.__path.stat().st_mode if self.exists else None
    @property
    def inode(self)->Optional[int]:
        return self.__path.stat().st_ino if self.exists else None
    @property
    def dev_id(self)->Optional[int]:
        return self.__path.stat().st_dev if self.exists else None
    @property
    def nlinks(self)->Optional[int]:
        return self.__path.stat().st_nlink
    @property
    def creation_time(self)->Optional[int]:
        return self.__path.stat().st_ctime if self.exists else None
    @property
    def creation_stime(self)->Optional[str]:
        return time.ctime(self.creation_time) if self.exists else None
    @property
    def accessed_time(self)->Optional[int]:
        return self.__path.stat().st_atime if self.exists else None
    @property
    def accessed_stime(self)->Optional[str]:
        return time.ctime(self.accessed_time) if self.exists else None
    @property
    def modified_time(self)->Optional[int]:
        return self.__path.stat().st_mtime if self.exists else None
    @property
    def modified_stime(self)->Optional[str]:
        return time.ctime(self.modified_time) if self.exists else None

class JsonParser(FileParser):
    _default_color = Fore.YELLOW
    def __init__(self, path : PathLike, **kw):
        """
        Prototype content parser for JSON files, abstracted from FileParser and contains methods for json data parsing and manipulating.

        **Arguments**
            -path: Path or string object absolute/relative path or name pointing to a file (file does not have to exist)
            -create: creates file object initially if it does not exist.
            -color: color of the name used in visualization methods in Directory and PathManager class, uses colorama's package colors by default.
            -hide: instance becomes invisible to references map and can be duplicated, by default, only one object can point to a certain file.

        **Example Usage**

        >>> file : JsonParser = JsonParser("data.json") #or FileParser("data.json")
        >>> data = file.read(overwrite=True, ...) #reads the file's content
        >>> file["Key"] = "hello" #modifying data
        >>> key_value = file["Key"] #getting data
        >>> file.write(...) #creating and rewriting file's data 
        """
        if hasattr(self, "_init"):
            return
        self._init : bool = True
        self.__path : Path = Path(path)
        self.__data : Dict[str, Any] = dict()
        self.__parent : Optional[Directory] = None
        self.__color : str = kw.get("color", JsonParser._default_color)

        if self.__path.suffix.lower() not in JSON_SUFFIX:
            logger.critical("Expected %s files, for %s file use %s instead.", JSON_SUFFIX, self.__path.suffix, FileParser._correct_parser(self.__path.suffix))
            raise Exception(f"Expected {JSON_SUFFIX} files, for {self.__path.suffix} file use {FileParser._correct_parser(self.__path.suffix)} instead.")
        if self.exists and self.__path.is_dir():
            logger.critical("Path must point to a file, not a directory.")
            raise Exception("Path must point to a file, not a directory.")
        
        self.__hidden : bool = kw.get("hide", False)
        if kw.get("write", False):
            self.__path.open("a").close()


    def read(self, overwrite : bool=True, *, encoding:Optional[str]=None, errors:Optional[str]=None, newline:Optional[str]=None, cls=None, object_hook=None, parse_float=None,
        parse_int=None, parse_constant=None, object_pairs_hook=None, **kw)->Dict[str, Any]:
        if self.exists:  
            if not self.__data or (self.__data and overwrite):
                with self.__path.open(encoding=encoding, errors=errors, newline=newline) as file:
                    try:
                        self.__data = json.load(file, cls=cls, object_hook=object_hook, parse_float=parse_float, parse_int=parse_int, parse_constant=parse_constant, object_pairs_hook=object_pairs_hook, **kw)
                    except json.decoder.JSONDecodeError:
                        logger.error("failed to read json file '%s'.", self)
                        return None
        return self.__data

    def write(self, *, only_create:bool=False, mode:str='w', encoding:Optional[str]=None, buffering:int=-1, errors:Optional[str]=None, newline:Optional[str]=None, skipkeys=False, 
              ensure_ascii=True, check_circular=True, allow_nan=True, cls=None, indent=None, separators=None, default=None, sort_keys=False, **kw)->None:
        if not self:
            logger.error("Missing path to json file '%s'.", self)
            return None

        if only_create:
            self.__path.open('a').close()
            return None

        with self.__path.open(mode=mode, buffering=buffering, encoding=encoding, errors=errors, newline=newline) as file:
            json.dump(self.__data, file, skipkeys=skipkeys, ensure_ascii=ensure_ascii, check_circular=check_circular,
                      allow_nan=allow_nan, cls=cls, indent=indent, separators=separators,
                      default=default, sort_keys=sort_keys, **kw)
        return
    
    def rename(self, new_name : Optional[str])->None:
        if not self:
            logger.error("Cannot modify the file '%s'", self)
            return None

        #string handling case       
        if isinstance(new_name, str): 
            if PathManager.os() == OperatingSystem.WINDOWS and PathManager.contains_illegal_chars_win(new_name):
                logger.error("name %s contains illegal characters.", new_name)
                return None            
            elif PathManager.os() in (OperatingSystem.LINUX, OperatingSystem.MACOS) and PathManager.contains_illegal_chars_psx(new_name):
                logger.error("name %s contains illegal characters.", new_name)
                return None            

            if len(Path(new_name).parts) != 1:
                logger.error("name must not include other directory parts.")

            #if file has parent
            if self.__parent and Directory._blocks_duplicates():
                new_name = self.__parent._numeric_name(new_name+self.suffix, True) #get correct name

            #no parent, add suffix to it
            else:
                new_name = new_name + self.suffix

        #requesting a duplicate rename
        elif new_name is None:
            if self.__parent is None:
                return
            new_name = self.__parent._numeric_name(self.name, True) #get correct name

        else: #new name is neither str or None
            logger.error("Expected a string or None type, got %s instead", type(new_name).__name__)
            return None

        #name did not change
        if new_name == self.name:
            logger.info("New name %s matches the old name, skipping the rest of the process", new_name)
            return None

        #new path conclusion
        path, old_name = PathManager.split(self.__path)
        new_path = Path(path) / new_name if path else new_name
        #check if new path is already occupied
        if FileParser.is_locked() and not self.__hidden:
            #check if new path already exists
            if FileParser._contains(new_path.absolute()):
                logger.error("File with the name '%s' already exists in this directory.", new_name)
                return None

        #updating name
            FileParser._FileParser__PERMISSION = True
            FileParser._update(self.absolute, new_path.absolute())
            if self.exists:
                self.__path.rename(new_path)

        self.__path = new_path
        logger.info("File '%s' updated to '%s'", old_name, new_name)
        return

    def encrypt_data(self, func : Callable, *args, **kw)->None:
        """
        Takes an encryption function with a return type of dict, can also be used to decrypt using this method.

        Function must take a dict variable as its first argument.
        """
        data = func(self.__data, *args, **kw)
        if not isinstance(data, dict):
            logger.error("Failed to encrypt data for file '%s'", self)
            return None
        self.__data = data
        return

    def pop_key(self, key : str)->Optional[Any]:
        """
        Removes specified key from json data.
        """
        return self.__data.pop(key, None)

    def clear(self)->None:
        """
        Clears object data, actual file is unchanged until write method is called.
        """
        self.__data.clear()

    def _redirect(self, new_master : Directory)->None:
        if not isinstance(new_master, Directory):
            logger.error("expected Directory object, got %s instead.", type(new_master).__name__)
            return None
        if not new_master:
            logger.error("Cannot modify the directory: %s.", new_master)
            return None
        if not new_master.permitted:
            logger.error("Missing permission to redirect '%s'", self)
            return None

        old_path = self.absolute
        if new_master is not self.__parent:
            FileParser._FileParser__PERMISSION = True
            if self.__parent is not None:
                self.__parent._remove_file(self)
            self.__path = new_master.path / self.name
            new_master._add_file(self)
            FileParser._FileParser__PERMISSION = False
        else:
            self.__path = new_master.path / self.name
            if self.__path and old_path == self.__path:
                #case when dir already exists in parent + path does not change
                logger.error("File with the name '%s' already exists in directory '%s'.", self.name, new_master)
                return None

        if FileParser.is_locked() and not self.__hidden:
            FileParser._FileParser__PERMISSION = True
            FileParser._update(old_path, self.__path.absolute())
        
        logger.info("Added file '%s' to directory '%s'", self.name, self.__parent)
        return

    def _extract_info(self, with_color : bool = True, *tags : KEYS)->List[Any]:
        """
        Returns an iterable of properties based on input tags, possible tags:

        **name, path, type, exists, hidden, parent, root, absolute, depth, size_b, size_kb, size_mb, ctime, lmtime, latime, mode, inode, dev_id, nlinks**
        """
        output = []
        if not self:
            return output

        for field in tags:
            match field.lower():
                case "name":
                    output.append(f"{self.color}{self.name}{RESET_COLOR}" if with_color else f"{self.name}")
                case "path":
                    output.append(f"{self.color}{self.path}{RESET_COLOR}" if with_color else f"{self.path}")
                case "type":
                    output.append(f"{self.color}{self.type}{RESET_COLOR}" if with_color else f"{self.type}")
                case "exists":
                    output.append(f"{self.color}{self.exists}{RESET_COLOR}" if with_color else f"{self.exists}")
                case "hidden":
                    output.append(f"{self.color}{self.hidden}{RESET_COLOR}" if with_color else f"{self.hidden}")
                case "root":
                    output.append(f"{self.color}{self.root}{RESET_COLOR}" if with_color else f"{self.root}")
                case "parent":
                    output.append(f"{self.color}{self.parent}{RESET_COLOR}" if with_color else f"{self.parent}")
                case "absolute":
                    output.append(f"{self.color}{self.absolute}{RESET_COLOR}" if with_color else f"{self.absolute}")
                case "depth":
                    output.append(f"{self.color}{self.depth}{RESET_COLOR}" if with_color else self.depth)
                case "size_b":
                    output.append(f"{self.color}{self.size_bytes} B{RESET_COLOR}" if with_color else self.size_bytes)
                case "size_kb":
                    output.append(f"{self.color}{self.size_kbytes} KB{RESET_COLOR}" if with_color else self.size_kbytes)
                case "size_mb":
                    output.append(f"{self.color}{self.size_mbytes} MB{RESET_COLOR}" if with_color else self.size_mbytes)
                case "ctime":
                    output.append(f"{self.color}{self.creation_stime}{RESET_COLOR}" if with_color else f"{self.creation_stime}")
                case "lmtime":
                    output.append(f"{self.color}{self.modified_stime}{RESET_COLOR}" if with_color else f"{self.modified_stime}")
                case "latime":
                    output.append(f"{self.color}{self.accessed_stime}{RESET_COLOR}" if with_color else f"{self.accessed_stime}")
                case "mode":
                    output.append(f"{self.color}{self.mode}{RESET_COLOR}" if with_color else self.mode)
                case "inode":
                    output.append(f"{self.color}{self.inode}{RESET_COLOR}" if with_color else self.inode)
                case "dev_id":
                    output.append(f"{self.color}{self.dev_id}{RESET_COLOR}" if with_color else self.dev_id)
                case "nlinks":
                    output.append(f"{self.color}{self.nlinks}{RESET_COLOR}" if with_color else self.nlinks)
                case _:
                    output.append(f"{self.color}{NOT_FOUND}{RESET_COLOR}" if with_color else f"{NOT_FOUND}")

        return output

    def __getitem__(self, key : str)->Optional[Any]:
        return self.__data.get(key, None)

    def __setitem__(self, key : str, value : Any)->None:
        self.__data[key] = value

    def __repr__(self)->str:
        return self.__path.name

    def __str__(self)->str:
        return str(self.__path)
    
    def __len__(self)->int:
        return len(self.__data)
    
    def __bool__(self)->bool:
        return bool(self.__path)

    def __del__(self)->None:
        try:
            if sys.meta_path is None:
                return
        except ImportError:
            return

        FileParser._FileParser__PERMISSION = True
        FileParser._del(self.absolute, self)

    def __eq__(self, other : Union[FileParser, PathLike])->bool:
        if isinstance(other, (FileParser, PathLike)):
            return PathManager.relative(self, other)
        elif other is None:
            return bool(self) == False
        logger.critical("Cannot compare %s type with %s type.", FileParser.__name__, type(other).__name__)
        raise Exception(f"Cannot compare {FileParser.__name__} type with {type(other).__name__} type.")

    @property
    def keys(self)->List[str]:
        """
        List of existing keys in data object.
        """
        return list(self.__data.keys())

    @property
    def path(self)->Path:
        return self.__path
    @property
    def name(self)->str:
        return self.__path.name
    @property
    def type(self)->str:
        return f"{self.suffix} File"
    @property
    def root(self)->Optional[Directory]:
        return self.__parent.root if self.__parent else None
    @property
    def suffix(self)->str:
        return self.__path.suffix
    @property
    def absolute(self)->Path:
        return self.__path.absolute()
    @property
    def exists(self)->bool:
        return self.__path.exists() if self else False
    @property
    def hidden(self)->bool:
        return self.__hidden
    @property
    def color(self)->str:
        return self.__color if self else FileParser._PATHLESS_COLOR
    @property
    def parent(self)->Optional[Directory]:
        return self.__parent
    @parent.setter
    def parent(self, new : Optional[Directory])->None:
        """
        Internally managed.
        """
        if not isinstance(new, Directory) and new is not None:
            logger.critical("Expected Directory object, got %s instead.", type(new).__name__)
            raise Exception(f"Expected Directory object, got {type(new).__name__} instead.")

        if new is not None and self.__parent is not None:
            logger.critical("Cannot assign parent to '%s' as it already has a parent '%s'.", self.name, self.__parent)
            raise Exception(f"Cannot assign parent to '{self.name}' as it already has a parent '{self.__parent}'.")

        if new is None and self.__parent is not None:
            if self.__parent.get_file(self.name):
                logger.critical("Failed to remove file '%s' from parent '%s'.", self.name, self.__parent)
                raise Exception(f"Failed to remove file '{self.name}' from parent '{self.__parent}'.")
            self.__parent = None
            self.__path = NullPath(self.absolute)

        elif isinstance(new, Directory) and self.__parent is None:
            if not new.get_file(self.name):
                logger.critical("Failed to assign '%s' to parent '%s'.", self.name, new)
                raise Exception(f"Failed to assign '{self.name}' to parent '{new}'.")

            self.__parent = new

    @property
    def permitted(self)->bool:
        return FileParser._FileParser__PERMISSION
    @property
    def depth(self)->int:
        return 0 if not self else len(PathManager.set(self.parent.root, self.parent, False).parts)
    @property
    def size_bytes(self)->int|Literal[-1]:
        return self.__path.stat().st_size if self.exists else -1
    @property
    def size_kbytes(self)->int|Literal[-1]:
        return round(self.size_bytes /1024, 3) if self.exists else -1
    @property
    def size_mbytes(self)->int|Literal[-1]:
        return round(self.size_bytes /(1024 * 1024), 3) if self.exists else -1
    @property
    def mode(self)->Optional[int]:
        return self.__path.stat().st_mode if self.exists else None
    @property
    def inode(self)->Optional[int]:
        return self.__path.stat().st_ino if self.exists else None
    @property
    def dev_id(self)->Optional[int]:
        return self.__path.stat().st_dev if self.exists else None
    @property
    def nlinks(self)->Optional[int]:
        return self.__path.stat().st_nlink
    @property
    def creation_time(self)->Optional[int]:
        return self.__path.stat().st_ctime if self.exists else None
    @property
    def creation_stime(self)->Optional[str]:
        return time.ctime(self.creation_time) if self.exists else None
    @property
    def accessed_time(self)->Optional[int]:
        return self.__path.stat().st_atime if self.exists else None
    @property
    def accessed_stime(self)->Optional[str]:
        return time.ctime(self.accessed_time) if self.exists else None
    @property
    def modified_time(self)->Optional[int]:
        return self.__path.stat().st_mtime if self.exists else None
    @property
    def modified_stime(self)->Optional[str]:
        return time.ctime(self.modified_time) if self.exists else None

class CsvParser(FileParser):
    _default_color = Fore.LIGHTGREEN_EX
    def __init__(self, path : PathLike, **kw):
        """
        Parser object for .csv files.

        **Arguments**
            -path: Path or string object absolute/relative path or name pointing to a file (file does not have to exist)
            -create: creates file object initially if it does not exist.
            -color: color of the name used in visualization methods in Directory and PathManager class, uses colorama's package colors by default.
            -hide: instance becomes invisible to references map and can be duplicated, by default, only one object can point to a certain file.
        
        **Example**
        >>> file = CsvParser("languages.csv") #or FileParser("languages.csv")
        >>> table = file.read(overwrite=True)
        >>> table
        [["Language", "publish year", "Popularity%"],
         ["Python", 1991, 23.37],
         ["c++", 1979, 8.95],
         ["java", 1995, 8.54],
         ["lua", 1993, 0.5]]
         >>> file.add_column("use case", ["data science", "high-performance computing", "enterprise applications", "video game dev"])
         >>> file.set_index("Language")
         >>> java_row = file.get_row("java")
         >>> python_row = file["Python"]
         >>> cpp_publish_year = file["c++", "publish year"] #or file[2, 1]
         >>> file["lua", 3] = "world of warcraft developement"
         >>> file.write()
        """
        if hasattr(self, "_init"):
            return
        self._init : bool = True
        self.__path : Path = Path(path)
        self.__data : List[List[Any]] = list()
        self.__parent : Optional[Directory] = None
        self.__color : str = kw.get("color", CsvParser._default_color)
        self.__index : Optional[int] = None
        if self.__path.suffix.lower() not in CSV_SUFFIX:
            logger.critical("Expected %s files, for %s file use %s instead.", CSV_SUFFIX, self.__path.suffix, FileParser._correct_parser(self.__path.suffix))
            raise Exception(f"Expected {CSV_SUFFIX} files, for {self.__path.suffix} file use {FileParser._correct_parser(self.__path.suffix)} instead.")
        if self.exists and self.__path.is_dir():
            logger.critical("Path must point to a file, not a directory.")
            raise Exception("Path must point to a file, not a directory.")
        
        self.__hidden : bool = kw.get("hide", False)
        if kw.get("create", False):
            self.__path.open("a").close()

    def read(self, overwrite : bool=True, *, encoding : Optional[str] = None, errors : Optional[str] = None, newline : Optional[str] = None, delimiter: str = ",", quotechar: Optional[str] = '"', 
             escapechar: Optional[str] = None, doublequote: bool = True, skipinitialspace: bool = False, lineterminator: str = "\r\n", quoting: csv._QuotingType = 0, strict: bool = False,)->List[Any]:
        if self.exists:
            if not self.__data or overwrite:
                with self.__path.open(encoding=encoding, errors=errors, newline=newline) as file:
                    self.__data = list(csv.reader(file, delimiter=delimiter, quotechar=quotechar, escapechar=escapechar, doublequote=doublequote, skipinitialspace=skipinitialspace, 
                                                  lineterminator=lineterminator, quoting=quoting, strict=strict))
                    if self.__data:
                        if len(self.columns) != len(set(self.columns)):
                            logger.error("Failed to load data: found duplicated columns.")
                            self.__data = list()
        return self.__data
    
    def write(self, *, only_create:bool=False, mode:str='w', buffering:int=-1, encoding : Optional[str] = None, errors : Optional[str] = None, newline : Optional[str] = None, delimiter: str = ",", 
              quotechar: Optional[str] = '"', escapechar: Optional[str] = None, doublequote: bool = True, skipinitialspace: bool = False, lineterminator: str = "\r\n", quoting: csv._QuotingType = 0, 
              strict: bool = False):
        if not self:
            logger.error("Cannot write into path '%s'.", self)
            return None
        
        if only_create:
            self.__path.open('a').close()
            return None

        with self.__path.open(mode=mode, buffering=buffering, encoding=encoding, errors=errors, newline=newline) as file:
            if self.__data:
                writer = csv.writer(file, delimiter=delimiter, quotechar=quotechar, escapechar=escapechar, doublequote=doublequote, skipinitialspace=skipinitialspace, 
                                                    lineterminator=lineterminator, quoting=quoting, strict=strict)
                writer.writerows(self.__data)
        
        return

    def rename(self, new_name : Optional[str])->None:
        if not self:
            logger.error("Cannot modify the file '%s'", self)
            return None

        #string handling case       
        if isinstance(new_name, str): 
            if PathManager.os() == OperatingSystem.WINDOWS and PathManager.contains_illegal_chars_win(new_name):
                logger.error("name %s contains illegal characters.", new_name)
                return None            
            elif PathManager.os() in (OperatingSystem.LINUX, OperatingSystem.MACOS) and PathManager.contains_illegal_chars_psx(new_name):
                logger.error("name %s contains illegal characters.", new_name)
                return None            

            if len(Path(new_name).parts) != 1:
                logger.error("name must not include other directory parts.")

            #if file has parent
            if self.__parent and Directory._blocks_duplicates():
                new_name = self.__parent._numeric_name(new_name+self.suffix, True) #get correct name

            #no parent, add suffix to it
            else:
                new_name = new_name + self.suffix

        #requesting a duplicate rename
        elif new_name is None:
            if self.__parent is None:
                return
            new_name = self.__parent._numeric_name(self.name, True) #get correct name

        else: #new name is neither str or None
            logger.error("Expected a string or None type, got %s instead", type(new_name).__name__)
            return None

        #name did not change
        if new_name == self.name:
            logger.info("New name %s matches the old name, skipping the rest of the process", new_name)
            return None

        #new path conclusion
        path, old_name = PathManager.split(self.__path)
        new_path = Path(path) / new_name if path else new_name
        #check if new path is already occupied
        if FileParser.is_locked() and not self.__hidden:
            #check if new path already exists
            if FileParser._contains(new_path.absolute()):
                logger.error("File with the name '%s' already exists in this directory.", new_name)
                return None

        #updating name
            FileParser._FileParser__PERMISSION = True
            FileParser._update(self.absolute, new_path.absolute())
            if self.exists:
                self.__path.rename(new_path)

        self.__path = new_path
        logger.info("File '%s' updated to '%s'", old_name, new_name)
        return

    def _redirect(self, new_master : Directory)->None:
        if not isinstance(new_master, Directory):
            logger.error("expected Directory object, got %s instead.", type(new_master).__name__)
            return None
        if not new_master:
            logger.error("Cannot modify the directory: %s.", new_master)
            return None
        if not new_master.permitted:
            logger.error("Missing permission to redirect '%s'", self)
            return None

        old_path = self.absolute
        if new_master is not self.__parent:
            FileParser._FileParser__PERMISSION = True
            if self.__parent is not None:
                self.__parent._remove_file(self)
            self.__path = new_master.path / self.name
            new_master._add_file(self)
            FileParser._FileParser__PERMISSION = False
        else:
            self.__path = new_master.path / self.name
            if self.__path and old_path == self.__path:
                #case when dir already exists in parent + path does not change
                logger.error("File with the name '%s' already exists in directory '%s'.", self.name, new_master)
                return None

        if FileParser.is_locked() and not self.__hidden:
            FileParser._FileParser__PERMISSION = True
            FileParser._update(old_path, self.__path.absolute())
        
        logger.info("Added file '%s' to directory '%s'", self.name, self.__parent)
        return

    def _extract_info(self, with_color : bool = True, *tags : KEYS)->List[Any]:
        """
        Returns an iterable of properties based on input tags, possible tags:

        **name, path, type, exists, hidden, parent, root, absolute, depth, size_b, size_kb, size_mb, ctime, lmtime, latime, mode, inode, dev_id, nlinks**
        """
        output = []
        if not self:
            return output

        for field in tags:
            match field.lower():
                case "name":
                    output.append(f"{self.color}{self.name}{RESET_COLOR}" if with_color else f"{self.name}")
                case "path":
                    output.append(f"{self.color}{self.path}{RESET_COLOR}" if with_color else f"{self.path}")
                case "type":
                    output.append(f"{self.color}{self.type}{RESET_COLOR}" if with_color else f"{self.type}")
                case "exists":
                    output.append(f"{self.color}{self.exists}{RESET_COLOR}" if with_color else f"{self.exists}")
                case "hidden":
                    output.append(f"{self.color}{self.hidden}{RESET_COLOR}" if with_color else f"{self.hidden}")
                case "root":
                    output.append(f"{self.color}{self.root}{RESET_COLOR}" if with_color else f"{self.root}")
                case "parent":
                    output.append(f"{self.color}{self.parent}{RESET_COLOR}" if with_color else f"{self.parent}")
                case "absolute":
                    output.append(f"{self.color}{self.absolute}{RESET_COLOR}" if with_color else f"{self.absolute}")
                case "depth":
                    output.append(f"{self.color}{self.depth}{RESET_COLOR}" if with_color else self.depth)
                case "size_b":
                    output.append(f"{self.color}{self.size_bytes} B{RESET_COLOR}" if with_color else self.size_bytes)
                case "size_kb":
                    output.append(f"{self.color}{self.size_kbytes} KB{RESET_COLOR}" if with_color else self.size_kbytes)
                case "size_mb":
                    output.append(f"{self.color}{self.size_mbytes} MB{RESET_COLOR}" if with_color else self.size_mbytes)
                case "ctime":
                    output.append(f"{self.color}{self.creation_stime}{RESET_COLOR}" if with_color else f"{self.creation_stime}")
                case "lmtime":
                    output.append(f"{self.color}{self.modified_stime}{RESET_COLOR}" if with_color else f"{self.modified_stime}")
                case "latime":
                    output.append(f"{self.color}{self.accessed_stime}{RESET_COLOR}" if with_color else f"{self.accessed_stime}")
                case "mode":
                    output.append(f"{self.color}{self.mode}{RESET_COLOR}" if with_color else self.mode)
                case "inode":
                    output.append(f"{self.color}{self.inode}{RESET_COLOR}" if with_color else self.inode)
                case "dev_id":
                    output.append(f"{self.color}{self.dev_id}{RESET_COLOR}" if with_color else self.dev_id)
                case "nlinks":
                    output.append(f"{self.color}{self.nlinks}{RESET_COLOR}" if with_color else self.nlinks)
                case _:
                    output.append(f"{self.color}{NOT_FOUND}{RESET_COLOR}" if with_color else f"{NOT_FOUND}")

        return output

    def set_index(self, key : str)->None:
        """
        Sets a column's values as indices for rows.
        """
        if key not in self.columns:
            logger.error("Failed to set index %s", key)
            return None
        self.__index = self.columns.index(key)
        logger.info("Index was set to %s", key)
        return

    def reset_index(self)->None:
        """
        Forgets indicing column.
        """
        self.__index = None

    def add_row(self, row_data : List[Any], index : Optional[int] = None)->None:
        """
        Adds a row to specified index, row must maintain the same size as the columns row.
        """
        if not isinstance(row_data, (list, tuple)):
            logger.error("Expected an iterable of type list or tuple, got %s.", type(row_data).__name__)
            return None
        if self.__data and len(row_data) != len(self.columns):
            logger.error("Row size does not match existing rows")
            return None
        
        if index is None:
            self.__data.append([data for data in row_data])
        elif index < -len(self):
            self.__data.insert(1, [data for data in row_data])

        return

    def get_row(self, index : Any)->List[Any]:
        """
        Returns rows based on index.
        """
        if len(self) == 0:
            return list()

        if isinstance(index, str):
            if self.__index is None:
                logger.error("No index column specified while receiving string key")
                return list()
            for row in self.__data:
                if row[self.__index] == index:
                    return [item for item in row]
        elif isinstance(index, int):
            if index == 0 or index < -len(self): #avoid accessing columns
                return [item for item in self.__data[1]]
            elif index >= len(self): #avoid index error
                return [item for item in self.__data[-1]]
            return [item for item in self.__data[index+1]]
        return list()

    def set_row(self, index : Union[int, Any], data_row : List[Any])->None:
        """
        Replaces existing row with data_row, data_row must respect the size of existing rows.
        """
        if len(self) == 0:
            return None
        if len(data_row) != len(self.columns):
            logger.error("Data count does not match the columns count %i != %i", len(data_row), len(self.columns))
            return None
        if isinstance(index, str):
            if self.__index is None:
                logger.error("No index column specified while receiving string key")
                return None
            for i, row in enumerate(self.__data):
                if row[self.__index] == index and i > 0:
                    break
            else:
                return None
            self.__data[i] = [data for data in data_row]

        elif isinstance(index, int):
            if index == 0 or index < -len(self):
                index = 1
            elif index > len(self):
                index = -1
            self.__data[index] = [data for data in data_row]
        return

    def remove_row(self, index : Any)->None:
        """
        Removes row from data based on index.
        """
        if not self.__data:
            return None

        row = self.get_row(index)
        if row and row != self.columns:
            self.__data.remove(row)
        return

    def add_column(self, key : str, data : List[Any])->None:
        """
        Adds a column at the far right of the matrix, key is the column name while data is an array that matches size the number of rows excluding the column row.
        """
        if key in self.columns:
            logger.error("Key %s already exists in columns.", key)
            return None

        data.insert(0, key) #adding column key to column
        if not self.__data:
            for row in data:
                self.__data.append([row])
        else:
            if len(data) != len(self) + 1:
                logger.error("Data size does not match data matrix' size %i != %i", len(data), len(self) + 1)
                return None

            for i, row in enumerate(self.__data):
                row.append(data[i])
        return

    def get_column(self, key : Union[int, str])->List[Any]:
        """
        Returns column values as an array based on key column name, or column index.
        """
        if not self.__data:
            return list()

        if isinstance(key, str):
            try:
                key = self.columns.index(key)
            except ValueError:
                logger.error("Could not find the key %s", key)
                return None

        if key >= len(self.columns):
            key = len(self.columns) - 1

        return [row[key] for row in self.__data]

    def set_column(self, key : Union[str, int], data_column : List[Any])->None:
        """
        Replaces the selected column values with data_column, data_column must match the size of the number of rows excluding the column row.
        """
        if len(data_column) != len(self):
            logger.error("Data size does not match matrix size %i != %i", len(data_column), len(self))
            return None

        if isinstance(key, str):
            try:
                key = self.columns.index(key)
            except ValueError:
                logger.error("Could not find the key %s", key)
                return None
        if key >= len(self.columns):
            logger.error("Index out of range %d >= %i", key, len(self.columns))
            return None

        for i, row in enumerate(self.__data):
            row[key] = data_column[i]
        return
    
    def remove_column(self, key : Union[str, int])->None:
        """
        Removes column key and its descending values, if column was set as the index column then resets the object's index as well.
        """
        if not self.__data:
            return None
        
        if isinstance(key, str):
            try:
                key = self.columns.index(key)
            except IndexError:
                logger.error("Could not find column %s.", key)
                return None

        if not isinstance(key, int):
            logger.error("Expected int type index, got %s instead.", type(key).__name__)
            return None

        for row in self.__data:
            row.pop(key)

        if self.__index:
            if key == self.__index:
                self.__index = None
            elif key < self.__index:
                self.__index -= 1
        return
    
    def clear(self)->None:
        self.__data.clear()
        self.__index = None

#MAGIC
    def __getitem__(self, key : int|str|Tuple[Union[int, str, None], Union[int, str, None]])->Optional[Any]:
        """
        Direct way to return either a row, a column, or one value

        Examples

        >>> row_1 = file[2] 
        """
        if isinstance(key, tuple):
            if len(key) < 2:
                logger.error("Only 2-dimensional indexing is supported.")
                return None
            
            row, column, *_ = key

            if row is None and column is not None:
                return self.get_column(column)
            
            elif row is not None and column is None:
                return self.get_row(row)
            
            elif row is not None and column is not None:
                data = self.get_row(row)
                if not data:
                    return None
                if isinstance(column, str):
                    try:
                        column = self.columns.index(column)
                    except ValueError:
                        return None

                if column >= len(data):
                    column = len(data) - 1
                return data[column]
                
            else:
                logger.error("Invalid index types: row=%s, column=%s", type(row).__name__, type(column).__name__)

        elif isinstance(key, int):
            if key >= len(self.__data):
                logger.error("Index out of range %i >= %i", key, len(self.__data))
                return None
            return self.__data[key]

        else:
            logger.error("Invalid index type. Must be an int or a tuple of two indices.")
        return

    def __setitem__(self, key : int|str|Tuple[Union[int, str, None], Union[int, str, None]], data : Union[Any, List[Any]])->None:
        if isinstance(key, tuple):
            if len(key) < 2:
                logger.error("Only 2-dimensional indexing is supported.")
                return None

            row, column, *_ = key
            if row is None and column is not None:
                self.set_column(column, data)
            
            elif row is not None and column is None:
                self.set_row(row, data)
            
            elif row is not None and column is not None:
                if self.__index is not None:
                    for i, r in enumerate(self.__data):
                        if r[self.__index] == row:
                            break
                    else: #did not find row
                        logger.error("Could not find row %s", row)
                        return None
                else: #row should be int
                    if not isinstance(row, int):
                        logger.error("Expected int type, got %s instead.", type(row).__name__)
                        return None
                    i = row
                if isinstance(column, str):
                    try:
                        j = self.columns.index(column)
                    except ValueError:
                        logger.error("Could not find the key %s", column)
                        return None
                else:
                    if not isinstance(column, int):
                        logger.error("Expected int type, got %s instead.", type(column).__name__)
                        return None
                    j = column
                if i >= len(self.__data) or j >= len(self.columns):
                    logger.error("One or more indices out of range (row=%i>=%i, column=%i>=%i)", i, len(self.__data), j, len(self.columns))
                    return None
                self.__data[i][j] = data

            else:
                logger.error("Invalid index types: row=%s, column=%s", type(row).__name__, type(column).__name__)

        elif isinstance(key, int):
            self.set_row(row, data)
        else:
            logger.error("Invalid index type. Must be an int or a tuple of two indices.")
        return
    def __repr__(self)->str:
        return self.__path.name

    def __str__(self)->str:
        return str(self.__path)
    
    def __len__(self)->int:
        """
        Number of rows excluding column row.
        """
        return max(len(self.__data) - 1, 0)
    
    def __bool__(self)->bool:
        return bool(self.__path)

    def __del__(self)->None:
        try:
            if sys.meta_path is None:
                return
        except ImportError:
            return

        FileParser._FileParser__PERMISSION = True
        FileParser._del(self.absolute, self)

    def __eq__(self, other : Union[FileParser, PathLike])->bool:
        if isinstance(other, (FileParser, PathLike)):
            return PathManager.relative(self, other)
        elif other is None:
            return bool(self) == False
        logger.critical("Cannot compare %s type with %s type.", FileParser.__name__, type(other).__name__)
        raise Exception(f"Cannot compare {FileParser.__name__} type with {type(other).__name__} type.")

#SPECIAL
    @property
    def columns(self)->List[str]:
        return self.__data[0] if self.__data else list()
    
    @property
    def index(self)->Optional[str]:
        return self.columns[self.__index] if self.__data or self.__index is not None else None

#PROPERTIES
    @property
    def path(self)->Path:
        return self.__path
    @property
    def name(self)->str:
        return self.__path.name
    @property
    def type(self)->str:
        return f"{self.suffix} File"
    @property
    def root(self)->Optional[Directory]:
        return self.__parent.root if self.__parent else None
    @property
    def suffix(self)->str:
        return self.__path.suffix
    @property
    def absolute(self)->Path:
        return self.__path.absolute()
    @property
    def exists(self)->bool:
        return self.__path.exists() if self else False
    @property
    def hidden(self)->bool:
        return self.__hidden
    @property
    def color(self)->str:
        return self.__color if self else FileParser._PATHLESS_COLOR
    @property
    def parent(self)->Optional[Directory]:
        return self.__parent
    @parent.setter
    def parent(self, new : Optional[Directory])->None:
        """
        Internally managed.
        """
        if not isinstance(new, Directory) and new is not None:
            logger.critical("Expected Directory object, got %s instead.", type(new).__name__)
            raise Exception(f"Expected Directory object, got {type(new).__name__} instead.")

        if new is not None and self.__parent is not None:
            logger.critical("Cannot assign parent to '%s' as it already has a parent '%s'.", self.name, self.__parent)
            raise Exception(f"Cannot assign parent to '{self.name}' as it already has a parent '{self.__parent}'.")

        if new is None and self.__parent is not None:
            if self.__parent.get_file(self.name):
                logger.critical("Failed to remove file '%s' from parent '%s'.", self.name, self.__parent)
                raise Exception(f"Failed to remove file '{self.name}' from parent '{self.__parent}'.")
            self.__parent = None
            self.__path = NullPath(self.absolute)

        elif isinstance(new, Directory) and self.__parent is None:
            if not new.get_file(self.name):
                logger.critical("Failed to assign '%s' to parent '%s'.", self.name, new)
                raise Exception(f"Failed to assign '{self.name}' to parent '{new}'.")

            self.__parent = new

    @property
    def permitted(self)->bool:
        return FileParser._FileParser__PERMISSION
    @property
    def depth(self)->int:
        return 0 if not self else len(PathManager.set(self.parent.root, self.parent, False).parts)
    @property
    def size_bytes(self)->int|Literal[-1]:
        return self.__path.stat().st_size if self.exists else -1
    @property
    def size_kbytes(self)->int|Literal[-1]:
        return round(self.size_bytes /1024, 3) if self.exists else -1
    @property
    def size_mbytes(self)->int|Literal[-1]:
        return round(self.size_bytes /(1024 * 1024), 3) if self.exists else -1
    @property
    def mode(self)->Optional[int]:
        return self.__path.stat().st_mode if self.exists else None
    @property
    def inode(self)->Optional[int]:
        return self.__path.stat().st_ino if self.exists else None
    @property
    def dev_id(self)->Optional[int]:
        return self.__path.stat().st_dev if self.exists else None
    @property
    def nlinks(self)->Optional[int]:
        return self.__path.stat().st_nlink
    @property
    def creation_time(self)->Optional[int]:
        return self.__path.stat().st_ctime if self.exists else None
    @property
    def creation_stime(self)->Optional[str]:
        return time.ctime(self.creation_time) if self.exists else None
    @property
    def accessed_time(self)->Optional[int]:
        return self.__path.stat().st_atime if self.exists else None
    @property
    def accessed_stime(self)->Optional[str]:
        return time.ctime(self.accessed_time) if self.exists else None
    @property
    def modified_time(self)->Optional[int]:
        return self.__path.stat().st_mtime if self.exists else None
    @property
    def modified_stime(self)->Optional[str]:
        return time.ctime(self.modified_time) if self.exists else None

class ImageParser(FileParser):
    _default_color = Fore.LIGHTMAGENTA_EX
    def __init__(self, path : PathLike, **kw):
        """
        Parser object for images and gifs, used for parsing images from urls and saving them to dedicated path.

        **Arguments**
            -path: Path or string object absolute/relative path or name pointing to a file (file does not have to exist)
            -create: creates file object initially if it does not exist.
            -color: color of the name used in visualization methods in Directory and PathManager class, uses colorama's package colors by default.
            -hide: instance becomes invisible to references map and can be duplicated, by default, only one object can point to a certain file.
        
        **Example**
            >>> img = ImageParser("image.png")
            >>> img.parse_url(url) #takes image from the web and saves it onto the object
            >>> img.write() #saves the image in image.png
        """
        if hasattr(self, "_init"):
            return
        self._init : bool = True
        self.__path : Path = Path(path)
        self.__img : Image = None
        self.__parent : Optional[Directory] = None
        self.__color : str = kw.get("color", ImageParser._default_color)

        if self.__path.suffix.lower() not in IMAGE_SUFFIX:
            logger.critical("Expected %s files, for %s file use %s instead.", IMAGE_SUFFIX, self.__path.suffix, FileParser._correct_parser(self.__path.suffix))
            raise Exception(f"Expected {IMAGE_SUFFIX} files, for {self.__path.suffix} file use {FileParser._correct_parser(self.__path.suffix)} instead.")
        if self.exists and self.__path.is_dir():
            logger.critical("Path must point to a file, not a directory.")
            raise Exception("Path must point to a file, not a directory.")

        self.__hidden : bool = kw.get("hide", False)
        if kw.get("create", False):
            self.__path.open("ab").close()

    def read(self, overwrite : bool=True)->Image:
        if not self.exists:
            logger.error("Could not read from path '%s'", self)
            return None
        if not self.__img or overwrite:
            self.__img = Image.open(self.__path)

        return self.__img

    def write(self, only_create : bool = False, *a, **kw)->None:
        if not self:
            logger.error("Cannot write image to path '%s'.", self)
            return None
        
        if only_create:
            self.__path.open("ab").close()

        self.__img.save(self.__path)
        return
    
    def parse_url(self, url : str)->None:
        """
        Loads an image data from the web, requires external management for the url.
        """
        if not isinstance(url, str):
            logger.error("Expected string type, got %s instead.", type(url).__name__)
            return None
        if not is_valid_url(url):
            logger.error("url is invalid.")
            return None
        try:
            self.__img = Image.open(urlopen(url))
        except RequestException:
            logger.error("Encountered error while handling the request to url '%s'.", url)
        except UnidentifiedImageError:
            logger.error("Encountered error while trying to open image.")
        return

    def rename(self, new_name : Optional[str])->None:
        if not self:
            logger.error("Cannot modify the file '%s'", self)
            return None

        #string handling case       
        if isinstance(new_name, str): 
            if PathManager.os() == OperatingSystem.WINDOWS and PathManager.contains_illegal_chars_win(new_name):
                logger.error("name %s contains illegal characters.", new_name)
                return None            
            elif PathManager.os() in (OperatingSystem.LINUX, OperatingSystem.MACOS) and PathManager.contains_illegal_chars_psx(new_name):
                logger.error("name %s contains illegal characters.", new_name)
                return None            

            if len(Path(new_name).parts) != 1:
                logger.error("name must not include other directory parts.")

            #if file has parent
            if self.__parent and Directory._blocks_duplicates():
                new_name = self.__parent._numeric_name(new_name+self.suffix, True) #get correct name

            #no parent, add suffix to it
            else:
                new_name = new_name + self.suffix

        #requesting a duplicate rename
        elif new_name is None:
            if self.__parent is None:
                return
            new_name = self.__parent._numeric_name(self.name, True) #get correct name

        else: #new name is neither str or None
            logger.error("Expected a string or None type, got %s instead", type(new_name).__name__)
            return None

        #name did not change
        if new_name == self.name:
            logger.info("New name %s matches the old name, skipping the rest of the process", new_name)
            return None

        #new path conclusion
        path, old_name = PathManager.split(self.__path)
        new_path = Path(path) / new_name if path else new_name
        #check if new path is already occupied
        if FileParser.is_locked() and not self.__hidden:
            #check if new path already exists
            if FileParser._contains(new_path.absolute()):
                logger.error("File with the name '%s' already exists in this directory.", new_name)
                return None

        #updating name
            FileParser._FileParser__PERMISSION = True
            FileParser._update(self.absolute, new_path.absolute())
            if self.exists:
                self.__path.rename(new_path)

        self.__path = new_path
        logger.info("File '%s' updated to '%s'", old_name, new_name)
        return

    def _redirect(self, new_master : Directory)->None:
        if not isinstance(new_master, Directory):
            logger.error("expected Directory object, got %s instead.", type(new_master).__name__)
            return None
        if not new_master:
            logger.error("Cannot modify the directory: %s.", new_master)
            return None
        if not new_master.permitted:
            logger.error("Missing permission to redirect '%s'", self)
            return None

        old_path = self.absolute
        if new_master is not self.__parent:
            FileParser._FileParser__PERMISSION = True
            if self.__parent is not None:
                self.__parent._remove_file(self)
            self.__path = new_master.path / self.name
            new_master._add_file(self)
            FileParser._FileParser__PERMISSION = False
        else:
            self.__path = new_master.path / self.name
            if self.__path and old_path == self.__path:
                #case when dir already exists in parent + path does not change
                logger.error("File with the name '%s' already exists in directory '%s'.", self.name, new_master)
                return None

        if FileParser.is_locked() and not self.__hidden:
            FileParser._FileParser__PERMISSION = True
            FileParser._update(old_path, self.__path.absolute())
        
        logger.info("Added file '%s' to directory '%s'", self.name, self.__parent)
        return

    def _extract_info(self, with_color : bool = True, *tags : KEYS)->List[Any]:
        """
        Returns an iterable of properties based on input tags, possible tags:

        **name, path, type, exists, hidden, parent, root, absolute, depth, size_b, size_kb, size_mb, ctime, lmtime, latime, mode, inode, dev_id, nlinks**
        """
        output = []
        if not self:
            return output

        for field in tags:
            match field.lower():
                case "name":
                    output.append(f"{self.color}{self.name}{RESET_COLOR}" if with_color else f"{self.name}")
                case "path":
                    output.append(f"{self.color}{self.path}{RESET_COLOR}" if with_color else f"{self.path}")
                case "type":
                    output.append(f"{self.color}{self.type}{RESET_COLOR}" if with_color else f"{self.type}")
                case "exists":
                    output.append(f"{self.color}{self.exists}{RESET_COLOR}" if with_color else f"{self.exists}")
                case "hidden":
                    output.append(f"{self.color}{self.hidden}{RESET_COLOR}" if with_color else f"{self.hidden}")
                case "root":
                    output.append(f"{self.color}{self.root}{RESET_COLOR}" if with_color else f"{self.root}")
                case "parent":
                    output.append(f"{self.color}{self.parent}{RESET_COLOR}" if with_color else f"{self.parent}")
                case "absolute":
                    output.append(f"{self.color}{self.absolute}{RESET_COLOR}" if with_color else f"{self.absolute}")
                case "depth":
                    output.append(f"{self.color}{self.depth}{RESET_COLOR}" if with_color else self.depth)
                case "size_b":
                    output.append(f"{self.color}{self.size_bytes} B{RESET_COLOR}" if with_color else self.size_bytes)
                case "size_kb":
                    output.append(f"{self.color}{self.size_kbytes} KB{RESET_COLOR}" if with_color else self.size_kbytes)
                case "size_mb":
                    output.append(f"{self.color}{self.size_mbytes} MB{RESET_COLOR}" if with_color else self.size_mbytes)
                case "ctime":
                    output.append(f"{self.color}{self.creation_stime}{RESET_COLOR}" if with_color else f"{self.creation_stime}")
                case "lmtime":
                    output.append(f"{self.color}{self.modified_stime}{RESET_COLOR}" if with_color else f"{self.modified_stime}")
                case "latime":
                    output.append(f"{self.color}{self.accessed_stime}{RESET_COLOR}" if with_color else f"{self.accessed_stime}")
                case "mode":
                    output.append(f"{self.color}{self.mode}{RESET_COLOR}" if with_color else self.mode)
                case "inode":
                    output.append(f"{self.color}{self.inode}{RESET_COLOR}" if with_color else self.inode)
                case "dev_id":
                    output.append(f"{self.color}{self.dev_id}{RESET_COLOR}" if with_color else self.dev_id)
                case "nlinks":
                    output.append(f"{self.color}{self.nlinks}{RESET_COLOR}" if with_color else self.nlinks)
                case _:
                    output.append(f"{self.color}{NOT_FOUND}{RESET_COLOR}" if with_color else f"{NOT_FOUND}")

        return output

    def __repr__(self)->str:
        return self.__path.name

    def __str__(self)->str:
        return str(self.__path)
    
    def __bool__(self)->bool:
        return bool(self.__path)

    def __del__(self)->None:
        try:
            if sys.meta_path is None:
                return
        except ImportError:
            return

        FileParser._FileParser__PERMISSION = True
        FileParser._del(self.absolute, self)

    def __eq__(self, other : Union[FileParser, PathLike])->bool:
        if isinstance(other, (FileParser, PathLike)):
            return PathManager.relative(self, other)
        elif other is None:
            return bool(self) == False
        logger.critical("Cannot compare %s type with %s type.", FileParser.__name__, type(other).__name__)
        raise Exception(f"Cannot compare {FileParser.__name__} type with {type(other).__name__} type.")

    @property
    def image(self)->Image:
        return self.__img

    @property
    def path(self)->Path:
        return self.__path
    @property
    def name(self)->str:
        return self.__path.name
    @property
    def type(self)->str:
        return f"{self.suffix} File"
    @property
    def root(self)->Optional[Directory]:
        return self.__parent.root if self.__parent else None
    @property
    def suffix(self)->str:
        return self.__path.suffix
    @property
    def absolute(self)->Path:
        return self.__path.absolute()
    @property
    def exists(self)->bool:
        return self.__path.exists() if self else False
    @property
    def hidden(self)->bool:
        return self.__hidden
    @property
    def color(self)->str:
        return self.__color if self else FileParser._PATHLESS_COLOR
    @property
    def parent(self)->Optional[Directory]:
        return self.__parent
    @parent.setter
    def parent(self, new : Optional[Directory])->None:
        """
        Internally managed.
        """
        if not isinstance(new, Directory) and new is not None:
            logger.critical("Expected Directory object, got %s instead.", type(new).__name__)
            raise Exception(f"Expected Directory object, got {type(new).__name__} instead.")

        if new is not None and self.__parent is not None:
            logger.critical("Cannot assign parent to '%s' as it already has a parent '%s'.", self.name, self.__parent)
            raise Exception(f"Cannot assign parent to '{self.name}' as it already has a parent '{self.__parent}'.")

        if new is None and self.__parent is not None:
            if self.__parent.get_file(self.name):
                logger.critical("Failed to remove file '%s' from parent '%s'.", self.name, self.__parent)
                raise Exception(f"Failed to remove file '{self.name}' from parent '{self.__parent}'.")
            self.__parent = None
            self.__path = NullPath(self.absolute)

        elif isinstance(new, Directory) and self.__parent is None:
            if not new.get_file(self.name):
                logger.critical("Failed to assign '%s' to parent '%s'.", self.name, new)
                raise Exception(f"Failed to assign '{self.name}' to parent '{new}'.")

            self.__parent = new

    @property
    def permitted(self)->bool:
        return FileParser._FileParser__PERMISSION
    @property
    def depth(self)->int:
        return 0 if not self else len(PathManager.set(self.parent.root, self.parent, False).parts)
    @property
    def size_bytes(self)->int|Literal[-1]:
        return self.__path.stat().st_size if self.exists else -1
    @property
    def size_kbytes(self)->int|Literal[-1]:
        return round(self.size_bytes /1024, 3) if self.exists else -1
    @property
    def size_mbytes(self)->int|Literal[-1]:
        return round(self.size_bytes /(1024 * 1024), 3) if self.exists else -1
    @property
    def mode(self)->Optional[int]:
        return self.__path.stat().st_mode if self.exists else None
    @property
    def inode(self)->Optional[int]:
        return self.__path.stat().st_ino if self.exists else None
    @property
    def dev_id(self)->Optional[int]:
        return self.__path.stat().st_dev if self.exists else None
    @property
    def nlinks(self)->Optional[int]:
        return self.__path.stat().st_nlink
    @property
    def creation_time(self)->Optional[int]:
        return self.__path.stat().st_ctime if self.exists else None
    @property
    def creation_stime(self)->Optional[str]:
        return time.ctime(self.creation_time) if self.exists else None
    @property
    def accessed_time(self)->Optional[int]:
        return self.__path.stat().st_atime if self.exists else None
    @property
    def accessed_stime(self)->Optional[str]:
        return time.ctime(self.accessed_time) if self.exists else None
    @property
    def modified_time(self)->Optional[int]:
        return self.__path.stat().st_mtime if self.exists else None
    @property
    def modified_stime(self)->Optional[str]:
        return time.ctime(self.modified_time) if self.exists else None

class WorkbookParser(FileParser):
    _default_color = Fore.GREEN
    def __init__(self, path : PathLike, **kw):
        """
        **Arguments**
            -path: Path or string object absolute/relative path or name pointing to a file (file does not have to exist)
            -create: creates file object initially if it does not exist.
            -color: color of the name used in visualization methods in Directory and PathManager class, uses colorama's package colors by default.
            -hide: instance becomes invisible to references map and can be duplicated, by default, only one object can point to a certain file.
        """
        if hasattr(self, "_init"):
            return
        self._init : bool = True
        self.__path : Path = Path(path)
        self.__data : Workbook = Workbook()
        self.__parent : Optional[Directory] = None
        self.__color : str = kw.get("color", WorkbookParser._default_color)
        if self.__path.suffix.lower() not in EXCEL_SUFFIX:
            logger.critical("Expected %s files, for %s file use %s instead.", EXCEL_SUFFIX, self.__path.suffix, FileParser._correct_parser(self.__path.suffix))
            raise Exception(f"Expected {EXCEL_SUFFIX} files, for {self.__path.suffix} file use {FileParser._correct_parser(self.__path.suffix)} instead.")
        if self.exists and self.__path.is_dir():
            logger.critical("Path must point to a file, not a directory.")
            raise Exception("Path must point to a file, not a directory.")

        self.__hidden : bool = kw.get("hide", False)
        if kw.get("create", False):
            self.__path.open("a").close()


    def read(self, overwrite : bool=True, *, read_only : bool = False, keep_vba : bool = False, data_only : bool = False, keep_links : bool = True,
            rich_text : bool = False, **kw)->List[List[Any]]:
        if not self.exists:
            logger.error("Could not open path '%s'", self)
            return list()

        if not self.__data or overwrite:
            self.__data = load_workbook(str(self.__path), read_only=read_only, keep_vba=keep_vba, data_only=data_only, keep_links=keep_links, rich_text=rich_text)
        sheet = self.__data.active
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))
        return data

    def write(self, only_create : bool = False, *a, **kw)->None:
        if not self:
            logger.error("Could not write to path '%s'", self)
            return None
        
        if only_create:
            self.__path.open("a").close()
            return None

        self.__data.save(str(self.__path))
        return
    
    def rename(self, new_name : Optional[str])->None:
        if not self:
            logger.error("Cannot modify the file '%s'", self)
            return None

        #string handling case       
        if isinstance(new_name, str): 
            if PathManager.os() == OperatingSystem.WINDOWS and PathManager.contains_illegal_chars_win(new_name):
                logger.error("name %s contains illegal characters.", new_name)
                return None            
            elif PathManager.os() in (OperatingSystem.LINUX, OperatingSystem.MACOS) and PathManager.contains_illegal_chars_psx(new_name):
                logger.error("name %s contains illegal characters.", new_name)
                return None            

            if len(Path(new_name).parts) != 1:
                logger.error("name must not include other directory parts.")

            #if file has parent
            if self.__parent and Directory._blocks_duplicates():
                new_name = self.__parent._numeric_name(new_name+self.suffix, True) #get correct name

            #no parent, add suffix to it
            else:
                new_name = new_name + self.suffix

        #requesting a duplicate rename
        elif new_name is None:
            if self.__parent is None:
                return
            new_name = self.__parent._numeric_name(self.name, True) #get correct name

        else: #new name is neither str or None
            logger.error("Expected a string or None type, got %s instead", type(new_name).__name__)
            return None

        #name did not change
        if new_name == self.name:
            logger.info("New name %s matches the old name, skipping the rest of the process", new_name)
            return None

        #new path conclusion
        path, old_name = PathManager.split(self.__path)
        new_path = Path(path) / new_name if path else new_name
        #check if new path is already occupied
        if FileParser.is_locked() and not self.__hidden:
            #check if new path already exists
            if FileParser._contains(new_path.absolute()):
                logger.error("File with the name '%s' already exists in this directory.", new_name)
                return None

        #updating name
            FileParser._FileParser__PERMISSION = True
            FileParser._update(self.absolute, new_path.absolute())
            if self.exists:
                self.__path.rename(new_path)

        self.__path = new_path
        logger.info("File '%s' updated to '%s'", old_name, new_name)
        return

    def _redirect(self, new_master : Directory)->None:
        if not isinstance(new_master, Directory):
            logger.error("expected Directory object, got %s instead.", type(new_master).__name__)
            return None
        if not new_master:
            logger.error("Cannot modify the directory: %s.", new_master)
            return None
        if not new_master.permitted:
            logger.error("Missing permission to redirect '%s'", self)
            return None

        old_path = self.absolute
        if new_master is not self.__parent:
            FileParser._FileParser__PERMISSION = True
            if self.__parent is not None:
                self.__parent._remove_file(self)
            self.__path = new_master.path / self.name
            new_master._add_file(self)
            FileParser._FileParser__PERMISSION = False
        else:
            self.__path = new_master.path / self.name
            if self.__path and old_path == self.__path:
                #case when dir already exists in parent + path does not change
                logger.error("File with the name '%s' already exists in directory '%s'.", self.name, new_master)
                return None

        if FileParser.is_locked() and not self.__hidden:
            FileParser._FileParser__PERMISSION = True
            FileParser._update(old_path, self.__path.absolute())
        
        logger.info("Added file '%s' to directory '%s'", self.name, self.__parent)
        return


    def _extract_info(self, with_color : bool = True, *tags : KEYS)->List[Any]:
        """
        Returns an iterable of properties based on input tags, possible tags:

        **name, path, type, exists, hidden, parent, root, absolute, depth, size_b, size_kb, size_mb, ctime, lmtime, latime, mode, inode, dev_id, nlinks**
        """
        output = []
        if not self:
            return output

        for field in tags:
            match field.lower():
                case "name":
                    output.append(f"{self.color}{self.name}{RESET_COLOR}" if with_color else f"{self.name}")
                case "path":
                    output.append(f"{self.color}{self.path}{RESET_COLOR}" if with_color else f"{self.path}")
                case "type":
                    output.append(f"{self.color}{self.type}{RESET_COLOR}" if with_color else f"{self.type}")
                case "exists":
                    output.append(f"{self.color}{self.exists}{RESET_COLOR}" if with_color else f"{self.exists}")
                case "hidden":
                    output.append(f"{self.color}{self.hidden}{RESET_COLOR}" if with_color else f"{self.hidden}")
                case "root":
                    output.append(f"{self.color}{self.root}{RESET_COLOR}" if with_color else f"{self.root}")
                case "parent":
                    output.append(f"{self.color}{self.parent}{RESET_COLOR}" if with_color else f"{self.parent}")
                case "absolute":
                    output.append(f"{self.color}{self.absolute}{RESET_COLOR}" if with_color else f"{self.absolute}")
                case "depth":
                    output.append(f"{self.color}{self.depth}{RESET_COLOR}" if with_color else self.depth)
                case "size_b":
                    output.append(f"{self.color}{self.size_bytes} B{RESET_COLOR}" if with_color else self.size_bytes)
                case "size_kb":
                    output.append(f"{self.color}{self.size_kbytes} KB{RESET_COLOR}" if with_color else self.size_kbytes)
                case "size_mb":
                    output.append(f"{self.color}{self.size_mbytes} MB{RESET_COLOR}" if with_color else self.size_mbytes)
                case "ctime":
                    output.append(f"{self.color}{self.creation_stime}{RESET_COLOR}" if with_color else f"{self.creation_stime}")
                case "lmtime":
                    output.append(f"{self.color}{self.modified_stime}{RESET_COLOR}" if with_color else f"{self.modified_stime}")
                case "latime":
                    output.append(f"{self.color}{self.accessed_stime}{RESET_COLOR}" if with_color else f"{self.accessed_stime}")
                case "mode":
                    output.append(f"{self.color}{self.mode}{RESET_COLOR}" if with_color else self.mode)
                case "inode":
                    output.append(f"{self.color}{self.inode}{RESET_COLOR}" if with_color else self.inode)
                case "dev_id":
                    output.append(f"{self.color}{self.dev_id}{RESET_COLOR}" if with_color else self.dev_id)
                case "nlinks":
                    output.append(f"{self.color}{self.nlinks}{RESET_COLOR}" if with_color else self.nlinks)
                case _:
                    output.append(f"{self.color}{NOT_FOUND}{RESET_COLOR}" if with_color else f"{NOT_FOUND}")

        return output

#MAGIC
    def __repr__(self)->str:
        return self.__path.name

    def __str__(self)->str:
        return str(self.__path)
    
    def __bool__(self)->bool:
        return bool(self.__path)

    def __del__(self)->None:
        try:
            if sys.meta_path is None:
                return
        except ImportError:
            return

        FileParser._FileParser__PERMISSION = True
        FileParser._del(self.absolute, self)

    def __eq__(self, other : Union[FileParser, PathLike])->bool:
        if isinstance(other, (FileParser, PathLike)):
            return PathManager.relative(self, other)
        elif other is None:
            return bool(self) == False
        logger.critical("Cannot compare %s type with %s type.", FileParser.__name__, type(other).__name__)
        raise Exception(f"Cannot compare {FileParser.__name__} type with {type(other).__name__} type.")

#SPECIAL
    @property
    def workbook(self)->Workbook:
        return self.__data
    
    @property
    def sheets(self)->List[str]:
        return self.__data.sheetnames

#PROPERTIES
    @property
    def path(self)->Path:
        return self.__path
    @property
    def name(self)->str:
        return self.__path.name
    @property
    def type(self)->str:
        return f"{self.suffix} File"
    @property
    def root(self)->Optional[Directory]:
        return self.__parent.root if self.__parent else None
    @property
    def suffix(self)->str:
        return self.__path.suffix
    @property
    def absolute(self)->Path:
        return self.__path.absolute()
    @property
    def exists(self)->bool:
        return self.__path.exists() if self else False
    @property
    def hidden(self)->bool:
        return self.__hidden
    @property
    def color(self)->str:
        return self.__color if self else FileParser._PATHLESS_COLOR
    @property
    def parent(self)->Optional[Directory]:
        return self.__parent
    @parent.setter
    def parent(self, new : Optional[Directory])->None:
        """
        Internally managed.
        """
        if not isinstance(new, Directory) and new is not None:
            logger.critical("Expected Directory object, got %s instead.", type(new).__name__)
            raise Exception(f"Expected Directory object, got {type(new).__name__} instead.")

        if new is not None and self.__parent is not None:
            logger.critical("Cannot assign parent to '%s' as it already has a parent '%s'.", self.name, self.__parent)
            raise Exception(f"Cannot assign parent to '{self.name}' as it already has a parent '{self.__parent}'.")

        if new is None and self.__parent is not None:
            if self.__parent.get_file(self.name):
                logger.critical("Failed to remove file '%s' from parent '%s'.", self.name, self.__parent)
                raise Exception(f"Failed to remove file '{self.name}' from parent '{self.__parent}'.")
            self.__parent = None
            self.__path = NullPath(self.absolute)

        elif isinstance(new, Directory) and self.__parent is None:
            if not new.get_file(self.name):
                logger.critical("Failed to assign '%s' to parent '%s'.", self.name, new)
                raise Exception(f"Failed to assign '{self.name}' to parent '{new}'.")

            self.__parent = new

    @property
    def permitted(self)->bool:
        return FileParser._FileParser__PERMISSION
    @property
    def depth(self)->int:
        return 0 if not self else len(PathManager.set(self.parent.root, self.parent, False).parts)
    @property
    def size_bytes(self)->int|Literal[-1]:
        return self.__path.stat().st_size if self.exists else -1
    @property
    def size_kbytes(self)->int|Literal[-1]:
        return round(self.size_bytes /1024, 3) if self.exists else -1
    @property
    def size_mbytes(self)->int|Literal[-1]:
        return round(self.size_bytes /(1024 * 1024), 3) if self.exists else -1
    @property
    def mode(self)->Optional[int]:
        return self.__path.stat().st_mode if self.exists else None
    @property
    def inode(self)->Optional[int]:
        return self.__path.stat().st_ino if self.exists else None
    @property
    def dev_id(self)->Optional[int]:
        return self.__path.stat().st_dev if self.exists else None
    @property
    def nlinks(self)->Optional[int]:
        return self.__path.stat().st_nlink
    @property
    def creation_time(self)->Optional[int]:
        return self.__path.stat().st_ctime if self.exists else None
    @property
    def creation_stime(self)->Optional[str]:
        return time.ctime(self.creation_time) if self.exists else None
    @property
    def accessed_time(self)->Optional[int]:
        return self.__path.stat().st_atime if self.exists else None
    @property
    def accessed_stime(self)->Optional[str]:
        return time.ctime(self.accessed_time) if self.exists else None
    @property
    def modified_time(self)->Optional[int]:
        return self.__path.stat().st_mtime if self.exists else None
    @property
    def modified_stime(self)->Optional[str]:
        return time.ctime(self.modified_time) if self.exists else None

class ExcelParser(FileParser):
    _default_color = Fore.GREEN
    def __init__(self, path : PathLike, **kw):
        """
        Prototype parser for excel type files using pandas library to store+modify data.

        **Arguments**
            -path: Path or string object absolute/relative path or name pointing to a file (file does not have to exist)
            -create: creates file object initially if it does not exist.
            -color: color of the name used in visualization methods in Directory and PathManager class, uses colorama's package colors by default.
            -hide: instance becomes invisible to references map and can be duplicated, by default, only one object can point to a certain file.

        **Example Usage**
        >>> file = ExcelParser("data.xlsx") #or FileParser("data.xlsx")
        >>> data = file.read(overwrite=True, ...) #reads file data if it exists
        >>> file.head #same as pandas.DataFrame.head
        >>> file.data["column_1"] = (...) #data assignment
        
        """
        if hasattr(self, '_init'):
            return
        self._init : bool = True
        self.__path : Path = Path(path)
        self.__data : DataFrame = DataFrame(None)
        self.__parent : Optional[Directory] = None
        self.__color : str = kw.get("color", ExcelParser._default_color)

        if self.__path.suffix.lower() not in EXCEL_SUFFIX:
            logger.critical("Expected %s files, for %s file use %s instead.", EXCEL_SUFFIX, self.__path.suffix, FileParser._correct_parser(self.__path.suffix))
            raise Exception(f"Expected {EXCEL_SUFFIX} files, for {self.__path.suffix} file use {FileParser._correct_parser(self.__path.suffix)} instead.")
        if self.exists and self.__path.is_dir():
            logger.critical("Path must point to a file, not a directory.")
            raise Exception("Path must point to a file, not a directory.")

        self.__hidden : bool = kw.get("hide", False)
        if kw.get("create", False):
            self.__path.open("a").close()

    def read(self, overwrite : bool=True, *args, **kw)->DataFrame:
        if not self.exists:
            logger.error("Could not read from path '%s'", self)
            return self.__data
        if not self.__data or (self.__data and overwrite):
            self.__data = read_excel(str(self.__path), *args, **kw)
        return self.__data

    def write(self, only_create : bool = False, engine : Optional[str] = None, date_format : Optional[str] = None, datetime_format : Optional[str] = None, mode : str = "w", storage_options: Optional[dict[str, Any]] = None, 
              if_sheet_exists: Optional[Literal["error", "new", "replace", "overlay"]] = None, engine_kwargs: Optional[Dict] = None, *args, **kw)->None:
        if not self.exists:
            logger.error("Could not write to path '%s'", self)
            return None

        if only_create:
            self.__path.open("a").close()
            return None

        with ExcelWriter(str(self.__path), engine=engine, date_format=date_format, datetime_format=datetime_format, mode=mode, storage_options=storage_options, 
                            if_sheet_exists=if_sheet_exists, engine_kwargs=engine_kwargs) as writer:
            self.__data.to_excel(writer, *args, **kw)
        return

    def rename(self, new_name : Optional[str])->None:
        if not self:
            logger.error("Cannot modify the file '%s'", self)
            return None

        #string handling case       
        if isinstance(new_name, str): 
            if PathManager.os() == OperatingSystem.WINDOWS and PathManager.contains_illegal_chars_win(new_name):
                logger.error("name %s contains illegal characters.", new_name)
                return None            
            elif PathManager.os() in (OperatingSystem.LINUX, OperatingSystem.MACOS) and PathManager.contains_illegal_chars_psx(new_name):
                logger.error("name %s contains illegal characters.", new_name)
                return None            

            if len(Path(new_name).parts) != 1:
                logger.error("name must not include other directory parts.")

            #if file has parent
            if self.__parent and Directory._blocks_duplicates():
                new_name = self.__parent._numeric_name(new_name+self.suffix, True) #get correct name

            #no parent, add suffix to it
            else:
                new_name = new_name + self.suffix

        #requesting a duplicate rename
        elif new_name is None:
            if self.__parent is None:
                return
            new_name = self.__parent._numeric_name(self.name, True) #get correct name

        else: #new name is neither str or None
            logger.error("Expected a string or None type, got %s instead", type(new_name).__name__)
            return None

        #name did not change
        if new_name == self.name:
            logger.info("New name %s matches the old name, skipping the rest of the process", new_name)
            return None

        #new path conclusion
        path, old_name = PathManager.split(self.__path)
        new_path = Path(path) / new_name if path else new_name
        #check if new path is already occupied
        if FileParser.is_locked() and not self.__hidden:
            #check if new path already exists
            if FileParser._contains(new_path.absolute()):
                logger.error("File with the name '%s' already exists in this directory.", new_name)
                return None

        #updating name
            FileParser._FileParser__PERMISSION = True
            FileParser._update(self.absolute, new_path.absolute())
            if self.exists:
                self.__path.rename(new_path)

        self.__path = new_path
        logger.info("File '%s' updated to '%s'", old_name, new_name)
        return

    def remove_column(self, column_name : str)->None:
        """
        Drops an existing column by name.
        """
        if column_name in self.__data.columns:
            self.__data.drop(columns=[column_name], inplace=True)
        else:
            logger.error("Column '%s' not found in the DataFrame.", column_name)
        return

    def filter_rows(self, column_name : str, condition : Callable)->None:
        """
        Filters data based on a condition in a specific column.

        Arguments:
            column_name (str): The name of the column to filter on.
            condition (lambda): A lambda function that defines the filtering condition.
        """
        if column_name in self.__data.columns:
            self.__data = self.__data[self.__data[column_name].apply(condition)]
        else:
            logger.error("Column '%s' not found in the DataFrame.", column_name)
        return

    def _redirect(self, new_master : Directory)->None:
        if not isinstance(new_master, Directory):
            logger.error("expected Directory object, got %s instead.", type(new_master).__name__)
            return None
        if not new_master:
            logger.error("Cannot modify the directory: %s.", new_master)
            return None
        if not new_master.permitted:
            logger.error("Missing permission to redirect '%s'", self)
            return None

        old_path = self.absolute
        if new_master is not self.__parent:
            FileParser._FileParser__PERMISSION = True
            if self.__parent is not None:
                self.__parent._remove_file(self)
            self.__path = new_master.path / self.name
            new_master._add_file(self)
            FileParser._FileParser__PERMISSION = False
        else:
            self.__path = new_master.path / self.name
            if self.__path and old_path == self.__path:
                #case when dir already exists in parent + path does not change
                logger.error("File with the name '%s' already exists in directory '%s'.", self.name, new_master)
                return None

        if FileParser.is_locked() and not self.__hidden:
            FileParser._FileParser__PERMISSION = True
            FileParser._update(old_path, self.__path.absolute())
        
        logger.info("Added file '%s' to directory '%s'", self.name, self.__parent)
        return

    def _extract_info(self, with_color : bool = True, *tags : KEYS)->List[Any]:
        """
        Returns an iterable of properties based on input tags, possible tags:

        **name, path, type, exists, hidden, parent, root, absolute, depth, size_b, size_kb, size_mb, ctime, lmtime, latime, mode, inode, dev_id, nlinks**
        """
        output = []
        if not self:
            return output

        for field in tags:
            match field.lower():
                case "name":
                    output.append(f"{self.color}{self.name}{RESET_COLOR}" if with_color else f"{self.name}")
                case "path":
                    output.append(f"{self.color}{self.path}{RESET_COLOR}" if with_color else f"{self.path}")
                case "type":
                    output.append(f"{self.color}{self.type}{RESET_COLOR}" if with_color else f"{self.type}")
                case "exists":
                    output.append(f"{self.color}{self.exists}{RESET_COLOR}" if with_color else f"{self.exists}")
                case "hidden":
                    output.append(f"{self.color}{self.hidden}{RESET_COLOR}" if with_color else f"{self.hidden}")
                case "root":
                    output.append(f"{self.color}{self.root}{RESET_COLOR}" if with_color else f"{self.root}")
                case "parent":
                    output.append(f"{self.color}{self.parent}{RESET_COLOR}" if with_color else f"{self.parent}")
                case "absolute":
                    output.append(f"{self.color}{self.absolute}{RESET_COLOR}" if with_color else f"{self.absolute}")
                case "depth":
                    output.append(f"{self.color}{self.depth}{RESET_COLOR}" if with_color else self.depth)
                case "size_b":
                    output.append(f"{self.color}{self.size_bytes} B{RESET_COLOR}" if with_color else self.size_bytes)
                case "size_kb":
                    output.append(f"{self.color}{self.size_kbytes} KB{RESET_COLOR}" if with_color else self.size_kbytes)
                case "size_mb":
                    output.append(f"{self.color}{self.size_mbytes} MB{RESET_COLOR}" if with_color else self.size_mbytes)
                case "ctime":
                    output.append(f"{self.color}{self.creation_stime}{RESET_COLOR}" if with_color else f"{self.creation_stime}")
                case "lmtime":
                    output.append(f"{self.color}{self.modified_stime}{RESET_COLOR}" if with_color else f"{self.modified_stime}")
                case "latime":
                    output.append(f"{self.color}{self.accessed_stime}{RESET_COLOR}" if with_color else f"{self.accessed_stime}")
                case "mode":
                    output.append(f"{self.color}{self.mode}{RESET_COLOR}" if with_color else self.mode)
                case "inode":
                    output.append(f"{self.color}{self.inode}{RESET_COLOR}" if with_color else self.inode)
                case "dev_id":
                    output.append(f"{self.color}{self.dev_id}{RESET_COLOR}" if with_color else self.dev_id)
                case "nlinks":
                    output.append(f"{self.color}{self.nlinks}{RESET_COLOR}" if with_color else self.nlinks)
                case _:
                    output.append(f"{self.color}{NOT_FOUND}{RESET_COLOR}" if with_color else f"{NOT_FOUND}")

        return output

    def __repr__(self)->str:
        return self.__path.name

    def __str__(self)->str:
        return str(self.__path)
    
    def __len__(self)->int:
        return len(self.__data)
    
    def __bool__(self)->bool:
        return bool(self.__path)

    def __eq__(self, other : Union[FileParser, PathLike])->bool:
        if isinstance(other, (FileParser, PathLike)):
            return PathManager.relative(self, other)
        elif other is None:
            return bool(self) == False
        logger.critical("Cannot compare %s type with %s type.", FileParser.__name__, type(other).__name__)
        raise Exception(f"Cannot compare {FileParser.__name__} type with {type(other).__name__} type.")

    def __del__(self)->None:
        try:
            if sys.meta_path is None:
                return
        except ImportError:
            return
        FileParser._FileParser__PERMISSION = True
        FileParser._del(self.absolute, self)

    @property
    def head(self)->DataFrame:
        return self.__data.head()

    @property
    def tail(self)->DataFrame:
        return self.__data.tail()

    @property
    def columns(self)->Index[str]:
        return self.__data.columns

    @property
    def data(self)->DataFrame:
        return self.__data

    @data.setter
    def data(self, new_data : DataFrame)->None:
        if not isinstance(new_data, DataFrame):
            raise TypeError
        self.__data = new_data

    @property
    def path(self)->Path:
        return self.__path
    @property
    def name(self)->str:
        return self.__path.name
    @property
    def type(self)->str:
        return f"{self.suffix} File"
    @property
    def suffix(self)->str:
        return self.__path.suffix
    @property
    def root(self)->Optional[Directory]:
        return self.__parent.root if self.__parent else None
    @property
    def depth(self)->int:
        return 0 if not self else len(PathManager.set(self.parent.root, self.parent, False).parts)
    @property
    def absolute(self)->Path:
        return self.__path.absolute()
    @property
    def exists(self)->bool:
        return self.__path.exists() if self else False
    @property
    def hidden(self)->bool:
        return self.__hidden
    @property
    def color(self)->str:
        return self.__color if self else FileParser._PATHLESS_COLOR
    @property
    def parent(self)->Optional[Directory]:
        """
        Equivalent of Path.parent but returns the Directory object that holds this instance's reference (root dir is fatherless).
        """
        return self.__parent
    @parent.setter
    def parent(self, new : Optional[Directory])->None:
        """
        Internally managed.
        """
        if not isinstance(new, Directory) and new is not None:
            logger.critical("Expected Directory object, got %s instead.", type(new).__name__)
            raise Exception(f"Expected Directory object, got {type(new).__name__} instead.")

        if new is not None and self.__parent is not None:
            logger.critical("Cannot assign parent to '%s' as it already has a parent '%s'.", self.name, self.__parent)
            raise Exception(f"Cannot assign parent to '{self.name}' as it already has a parent '{self.__parent}'.")

        if new is None and self.__parent is not None:
            if self.__parent.get_file(self.name):
                logger.critical("Failed to remove file '%s' from parent '%s'.", self.name, self.__parent)
                raise Exception(f"Failed to remove file '{self.name}' from parent '{self.__parent}'.")
            self.__parent = None
            self.__path = NullPath(self.absolute)

        elif isinstance(new, Directory) and self.__parent is None:
            if not new.get_file(self.name):
                logger.critical("Failed to assign '%s' to parent '%s'.", self.name, new)
                raise Exception(f"Failed to assign '{self.name}' to parent '{new}'.")

            self.__parent = new

    @property
    def permitted(self)->bool:
        return FileParser._FileParser__PERMISSION
    @property
    def size_bytes(self)->int|Literal[-1]:
        return self.__path.stat().st_size if self.exists else -1
    @property
    def size_kbytes(self)->int|Literal[-1]:
        return round(self.size_bytes /1024, 3) if self.exists else -1
    @property
    def size_mbytes(self)->int|Literal[-1]:
        return round(self.size_bytes /(1024 * 1024), 3) if self.exists else -1
    @property
    def mode(self)->Optional[int]:
        return self.__path.stat().st_mode if self.exists else None
    @property
    def inode(self)->Optional[int]:
        return self.__path.stat().st_ino if self.exists else None
    @property
    def dev_id(self)->Optional[int]:
        return self.__path.stat().st_dev if self.exists else None
    @property
    def nlinks(self)->Optional[int]:
        return self.__path.stat().st_nlink
    @property
    def creation_time(self)->Optional[int]:
        return self.__path.stat().st_ctime if self.exists else None
    @property
    def creation_stime(self)->Optional[str]:
        return time.ctime(self.creation_time) if self.exists else None
    @property
    def accessed_time(self)->Optional[int]:
        return self.__path.stat().st_atime if self.exists else None
    @property
    def accessed_stime(self)->Optional[str]:
        return time.ctime(self.accessed_time) if self.exists else None
    @property
    def modified_time(self)->Optional[int]:
        return self.__path.stat().st_mtime if self.exists else None
    @property
    def modified_stime(self)->Optional[str]:
        return time.ctime(self.modified_time) if self.exists else None

#########################################
    #MANAGERS
#########################################
DirFile = Union[Directory, FileParser]

class OperatingSystem(Enum):
    """
    Enum class used in PathManager to differ between operating systems.
    """
    WINDOWS = "windows"
    LINUX = "linux"
    MACOS = "macos"
    OTHER = "other"

class PathManager:
    """
    Utility methods for Path and Directory objects.
    """
    MAX_LEVEL = 10
    PROHIBITED_PATHS_WIN = (Path("C:\\Windows"), Path("C:\\Program Files"), Path("C:\\Program Files (x86)"), Path("C:\\System Volume Information"))
    ILLEGAL_CHARS_WIN = ('<', '>', ':', '"', '|', '?', '*')
    ILLEGAL_CHARS_PSX = ('*', '?', '[', ']', '&', '|', '<', '>', '(', ')', '{', '}', ';', '!', '$', '#', '~', '^', '\\0')
    _ALIASES = {
        "win": OperatingSystem.WINDOWS,
        "windows": OperatingSystem.WINDOWS,
        "mswin": OperatingSystem.WINDOWS,
        "microsoft": OperatingSystem.WINDOWS,
        "nt" : OperatingSystem.WINDOWS,
        "linux": OperatingSystem.LINUX,
        "gnu/linux": OperatingSystem.LINUX,
        "posix" : {"linux" : OperatingSystem.LINUX, "darwin" : OperatingSystem.MACOS},
        "mac": OperatingSystem.MACOS,
        "macos": OperatingSystem.MACOS,
        "osx": OperatingSystem.MACOS,
        "darwin": OperatingSystem.MACOS,
        "apple": OperatingSystem.MACOS,
        "other": OperatingSystem.OTHER,
        "unknown": OperatingSystem.OTHER,
        "": OperatingSystem.OTHER,
        None: OperatingSystem.OTHER,
    }
    @staticmethod
    @functools.lru_cache(maxsize=1)
    def os()->OperatingSystem:
        """
        Returns the string indicator of the operating system used to run the program as ("windows", "linux", "macos", "other") 
        """
        operating_system = PathManager._ALIASES.get(os.name, OperatingSystem.OTHER)
        if isinstance(operating_system, dict):
            return operating_system.get(sys.platform, OperatingSystem.OTHER)
        return operating_system

    @staticmethod
    def contains_illegal_chars_win(name : str)->bool:
        """
        For windows, check if input has an illegal character.
        """
        if not isinstance(name, str):
            logger.critical("Expected string, got %s.", type(name).__name__)
            raise Exception(f"Expected string, got {type(name).__name__}.")
        return any(c in name for c in PathManager.ILLEGAL_CHARS_WIN)

    @staticmethod
    def contains_illegal_chars_psx(name : str)->bool:
        """
        For windows, check if input has an illegal character.
        """
        if not isinstance(name, str):
            logger.critical("Expected string, got %s.", type(name).__name__)
            raise Exception(f"Expected string, got {type(name).__name__}.")
        return any(c in name for c in PathManager.ILLEGAL_CHARS_PSX)

    @staticmethod
    def path_os(path : Union[DirFile, PathLike])->OperatingSystem:
        """
        Returns the operating system of path, returns the machine's OS by default.
        """
        if not path:
            raise ValueError

        if isinstance(path, DirFile):
            path = path.path

        if not isinstance(path, PathLike):
            raise TypeError
        
        s = str(path).strip()
        if not s:
            return OperatingSystem.OTHER
        
        if re.match(r"^[A-Za-z]:[\\/]", s) or re.match(r"^[A-Za-z]:$", s) or re.match(r"^[A-Za-z]:[^\\/]", s):
            return OperatingSystem.WINDOWS

        if s.startswith((r"\\", r"//", r"\\?\\", r"\\\\?\\", r"\\\\.\\", r"\\.\\")):
            return OperatingSystem.WINDOWS
        
        if s.startswith('/'):
            return OperatingSystem.LINUX
        
        if s.count("/") > s.count("\\"):
            return OperatingSystem.LINUX
        if s.count("\\") > s.count("/"):
            return OperatingSystem.WINDOWS

        return PathManager.os()

    @staticmethod
    def normalize_path(path : PathLike, return_abs : bool = False)->str:
        """
        Returns normalized version of path as absolute if return_abs is True, otherwise returns the path normalized.
        """
        if not isinstance(path, PathLike):
            raise TypeError
        return os.path.normcase(os.path.abspath(os.path.normpath(str(path)))) if return_abs else os.path.normcase(os.path.normpath(str(path)))

    @staticmethod
    def strip(name : Union[PathLike, DirFile])->Union[Tuple[str, int], Tuple[str, Literal[-1]]]:
        """
        Strips (n) pattern off name and returns the stripped version + n if n is a digit otherwise -1
        Returns name, -1 if no pattern was found.

        Example:
        >>> PathManager.strip("file(1).txt") #->file.txt, 1
        """
        if isinstance(name, DirFile):
            name = name.name
        else:
            name = Path(name).name

        end_pos = name.rfind(')')
        if end_pos == -1:
            return name, -1
        
        start_pos = name.rfind('(', 0, end_pos)
        if start_pos == -1:
            return name, -1

        content = name[start_pos + 1:end_pos]
        n = -1
        if content.isdigit() and int(content) > 0:
            n = int(content)
        
        stripped_name = name[:start_pos] + name[end_pos+1:]
        return stripped_name, n

    @staticmethod
    def split(path : Union[DirFile, PathLike])->Union[Tuple[Literal[''], str], Tuple[Path, str]]:
        """
        Splits path and returns relative path and name separately, returns None+name if path is one part long.
        """
        if isinstance(path, DirFile):
            if not path:
                return '', path.name

            path = path.path
        
        if not path or not isinstance(path, PathLike):
            raise TypeError
        path = Path(PathManager.normalize_path(path))
        if len(path.parts) == 1:
            return '', path.name
        
        return Path(*path.parts[:-1]), path.name

    @staticmethod
    def set(parent : Union[Directory, PathLike], child : Union[DirFile, PathLike], return_root : bool = False)->Path:
        """
        Joins parent path with child path together and returns the path presenting child path after parent directory.
        """
        if not parent or not child:
            logger.critical("one or more paths missing: '%s', '%s'.", parent, child)
            raise Exception(f"one or more paths missing: '{parent}', '{child}'.")

        if isinstance(parent, Directory):
            parent = parent.path
        if isinstance(child, DirFile):
            child = child.path
        
        if not isinstance(parent, PathLike) or not isinstance(child, PathLike):
            logger.critical("Expected 2 paths, got %s, %s instead", type(parent).__name__, type(child).__name__)
            raise Exception(f"Expected 2 paths, got {type(parent).__name__}, {type(child).__name__} instead.")

        parent = Path(PathManager.normalize_path(parent))
        child = Path(PathManager.normalize_path(child))
        partitionned_parent = parent.parts
        partitionned_child = child.parts
        Lp = len(partitionned_parent)
        Lc = len(partitionned_child)
        similar_idx_child : int = None
        similar_idx_parent : int = None
        parent_name : str = parent.name
        if not Lc or not Lp:
            return None
        #if child is just name
        if Lc == 1:
            return parent_name / child if return_root else child
        #if child is relative but does not include parent
        for similar_idx_child, part in enumerate(partitionned_child[:-1]):
            if part in partitionned_parent:
                similar_idx_parent = partitionned_parent.index(part)
                break
        
        #if there no similar points, child is relative that starts with parent dir
        if similar_idx_parent is None:
            if not child.is_absolute():
                return parent_name / child if return_root else child
            logger.critical("Child path cannot exist on top of parent directory.")
            raise Exception("Child path cannot exist on top of parent directory.")

        #similar point, start building from it, if similarity breaks return parent / child, if similarity continues negate all similar points from child path and return it
        else:
            i = 0
            while True:
                try:
                    child_part = partitionned_child[similar_idx_child + i + 1]
                except IndexError:
                    return parent_name / child if return_root else child
                try:
                    parent_part = partitionned_parent[similar_idx_parent + i + 1]
                except IndexError:
                    return parent_name / Path(*partitionned_child[similar_idx_child + i + 1:]) if return_root else Path(*partitionned_child[similar_idx_child + i + 1:])
                if child_part != parent_part:
                    if not child.is_absolute():
                        return parent_name / child if return_root else child
                    logger.critical("Child path cannot exist on top of parent directory.")
                    raise Exception("Child path cannot exist on top of parent directory.")
                i += 1

    @staticmethod
    def join_paths(parent : Union[Directory, PathLike], child : Union[DirFile, PathLike])->Path:
        """
        Joins parent path with child path together and returns the path presenting child path after parent directory.
        """
        if not parent or not child:
            logger.critical("one or more paths missing: '%s', '%s'.", parent, child)
            raise Exception(f"one or more paths missing: '{parent}', '{child}'.")

        if isinstance(parent, Directory):
            parent = parent.path
        if isinstance(child, DirFile):
            child = child.path
        
        if not isinstance(parent, PathLike) or not isinstance(child, PathLike):
            logger.critical("Expected 2 paths, got %s, %s instead", type(parent).__name__, type(child).__name__)
            raise Exception(f"Expected 2 paths, got {type(parent).__name__}, {type(child).__name__} instead.")

        parent_path = Path(PathManager.normalize_path(parent))
        child_path = Path(PathManager.normalize_path(child))
        if PathManager.relative(parent_path, child_path):
            return parent_path / child_path
        parent_parts : Tuple[str] = parent_path.parts
        child_parts : Tuple[str] = child_path.parts
        similar_idx_child : int = None
        similar_idx_parent : int = None

        for similar_idx_child, part in enumerate(child_parts[:-1]):
            if part in parent_parts:
                similar_idx_parent = parent_parts.index(part)
                break
        if similar_idx_parent is None:
            if not child_path.is_absolute():
                return parent_path / child_path
            logger.warning("Child path cannot belong to parent directory.")
            return child_path
        else:
            offset = 0
            while True:
                try:
                    child_part = child_parts[similar_idx_child + offset + 1]
                except IndexError: #similar and child has no extra path parts
                    return parent_path / child_path
                try:
                    parent_part = parent_parts[similar_idx_parent + offset + 1]
                except IndexError: #similar and child has extra parts
                    break
                if child_part != parent_part:
                    break
                offset += 1
        #not fully similar, starts from last similar part for child
        if similar_idx_parent == 0:
                return child_path
        return Path(*parent_parts[:similar_idx_parent + offset]) / Path(*child_parts[similar_idx_child + offset:])

    @staticmethod
    def relative(original_path : Union[DirFile, PathLike], new_path : Union[DirFile, PathLike])->bool:
        """
        Checks whether both paths point at the same object.
        """
        if isinstance(original_path, DirFile):
            original_path = original_path.path
        if isinstance(new_path, DirFile):
            new_path = new_path.path
        
        if not isinstance(original_path, PathLike) or not isinstance(new_path, PathLike):
            raise TypeError

        i_path = Path(PathManager.normalize_path(original_path))
        j_path = Path(PathManager.normalize_path(new_path))
        if i_path == j_path:
            return True
        i_parts : Tuple[str] = i_path.parts
        j_parts : Tuple[str] = j_path.parts
        
        len_i = len(i_parts)
        len_j = len(j_parts)
        for offset in range(1, min(len_i, len_j) + 1):
            if i_parts[-offset] != j_parts[-offset]:
                return False
    
        return True

    @staticmethod
    def is_subpath(parent : Union[Directory, PathLike], child : Union[DirFile, PathLike], True_if_same : bool = False)->bool:
        """
        Tests whether child path is connected with parent path, returns True if child is the same as father if True_if_same is set to True.
            True => you ARE the father!
            False => you are NOT the father!
        """
        if not parent or not child:
            logger.error("One or more paths missing: '%s', '%s'.", parent, child)
            return False

        if isinstance(parent, Directory):
            parent = parent.path
        if isinstance(child, DirFile):
            child = child.path

        parent = PathManager.normalize_path(parent, True)
        child  = PathManager.normalize_path(child, True)
        try:
            common_path = os.path.commonpath((parent, child))
        except ValueError:
            return False

        return PathManager.relative(parent, common_path) if True_if_same else common_path == parent and parent != child

    @staticmethod
    def indent_directory_legacy(directory : Directory, indent_level : int = 0, indent_space : int = 4, max_level : Optional[int] = None, list_files : bool = False)->str:
        """
        pprint but for Directory object.
        """
        indent = ' ' * indent_space * indent_level
        s = f"{indent}{directory.color}{directory.name}{RESET_COLOR}"
        if max_level is None:
            max_level = PathManager.MAX_LEVEL
        if indent_level >= max_level:
            s += f"\n{indent + ' ' * (indent_space-2)}" + "{...}"
        else:
            if list_files:
                s += f"\n{indent + ' ' * (indent_space-2)}Files: {', '.join([f"{file.color}{file.name}{RESET_COLOR}" for file in directory.files])}" if directory.files else ''
            else:
                for file in directory.files:
                    s += f"\n{indent + ' ' * (indent_space-2)}{file.color}{file.name}{RESET_COLOR}"
            for subdir in directory.directories:

                s += f"\n{PathManager.indent_directory_legacy(subdir, indent_level + 1, indent_space=indent_space, max_level=max_level, list_files=list_files)}"
        return s
     
    @staticmethod
    def indent_directory(directory: Directory, indent_level: int = 0, indent_space: int = 4, max_level: Optional[int] = None, list_files: bool = False) -> str:
        """
        pprint but for Directory object.
        """
        initial_indent = ' ' * indent_space * indent_level
        s = f"\n{initial_indent}{directory.color}{directory.name}{RESET_COLOR}"

        s += PathManager.__indent_subs(directory, indent=initial_indent, indent_space=indent_space, max_level=max_level or PathManager.MAX_LEVEL, list_files=list_files, current_level=0)
        return s

    @staticmethod
    def __indent_subs(directory: Directory, indent: str, indent_space: int, max_level: int, list_files: bool, current_level: int) -> str:
        """
        Recursively builds the string for the contents of a directory.
        """
        s = ''
        sub_items = directory.files if list_files else directory.files + directory.directories
        if list_files:
            sub_items = (["files"] if directory.files else []) + directory.directories

        subs_count = len(sub_items)

        if current_level >= max_level:
            if subs_count > 0:
                s += f"\n{indent}└── ..."
            return s

        for i, item in enumerate(sub_items):
            last_condition = (i == subs_count - 1)

            connector = ('└' + '─' * (indent_space - 1)) if last_condition else ('├' + '─' * (indent_space - 1))
            child_indent = (' ' * indent_space) if last_condition else ('│' + ' ' * (indent_space - 1))

            if isinstance(item, FileParser) and item == "files" and list_files:
                s += f"\n{indent}{connector}Files: {', '.join([f'{file.color}{file.name}{RESET_COLOR}' for file in directory.files])}"

            else:
                s += f"\n{indent}{connector}{item.color}{item.name}{RESET_COLOR}"
                if isinstance(item, Directory):
                    s += PathManager.__indent_subs(
                        item,
                        indent=indent + child_indent,
                        indent_space=indent_space,
                        max_level=max_level,
                        list_files=list_files,
                        current_level=current_level + 1
                    )
        return s

def cli_tree()->None:
    path = sys.argv[1]
    TOGGLE_CACHE()
    d = Directory(path)
    d.image()
    return PathManager.indent_directory(d)

def cli_table_keys()->None:
    keys = "name: name of the object\npath: relative path\nabsolute: absolute path\nexists: exists in files system\ntype: Directory or File\ncount: sub objects count\n"\
    "totalcount: tree objects count\nhidden: visible to membership map\nis_root: Directory is root object\nroot: object\'s root\nparent: object\'s parent\ndepth: layer depth from root object\n"\
    "ctime: creation time\nlmtime: last modified time\nlatime: last accessed time\nsize_b: size in bytes\nsize_kb: size in Kilo Bytes\nsize_mb: size in Mega Bytes\nmode: File type and mode bits\n"\
    "inode: Inode number\ndev_id: Device identifier\nnlinks: Number of hard links"
    return keys

def cli_table()->None:
    TOGGLE_CACHE()
    d = Directory(sys.argv[1])
    d.image()
    table = pt.PrettyTable(field_names=sys.argv[2:])
    return d.info(table)

if __name__ == "__main__":
    print(__doc__)
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.cell import get_column_letter
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.formula import ArrayFormula
from typing import Optional


class ExcelUtils:
    def __init__(self, path: str, data_only: bool = False) -> None:
        """
        :param path: path to Excel file
        :param data_only: if any formula exists in workbook,
            read operation will return calculated value of each cell.
            else, read operation will return the formula itself.
            (openpyxl's default value is False.)
        """
        self._path = path
        self._workbook = load_workbook(filename=path, data_only=data_only)
        self._worksheet: Worksheet = None

    def save(self):
        self._workbook.save(filename=self._path)

    def select_sheet(self, sheet_name: str):
        self._worksheet = self._workbook[sheet_name]

    def get_cell(self, address: str):
        cell: Cell = self._worksheet[address]
        return cell.value

    def get_cell_by_row_column(self, row: int, column: int):
        cell: Cell = self._worksheet.cell(row=row, column=column)
        return cell.value

    def get_range(self, address1: str, address2: str):
        cells: tuple[tuple[Cell]] = self._worksheet[address1:address2]
        result = [[cell.value for cell in row] for row in cells]
        return result

    def get_range_by_rows(self):
        result = [[cell.value for cell in row] for row in self._worksheet.rows]
        return result

    def set_cell(self, address: str, value, array_formula: bool = False, array_ref: Optional[str] = None):
        if array_formula:
            value = ArrayFormula(ref=address if array_ref is None else array_ref, text=value)
        self._worksheet[address] = value

    def set_cell_by_row_column(self, row: int, column: int, value):
        self._worksheet.cell(row=row, column=column, value=value)

    def append_row(self, values: list | dict):
        self._worksheet.append(values)

    def add_filter(self, row_size: int, column_size: int):
        """
        Add whole-range filter in active worksheet.
        Asserts data starts from Cell A1.
        :param row_size: number of rows (int, >0)
        :param column_size: number of columns (int, >0)
        """
        # 240610 Dongha Kim - to add filter in 'calculate' and 'raw data'
        def number_to_coordinate(row_idx, col_idx):
            assert all([
                isinstance(row_idx, int),
                isinstance(col_idx, int),
                row_idx > 0,
                col_idx > 0
            ])
            col_string = get_column_letter(col_idx)
            return f'{col_string}{row_idx}'
        start = number_to_coordinate(1, 1)
        end = number_to_coordinate(row_size, column_size)
        self._worksheet.auto_filter.ref = f'{start}:{end}'
        self._worksheet.auto_filter.add_filter_column(0, [])


if __name__ == "__main__":
    import pathlib
    import shutil

    absolute_dir = pathlib.Path(__file__).resolve()  # ABLE-Elba\robot\src\utils\excel_utils.py
    root = absolute_dir.parent.parent.parent.parent
    dir_path = root / "robot" / "resources" / "platform_autocal" / "template"
    source_file = f"{dir_path}/template_(1)_cal.xlsx"
    target_file = f"{dir_path}/result_(1)_cal.xlsx"
    shutil.copyfile(source_file, target_file)

    excel = ExcelUtils(target_file)
    excel.select_sheet("raw data")

    # cell = excel.get_cell("A2")
    # print(cell)
    # cell_range = excel.get_range("A1", "K1")
    # print(cell_range)
    # cell_range = excel.get_range_by_rows()
    # print(cell_range)

    volumes = [200, 100, 20]

    # 엑셀에 raw data 쓰기.
    import datetime
    import random

    index = 1
    reuse_tip = 1
    reuse_tip_count = 5
    tip_position_index = 1
    tip_position = f"A{tip_position_index}"
    for volume in volumes:
        for channel in range(1, 9):
            for repeat in range(1, 6):
                weight = volume * (1 + 1 * (0.5 - random.random()))
                pressure = 1000 + 5 * (0.5 - random.random())
                humidity = 50 + 1 * (0.5 - random.random())
                temperature = 20 + 0.2 * (0.5 - random.random())
                values = [
                    index,
                    datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    channel,
                    repeat,
                    reuse_tip,
                    tip_position,
                    volume,
                    weight,
                    pressure,
                    humidity,
                    temperature,
                ]
                excel.append_row(values)
                # excel.save()
                print(values)

                index += 1
                reuse_tip += 1
                if reuse_tip > reuse_tip_count:
                    reuse_tip = 1
                    tip_position_index += 1
                    tip_position = f"A{tip_position_index}"
    # A ~ K 열, 1 ~ 121 행 Filter 추가
    excel.add_filter(row_size=121, column_size=11)
    excel.save()

    # 엑셀 읽어서 calculate 계산.
    import pandas as pd
    from pandas import DataFrame
    import numpy as np

    raw_data = pd.read_excel(target_file, sheet_name="raw data")
    len_raw_data = len(raw_data)
    print(raw_data)

    volume_channel_weights = {
        volume: raw_data.loc[
            raw_data["volume (μL)"] == volume, ["channel", "weight (mg)"]
        ]
        for volume in volumes
    }
    volume_calculates: dict[float, DataFrame] = {}
    for volume, channel_weights in volume_channel_weights.items():
        print(f"\nvolume={volume}")
        print(channel_weights)
        grouped_weights = channel_weights.groupby("channel")
        calculate = grouped_weights.agg(
            [
                "mean",
                # ddof=0: 모집단 표준편차(STDEV.P) / ddof=1(default): 표본 표준편차(STDEV.S)
                lambda x: np.std(x, ddof=0),
                lambda x: np.std(x) / np.mean(x),
                lambda x: (volume - np.mean(x)) / volume,
            ]
        )
        calculate.columns = ["avg", "std", "cv", "acc"]
        max_cv_idx = calculate["cv"].idxmax()
        max_acc_idx = calculate["acc"].abs().idxmax()
        calculate["max cv"] = ""
        calculate["max acc"] = ""
        calculate.loc[max_cv_idx, "max cv"] = calculate.loc[max_cv_idx, "cv"]
        calculate.loc[max_acc_idx, "max acc"] = calculate.loc[max_acc_idx, "acc"]
        calculate = calculate.reset_index()
        calculate.insert(0, "volume", volume)
        print(calculate)
        volume_calculates[volume] = calculate

    # 엑셀에 calculate 쓰기.
    # df.to_excel()은 새로 쓰기 때문에 기존 데이터가 제거됨.
    # for calculate in volume_calculates.values():
    #     calculate.to_excel(
    #         target_file,
    #         sheet_name="calculate",
    #         startcol=ord("J") - ord("A"),
    #         startrow=1,
    #     )

    excel.select_sheet("calculate")
    r = 2
    for calculate in volume_calculates.values():
        for i, row in calculate.iterrows():
            (volume, channel, avg, std, cv, acc, max_cv, max_acc) = row
            excel.select_sheet("calculate")
            excel.set_cell(f"A{r}", volume)
            excel.set_cell(f"B{r}", channel)
            excel.set_cell(f"C{r}", avg)
            excel.set_cell(f"D{r}", std)
            excel.set_cell(f"E{r}", cv)
            excel.set_cell(f"F{r}", acc)
            excel.set_cell(f"G{r}", max_cv)
            excel.set_cell(f"H{r}", max_acc)
            excel.select_sheet("calculate_f")
            excel.set_cell(f"A{r}", volume)
            excel.set_cell(f"B{r}", channel)
            r += 1
    r -= 1
    excel.select_sheet("calculate_f")
    for rr in range(2, r + 1):
        excel.set_cell(
            f"C{rr}",
            f"=AVERAGEIFS("
            f"'raw data'!$H$2:$H${len_raw_data + 1}, "
            f"'raw data'!$C$2:$C${len_raw_data + 1}, $B{rr}, 'raw data'!$G$2:$G${len_raw_data + 1}, $A{rr}"
            f")",
            array_formula=True
        )
        excel.set_cell(
            f"D{rr}",
            f"=_xlfn.STDEV.P(IF(("
            f"'raw data'!$C$2:$C${len_raw_data + 1}=$B{rr})*('raw data'!$G$2:$G${len_raw_data + 1}=$A{rr}), "
            f"'raw data'!$H$2:$H${len_raw_data + 1}"
            f"))",
            array_formula=True
        )
        excel.set_cell(f"E{rr}", f"=$D{rr}/$C{rr}")
        excel.set_cell(f"F{rr}", f"=($A{rr}-$C{rr})/$A{rr}")
        excel.set_cell(
            f"G{rr}",
            f"=IF(ABS($E{rr})=MAX(IF($A$2:$A${r}=$A{rr}, ABS($E$2:$E${r}))), ABS($E{rr}), \"\")",
            array_formula=True
        )
        excel.set_cell(
            f"H{rr}",
            f"=IF(ABS($F{rr})=MAX(IF($A$2:$A${r}=$A{rr}, ABS($F$2:$F${r}))), ABS($F{rr}), \"\")",
            array_formula=True
        )
    for sheet_name in ("calculate_f", "calculate"):
        excel.select_sheet(sheet_name)
        excel.add_filter(r, 8)
    excel.save()

    # calibartion 계산.
    from collections import defaultdict

    cal_x = []
    cal_y = []
    for volume, calculate in volume_calculates.items():
        avg_weight_of_volume = calculate["avg"].mean()
        cal_x.append(avg_weight_of_volume)  # 측정값
        cal_y.append(volume)  # 목표값
    _cal_x = [int(x * 10) / 10 for x in cal_x]
    print(f"cal_x={_cal_x}")
    print(f"cal_y={cal_y}")

    from sklearn.linear_model import LinearRegression

    reg = LinearRegression()
    reg_X = [[x] for x in cal_x]
    reg_y = cal_y
    reg.fit(X=reg_X, y=reg_y)
    r2 = reg.score(X=reg_X, y=reg_y)
    a = reg.coef_[0]
    b = reg.intercept_
    print(f"a={a:.4f}  b={b:.4f}  r2={r2:.4f}")

    # 엑셀에 calibartion 쓰기.
    for i in range(len(cal_x)):
        rr = 7 + i
        excel.select_sheet("calculate")
        excel.set_cell(f"J{rr}", cal_x[i])
        excel.set_cell(f"K{rr}", cal_y[i])
        excel.select_sheet("calculate_f")
        excel.set_cell(f"K{rr}", cal_y[i])
        excel.set_cell(
            f"J{rr}",
            f"=AVERAGEIFS($C$2:$C${r}, $A$2:$A${r}, $K{rr})",
            array_formula=True
        )
    excel.select_sheet("calculate")
    excel.set_cell("K2", a)
    excel.set_cell("K3", b)
    excel.set_cell("K4", r2)
    fmt = {"end_row": 6 + len(reg_X)}
    excel.set_cell(
        "L2",
        r"=INDEX(LINEST($K$7:$K$%(end_row)d,$J$7:$J$%(end_row)d^{1},1,0),1)" % fmt,
        array_formula=True,
    )
    excel.set_cell(
        "L3",
        r"=INDEX(LINEST($K$7:$K$%(end_row)d,$J$7:$J$%(end_row)d^{1},1,0),2)" % fmt,
        array_formula=True,
    )
    excel.set_cell(
        "L4",
        r"=INDEX(LINEST($K$7:$K$%(end_row)d,$J$7:$J$%(end_row)d^{1},1,1),3,1)" % fmt,
        array_formula=True,
    )
    excel.select_sheet("calculate_f")
    excel.set_cell(
        "K2",
        r"=INDEX(LINEST($K$7:$K$%(end_row)d,$J$7:$J$%(end_row)d^{1},1,0),1)" % fmt,
    )
    excel.set_cell(
        "K3",
        r"=INDEX(LINEST($K$7:$K$%(end_row)d,$J$7:$J$%(end_row)d^{1},1,0),2)" % fmt,
        array_formula=True
    )
    excel.set_cell(
        "K4",
        r"=INDEX(LINEST($K$7:$K$%(end_row)d,$J$7:$J$%(end_row)d^{1},1,1),3,1)" % fmt,
        array_formula=True
    )
    excel.set_cell("L2", None)
    excel.set_cell("L3", None)
    excel.set_cell("L4", None)
    excel.save()
